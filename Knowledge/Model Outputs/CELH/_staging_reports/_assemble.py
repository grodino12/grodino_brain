import json, glob, re, os
import openpyxl
from openpyxl.styles import Font, PatternFill

STAGING = r'C:\Users\rodin\Desktop\Brain\Knowledge\Model Outputs\CELH\_staging_reports'
XLSX = r'C:\Users\rodin\Desktop\Brain\Knowledge\Model Outputs\CELH\CELH_disclosures.xlsx'

frags = []
for f in glob.glob(os.path.join(STAGING, '*.json')):
    frags.append(json.load(open(f, encoding='utf-8')))
frags.sort(key=lambda d: (d['date'], d['index']))

wb = openpyxl.load_workbook(XLSX)
orig_sheets = list(wb.sheetnames)

ILLEGAL = re.compile(r'[\[\]:\*\?/\\]')
def clean_label(sn):
    s = re.sub(r'^\d{4}-\d{2}-\d{2}\s+', '', sn)
    s = re.sub(r'^\d{1,3}\s+', '', s)
    return s.strip()

def cell(v):
    if v is None or isinstance(v, (str, int, float)):
        return v
    return str(v)

SECTION = {'QUANTITATIVE', 'QUALITATIVE TAKEAWAYS', 'Q&A SUMMARY'}
hdr_fill = PatternFill('solid', fgColor='D9D9D9')
used = set()
new_titles = []

for r, d in enumerate(frags, 1):
    label = clean_label(d['sheet_name'])
    name = ILLEGAL.sub('', f'{r:02d} {label}')[:31].strip()
    base, k = name, 2
    while name.lower() in used:
        suf = f' ({k})'
        name = base[:31 - len(suf)] + suf
        k += 1
    used.add(name.lower())
    new_titles.append(name)
    ws = wb.create_sheet(title=name)
    for row in d['rows']:
        ws.append([cell(x) for x in row] if row else [])
    for cr in ws.iter_rows():
        v = cr[0].value
        if v in SECTION:
            for c in cr:
                c.font = Font(bold=True, size=11)
                c.fill = hdr_fill
        elif v in ('Event', 'Date'):
            cr[0].font = Font(bold=True)
    ws.column_dimensions['A'].width = 46
    for col in 'BCDEFGH':
        ws.column_dimensions[col].width = 15

wb._sheets = [wb[t] for t in new_titles] + [wb[t] for t in orig_sheets]
wb.save(XLSX)
print('Saved', XLSX)
print('total sheets:', len(wb.sheetnames))
print('first 3:', wb.sheetnames[:3])
print('last 3:', wb.sheetnames[-3:])
