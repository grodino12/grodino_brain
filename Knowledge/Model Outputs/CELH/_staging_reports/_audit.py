"""Audit the KPI Consolidated sheet against the .md STEP-5 KPI blocks.

For every DIRECTLY-REPORTED cell (blue, non-italic — not a prior-year backfill
or a reconstructed formula), this finds the source transcript from the cell's
Shift+F2 note, re-parses that transcript's .md STEP-5 KPIs, and checks the cell
value against them. Writes _audit_report.md and prints a summary.

What it CAN catch: digest mis-copies, wrong-column / wrong-period placement,
value-vs-prior-year swaps, metric mismatches — i.e. errors in the
.md -> digest -> sheet layers.

What it CANNOT catch: errors in the .md itself vs the source PDF (Layer 1).
Also, STEP 5 does not cover channel / distribution / market-share metrics, so
those cells come back 'unverifiable — metric not in STEP 5'.
"""
import os
import re
import sys
import glob
import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import _consolidate as C          # reuse the .md parser + paths
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPORT = os.path.join(HERE, "_audit_report.md")
TOL = 0.025                       # relative tolerance for a value match


def vmatch(a, b):
    if a is None or b is None or not isinstance(a, (int, float)) \
            or not isinstance(b, (int, float)):
        return False
    if a == 0 and b == 0:
        return True
    return abs(a - b) <= TOL * max(abs(a), abs(b)) + 1e-6


def main():
    wb = openpyxl.load_workbook(C.WB_PATH)
    ws = wb[C.SHEET]
    hdr = [str(c.value) if c.value is not None else '' for c in ws[4]]

    md_by_date = {}
    for p in glob.glob(os.path.join(C.SOURCES, "**", "transcripts", "CELH_*.md"),
                       recursive=True):
        mm = re.match(r'CELH_(\d{4}-\d{2}-\d{2})_', os.path.basename(p))
        if mm:
            md_by_date.setdefault(mm.group(1), []).append(p)

    counts = {'ok': 0, 'ok_loose': 0, 'mismatch': 0, 'misplaced': 0,
              'no_metric': 0, 'no_period': 0, 'no_source': 0}
    discrepancies = []
    unverifiable = []

    for row in ws.iter_rows(min_row=5):
        metric = row[0].value
        if not metric or str(metric).startswith('—'):
            continue
        metric = str(metric)
        for c in row[1:]:
            if c.value is None:
                continue
            # audit only directly-reported cells (blue, non-italic, not a formula)
            if isinstance(c.value, str) and c.value.startswith('='):
                continue
            if c.font is not None and c.font.italic:
                continue
            period = hdr[c.column - 1]

            # source transcript date from the Shift+F2 note
            ct = c.comment.text if c.comment else ''
            dm = re.search(r'Date:\s*(\d{4}-\d{2}-\d{2})', ct) \
                or re.search(r'►.*?\((\d{4}-\d{2}-\d{2})\)', ct)
            mds = md_by_date.get(dm.group(1), []) if dm else []
            if not mds:
                counts['no_source'] += 1
                unverifiable.append((metric, period, c.value, 'no source .md'))
                continue

            # gather STEP-5 KPIs for this period across the source .md(s)
            per_kpis, all_lines = [], []
            for md in mds:
                kpis, _, _, _ = C.parse_md_kpis(md)
                for k in kpis:
                    k = dict(k, _md=md)
                    all_lines.append(k)
                    if k['period'] == period:
                        per_kpis.append(k)

            if not per_kpis:
                counts['no_period'] += 1
                unverifiable.append((metric, period, c.value,
                                     'period not in source STEP 5'))
                continue

            mk = [k for k in per_kpis if k['metric'].lower() == metric.lower()]
            if mk:
                if any(vmatch(c.value, k['value']) for k in mk):
                    counts['ok'] += 1
                elif any(vmatch(c.value, k['prior']) for k in mk):
                    counts['misplaced'] += 1
                    discrepancies.append(
                        ('MISPLACED — cell holds the prior-year figure', metric,
                         period, c.value, mk[0]['value'], mk[0]['_md'], mk[0]['line']))
                else:
                    counts['mismatch'] += 1
                    discrepancies.append(
                        ('VALUE MISMATCH', metric, period, c.value,
                         mk[0]['value'], mk[0]['_md'], mk[0]['line']))
            else:
                vk = [k for k in per_kpis if vmatch(c.value, k['value'])]
                pk = [k for k in per_kpis if vmatch(c.value, k['prior'])]
                if vk:
                    counts['ok_loose'] += 1
                elif pk:
                    counts['misplaced'] += 1
                    discrepancies.append(
                        ('MISPLACED — value equals a prior-year figure', metric,
                         period, c.value, pk[0]['value'], pk[0]['_md'], pk[0]['line']))
                else:
                    counts['no_metric'] += 1
                    unverifiable.append((metric, period, c.value,
                                         'metric not in STEP 5 for this period'))

    total = sum(counts.values())
    checked_ok = counts['ok'] + counts['ok_loose']
    flagged = counts['mismatch'] + counts['misplaced']

    lines = []
    lines.append("# KPI Consolidated — Audit vs `.md` STEP-5 KPIs")
    lines.append("")
    lines.append(f"Generated: {datetime.date.today().isoformat()}  ·  "
                 f"{total} directly-reported cells checked")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append("| Result | Count |")
    lines.append("|---|---|")
    lines.append(f"| Confirmed — value matches a STEP-5 KPI | {counts['ok']} |")
    lines.append(f"| Confirmed — value matches, metric-name differs | {counts['ok_loose']} |")
    lines.append(f"| **VALUE MISMATCH** — disagrees with STEP-5 | {counts['mismatch']} |")
    lines.append(f"| **MISPLACED** — cell holds a prior-year figure | {counts['misplaced']} |")
    lines.append(f"| Unverifiable — metric not in STEP 5 for that period | {counts['no_metric']} |")
    lines.append(f"| Unverifiable — period not in the source STEP 5 | {counts['no_period']} |")
    lines.append(f"| No source `.md` found | {counts['no_source']} |")
    lines.append("")
    lines.append(f"**{checked_ok} confirmed · {flagged} flagged · "
                 f"{total - checked_ok - flagged} unverifiable.**")
    lines.append("")

    if discrepancies:
        lines.append(f"## Flagged — review these ({len(discrepancies)})")
        lines.append("")
        for kind, met, per, sv, mv, md, ln in sorted(discrepancies):
            rel = os.path.relpath(md, C.BRAIN_ROOT).replace(os.sep, '/')
            lines.append(f"- **{kind}** — `{met}` / {per}: sheet = `{sv}`, "
                         f".md STEP-5 = `{mv}`  ·  {rel}:{ln}")
        lines.append("")

    if unverifiable:
        lines.append(f"## Unverifiable ({len(unverifiable)})")
        lines.append("")
        lines.append("Mostly channel / distribution / market-share metrics that the "
                     "transcript analysis never formalized as STEP-5 KPIs.")
        lines.append("")
        for met, per, sv, why in sorted(unverifiable):
            lines.append(f"- `{met}` / {per} (`{sv}`) — {why}")
        lines.append("")

    lines.append("## Limits")
    lines.append("")
    lines.append("- Cannot detect errors in the `.md` itself vs the source PDF.")
    lines.append("- STEP 5 omits channel / distribution / market-share metrics.")
    lines.append("- Prior-year backfilled cells and reconstructed formulas are NOT "
                 "audited here (they are derived, not directly reported).")

    with open(REPORT, 'w', encoding='utf-8') as fh:
        fh.write("\n".join(lines) + "\n")

    print(f"Audited {total} directly-reported cells")
    print(f"  confirmed: {checked_ok}  (exact {counts['ok']}, "
          f"value-only {counts['ok_loose']})")
    print(f"  FLAGGED:   {flagged}  (mismatch {counts['mismatch']}, "
          f"misplaced {counts['misplaced']})")
    print(f"  unverifiable: {counts['no_metric'] + counts['no_period']}  "
          f"(no metric in STEP5 {counts['no_metric']}, no period {counts['no_period']})")
    print(f"  no source: {counts['no_source']}")
    print(f"  report -> {REPORT}")


if __name__ == "__main__":
    main()
