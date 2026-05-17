"""Build a 'KPI Consolidated' sheet in CELH_disclosures.xlsx.

Reads the 69 staging digest JSONs, normalizes metric + period, and writes a
metric-time-series matrix: rows = normalized KPI, columns = period (chrono).

ALL quantitative datapoints are kept — standard three-statement lines as well
as transcript-only granularity, and every transcript that reports a given
(metric, period) is recorded. The cell shows the earnings-call figure (or the
earliest source if none); the Shift+F2 note lists every source with its value.

Period can live in the row label ('Total Revenue Q4 2018') or in the column
header ('Q1 2022' / 'Q1 2021'); both styles are handled.
"""
import json, glob, re, os
from collections import Counter
from urllib.parse import quote
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
WB_PATH = os.path.join(os.path.dirname(HERE), "CELH_disclosures.xlsx")
BRAIN_ROOT = os.path.abspath(os.path.join(HERE, "..", "..", "..", ".."))
SOURCES = os.path.join(BRAIN_ROOT, "Sources", "CELH")
VAULT = "Brain"                       # Obsidian vault name (folder basename)
SHEET = "KPI Consolidated"

PER_RE = re.compile(
    r'\b(Q[1-4])\s*[\'’]?\s*(20\d\d)\b'            # Q1 2022
    r'|\bFY\s*(20\d\d)\b'                          # FY2018
    r'|\b(9M)\s*(20\d\d)\b'                        # 9M 2018
    r'|\b(H[12])\s*(20\d\d)\b'                     # H1 2019
    r'|\bDec(?:ember)?\s*31?,?\s*(20\d\d)\b',      # Dec 31 2018 -> FYxxxx
    re.I)

ORD = {'Q1': 1, 'Q2': 2, 'H1': 2.5, 'Q3': 3, '9M': 3.5, 'H2': 3.5, 'Q4': 4, 'FY': 5}

# trailing unit suffix on a label, e.g. '($M)', '($B)', '($000)', '(%)', '(bps)'
UNIT_RE = re.compile(
    r'\s*\(\s*(?:in\s*)?'
    r'(?:\$\s*[MBK]?|\$?\s*000s?|[MBK]|%|bps|pts?|pp|x|'
    r'millions?|billions?|thousands?)'
    r'\s*\)\s*$', re.I)

SCALE_RE = re.compile(r'\(\s*(?:in\s*)?\$?\s*(B|K|000|millions?|billions?|thousands?)\s*\)\s*$', re.I)

# bare period qualifiers with no year ('9M Revenue', 'FY G&A', 'Revenue Q4')
LEAD_PER = re.compile(r'^\s*(?:Q[1-4]|FY|9M|H[12]|1H|2H|YTD)\b[\s.:/-]*', re.I)
TRAIL_PER = re.compile(r'[\s.:/-]*\b(?:Q[1-4]|FY|9M|H[12]|1H|2H|YTD)\s*$', re.I)
# parenthetical period qualifier, e.g. '(FY)', '(Q4)', '(Full Year)', '(Nine Months)'
_PAREN_PERIOD = re.compile(
    r'\s*\((?:Q[1-4]|FY|H[12]|9M|Full[ -]?Year|Nine Months(?: YTD)?|'
    r'Six Months(?: YTD)?|Three Months(?: YTD)?|YTD|First Half|Second Half|'
    r'Quarter|Annual)\)\s*', re.I)

# light canonicalization of obvious synonyms
ALIAS = {
    'domestic / north american revenue': 'North America Revenue',
    'domestic / north america revenue': 'North America Revenue',
    'north american revenue': 'North America Revenue',
    'north america revenue': 'North America Revenue',
    'us / north america revenue': 'North America Revenue',
    'international revenue (total)': 'International Revenue',
    'total international revenue': 'International Revenue',
    'europe / nordics': 'Europe Revenue',
    'europe / nordics revenue': 'Europe Revenue',
    'europe revenue': 'Europe Revenue',
    'asia (incl. china royalties)': 'Asia Revenue',
    'asia revenue': 'Asia Revenue',
    'total revenue': 'Total Revenue',
    'revenue': 'Total Revenue',
    'gross profit': 'Gross Profit',
    'gross margin': 'Gross Margin',
    'net income': 'Net Income',
    'net income (loss)': 'Net Income',
    'adjusted ebitda': 'Adjusted EBITDA',
    'adj. ebitda': 'Adjusted EBITDA',
    'selling & marketing': 'Selling & Marketing',
    'sales & marketing': 'Selling & Marketing',
    'g&a': 'General & Administrative',
    'general & administrative': 'General & Administrative',
}

# Canonical-name map: collapses near-duplicate labels onto one row.
# Keys are lowercased post-norm_metric metric names. Genuinely distinct metrics
# (basic vs diluted EPS, GAAP vs non-GAAP, $ vs %, actual vs long-term target)
# are deliberately NOT merged.
CANON = {}


def _canon(name, *variants):
    for v in variants:
        CANON[v.lower()] = name


_canon('Total Revenue', 'total revenue', 'net revenue', 'net revenues', 'net sales',
       'consolidated revenue', 'reported revenue (books & records)')
_canon('North America Revenue', 'north america', 'north america revenue',
       'north america / domestic', 'north america / domestic revenue',
       'north american / domestic revenue', 'north america (h1)')
_canon('International Revenue', 'international', 'international revenue',
       'international sales', 'international (h1)')
_canon('Europe Revenue', 'europe revenue', 'european revenue', 'europe (nordic)',
       'europe / nordic', 'nordic revenue')
_canon('Asia Revenue', 'asia revenue', 'asia', 'asian revenue', 'asia (china royalty)',
       'asia (china royalties)', 'asia (incl. china royalties)')
_canon('Other International Markets', 'other international', 'other international markets')
_canon('Gross Profit', 'gross profit')
_canon('Gross Margin', 'gross margin', 'gross profit margin', 'consolidated gross margin',
       'gaap gross margin', 'standalone gross margin')
_canon('Selling & Marketing', 'selling & marketing', 'sales & marketing expense',
       'sales & marketing expenses', 's&m expense')
_canon('General & Administrative', 'general & administrative', 'g&a expense', 'g&a expenses')
_canon('SG&A', 'sg&a', 'sg&a expense')
_canon('Operating Income', 'operating income')
_canon('Net Income', 'net income', 'net income (gaap)', 'gaap net income',
       'net income (common shareholders)', 'net income (to common)', 'net income / (loss)',
       'net income to common', 'net income to common shareholders', 'net loss to common')
_canon('Adjusted EBITDA', 'adjusted ebitda', 'adj ebitda', 'adjusted ebitda (non-gaap)',
       'adjusted ebitda ($m, non-gaap)', 'non-gaap adjusted ebitda',
       'net non-gaap adjusted ebitda')
_canon('Adjusted EBITDA Margin', 'adjusted ebitda margin', 'adjusted ebitda margin (non-gaap)',
       'ebitda margin')
_canon('Amazon Revenue', 'amazon revenue', 'amazon sales', 'amazon sales ytd 2022')
_canon('Cash', 'cash', 'cash & cash equivalents', 'cash & equivalents',
       'cash and cash equivalents', 'cash balance', 'cash position', 'cash (year-end)',
       'cash (mar 31, 2023)', 'cash on hand (mar 31, 2024)')
_canon('Operating Cash Flow', 'operating cash flow', 'cash from operations',
       'cash flow from operating activities', 'cash provided by operations',
       'cash used in operations', 'operating cash flow (6m)', 'operating cash flow (9m)',
       'operating cash flow fy')
_canon('Working Capital', 'working capital', 'net working capital',
       'working capital (year-end)', 'net working capital (mar 31, 2023)')
_canon('Diluted EPS', 'diluted eps', 'diluted eps (gaap)', 'gaap diluted eps')
_canon('Adjusted Diluted EPS', 'adjusted diluted eps', 'adjusted diluted eps (non-gaap)',
       'non-gaap adjusted diluted eps')
_canon('S&M % of Revenue', 's&m % of revenue', 's&m % of sales',
       'sales & marketing (% of revenue)', 'sales & marketing (% of sales)',
       'sales & marketing as % of revenue', 's&m expense % of sales')
_canon('G&A % of Revenue', 'g&a % of revenue', 'g&a % of sales', 'g&a (% of revenue)',
       'g&a (% of sales)')
_canon('Adjusted SG&A % of Revenue', 'adjusted sg&a % of revenue',
       'adjusted sg&a as % of revenue')
# convenience-channel: keep sales growth, store count and ACV as 3 DISTINCT metrics;
# only merge exact same-metric label variants within each
_canon('Convenience Channel ACV', 'convenience acv', 'convenience channel acv')
_canon('Convenience Channel Store Locations', 'convenience channel store locations',
       'convenience store locations')

# key metrics surfaced (bold) at the top of the sheet, in this order
CORE = [
    'Total Revenue', 'North America Revenue', 'International Revenue',
    'Europe Revenue', 'Asia Revenue', 'Amazon Revenue',
    'Gross Profit', 'Gross Margin', 'Selling & Marketing',
    'General & Administrative', 'Operating Income', 'Net Income',
    'Adjusted EBITDA', 'Operating Cash Flow', 'Cash', 'Working Capital',
]

HEADER_TOKENS = {'prior yr', 'prior year', 'current', 'change', 'yoy', 'yoy %',
                 'value', 'period', 'qoq', 'qoq %', 'metric'}


def label_scale(label):
    """Multiplier to bring a value into $M, read from a unit suffix on the label."""
    m = SCALE_RE.search(str(label))
    if not m:
        return 1.0
    t = m.group(1).lower()
    if t in ('b', 'billion', 'billions'):
        return 1000.0
    if t in ('k', '000', 'thousand', 'thousands'):
        return 0.001
    return 1.0


def parse_period(text):
    """Return (canonical_label, sort_key) or None."""
    if text is None:
        return None
    m = PER_RE.search(str(text))
    if not m:
        return None
    g = m.groups()
    if g[0]:                       # Qn YYYY
        q, y = g[0].upper(), int(g[1])
        return (f"{q} {y}", y * 10 + ORD[q])
    if g[2]:                       # FY YYYY
        y = int(g[2]); return (f"FY{y}", y * 10 + ORD['FY'])
    if g[3]:                       # 9M YYYY
        y = int(g[4]); return (f"9M {y}", y * 10 + ORD['9M'])
    if g[5]:                       # Hn YYYY
        h, y = g[5].upper(), int(g[6])
        return (f"{h} {y}", y * 10 + ORD.get(h, 2.5))
    if g[7]:                       # Dec 31 YYYY -> fiscal year end
        y = int(g[7]); return (f"FY{y}", y * 10 + ORD['FY'])
    return None


def is_period_cell(x):
    """True only if the cell is essentially a period LABEL (a column header) —
    not a data value that merely contains a period annotation, e.g. the cell
    '18.6% (Q2 2023)' is a value, not a header."""
    s = str(x).strip()
    p = PER_RE.search(s)
    if not p:
        return False
    remainder = (s[:p.start()] + s[p.end():]).strip(" .,;:()-–—'\"E")
    return len(remainder) <= 1


def norm_metric(label):
    s = str(label)
    s = PER_RE.sub('', s)                       # drop period tokens with a year
    s = _PAREN_PERIOD.sub(' ', s)               # drop '(FY)' / '(Q4)' / '(Full Year)'
    s = LEAD_PER.sub('', s)                     # drop bare leading period ('9M ', 'FY ')
    s = TRAIL_PER.sub('', s)                    # drop bare trailing period ('... Q4')
    s = UNIT_RE.sub('', s)
    s = re.sub(r'\bex[- ]?(China|Asia|one-time|outbound freight)\b', '', s, flags=re.I)
    s = re.sub(r'\s{2,}', ' ', s).strip(' -/,')
    key = s.lower().strip()
    return ALIAS.get(key, s)


# a value is quantitative only if it is a bare number or a number + a unit token
# (%, pp, bps, x); phrases ('Going forward', 'Energy Category') and ranges are out
QUANT_RE = re.compile(
    r'[~≈<>]?\s*[+-]?\$?[\d,]+\.?\d*\s*(?:%|pp|ppt|pts?|bps|bp|x)?\+?\s*$', re.I)


def finalize(v):
    """Return (number, is_percent) for a quantitative value, else None.

    Percentages are converted to a fraction (40.4% -> 0.404) so every cell
    holds a real number; text phrases and ranges return None and are dropped.
    """
    if isinstance(v, (int, float)):
        return (float(v), False)
    if not isinstance(v, str):
        return None
    s = v.strip()
    if not s or not QUANT_RE.fullmatch(s):
        return None
    m = re.search(r'[+-]?[\d,]+\.?\d*', s)
    if not m:
        return None
    try:
        num = float(m.group().replace(',', ''))
    except ValueError:
        return None
    if '%' in s:
        return (num / 100.0, True)
    return (num, False)


def clean_value(v):
    """Coerce '$15.1M' / '~$400M' / '1,200' to a float; leave %/text as-is."""
    if isinstance(v, (int, float)):
        return v
    if not isinstance(v, str):
        return v
    s = v.strip().lstrip('~≈').strip().replace('$', '').strip()
    m = re.fullmatch(r'([+-]?[\d,]+\.?\d*)\s*([MmBbKk]?)', s)
    if m:
        try:
            num = float(m.group(1).replace(',', '').lstrip('+'))
            suf = m.group(2).upper()
            if suf == 'B':
                num *= 1000          # billions -> $M
            elif suf == 'K':
                num /= 1000          # thousands -> $M
            return num
        except ValueError:
            pass
    return v.strip()


def canon_metric(raw):
    """Canonicalize a STEP-5 KPI metric name so it lines up with sheet rows."""
    m = norm_metric(raw)
    return CANON.get(m.lower(), m)


def md_value(s):
    """Parse a STEP-5 '**Value**' field to a $M number (fraction for %).

    Handles leading-dot billions notation ('.0282B' = $28.2M) and '$28.2M' style.
    """
    s = str(s).strip()
    m = re.search(r'[+-]?\$?\s*[+-]?[\d,]*\.?\d+', s)
    if not m:
        return None
    try:
        num = float(m.group().replace('$', '').replace(' ', '')
                    .replace(',', '').lstrip('+'))
    except ValueError:
        return None
    tail = s[m.end():m.end() + 1].upper()
    if tail == 'B':
        num *= 1000
    elif tail == 'K':
        num /= 1000
    if '%' in s:
        num /= 100.0
    return num


_MD_CACHE = {}


def parse_md_kpis(path):
    """Parse a transcript .md -> (kpis, step5_line, event, date).
    Each kpi: {line, metric, metric_raw, period, value, prior, pct}."""
    if path in _MD_CACHE:
        return _MD_CACHE[path]
    kpis, step5_line, ev, dt = [], None, '', ''
    try:
        lines = open(path, encoding='utf-8').read().split('\n')
    except OSError:
        _MD_CACHE[path] = (kpis, step5_line, ev, dt)
        return _MD_CACHE[path]
    in5 = False
    for i, ln in enumerate(lines, 1):
        if not ev and ln.startswith('event_title:'):
            ev = ln.split(':', 1)[1].strip()
        if not dt and ln.startswith('date:'):
            dt = ln.split(':', 1)[1].strip()
        if re.match(r'##\s+STEP\s*5\b', ln):
            in5, step5_line = True, i
            continue
        if in5 and re.match(r'##\s+STEP\s*6\b', ln):
            break
        if in5 and ln.startswith('##### '):
            mm = re.match(r'#####\s+\*\*(.+?)\*\*', ln)
            if not mm:
                continue
            vm = re.search(r'\*\*Value\*\*:\s*([^|]+)', ln)
            pvm = re.search(r'PriorYearValue:\s*([^|]+)', ln)
            ym = re.search(r'YoYChangePct:\s*([^|]+)', ln)
            yoy = None
            if ym:
                yn = re.search(r'[+-]?\d[\d.]*', ym.group(1))
                if yn:
                    try:
                        yoy = float(yn.group())
                    except ValueError:
                        pass
            fy = re.search(r'FiscalYear:\s*(\d{4})', ln)
            fq = re.search(r'FiscalQuarter:\s*([A-Za-z0-9]+)', ln)
            period = None
            if fy and fq:
                y, q = fy.group(1), fq.group(1).upper()
                if q in ('Q1', 'Q2', 'Q3', 'Q4', 'H1', 'H2'):
                    period = f"{q} {y}"
                elif q == 'FY':
                    period = f"FY{y}"
                elif q == '9M':
                    period = f"9M {y}"
            kpis.append({'line': i, 'metric': canon_metric(mm.group(1)),
                         'metric_raw': mm.group(1).strip(), 'period': period,
                         'value': md_value(vm.group(1)) if vm else None,
                         'prior': md_value(pvm.group(1)) if pvm else None,
                         'yoy': yoy,
                         'pct': bool(vm) and '%' in vm.group(1)})
    _MD_CACHE[path] = (kpis, step5_line, ev, dt)
    return _MD_CACHE[path]


def prior_consistent(value, prior, yoy):
    """Sanity-check a STEP-5 PriorYearValue against the line's own YoYChangePct.
    Catches .md typos (e.g. '4.02B' where '.402B' was meant). True = trust it."""
    if value is None or prior is None or yoy is None:
        return True                       # can't check -> allow
    denom = 1.0 + yoy / 100.0
    if abs(denom) < 0.05:
        return True
    expected = value / denom
    if abs(expected) < 1e-6:
        return True
    return abs(prior - expected) / abs(expected) <= 0.4


def prior_year_period(plabel):
    """Prior-year period label: 'Q3 2023' -> 'Q3 2022', 'FY2018' -> 'FY2017'."""
    s = str(plabel).strip()
    m = re.match(r'(Q[1-4]|9M|H[12])\s*(20\d\d)$', s)
    if m:
        return f"{m.group(1)} {int(m.group(2)) - 1}"
    m = re.match(r'FY\s*(20\d\d)$', s)
    if m:
        return f"FY{int(m.group(1)) - 1}"
    return None


def kpi_line(path, metric, period, value):
    """Best STEP-5 line for an (metric, period, value) datapoint -> (line, kind).

    A specific KPI line is only used when the datapoint can be confirmed against
    it — by matching value, or by an exact metric-name match. A loose name
    resemblance is NOT enough (it would mislink e.g. a sales-growth % onto a
    store-count KPI); those fall back to the STEP-5 section header instead.
    """
    kpis, step5, _, _ = parse_md_kpis(path)
    cands = [k for k in kpis if k['period'] == period]
    section = (step5, 'section' if step5 else 'none')
    if not cands:
        return section

    def vclose(k):
        if k['value'] is None or value is None:
            return False
        return abs(k['value'] - value) <= 0.02 * max(abs(k['value']), abs(value), 1e-9) + 1e-6

    ml = str(metric).lower()
    vmatch = [k for k in cands if vclose(k)]
    if vmatch:                                   # value confirmed -> trust it
        if len(vmatch) == 1:
            return (vmatch[0]['line'], 'kpi')
        exact = [k for k in vmatch if k['metric'].lower() == ml]
        if exact:
            return (exact[0]['line'], 'kpi')
        mt = set(ml.split())
        best = max(vmatch, key=lambda k: len(mt & set(k['metric'].lower().split())))
        return (best['line'], 'kpi')
    exact = [k for k in cands if k['metric'].lower() == ml]   # no value match
    if len(exact) == 1:                          # exact name is still trustworthy
        return (exact[0]['line'], 'kpi')
    return section                               # uncertain -> KPI section, not a guess


def obsidian_uri(path, line):
    rel = os.path.relpath(path, BRAIN_ROOT).replace(os.sep, '/')
    uri = f"obsidian://adv-uri?vault={quote(VAULT)}&filepath={quote(rel)}"
    if line:
        uri += f"&line={line}"
    return uri


_FNAME_PERIOD = re.compile(r'(20\d\d)\s*(Q[1-4]|FY)', re.I)
_EVENT_PERIOD = re.compile(r'\b(Q[1-4])\b.{0,14}?\b(20\d\d)\b')


def transcript_period(fname, event):
    """The transcript's own (year, fiscal_quarter), for period fallback. Or None."""
    m = _FNAME_PERIOD.search(os.path.basename(fname))
    if m:
        return (int(m.group(1)), m.group(2).upper())
    m = _EVENT_PERIOD.search(event or '')
    if m:
        return (int(m.group(2)), m.group(1).upper())
    return None


def fallback_period(label, tp):
    """Period for a row with none in its label/header: a bare leading qualifier
    (9M/FY/Qn) on the label + the transcript's year, else the transcript's period."""
    year, tq = tp
    lm = LEAD_PER.match(str(label))
    if lm:
        tk = re.match(r'(Q[1-4]|FY|9M|H[12]|1H|2H|YTD)', lm.group(0).strip(), re.I)
        tok = tk.group(1).upper() if tk else ''
        if tok in ('Q1', 'Q2', 'Q3', 'Q4'):
            return f"{tok} {year}"
        if tok == 'FY':
            return f"FY{year}"
        if tok == '9M':
            return f"9M {year}"
        if tok in ('H1', '1H'):
            return f"H1 {year}"
        if tok in ('H2', '2H'):
            return f"H2 {year}"
        if tok == 'YTD':
            return {'Q2': f"H1 {year}", 'Q3': f"9M {year}",
                    'Q4': f"FY{year}", 'FY': f"FY{year}"}.get(tq, f"{tq} {year}")
    return f"FY{year}" if tq == 'FY' else f"{tq} {year}"


def main():
    files = sorted(glob.glob(os.path.join(HERE, "*.json")))
    # matrix[metric][period] = list of source records
    matrix = {}
    periods = {}          # label -> sort_key
    skipped_no_period = 0
    fallback_used = 0
    dropped_text = 0
    placed = 0

    for f in files:
        d = json.load(open(f, encoding='utf-8'))
        rows = d.get('rows', [])
        # use the tab's own Event/Date cells (rows[0]/rows[1]) as the source key
        event = str(d.get('event', ''))
        date = str(d.get('date', ''))
        if rows and len(rows[0]) > 1 and str(rows[0][0]).strip().lower() == 'event':
            event = str(rows[0][1]).strip()
        if len(rows) > 1 and len(rows[1]) > 1 and str(rows[1][0]).strip().lower() == 'date':
            date = str(rows[1][1]).strip()
        is_earn = 'earning' in event.lower()
        tp = transcript_period(f, event)        # for no-period row fallback
        in_quant = False
        cur_col = None        # column index of the 'Current' value
        hdr_period = None     # period from a period-style header

        for r in rows:
            if not r:
                continue
            c0 = str(r[0]).strip()
            if c0 == 'QUANTITATIVE':
                in_quant = True
                continue
            if c0.startswith('QUALITATIVE') or c0.startswith('Q&A'):
                in_quant = False
                continue
            if not in_quant or len(r) < 2:
                continue

            vals = r[1:]
            low = [str(x).strip().lower() for x in vals]

            # --- header row detection ---
            is_header = any(t in HEADER_TOKENS for t in low)
            # only treat a cell as a period header if it IS a period label,
            # not a value that merely contains one (e.g. '18.6% (Q2 2023)')
            per_in_hdr = [parse_period(x) if is_period_cell(x) else None for x in vals]
            has_hdr_period = any(per_in_hdr)

            if is_header or has_hdr_period:
                cur_col = None
                hdr_period = None
                meaningful = [t for t in low if t]
                if meaningful and set(meaningful) <= {'value', 'period'}:
                    continue
                if 'current' in low:                       # Prior/Current style
                    cur_col = low.index('current') + 1
                elif has_hdr_period:                        # period-header style
                    best = None
                    for i, p in enumerate(per_in_hdr):
                        if p and (best is None or p[1] > best[1]):
                            best = p; cur_col = i + 1
                    hdr_period = best[0] if best else None
                continue

            # --- data row ---
            if cur_col is None or cur_col >= len(r):
                continue
            raw = r[cur_col]
            if raw in (None, '', 'n/a', 'N/A', 'n/d'):
                continue
            value = clean_value(raw)
            # scale to $M from a label unit suffix, unless the raw value already
            # carried its own M/B/K suffix (clean_value handled that case)
            if isinstance(value, (int, float)) and not (
                    isinstance(raw, str) and re.search(r'[MBK]\s*$', raw.strip(), re.I)):
                value *= label_scale(c0)
            fin = finalize(value)
            if fin is None:                  # text phrase / range -> not quantitative
                dropped_text += 1
                continue
            value, is_pct = fin

            per = parse_period(c0)
            plabel = per[0] if per else hdr_period
            if not plabel and tp:               # no period anywhere -> transcript's own
                plabel = fallback_period(c0, tp)
                fallback_used += 1
            if not plabel:
                skipped_no_period += 1
                continue
            pp = parse_period(plabel)
            pkey = pp[1] if pp else 0

            metric = norm_metric(c0)
            if not metric:
                continue

            periods[plabel] = pkey
            rec = {'value': value, 'pct': is_pct, 'earn': is_earn, 'event': event,
                   'date': date, 'label': c0}
            matrix.setdefault(metric, {}).setdefault(plabel, []).append(rec)
            placed += 1

    # --- consolidate near-duplicate metric rows ---
    groups = {}
    for m in matrix:
        groups.setdefault(m.lower(), []).append(m)
    casefold = {}
    for low, variants in groups.items():
        if len(variants) == 1:
            casefold[low] = variants[0]
        else:                       # pick the casing with the most datapoints
            score = Counter({v: sum(len(r) for r in matrix[v].values()) for v in variants})
            casefold[low] = score.most_common(1)[0][0]

    def canon(m):
        return CANON.get(m.lower(), casefold.get(m.lower(), m))

    n_before = len(matrix)
    merged = {}
    for m, pm in matrix.items():
        dst = merged.setdefault(canon(m), {})
        for p, recs in pm.items():
            dst.setdefault(p, []).extend(recs)
    matrix = merged
    n_after = len(matrix)

    # --- backfill prior-year values from .md STEP 5 into otherwise-empty cells ---
    # consistent priors -> the .md figure directly; priors that conflict with the
    # line's own YoY% -> a live formula reconstructing it from the current cell.
    backfilled = 0
    computed_cnt = 0
    bad_prior = 0
    matrix_lc = {k.lower(): k for k in matrix}
    for mdp in glob.glob(os.path.join(SOURCES, "**", "transcripts", "CELH_*.md"),
                         recursive=True):
        kpis, _, ev, dt = parse_md_kpis(mdp)
        for k in kpis:
            if k['prior'] is None or not k['period']:
                continue
            pp = prior_year_period(k['period'])
            cmk = matrix_lc.get(k['metric'].lower())
            if not pp or cmk is None or pp in matrix[cmk]:
                continue                       # only fill EXISTING rows / EMPTY cells
            base = {'pct': k['pct'], 'event': ev, 'date': dt,
                    'label': k['metric_raw'], 'md': mdp, 'line': k['line']}
            if k['pct'] or prior_consistent(k['value'], k['prior'], k['yoy']):
                matrix[cmk][pp] = [dict(base, value=k['prior'], earn=False,
                                        derived=True)]
                backfilled += 1
            elif (k['yoy'] is not None and k['period'] in matrix[cmk]
                  and any(not r.get('derived') and not r.get('computed')
                          for r in matrix[cmk][k['period']])):
                matrix[cmk][pp] = [dict(base, computed=True,
                                        current_period=k['period'], yoy=k['yoy'])]
                computed_cnt += 1
            else:
                bad_prior += 1                 # no reliable basis — leave blank
                continue
            pk = parse_period(pp)
            periods.setdefault(pp, pk[1] if pk else 0)

    # --- order axes ---
    ordered_periods = sorted(periods, key=lambda p: periods[p])
    core_present = [m for m in CORE if m in matrix]
    rest = sorted((m for m in matrix if m not in CORE), key=str.lower)

    # --- write sheet ---
    wb = openpyxl.load_workbook(WB_PATH)

    # map (event, date) -> real workbook tab name, read from each transcript tab
    tab_map = {}
    for sn in wb.sheetnames:
        if sn in (SHEET, ' Transcript Reports', 'Disclosures'):
            continue
        sh = wb[sn]
        if str(sh.cell(1, 1).value).strip().lower() != 'event':
            continue
        ev = str(sh.cell(1, 2).value).strip()
        dt = str(sh.cell(2, 2).value).strip()[:10]
        tab_map[(ev, dt)] = sn

    # map transcript date -> source .md analysis file (for cell hyperlinks)
    md_by_date = {}
    for p in glob.glob(os.path.join(SOURCES, "**", "transcripts", "CELH_*.md"),
                       recursive=True):
        mm = re.match(r'CELH_(\d{4}-\d{2}-\d{2})_(.+)\.md$', os.path.basename(p))
        if mm:
            md_by_date.setdefault(mm.group(1), []).append((mm.group(2), p))

    def find_md(event, date):
        cands = md_by_date.get(str(date).strip()[:10], [])
        if not cands:
            return None
        if len(cands) == 1:
            return cands[0][1]
        ev = str(event).lower()                       # disambiguate same-date events
        return max(cands, key=lambda c: sum(w in ev for w in c[0].lower().split()))[1]

    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, 0)   # first tab

    hdr_fill = PatternFill('solid', fgColor='1F4E78')
    hdr_font = Font(bold=True, color='FFFFFF')
    core_font = Font(bold=True)
    prior_font = Font(italic=True, color='808080')   # prior-year backfilled cells
    computed_font = Font(italic=True, color='2F6FA8')  # reconstructed-formula cells

    ws.cell(row=1, column=1, value="CELH Transcript KPI Consolidation — reported value per period")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.cell(row=2, column=1,
            value="Every quantitative datapoint disclosed across the 69 transcripts — including items also in "
                  "SEC filings, and metrics reported by more than one transcript. Cell shows the earnings-call "
                  "figure (else earliest source); Shift+F2 note lists every source with its value. "
                  "Values normalized to $M; period read from row label or column header.")
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color='808080')

    hrow = 4
    ws.cell(row=hrow, column=1, value="Metric")
    for j, p in enumerate(ordered_periods):
        ws.cell(row=hrow, column=2 + j, value=p)
    for c in ws[hrow]:
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center')

    unlocated = set()
    no_md = []
    stats = {'kpi': 0, 'section': 0, 'none': 0, 'prior': 0, 'computed': 0}

    def tab_of(rec):
        return tab_map.get((str(rec['event']).strip(), str(rec['date']).strip()[:10]))

    def vfmt(r):
        return f"{r['value']:.1%}" if r['pct'] else f"{r['value']:g}"

    def place(row, j, metric, period, recs):
        recs_s = sorted(recs, key=lambda r: str(r['date']))
        if recs_s[0].get('derived'):              # prior-year backfilled cell
            r = recs_s[0]
            c = ws.cell(row=row, column=2 + j, value=r['value'])
            if r['pct']:
                c.number_format = '0.0%'
            c.font = prior_font
            c.hyperlink = obsidian_uri(r['md'], r['line'])
            stats['prior'] += 1
            cm = Comment(
                "Prior-year comparative — no transcript reported this period "
                f"directly.\nFigure as stated in {r['event']} ({r['date']}),\n"
                f"STEP-5 KPI “{r['label']}” (PriorYearValue field).",
                "transcript consolidation")
            cm.width, cm.height = 330, 100
            c.comment = cm
            return
        if recs_s[0].get('computed'):             # reconstructed as a live formula
            r = recs_s[0]
            ref = f"{get_column_letter(2 + ordered_periods.index(r['current_period']))}{row}"
            c = ws.cell(row=row, column=2 + j,
                        value=f"={ref}/(1+{r['yoy'] / 100:g})")
            c.font = computed_font
            c.hyperlink = obsidian_uri(r['md'], r['line'])
            stats['computed'] += 1
            cm = Comment(
                f"Computed prior-year value: ={ref}/(1+{r['yoy']:g}%).\n"
                "The .md PriorYearValue conflicted with the line's stated YoY%, so "
                f"it is reconstructed from the {r['current_period']} figure.\n"
                f"YoY source: {r['event']} ({r['date']}) — “{r['label']}”.",
                "transcript consolidation")
            cm.width, cm.height = 340, 116
            c.comment = cm
            return
        prim = next((r for r in recs_s if r['earn']), recs_s[0])
        c = ws.cell(row=row, column=2 + j, value=prim['value'])
        if prim['pct']:
            c.number_format = '0.0%'
        md = find_md(prim['event'], prim['date'])    # deep-link to the source .md
        if md:
            line, kind = kpi_line(md, metric, period, prim['value'])
            c.hyperlink = obsidian_uri(md, line)
            stats[kind] = stats.get(kind, 0) + 1
        else:
            no_md.append((prim['event'], prim['date']))
        for r in recs_s:
            if tab_of(r) is None:
                unlocated.add((r['event'], r['date']))
        if len(recs_s) == 1:
            r = recs_s[0]
            txt = (f"Source: {r['event']}\n"
                   f"Date: {r['date']}\n"
                   f"Workbook tab: {tab_of(r) or '(see Event/Date)'}\n"
                   f"Digest label: “{r['label']}”")
        else:
            lines = [f"Reported in {len(recs_s)} transcripts "
                     f"(► = figure shown in cell):"]
            for r in recs_s:
                mark = "►" if r is prim else "•"
                lines.append(f"{mark} {vfmt(r)}  —  {r['event']} ({r['date']})")
                lines.append(f"   tab: {tab_of(r) or '(see Event/Date)'}  |  “{r['label']}”")
            txt = "\n".join(lines)
        cm = Comment(txt, "transcript consolidation")
        cm.width = 360
        cm.height = min(430, 46 + 15 * (len(recs_s) * 2 + 1))
        c.comment = cm

    div_fill = PatternFill('solid', fgColor='D9D9D9')
    rr = hrow
    for m in core_present:
        rr += 1
        ws.cell(row=rr, column=1, value=m).font = core_font
        for j, p in enumerate(ordered_periods):
            recs = matrix[m].get(p)
            if recs:
                place(rr, j, m, p, recs)
    rr += 1
    dc = ws.cell(row=rr, column=1, value="— Other metrics —")
    dc.font = Font(bold=True, italic=True, size=9)
    for j in range(len(ordered_periods) + 1):
        ws.cell(row=rr, column=1 + j).fill = div_fill
    for m in rest:
        rr += 1
        ws.cell(row=rr, column=1, value=m)
        for j, p in enumerate(ordered_periods):
            recs = matrix[m].get(p)
            if recs:
                place(rr, j, m, p, recs)

    ws.freeze_panes = "B5"
    ws.column_dimensions['A'].width = 38
    for j in range(len(ordered_periods)):
        ws.column_dimensions[ws.cell(row=hrow, column=2 + j).column_letter].width = 11

    # newest transcript tabs to the left (descending NN prefix); KPI sheet stays first
    transcript_tabs, other_tabs = [], []
    for sn in wb.sheetnames:
        if sn == SHEET:
            continue
        sh = wb[sn]
        if str(sh.cell(1, 1).value).strip().lower() == 'event':
            m = re.match(r'(\d+)', sn)
            transcript_tabs.append((int(m.group(1)) if m else 0, sn))
        else:
            other_tabs.append(sn)
    transcript_tabs.sort(key=lambda x: x[0], reverse=True)
    wb._sheets = ([wb[SHEET]] + [wb[sn] for _, sn in transcript_tabs]
                  + [wb[sn] for sn in other_tabs])

    wb.save(WB_PATH)
    n_cells = sum(len(pm) for pm in matrix.values())
    n_multi = sum(1 for pm in matrix.values() for recs in pm.values() if len(recs) > 1)
    print(f"Wrote '{SHEET}': {len(matrix)} metrics x {len(ordered_periods)} periods")
    print(f"  metric rows consolidated: {n_before} -> {n_after}")
    print(f"  numeric datapoints kept: {placed}; non-numeric dropped: {dropped_text}")
    print(f"  matrix cells: {n_cells}; cells with multiple sources: {n_multi}")
    print(f"  prior-year cells: {backfilled} backfilled from .md, "
          f"{computed_cnt} reconstructed as formulas (.md-inconsistent), "
          f"{bad_prior} left blank (no basis)")
    print(f"  cells deep-linked to .md: {n_cells - len(no_md)} "
          f"(exact KPI line: {stats['kpi']}, STEP-5 section: {stats['section']}, "
          f"prior-year: {stats['prior']}, computed: {stats['computed']}, "
          f"file top: {stats['none']}); no .md found: {len(no_md)}")
    print(f"  source notes attached: {n_cells}; datapoints with no tab match: {len(unlocated)}")
    print(f"  rows placed via transcript-period fallback: {fallback_used}")
    print(f"  rows skipped (no period, no fallback): {skipped_no_period}")
    print(f"  transcript tabs reordered newest-first; total sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
