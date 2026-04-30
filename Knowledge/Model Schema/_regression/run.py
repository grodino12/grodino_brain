"""
Regression harness for the financials pipeline.

Locks current CELH (12 filings) + PG (14 filings) outputs as goldens. Any
future framework change that produces different outputs surfaces as an
explicit diff. Replaces the manual "rerun both tickers and eyeball" process
with a mechanical detector.

Workflow:
    python run.py --bootstrap          # one-time: snapshot current Model Output to goldens
    python run.py                      # rerun pipeline from source, diff against goldens
    python run.py --ticker CELH        # only one ticker
    python run.py --accept             # rerun + overwrite goldens with fresh output
    python run.py --keep-temp          # keep the temp pipeline output for inspection

Exit codes:
    0   no diffs (or bootstrap/accept completed)
    1   diffs found
    2   pipeline failure (extract/reconcile/validate/model-write returned nonzero)
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

# Force utf-8 stdout so diffs containing Greek delta / em-dashes / etc. don't
# crash on Windows cp1252.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ============================================================================
# Paths (all absolute, anchored to the user's repo layout)
# ============================================================================

USER_HOME = Path.home()
BRAIN = USER_HOME / "Desktop" / "Brain"
MODEL_SCHEMA = BRAIN / "Knowledge" / "Model Schema"
TICKER_LIBRARIES = MODEL_SCHEMA / "Ticker Libraries"
MODEL_OUTPUTS = BRAIN / "Knowledge" / "Model Outputs"
SCHEMA_PKG = MODEL_SCHEMA / "financials-schema"
LIBRARY = MODEL_SCHEMA / "pattern_libraries" / "generic_line_item_mappings.json"

# Project-scoped skills as of 2026-04-30. Skills used by this pipeline live
# under MODEL_SCHEMA/.claude/skills/ (Claude Code project-level scope) so they
# don't comingle with general-purpose user-level skills at ~/.claude/skills/.
SKILLS_ROOT = MODEL_SCHEMA / ".claude" / "skills"
EXTRACT_CLI = SKILLS_ROOT / "financials-extract" / "scripts" / "extract.py"
RECONCILE_CLI = SKILLS_ROOT / "financials-reconcile" / "scripts" / "reconcile.py"
VALIDATE_CLI = SKILLS_ROOT / "financials-validate" / "scripts" / "validate.py"
MODELWRITE_CLI = SKILLS_ROOT / "model-write" / "scripts" / "write.py"

GOLDENS_ROOT = MODEL_SCHEMA / "_regression" / "goldens"
EXTRACT_CACHE_DIR = MODEL_SCHEMA / "_regression" / "_extract_cache"

# Latest workbook filename per ticker (the canonical artifact to snapshot).
WORKBOOK_FILENAME = {
    "CELH": "CELH_model.xlsx",
    "PG":   "PG_model.xlsx",
    "PEP":  "PEP_model.xlsx",
    "MNST": "MNST_model.xlsx",
    "GOOG": "GOOG_model.xlsx",
}

# Tolerance for numeric diffs ($1, since statements are in thousands).
NUMERIC_TOLERANCE = Decimal("1")


# ============================================================================
# Manifest: discover filings per ticker from existing raw_*.json files
# ============================================================================

def ticker_root(ticker: str) -> Path:
    return TICKER_LIBRARIES / ticker


def model_output_dir(ticker: str) -> Path:
    """Where the produced workbook lives. Only the .xlsx sits here."""
    return MODEL_OUTPUTS / ticker


def cache_dir(ticker: str) -> Path:
    """raw_/mapped_/novels_ JSONs live alongside the ticker library."""
    return ticker_root(ticker) / ".cache"


def validated_dir(ticker: str) -> Path:
    """validated_*.json lives in the ticker library, not Model Outputs."""
    return ticker_root(ticker)


def discover_filings(ticker: str) -> list[dict]:
    """For each existing raw_*.json in the ticker's .cache, return the source
    path embedded in it. That's the manifest entry.
    """
    out = []
    for raw_path in sorted(cache_dir(ticker).glob("raw_*.json")):
        period = raw_path.stem.removeprefix("raw_")  # e.g. "2024-FY"
        data = json.loads(raw_path.read_text(encoding="utf-8"))
        # source_path is stored as e.g. "Desktop\\Brain\\Sources\\CELH\\..."
        # relative to USER_HOME. Resolve.
        rel = data["source_path"].replace("\\", "/")
        source = USER_HOME / rel
        if not source.exists():
            raise SystemExit(f"ERROR: source for {ticker}/{period} not found at {source}")
        out.append({
            "ticker": ticker,
            "period": period,
            "source": source,
            "filing_type": data["filing_type"],
        })
    return out


# ============================================================================
# Pipeline invocation (subprocess)
# ============================================================================

def _env_for_skills() -> dict:
    """PYTHONPATH must include the financials-schema package directory."""
    env = os.environ.copy()
    existing = env.get("PYTHONPATH", "")
    parts = [str(SCHEMA_PKG)]
    if existing:
        parts.append(existing)
    env["PYTHONPATH"] = os.pathsep.join(parts)
    env["PYTHONIOENCODING"] = "utf-8"
    return env


# ============================================================================
# Extract cache: skip re-running extract.py when (source, library, extract code)
# are unchanged. Extract is the slow step (PDF/HTM parsing) and dominates the
# inner loop when iterating on reconcile/validate/library tweaks. Reconcile,
# validate, and model-write always run fresh — they're cheap and depend on
# state the cache key doesn't capture (ticker ledger, schema package).
# ============================================================================

def _file_sha(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


def extract_cache_key(source: Path) -> str:
    """16-hex-char digest of (source bytes, library JSON, extract.py).
    Any change to one of those invalidates the cache for this filing."""
    h = hashlib.sha256()
    h.update(_file_sha(source).encode())
    h.update(_file_sha(LIBRARY).encode())
    h.update(_file_sha(EXTRACT_CLI).encode())
    return h.hexdigest()[:32]


# Per-ticker hit/miss tally — written by run_pipeline, summarized in main().
_CACHE_STATS: dict[str, dict[str, int]] = {}


def _run(cmd: list, label: str) -> None:
    """Run a subprocess; raise SystemExit(2) on nonzero return."""
    result = subprocess.run(
        [str(c) for c in cmd],
        env=_env_for_skills(),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    if result.returncode != 0:
        sys.stderr.write(f"\n[{label}] FAILED (rc={result.returncode})\n")
        sys.stderr.write(f"  cmd: {' '.join(str(c) for c in cmd)}\n")
        if result.stdout:
            sys.stderr.write(f"  stdout (last 500): ...{result.stdout[-500:]}\n")
        if result.stderr:
            sys.stderr.write(f"  stderr (last 500): ...{result.stderr[-500:]}\n")
        raise SystemExit(2)


def run_pipeline(ticker: str, filings: list[dict], out_dir: Path, *, use_cache: bool = True) -> Path:
    """Run extract -> reconcile -> validate per filing, then model-write across
    all validated outputs. Returns the path to the produced workbook.

    Extract output is cached by sha256(source + library + extract.py); cache
    hits skip the extract subprocess and copy the prior output directly.
    Pass use_cache=False to force fresh extract on all filings.
    """
    troot = ticker_root(ticker)
    cache = out_dir / ".cache"
    cache.mkdir(parents=True, exist_ok=True)
    if use_cache:
        EXTRACT_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    validated_paths: list[Path] = []
    hits = misses = 0

    for f in filings:
        period = f["period"]
        raw = cache / f"raw_{period}.json"
        mapped = cache / f"mapped_{period}.json"
        novels = cache / f"novels_{period}.json"
        validated = out_dir / f"validated_{period}.json"

        cache_path = EXTRACT_CACHE_DIR / f"{extract_cache_key(f['source'])}.json" if use_cache else None
        if cache_path is not None and cache_path.exists():
            shutil.copy2(cache_path, raw)
            hits += 1
        else:
            _run([
                "python", EXTRACT_CLI,
                "--ticker-root", troot,
                "--source", f["source"],
                "--out", raw,
                "--library", LIBRARY,
            ], label=f"{ticker}/{period} extract")
            if cache_path is not None:
                # Copy via temp then rename to keep cache writes atomic across
                # parallel ticker workers.
                tmp = cache_path.with_suffix(".tmp")
                shutil.copy2(raw, tmp)
                tmp.replace(cache_path)
            misses += 1

        _run([
            "python", RECONCILE_CLI,
            "--ticker-root", troot,
            "--in", raw,
            "--out", mapped,
            "--novels-out", novels,
        ], label=f"{ticker}/{period} reconcile")

        _run([
            "python", VALIDATE_CLI,
            "--ticker-root", troot,
            "--in", mapped,
            "--out", validated,
        ], label=f"{ticker}/{period} validate")

        validated_paths.append(validated)

    workbook_out = out_dir / WORKBOOK_FILENAME[ticker]
    cmd = ["python", MODELWRITE_CLI, "--ticker-root", troot]
    for v in validated_paths:
        cmd += ["--in", v]
    cmd += ["--out", workbook_out]
    _run(cmd, label=f"{ticker} model-write")

    _CACHE_STATS[ticker] = {"hits": hits, "misses": misses}
    return workbook_out


# ============================================================================
# Snapshots: workbook -> cell-map JSON for deterministic comparison
# ============================================================================

def snapshot_workbook(xlsx_path: Path) -> dict:
    """Load workbook, return {sheet: {row_label: {col_label: cell}}}.

    Row label = column-A value of each row. Column label = row-1 value of
    each column. Cells are (formula or number) as a string.
    """
    wb = load_workbook(xlsx_path, data_only=False)
    snap: dict = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 1 or ws.max_column < 1:
            continue
        # Column headers from row 1.
        col_labels: dict[int, str] = {}
        for cell in ws[1]:
            if cell.value is not None:
                col_labels[cell.column] = str(cell.value)
        sheet_snap: dict = {}
        for row_idx in range(2, ws.max_row + 1):
            row_label_cell = ws.cell(row=row_idx, column=1)
            if row_label_cell.value is None:
                continue
            row_label = str(row_label_cell.value)
            row_snap: dict = {}
            for col_idx in range(2, ws.max_column + 1):
                col_label = col_labels.get(col_idx)
                if col_label is None:
                    continue
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                row_snap[col_label] = str(v)
            if row_snap:
                # If the same row label appears twice on a sheet (rare), keep
                # the later occurrence under a suffixed key so neither is lost.
                key = row_label
                suffix = 2
                while key in sheet_snap:
                    key = f"{row_label} #{suffix}"
                    suffix += 1
                sheet_snap[key] = row_snap
        snap[sheet_name] = sheet_snap
    return snap


# ============================================================================
# Diff
# ============================================================================

def _try_decimal(s) -> Decimal | None:
    if isinstance(s, (int, float, Decimal)):
        try:
            return Decimal(str(s))
        except (InvalidOperation, ValueError):
            return None
    if isinstance(s, str):
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError):
            return None
    return None


def _values_equal(a, b) -> bool:
    if a == b:
        return True
    da, db = _try_decimal(a), _try_decimal(b)
    if da is not None and db is not None:
        return abs(da - db) <= NUMERIC_TOLERANCE
    return False


# Path-key patterns that should be ignored during diff. Source paths can
# vary by absolute-vs-relative form depending on how the pipeline was
# invoked; that's representational, not a data regression.
_DIFF_SKIP_KEYS = {"source_path"}


def diff_json(golden, current, path: str = "") -> list[str]:
    """Recursively diff two JSON-decoded structures. Numeric values within
    NUMERIC_TOLERANCE are considered equal. Keys in _DIFF_SKIP_KEYS are
    skipped (representational metadata, not data).
    """
    diffs: list[str] = []
    if type(golden) != type(current):
        diffs.append(f"{path}: type {type(golden).__name__} -> {type(current).__name__}")
        return diffs
    if isinstance(golden, dict):
        for k in sorted(set(golden) | set(current)):
            if k in _DIFF_SKIP_KEYS:
                continue
            if k not in golden:
                diffs.append(f"{path}/{k}: ADDED")
            elif k not in current:
                diffs.append(f"{path}/{k}: REMOVED")
            else:
                diffs.extend(diff_json(golden[k], current[k], f"{path}/{k}"))
    elif isinstance(golden, list):
        if len(golden) != len(current):
            diffs.append(f"{path}: list length {len(golden)} -> {len(current)}")
        for i, (g, c) in enumerate(zip(golden, current)):
            diffs.extend(diff_json(g, c, f"{path}[{i}]"))
    else:
        if not _values_equal(golden, current):
            diffs.append(f"{path}: {golden!r} -> {current!r}")
    return diffs


# ============================================================================
# Bootstrap / accept (copy current outputs into goldens)
# ============================================================================

def write_goldens_from_current(ticker: str) -> None:
    """Snapshot the current outputs for `ticker` into goldens/. Validated JSONs
    live under the ticker library; the workbook lives under Model Outputs."""
    dst = GOLDENS_ROOT / ticker
    dst.mkdir(parents=True, exist_ok=True)

    for vpath in sorted(validated_dir(ticker).glob("validated_*.json")):
        shutil.copy2(vpath, dst / vpath.name)

    wb_src = model_output_dir(ticker) / WORKBOOK_FILENAME[ticker]
    if not wb_src.exists():
        raise SystemExit(f"ERROR: workbook {wb_src} not found")
    snap = snapshot_workbook(wb_src)
    (dst / "workbook.json").write_text(
        json.dumps(snap, indent=2, sort_keys=True), encoding="utf-8"
    )
    print(f"  [{ticker}] wrote {len(list(dst.glob('validated_*.json')))} validated files + workbook.json")


def write_goldens_from_temp(ticker: str, temp_dir: Path) -> None:
    """After a successful pipeline run in temp_dir, copy results into goldens/."""
    dst = GOLDENS_ROOT / ticker
    dst.mkdir(parents=True, exist_ok=True)

    # Wipe stale goldens for this ticker first to mirror the new state exactly.
    for old in list(dst.glob("validated_*.json")):
        old.unlink()
    for old in list(dst.glob("workbook.json")):
        old.unlink()

    for vpath in sorted(temp_dir.glob("validated_*.json")):
        shutil.copy2(vpath, dst / vpath.name)
    snap = snapshot_workbook(temp_dir / WORKBOOK_FILENAME[ticker])
    (dst / "workbook.json").write_text(
        json.dumps(snap, indent=2, sort_keys=True), encoding="utf-8"
    )
    print(f"  [{ticker}] wrote {len(list(dst.glob('validated_*.json')))} validated files + workbook.json")


# ============================================================================
# Compare temp run against goldens
# ============================================================================

def compare_against_goldens(ticker: str, temp_dir: Path) -> list[str]:
    """Return a list of diff strings. Empty list = clean."""
    diffs: list[str] = []
    g_dir = GOLDENS_ROOT / ticker

    # Validated files: must match by filename set; per-file deep diff.
    g_files = {p.name for p in g_dir.glob("validated_*.json")}
    t_files = {p.name for p in temp_dir.glob("validated_*.json")}
    for missing in sorted(g_files - t_files):
        diffs.append(f"[{ticker}] validated MISSING in current run: {missing}")
    for extra in sorted(t_files - g_files):
        diffs.append(f"[{ticker}] validated UNEXPECTED in current run: {extra}")
    for name in sorted(g_files & t_files):
        g = json.loads((g_dir / name).read_text(encoding="utf-8"))
        c = json.loads((temp_dir / name).read_text(encoding="utf-8"))
        for d in diff_json(g, c, path=f"[{ticker}] {name}"):
            diffs.append(d)

    # Workbook cell-map diff.
    g_wb_path = g_dir / "workbook.json"
    if not g_wb_path.exists():
        diffs.append(f"[{ticker}] golden workbook.json missing -- run --bootstrap")
        return diffs
    g_wb = json.loads(g_wb_path.read_text(encoding="utf-8"))
    c_wb = snapshot_workbook(temp_dir / WORKBOOK_FILENAME[ticker])
    for d in diff_json(g_wb, c_wb, path=f"[{ticker}] workbook"):
        diffs.append(d)

    return diffs


# ============================================================================
# CLI
# ============================================================================

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--ticker", choices=["CELH", "PG", "PEP", "MNST", "GOOG"], help="restrict to one ticker")
    ap.add_argument("--bootstrap", action="store_true",
                    help="copy current Model Output to goldens (one-time)")
    ap.add_argument("--accept", action="store_true",
                    help="rerun pipeline and overwrite goldens with fresh output")
    ap.add_argument("--keep-temp", action="store_true",
                    help="keep the temp pipeline directory for inspection")
    ap.add_argument("--max-workers", type=int, default=0,
                    help="parallel ticker workers (default: min(cpu_count, n_tickers); 1 = serial)")
    ap.add_argument("--no-cache", action="store_true",
                    help="force fresh extract on every filing (bypass _extract_cache/)")
    args = ap.parse_args()

    tickers = [args.ticker] if args.ticker else ["CELH", "PG", "PEP", "MNST", "GOOG"]
    max_workers = args.max_workers if args.max_workers > 0 else min(os.cpu_count() or 1, len(tickers))

    GOLDENS_ROOT.mkdir(parents=True, exist_ok=True)

    if args.bootstrap:
        print("Bootstrapping goldens from current Model Output...")
        for t in tickers:
            write_goldens_from_current(t)
        print("Done. Re-run without --bootstrap to verify clean diff.")
        return 0

    # For every other mode we need a fresh pipeline run in a temp dir.
    with tempfile.TemporaryDirectory(prefix="regression_") as tmp_root:
        tmp_root = Path(tmp_root)
        all_diffs: dict[str, list[str]] = {}

        # Phase A: parallel per-ticker pipeline runs.
        # Each ticker's run_pipeline shells out to extract/reconcile/validate/
        # model-write subprocesses. The Python threads spend their time waiting
        # on subprocess.run, so they release the GIL — ThreadPoolExecutor gives
        # us the parallelism without multiprocessing's pickling overhead.
        # Tickers are fully isolated (own ticker_root, own out_dir under tmp_root).
        ticker_dirs: dict[str, Path] = {}
        ticker_errors: dict[str, BaseException] = {}

        def _one_ticker(t: str) -> tuple[str, Path]:
            print(f"[{t}] discovering filings...")
            filings = discover_filings(t)
            print(f"[{t}] running pipeline on {len(filings)} filings...")
            t_dir = tmp_root / t
            t_dir.mkdir(parents=True, exist_ok=True)
            t0 = time.monotonic()
            run_pipeline(t, filings, t_dir, use_cache=not args.no_cache)
            elapsed = time.monotonic() - t0
            stats = _CACHE_STATS.get(t, {"hits": 0, "misses": 0})
            print(f"[{t}] pipeline done ({elapsed:.1f}s; "
                  f"extract cache {stats['hits']} hit / {stats['misses']} miss)")
            return t, t_dir

        wall_start = time.monotonic()
        print(f"Running {len(tickers)} ticker(s) with max_workers={max_workers}...")
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = {pool.submit(_one_ticker, t): t for t in tickers}
            for fut in as_completed(futures):
                t = futures[fut]
                try:
                    _, t_dir = fut.result()
                    ticker_dirs[t] = t_dir
                except BaseException as exc:
                    ticker_errors[t] = exc
                    sys.stderr.write(f"[{t}] pipeline FAILED: {exc!r}\n")
        print(f"All pipelines finished in {time.monotonic() - wall_start:.1f}s.")

        # Phase B: serial post-processing (accept goldens or diff). Fast, and
        # avoids any race on goldens/ writes.
        for t in tickers:
            if t in ticker_errors:
                continue
            t_dir = ticker_dirs[t]
            if args.accept:
                print(f"[{t}] accepting fresh output as new golden")
                write_goldens_from_temp(t, t_dir)
            else:
                print(f"[{t}] diffing against goldens...")
                all_diffs[t] = compare_against_goldens(t, t_dir)

        if args.keep_temp:
            keep = MODEL_SCHEMA / "_regression" / "_last_run"
            if keep.exists():
                shutil.rmtree(keep)
            shutil.copytree(tmp_root, keep)
            print(f"[debug] temp run preserved at {keep}")

    if ticker_errors:
        print(f"\n{len(ticker_errors)} ticker(s) FAILED:")
        for t, exc in ticker_errors.items():
            print(f"  [{t}] {exc!r}")
        return 2

    if args.accept:
        print("Goldens overwritten with fresh output.")
        return 0

    # Report diffs.
    total = sum(len(d) for d in all_diffs.values())
    if total == 0:
        print("\nALL CLEAN -- no regressions.")
        return 0
    print(f"\n{total} DIFFS FOUND across {sum(1 for d in all_diffs.values() if d)} ticker(s):\n")
    # Cluster diffs by anonymized signature so repeated patterns collapse to
    # a single line + count, while one-off anomalies remain individually
    # visible. Signature = the diff string with array indices ([0], [42], ...)
    # replaced by [*]. Two diffs that differ only in which line-item index
    # they hit share a signature; a singleton anomaly does not.
    #
    # Why not a fixed cap on examples printed: a 1000-diff regression with
    # one expected pattern hides any 1-3-diff anomaly in the tail. Clustering
    # surfaces both kinds of change without burying either.
    _IDX_RE = re.compile(r"\[\d+\]")
    for t, diffs in all_diffs.items():
        if not diffs:
            continue
        groups: dict[str, list[str]] = {}
        for d in diffs:
            sig = _IDX_RE.sub("[*]", d)
            groups.setdefault(sig, []).append(d)
        print(f"--- {t} ({len(diffs)} diffs across {len(groups)} unique pattern(s)) ---")
        # Sort by descending occurrence count — biggest patterns first, then
        # singleton anomalies at the bottom of the list.
        for sig, examples in sorted(groups.items(), key=lambda kv: -len(kv[1])):
            print(f"  {len(examples):>5}x  {examples[0]}")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
