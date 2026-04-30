"""depreciation-amortization-impairment-projections — reconcile.

Cross-checks the extracted asset_depreciation.json values against the SEC's
auto-generated R-file (`*_financial_report.xlsx`) for every filing on disk.

For each filing, opens the standard footnote sheets, finds well-known total
rows (PP&E net, accumulated depreciation, goodwill, intangibles net), maps the
column headers to fiscal-period labels, and prints any drift between the
R-file values and the extracted asset_depreciation.json values.

This is a SANITY CHECK, not a primary data path. Pattern-matching row labels
is acceptable here because it's defensive — when reconcile passes, we trust
the data; when it fails, we investigate the source.

CLI:
    python reconcile.py \
        --ticker        CELH \
        --asset-deprec  "Brain/Knowledge/Model Schema/Ticker Libraries/CELH/asset_depreciation.json" \
        --filings-root  "Brain/Sources/CELH"
"""
from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# R-file row-label patterns
# ---------------------------------------------------------------------------

# Sheet names are truncated by Excel at 31 chars, so we match by row CONTENT
# instead. For each field, scan every sheet and find a row whose label matches
# the row pattern AND whose sheet has date-formatted column headers.
RECONCILE_PATTERNS = [
    # PP&E net — try "net"-explicit first (BS-style), fall back to bare label
    # (footnote-style at filers like CELH that have only one PP&E row in the
    # schedule). Dedup ensures the first match wins per (field, period).
    {
        "row_pattern": r"^property,?\s+plant\s+and\s+equipment,?\s+net$",
        "field":       "ppe_net",
        "tolerance":   Decimal("1"),
    },
    {
        "row_pattern": r"^property,?\s+plant\s+and\s+equipment$",
        "field":       "ppe_net",
        "tolerance":   Decimal("1"),
    },
    # Accumulated depreciation — sign-flipped (R-file negative, field positive)
    {
        "row_pattern": r"less:?\s+accumulated\s+depreciation|^accumulated\s+depreciation",
        "field":       "ppe_accumulated_depreciation",
        "tolerance":   Decimal("1"),
        "sign_flip":   True,
    },
    # Goodwill total
    {
        "row_pattern": r"^goodwill$",
        "field":       "goodwill_balance",
        "tolerance":   Decimal("1"),
    },
    # Intangibles net — broader total preferred (matches "Intangibles-net" or
    # broader "Intangible Assets, Net"); falls back to definite-lived only
    {
        "row_pattern": r"^intangibles[\s-]+net$|^intangible\s+assets,?\s+net$",
        "field":       "intangibles_net",
        "tolerance":   Decimal("1"),
    },
    {
        "row_pattern": r"definite-lived intangible assets,?\s+net",
        "field":       "intangibles_net",
        "tolerance":   Decimal("1"),
    },
]


# ---------------------------------------------------------------------------
# Period-date mapping
# ---------------------------------------------------------------------------

DATE_RE = re.compile(r"(\w+)\.?\s+(\d+),?\s+(\d{4})")
MONTH_NAMES = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "january": 1, "february": 2, "march": 3, "april": 4, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}


def parse_date_header(s: str) -> Optional[tuple[int, int, int]]:
    """'Dec. 31, 2024' -> (2024, 12, 31). Returns None if unparseable."""
    if not s:
        return None
    s_clean = s.replace("\n", " ").strip()
    m = DATE_RE.search(s_clean)
    if not m:
        return None
    month_str = m.group(1).lower()[:3] if m.group(1) else ""
    month = MONTH_NAMES.get(month_str) or MONTH_NAMES.get(m.group(1).lower())
    if month is None:
        return None
    try:
        day = int(m.group(2))
        year = int(m.group(3))
    except ValueError:
        return None
    return (year, month, day)


def derive_fy_end_month(asset_depreciation: dict) -> int:
    """Read filer's fiscal-year-end month from the asset_depreciation.json
    `fiscal_year_end_month` field (set during extract). Falls back to 12."""
    return asset_depreciation.get("fiscal_year_end_month") or 12


def date_to_period_label(year: int, month: int, day: int, fy_end_month: int = 12) -> str:
    """Map a calendar end-date to a fiscal-period label, accounting for
    non-calendar fiscal years (PG fiscal year ends June, etc.).

    Fiscal-year labeling: a date is in fiscal year Y if it falls within the
    12 months ending at month `fy_end_month` of year Y. Calendar months <= fy_end
    use same year; later months belong to next fiscal year.

    Examples for fy_end_month=12 (calendar):
        2024-12-31 -> FY2024;  2024-09-30 -> Q3 FY2024;  2024-03-31 -> Q1 FY2024
    Examples for fy_end_month=6 (PG):
        2024-06-30 -> FY2024;  2023-09-30 -> Q1 FY2024;  2023-12-31 -> Q2 FY2024;
        2024-03-31 -> Q3 FY2024
    """
    fy_year = year if month <= fy_end_month else year + 1
    months_after_prev_fye = (month - fy_end_month) % 12

    # Fiscal-year-end: month matches fy_end and day is at month-end (>= 28)
    if month == fy_end_month and day >= 28:
        return f"FY{fy_year}"
    if months_after_prev_fye == 3:
        return f"Q1 FY{fy_year}"
    if months_after_prev_fye == 6:
        return f"Q2 FY{fy_year}"
    if months_after_prev_fye == 9:
        return f"Q3 FY{fy_year}"
    if months_after_prev_fye == 0:
        return f"FY{fy_year}"
    return f"FY{fy_year}-M{month:02d}"  # uncategorizable; tag for manual review


# ---------------------------------------------------------------------------
# R-file walking
# ---------------------------------------------------------------------------

def find_filings(filings_root: Path) -> list[Path]:
    """Return every *_financial_report.xlsx under filings_root."""
    return sorted(filings_root.glob("*/filings/*_financial_report.xlsx"))


def reconcile_filing(
    xlsx_path: Path,
    asset_deprec: dict,
    fy_end_month: int = 12,
) -> list[dict]:
    """Open the R-file, scan target sheets, return a list of reconciliation
    results: {field, period, rfile_value, extract_value, drift, status}."""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        return [{"field": "OPEN_ERROR", "period": "", "rfile_value": None,
                 "extract_value": None, "drift": None, "status": f"ERR: {e}",
                 "filing": xlsx_path.name}]

    results = []
    seen_field_period: set[tuple[str, str]] = set()  # dedupe across sheets

    for pattern in RECONCILE_PATTERNS:
        row_pat = re.compile(pattern["row_pattern"], re.IGNORECASE)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            # Parse column headers (row 0) for period dates
            header_row = rows[0]
            col_periods: dict[int, str] = {}
            for col_idx, cell in enumerate(header_row):
                if cell is None or col_idx == 0:
                    continue
                parsed = parse_date_header(str(cell))
                if parsed:
                    col_periods[col_idx] = date_to_period_label(*parsed, fy_end_month=fy_end_month)

            if not col_periods:
                continue

            # Find a row matching the pattern with at least one numeric value
            for row in rows[1:]:
                if not row or row[0] is None:
                    continue
                label = str(row[0]).strip()
                if not row_pat.search(label):
                    continue
                # Confirm this row has at least one numeric value in a period column
                # (filters out spurious matches like "[Line Items]" header rows)
                has_numeric = any(
                    isinstance(row[col_idx], (int, float))
                    for col_idx in col_periods if col_idx < len(row)
                )
                if not has_numeric:
                    continue

                # Extract values from each period column
                for col_idx, period in col_periods.items():
                    if col_idx >= len(row):
                        continue
                    cell_val = row[col_idx]
                    if cell_val is None or not isinstance(cell_val, (int, float)):
                        continue
                    if (pattern["field"], period) in seen_field_period:
                        continue  # already reconciled this (field, period) on a different sheet
                    try:
                        rfile_val = Decimal(str(cell_val))
                    except Exception:
                        continue
                    if pattern.get("sign_flip"):
                        rfile_val = -rfile_val

                    extract_map = asset_deprec.get(pattern["field"], {})
                    extract_val = extract_map.get(period)
                    if extract_val is None:
                        status = "MISSING_IN_EXTRACT"
                        drift = None
                    else:
                        extract_dec = Decimal(str(extract_val))
                        drift = rfile_val - extract_dec
                        status = "OK" if abs(drift) <= pattern["tolerance"] else "DRIFT"

                    results.append({
                        "filing":        xlsx_path.name,
                        "sheet":         sheet_name,
                        "field":         pattern["field"],
                        "period":        period,
                        "rfile_value":   rfile_val,
                        "extract_value": Decimal(str(extract_val)) if extract_val is not None else None,
                        "drift":         drift,
                        "status":        status,
                    })
                    seen_field_period.add((pattern["field"], period))
                break  # only first matching row per sheet
    return results


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--ticker", required=True)
    p.add_argument("--asset-deprec", required=True, type=Path,
                   help="path to asset_depreciation.json (output of extract.py)")
    p.add_argument("--filings-root", required=True, type=Path,
                   help="path to Brain/Sources/{TICKER} (parent of period subdirs)")
    p.add_argument("--summary-only", action="store_true",
                   help="print summary table only, not per-row results")
    args = p.parse_args()

    asset_deprec = json.loads(args.asset_deprec.read_text(encoding="utf-8"))
    fy_end = derive_fy_end_month(asset_deprec)

    filings = find_filings(args.filings_root)
    print(f"[{args.ticker}] reconciling {len(filings)} filings against {args.asset_deprec.name} "
          f"(fy_end_month={fy_end}, unit={asset_deprec.get('reporting_unit')})")

    all_results: list[dict] = []
    for f in filings:
        r = reconcile_filing(f, asset_deprec, fy_end_month=fy_end)
        all_results.extend(r)

    # Summary
    by_status = defaultdict(int)
    by_field_status: dict[tuple[str, str], int] = defaultdict(int)
    for r in all_results:
        by_status[r["status"]] += 1
        by_field_status[(r["field"], r["status"])] += 1

    print(f"\n[{args.ticker}] summary across {len(filings)} filings ({len(all_results)} checks):")
    for status, count in sorted(by_status.items(), key=lambda x: -x[1]):
        print(f"  {status:25s} {count}")

    print(f"\n[{args.ticker}] per-field status:")
    fields = sorted(set(k[0] for k in by_field_status))
    for fld in fields:
        ok = by_field_status.get((fld, "OK"), 0)
        drift = by_field_status.get((fld, "DRIFT"), 0)
        missing = by_field_status.get((fld, "MISSING_IN_EXTRACT"), 0)
        print(f"  {fld:35s} OK={ok}  DRIFT={drift}  MISSING={missing}")

    if args.summary_only:
        return

    # Show all DRIFTs and MISSINGs (the actionable part)
    drifts = [r for r in all_results if r["status"] in ("DRIFT", "MISSING_IN_EXTRACT")]
    if drifts:
        print(f"\n[{args.ticker}] {len(drifts)} actionable rows:")
        for r in drifts:
            print(f"  [{r['status']:18s}] {r['filing']:50s} {r['field']:30s} {r['period']:12s} "
                  f"rfile={r['rfile_value']!s:>12} extract={r['extract_value']!s:>12} drift={r['drift']!s}")


if __name__ == "__main__":
    main()
