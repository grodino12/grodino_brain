"""model-calc CLI — add ASSUMPTIONS + IS/BS/CF driver tabs to a built financial model xlsx.

Historical driver cells compute live formulas from the ANNL P&L / BALANCE SHEET /
CASH FLOW tabs. Projection driver cells follow each driver's forecast rule:
  - "hold_last":       =prior column (chains through forecast years)
  - "input":           left blank for user input (e.g. Revenue Growth %)
  - "assumption_ref":  =ASSUMPTIONS!$B${row} (e.g. Share Repurchases $ = 0)
  - "derived":         same formula as historical applied to forecast columns

ASSUMPTIONS tab holds constants referenced from driver formulas (Days in Year,
Share Repurchases $, etc.) so formulas don't embed magic numbers.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================================
# Constants
# ============================================================================

FORECAST_LABELS = [f"FY{y}E" for y in range(2025, 2031)]

SOURCE_SHEETS = ("ANNL P&L", "BALANCE SHEET", "CASH FLOW")
DRIVER_SHEETS = ("IS DRIVERS", "BS DRIVERS", "CF DRIVERS")
ASSUMPTIONS_SHEET = "ASSUMPTIONS"

# Number formats
RATIO_FMT   = '0.0%;(0.0%);"--"'
DAYS_FMT    = '0.0;(0.0);"--"'
DOLLAR_FMT  = '#,##0;(#,##0);"--"'

# Styling
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", start_color="203864", end_color="203864")
HEADER_ALIGN = Alignment(horizontal="center")
LABEL_FONT = Font(bold=True)
FORECAST_FILL = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
INPUT_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")  # user-input highlight
SECTION_FONT = Font(bold=True, italic=True, color="595959")
SECTION_FILL = PatternFill(fill_type="solid", start_color="EAEAEA", end_color="EAEAEA")
ASSUMPTION_LABEL_FONT = Font(bold=True)

# ============================================================================
# ASSUMPTIONS constants
# ============================================================================
# Central constants referenced from driver / statement formulas. Raw values
# here are the ONLY hardcoded numbers in the entire model — formulas reference
# these cells by label.

ASSUMPTIONS: list[dict] = [
    {"label": "Days in Year",         "value": 365, "note": "Used in DSO, DIO, DPO formulas"},
    {"label": "Share Repurchases $",  "value": 0,   "note": "Per-period repurchase assumption for projection"},
]


# ============================================================================
# Driver specs
# ============================================================================
#
# Kinds (computation shape):
#   section        — header row (label only)
#   growth         — YoY on a single ref: (cur - prev) / prev
#   ratio          — num / den (same period)
#   lagged_ratio   — num[t] / den[t-1] (prior-period denominator)
#   days_ratio     — num / den * ASSUMPTIONS!Days_in_Year
#   dollar         — direct reference to one ref
#   dollar_sum     — sum of refs
#   net_debt       — sum(debt refs) - sum(cash refs) (same period)
#
# Forecast rules:
#   hold_last      — =prior column (chains through projection)
#   input          — leave blank for user input (highlighted)
#   assumption_ref — =ASSUMPTIONS!$B${row} with assumption label lookup
#   derived        — same formula as historical applied to forecast columns

REV = ("ANNL P&L", "Net Sales / Revenue")
COGS = ("ANNL P&L", "COGS")

DRIVER_SPECS: dict[str, list[dict]] = {
    "IS DRIVERS": [
        {"label": "Revenue Growth %",              "kind": "growth",     "forecast": "input",
         "num": REV},
        {"label": "COGS % of Revenue",             "kind": "ratio",      "forecast": "hold_last",
         "num": COGS, "den": REV},
        {"label": "SG&A % of Revenue",             "kind": "ratio",      "forecast": "hold_last",
         "num": ("ANNL P&L", "SG&A"), "den": REV},
        {"label": "Interest Income (Expense) $",   "kind": "dollar",     "forecast": "hold_last",
         "refs": [("ANNL P&L", "Interest Income (Expense)")]},
        {"label": "Foreign Currency Gain (Loss) $","kind": "dollar",     "forecast": "hold_last",
         "refs": [("ANNL P&L", "Foreign Currency Gain (Loss)")]},
        {"label": "Other Income (Expense) $",      "kind": "dollar",     "forecast": "hold_last",
         "refs": [("ANNL P&L", "Other Income (Expense)")]},
        {"label": "Effective Tax Rate %",          "kind": "ratio",      "forecast": "hold_last",
         "num": ("ANNL P&L", "Income Tax (Benefit) Expense"),
         "den": ("ANNL P&L", "Pre-Tax Income (Loss)")},
    ],
    "BS DRIVERS": [
        {"kind": "section", "label": "Current Assets"},
        {"label": "DSO (days)",                                  "kind": "days_ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Accounts Receivable"), "den": REV},
        {"label": "DIO (days)",                                  "kind": "days_ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Inventories"), "den": COGS},
        {"label": "Note Receivable - Current % of Rev",          "kind": "ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Note Receivable - Current"), "den": REV},
        {"label": "Deferred Other Costs - Current % of Rev",     "kind": "ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Deferred Other Costs - Current"), "den": REV},
        {"label": "Prepaid Expenses % of Rev",                   "kind": "ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Prepaid Expenses"), "den": REV},

        {"kind": "section", "label": "Non-Current Assets"},
        {"label": "Other Non-Current Assets % of Rev",           "kind": "ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Other Non-Current Assets"), "den": REV},

        {"kind": "section", "label": "Current Liabilities"},
        {"label": "DPO (days)",                                  "kind": "days_ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Accounts Payable"), "den": COGS},
        {"label": "Accrued Expenses % of Rev",                   "kind": "ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Accrued Expenses"), "den": REV},
        {"label": "Income Taxes Payable % of Rev",               "kind": "ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Income Taxes Payable"), "den": REV},
        {"label": "Accrued Distributor Termination Fees % of Rev","kind": "ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Accrued Distributor Termination Fees"), "den": REV},
        {"label": "Accrued Promotional Allowance % of Rev",      "kind": "ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Accrued Promotional Allowance"), "den": REV},
        {"label": "Lease Liability - Operating - Current % of Rev","kind": "ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Lease Liability - Operating - Current"), "den": REV},
        {"label": "Lease Liability - Finance - Current % of Rev","kind": "ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Lease Liability - Finance - Current"), "den": REV},
        {"label": "Deferred Revenue - Current % of Rev",         "kind": "ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Deferred Revenue - Current"), "den": REV},
        {"label": "Other Current Liabilities % of Rev",          "kind": "ratio", "forecast": "hold_last",
         "num": ("BALANCE SHEET", "Other Current Liabilities"), "den": REV},
    ],
    "CF DRIVERS": [
        {"label": "Allowance for Credit Losses % of Revenue", "kind": "ratio", "forecast": "hold_last",
         "num": ("CASH FLOW", "Allowance for Credit Losses"), "den": REV},
        {"label": "D&A % of PP&E",          "kind": "lagged_ratio", "forecast": "hold_last",
         "num": ("CASH FLOW", "Depreciation & Amortization"),
         "den": ("BALANCE SHEET", "Net PP&E")},
        {"label": "SBC $",                  "kind": "dollar",       "forecast": "hold_last",
         "refs": [("CASH FLOW", "Stock-Based Compensation")]},
        {"label": "CapEx % of PP&E",        "kind": "lagged_ratio", "forecast": "hold_last",
         "num": ("CASH FLOW", "Purchase of PP&E"),
         "den": ("BALANCE SHEET", "Net PP&E")},
        {"label": "Preferred Dividends % of Preferred Balance", "kind": "ratio", "forecast": "hold_last",
         "num": ("CASH FLOW", "Preferred Dividends"),
         "den": ("BALANCE SHEET", "Convertible Preferred Stock")},
        {"label": "Common Dividends % of Net Income", "kind": "ratio", "forecast": "hold_last",
         "num": ("CASH FLOW", "Common Dividends"),
         "den": ("CASH FLOW", "Net Income (Loss)")},
        {"label": "Share Repurchases $",    "kind": "dollar",       "forecast": "assumption_ref",
         "refs": [("CASH FLOW", "Share Repurchases")],
         "assumption_label": "Share Repurchases $"},
        {"label": "Net Debt $",             "kind": "net_debt",     "forecast": "derived",
         "debt": [
             ("BALANCE SHEET", "Lease Liability - Operating - Current"),
             ("BALANCE SHEET", "Lease Liability - Finance - Current"),
             ("BALANCE SHEET", "Lease Liability - Operating - Non-Current"),
             ("BALANCE SHEET", "Lease Liability - Finance - Non-Current"),
         ],
         "cash": [
             ("BALANCE SHEET", "Cash & Cash Equivalents"),
             ("BALANCE SHEET", "Restricted Cash"),
         ]},
    ],
}


# ============================================================================
# ASSUMPTIONS tab
# ============================================================================

def build_assumptions_sheet(wb) -> dict[str, int]:
    """Write the ASSUMPTIONS sheet and return a {label: excel_row} map.
    Columns: A=Assumption label, B=Value, C=Note.
    """
    if ASSUMPTIONS_SHEET in wb.sheetnames:
        del wb[ASSUMPTIONS_SHEET]
    ws = wb.create_sheet(ASSUMPTIONS_SHEET)

    # Header
    for col_idx, label in enumerate(("Assumption", "Value", "Note"), start=1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN

    rows: dict[str, int] = {}
    for i, entry in enumerate(ASSUMPTIONS, start=2):
        ws.cell(row=i, column=1, value=entry["label"]).font = ASSUMPTION_LABEL_FONT
        value_cell = ws.cell(row=i, column=2, value=entry["value"])
        value_cell.fill = INPUT_FILL  # user-editable highlight
        # Apply a sensible number format based on value type
        if isinstance(entry["value"], int) and entry["label"] != "Days in Year":
            value_cell.number_format = DOLLAR_FMT
        ws.cell(row=i, column=3, value=entry.get("note", ""))
        rows[entry["label"]] = i

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A2"
    return rows


def assumption_ref(label: str, rows: dict[str, int]) -> str:
    """Absolute cell reference to an assumption value, e.g. ASSUMPTIONS!$B$2."""
    row = rows[label]
    return f"{ASSUMPTIONS_SHEET}!$B${row}"


# ============================================================================
# Source-sheet indexing
# ============================================================================

def build_source_maps(
    wb,
) -> tuple[dict[str, dict[str, int]], dict[str, dict[str, int]]]:
    period_col_by_sheet: dict[str, dict[str, int]] = {}
    row_by_sheet: dict[str, dict[str, int]] = {}
    for sheet_name in SOURCE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        periods: dict[str, int] = {}
        for col in range(2, ws.max_column + 1):
            label = ws.cell(row=1, column=col).value
            if label:
                periods[str(label).strip()] = col
        period_col_by_sheet[sheet_name] = periods

        rows: dict[str, int] = {}
        for row in range(2, ws.max_row + 1):
            label = ws.cell(row=row, column=1).value
            if label:
                rows[str(label).strip()] = row
        row_by_sheet[sheet_name] = rows
    return period_col_by_sheet, row_by_sheet


def build_driver_period_labels(period_col_by_sheet: dict[str, dict[str, int]]) -> list[str]:
    historical: set[str] = set()
    for _sheet, cols in period_col_by_sheet.items():
        for label in cols:
            if not label.endswith("E"):
                historical.add(label)

    def sort_key(lbl: str) -> tuple[int, int]:
        m = re.match(r"^Q(\d+)\s+FY(\d{4})$", lbl)
        if m:
            return (int(m.group(2)), int(m.group(1)))
        m = re.match(r"^FY(\d{4})$", lbl)
        if m:
            return (int(m.group(1)), 0)
        return (9999, 0)

    return sorted(historical, key=sort_key) + list(FORECAST_LABELS)


# ============================================================================
# Formula builders
# ============================================================================

def _cell_ref(sheet: str, row: int, col: int) -> str:
    return f"'{sheet}'!{get_column_letter(col)}{row}"


def _resolve(
    ref: tuple[str, str],
    period_label: str,
    period_col_by_sheet: dict[str, dict[str, int]],
    row_by_sheet: dict[str, dict[str, int]],
) -> tuple[str, int, int] | None:
    sheet, label = ref
    rows = row_by_sheet.get(sheet, {})
    periods = period_col_by_sheet.get(sheet, {})
    row = rows.get(label)
    col = periods.get(period_label)
    if row is None or col is None:
        return None
    return (sheet, row, col)


def build_formula(
    spec: dict,
    period_label: str,
    period_labels: list[str],
    period_col_by_sheet: dict[str, dict[str, int]],
    row_by_sheet: dict[str, dict[str, int]],
    assumptions_rows: dict[str, int],
) -> str | None:
    """Return an Excel formula for this driver at this period, or None if the
    required source cells are not resolvable."""
    kind = spec["kind"]

    if kind == "growth":
        idx = period_labels.index(period_label)
        if idx == 0:
            return None
        prev_label = period_labels[idx - 1]
        cur  = _resolve(spec["num"], period_label, period_col_by_sheet, row_by_sheet)
        prev = _resolve(spec["num"], prev_label,   period_col_by_sheet, row_by_sheet)
        if cur is None or prev is None:
            return None
        return f"=IFERROR(({_cell_ref(*cur)}-{_cell_ref(*prev)})/{_cell_ref(*prev)},0)"

    if kind == "ratio":
        num = _resolve(spec["num"], period_label, period_col_by_sheet, row_by_sheet)
        den = _resolve(spec["den"], period_label, period_col_by_sheet, row_by_sheet)
        if num is None or den is None:
            return None
        return f"=IFERROR({_cell_ref(*num)}/{_cell_ref(*den)},0)"

    if kind == "lagged_ratio":
        idx = period_labels.index(period_label)
        if idx == 0:
            return None
        prev_label = period_labels[idx - 1]
        num = _resolve(spec["num"], period_label, period_col_by_sheet, row_by_sheet)
        den = _resolve(spec["den"], prev_label,   period_col_by_sheet, row_by_sheet)
        if num is None or den is None:
            return None
        return f"=IFERROR({_cell_ref(*num)}/{_cell_ref(*den)},0)"

    if kind == "days_ratio":
        num = _resolve(spec["num"], period_label, period_col_by_sheet, row_by_sheet)
        den = _resolve(spec["den"], period_label, period_col_by_sheet, row_by_sheet)
        if num is None or den is None:
            return None
        days_ref = assumption_ref("Days in Year", assumptions_rows)
        return f"=IFERROR({_cell_ref(*num)}/{_cell_ref(*den)}*{days_ref},0)"

    if kind == "dollar":
        refs = [_resolve(r, period_label, period_col_by_sheet, row_by_sheet) for r in spec["refs"]]
        if any(r is None for r in refs):
            return None
        return "=" + "+".join(_cell_ref(*r) for r in refs)  # type: ignore[misc]

    if kind == "dollar_sum":
        refs = [_resolve(r, period_label, period_col_by_sheet, row_by_sheet) for r in spec["refs"]]
        refs = [r for r in refs if r is not None]
        if not refs:
            return None
        return "=" + "+".join(_cell_ref(*r) for r in refs)

    if kind == "net_debt":
        debts = [_resolve(r, period_label, period_col_by_sheet, row_by_sheet) for r in spec["debt"]]
        cashes = [_resolve(r, period_label, period_col_by_sheet, row_by_sheet) for r in spec["cash"]]
        debts  = [r for r in debts  if r is not None]
        cashes = [r for r in cashes if r is not None]
        if not debts or not cashes:
            return None
        debt_expr = "+".join(_cell_ref(*r) for r in debts)
        cash_expr = "+".join(_cell_ref(*r) for r in cashes)
        return f"=({debt_expr})-({cash_expr})"

    return None


def build_forecast_formula(
    spec: dict,
    period_label: str,
    period_labels: list[str],
    driver_row: int,
    driver_col: int,
    period_col_by_sheet: dict[str, dict[str, int]],
    row_by_sheet: dict[str, dict[str, int]],
    assumptions_rows: dict[str, int],
) -> str | None:
    """Projection-period formula dispatch per the spec's 'forecast' rule."""
    rule = spec.get("forecast", "hold_last")

    if rule == "input":
        return None  # blank cell for user input

    if rule == "hold_last":
        prev_col_letter = get_column_letter(driver_col - 1)
        return f"={prev_col_letter}{driver_row}"

    if rule == "assumption_ref":
        label = spec["assumption_label"]
        return f"={assumption_ref(label, assumptions_rows)}"

    if rule == "derived":
        # Apply the historical formula to the forecast period (refs to forecast
        # columns on the source sheets — those cells get populated downstream).
        return build_formula(
            spec, period_label, period_labels,
            period_col_by_sheet, row_by_sheet, assumptions_rows,
        )

    return None


def format_for_kind(kind: str) -> str | None:
    if kind in ("growth", "ratio", "lagged_ratio"):
        return RATIO_FMT
    if kind == "days_ratio":
        return DAYS_FMT
    if kind in ("dollar", "dollar_sum", "net_debt"):
        return DOLLAR_FMT
    return None


# ============================================================================
# Writing driver sheets
# ============================================================================

def write_driver_sheet(
    wb,
    sheet_name: str,
    specs: list[dict],
    period_labels: list[str],
    period_col_by_sheet: dict[str, dict[str, int]],
    row_by_sheet: dict[str, dict[str, int]],
    assumptions_rows: dict[str, int],
) -> int:
    """Write (or replace) a driver sheet. Returns total formula cells written
    (historical + forecast)."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Header
    hdr_a = ws.cell(row=1, column=1, value="")
    hdr_a.fill = HEADER_FILL
    hdr_a.font = HEADER_FONT
    hist_cols: list[tuple[int, str]] = []
    forecast_cols: list[tuple[int, str]] = []
    for col_idx, label in enumerate(period_labels, start=2):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        if label.endswith("E"):
            forecast_cols.append((col_idx, label))
        else:
            hist_cols.append((col_idx, label))

    formulas_written = 0
    excel_row = 2
    for spec in specs:
        kind = spec["kind"]
        label = spec["label"]

        if kind == "section":
            cell = ws.cell(row=excel_row, column=1, value=label)
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            for col_idx in range(2, 2 + len(period_labels)):
                ws.cell(row=excel_row, column=col_idx).fill = SECTION_FILL
            excel_row += 1
            continue

        ws.cell(row=excel_row, column=1, value=label).font = LABEL_FONT
        fmt = format_for_kind(kind)

        # --- Historical columns ---
        for col_idx, period_label in hist_cols:
            formula = build_formula(
                spec, period_label, period_labels,
                period_col_by_sheet, row_by_sheet, assumptions_rows,
            )
            if formula is None:
                continue
            cell = ws.cell(row=excel_row, column=col_idx, value=formula)
            if fmt:
                cell.number_format = fmt
            formulas_written += 1

        # --- Forecast columns ---
        for col_idx, period_label in forecast_cols:
            cell = ws.cell(row=excel_row, column=col_idx)
            if fmt:
                cell.number_format = fmt
            # Tint: yellow for user-input cells, grey for formula forecasts
            if spec.get("forecast") == "input":
                cell.fill = INPUT_FILL
                continue
            formula = build_forecast_formula(
                spec, period_label, period_labels, excel_row, col_idx,
                period_col_by_sheet, row_by_sheet, assumptions_rows,
            )
            if formula is None:
                if cell.fill.fill_type is None:
                    cell.fill = FORECAST_FILL
                continue
            cell.value = formula
            if cell.fill.fill_type is None:
                cell.fill = FORECAST_FILL
            formulas_written += 1

        excel_row += 1

    ws.column_dimensions["A"].width = 48
    for col_idx in range(2, 2 + len(period_labels)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13
    ws.freeze_panes = "B2"
    return formulas_written


# ============================================================================
# Statement forecast specs
# ============================================================================
#
# Each entry is a dict with "label" (canonical row on the target sheet) and
# "kind" (forecast rule). Additional fields per kind below.
#
# Kinds:
#   revenue_growth     — Rev[t] = Rev[t-1] × (1 + driver[t])
#   ratio_of_rev       — line[t] = driver[t] × Rev[t]
#   ratio_of_cogs      — line[t] = driver[t] × COGS[t]
#   days_driven_rev    — line[t] = driver[t] × Rev[t] / Days_in_Year
#   days_driven_cogs   — line[t] = driver[t] × COGS[t] / Days_in_Year
#   dollar_driver      — line[t] = driver[t]
#   tax                — line[t] = Pre-Tax[t] × driver[t]
#   flat               — line[t] = line[t-1]
#   subtotal           — skip (already has a live formula from model-write)
#   cf_from_pl         — line[t] = P&L[pl_label][t]
#   cf_from_cf         — line[t] = CF[cf_label][t]
#   cash_rollforward   — line[t] = line[t-1] + CF!Net Change in Cash[t]
#   pp_e_rollforward   — line[t] = line[t-1] - CF!Purchase of PP&E[t] - CF!D&A[t]
#   apic_rollforward   — line[t] = line[t-1] + CF!SBC[t]
#   re_rollforward     — line[t] = line[t-1] + P&L!NI[t] + CF!Preferred Div + (CF!Common Div if present)
#   amortize           — line[t] = line[t-1] - CF!Amortization of Deferred Other Costs[t]
#   ni_attrib_common   — line[t] = P&L!NI[t] + P&L!Preferred Div + P&L!Income Allocated
#   cf_wc_asset        — line[t] = -(BS!label[t] - BS!label[t-1])
#   cf_wc_liability    — line[t] =  (BS!label[t] - BS!label[t-1])
#   cf_wc_combined     — line[t] =  (SUM(BS!labels[t]) - SUM(BS!labels[t-1]))
#                                    * sign (1 for liability-like, -1 for asset-like)
#   capex              — line[t] = driver[t] × BS!PP&E[t-1]  (driver stored negative → result negative)
#   d_a                — line[t] = driver[t] × BS!PP&E[t-1]  (positive ratio × positive PP&E = positive)
#   dividends_preferred — line[t] = driver[t] × BS!Convertible Preferred Stock[t]
#   dividends_common    — line[t] = driver[t] × P&L!Net Income[t]
#   cf_net_change      — line[t] = CFO[t] + CFI[t] + CFF[t] + FX[t]
#   cash_beg           — line[t] = CF!Cash at End of Period[t-1]
#   cash_end           — line[t] = CF!Cash at Beginning of Period[t] + CF!Net Change[t]
#   aoci_rollforward   — line[t] = line[t-1] + CF!FX Effect on Cash[t]
#   zero               — line[t] = 0  (non-recurring items in steady-state forecast)
#   skip               — leave blank
# ============================================================================

FORECAST_STATEMENT_SPECS: dict[str, list[dict]] = {
    "ANNL P&L": [
        {"label": "Net Sales / Revenue",                     "kind": "revenue_growth",
         "driver": ("IS DRIVERS", "Revenue Growth %")},
        {"label": "COGS",                                     "kind": "ratio_of_rev",
         "driver": ("IS DRIVERS", "COGS % of Revenue")},
        {"label": "Gross Profit (Loss)",                      "kind": "subtotal"},
        {"label": "SG&A",                                     "kind": "ratio_of_rev",
         "driver": ("IS DRIVERS", "SG&A % of Revenue")},
        {"label": "Income (Loss) from Operations",            "kind": "subtotal"},
        {"label": "Interest Income (Expense)",                "kind": "dollar_driver",
         "driver": ("IS DRIVERS", "Interest Income (Expense) $")},
        {"label": "Foreign Currency Gain (Loss)",             "kind": "zero"},
        {"label": "Other Income (Expense)",                   "kind": "dollar_driver",
         "driver": ("IS DRIVERS", "Other Income (Expense) $")},
        {"label": "Pre-Tax Income (Loss)",                    "kind": "subtotal"},
        {"label": "Income Tax (Benefit) Expense",             "kind": "tax",
         "driver": ("IS DRIVERS", "Effective Tax Rate %")},
        {"label": "Net Income (Loss)",                        "kind": "subtotal"},
        {"label": "Preferred Dividends",                      "kind": "cf_from_cf",
         "cf_label": "Preferred Dividends"},
        {"label": "Income Allocated to Participating Preferred", "kind": "flat"},
        {"label": "Net Income (Loss) Attributable to Common Shareholders", "kind": "ni_attrib_common"},
        {"label": "Basic Earnings (Loss) per Share",          "kind": "skip"},
        {"label": "Diluted Earnings (Loss) per Share",        "kind": "skip"},
        {"label": "Weighted Average Shares Outstanding (Basic)", "kind": "flat"},
        {"label": "Weighted Average Shares Outstanding (Diluted)", "kind": "flat"},
    ],
    "BALANCE SHEET": [
        {"label": "Cash & Cash Equivalents",                  "kind": "cash_rollforward"},
        {"label": "Restricted Cash",                          "kind": "flat"},
        {"label": "Accounts Receivable",                      "kind": "days_driven_rev",
         "driver": ("BS DRIVERS", "DSO (days)")},
        {"label": "Note Receivable - Current",                "kind": "ratio_of_rev",
         "driver": ("BS DRIVERS", "Note Receivable - Current % of Rev")},
        {"label": "Inventories",                              "kind": "days_driven_cogs",
         "driver": ("BS DRIVERS", "DIO (days)")},
        {"label": "Deferred Other Costs - Current",           "kind": "ratio_of_rev",
         "driver": ("BS DRIVERS", "Deferred Other Costs - Current % of Rev")},
        {"label": "Prepaid Expenses",                         "kind": "ratio_of_rev",
         "driver": ("BS DRIVERS", "Prepaid Expenses % of Rev")},
        {"label": "Total Current Assets",                     "kind": "subtotal"},
        {"label": "Note Receivable - Non-Current",            "kind": "flat"},
        {"label": "Net PP&E",                                 "kind": "pp_e_rollforward"},
        {"label": "ROU Assets - Operating - Non-Current",     "kind": "flat"},
        {"label": "ROU Assets - Finance - Non-Current",       "kind": "flat"},
        {"label": "Intangible Assets",                        "kind": "flat"},
        {"label": "Goodwill",                                 "kind": "flat"},
        {"label": "Deferred Other Costs - Non-Current",       "kind": "amortize"},
        {"label": "Deferred Tax Assets",                      "kind": "flat"},
        {"label": "Other Non-Current Assets",                 "kind": "ratio_of_rev",
         "driver": ("BS DRIVERS", "Other Non-Current Assets % of Rev")},
        {"label": "Total Assets",                             "kind": "subtotal"},
        {"label": "Accounts Payable",                         "kind": "days_driven_cogs",
         "driver": ("BS DRIVERS", "DPO (days)")},
        {"label": "Accrued Expenses",                         "kind": "ratio_of_rev",
         "driver": ("BS DRIVERS", "Accrued Expenses % of Rev")},
        {"label": "Income Taxes Payable",                     "kind": "ratio_of_rev",
         "driver": ("BS DRIVERS", "Income Taxes Payable % of Rev")},
        {"label": "Accrued Distributor Termination Fees",     "kind": "ratio_of_rev",
         "driver": ("BS DRIVERS", "Accrued Distributor Termination Fees % of Rev")},
        {"label": "Accrued Promotional Allowance",            "kind": "ratio_of_rev",
         "driver": ("BS DRIVERS", "Accrued Promotional Allowance % of Rev")},
        {"label": "Lease Liability - Operating - Current",    "kind": "ratio_of_rev",
         "driver": ("BS DRIVERS", "Lease Liability - Operating - Current % of Rev")},
        {"label": "Lease Liability - Finance - Current",      "kind": "ratio_of_rev",
         "driver": ("BS DRIVERS", "Lease Liability - Finance - Current % of Rev")},
        {"label": "Deferred Revenue - Current",               "kind": "ratio_of_rev",
         "driver": ("BS DRIVERS", "Deferred Revenue - Current % of Rev")},
        {"label": "Other Current Liabilities",                "kind": "ratio_of_rev",
         "driver": ("BS DRIVERS", "Other Current Liabilities % of Rev")},
        {"label": "Total Current Liabilities",                "kind": "subtotal"},
        {"label": "Lease Liability - Operating - Non-Current","kind": "flat"},
        {"label": "Lease Liability - Finance - Non-Current",  "kind": "flat"},
        {"label": "Deferred Tax Liability",                   "kind": "flat"},
        {"label": "Deferred Revenue - Non-Current",           "kind": "flat"},
        {"label": "Total Liabilities",                        "kind": "subtotal"},
        {"label": "Convertible Preferred Stock",              "kind": "flat"},
        {"label": "Common Stock",                             "kind": "flat"},
        {"label": "Additional Paid-in Capital",               "kind": "apic_rollforward"},
        {"label": "Accumulated Other Comprehensive Income (Loss)", "kind": "aoci_rollforward"},
        {"label": "Retained Earnings (Accumulated Deficit)",  "kind": "re_rollforward"},
        {"label": "Total Stockholders' Equity",               "kind": "subtotal"},
        {"label": "Total Liabilities, Mezzanine & Stockholders' Equity", "kind": "subtotal"},
    ],
    "CASH FLOW": [
        {"label": "Net Income (Loss)",                        "kind": "cf_from_pl",
         "pl_label": "Net Income (Loss)"},
        {"label": "Depreciation & Amortization",              "kind": "d_a",
         "driver": ("CF DRIVERS", "D&A % of PP&E")},
        {"label": "Impairment of Intangibles",                "kind": "flat"},
        {"label": "Allowance for Credit Losses",              "kind": "ratio_of_rev",
         "driver": ("CF DRIVERS", "Allowance for Credit Losses % of Revenue")},
        {"label": "Amortization of Deferred Other Costs",     "kind": "flat"},
        {"label": "Inventory Write-Down",                     "kind": "zero"},
        {"label": "Gain (Loss) on Disposal of PP&E",          "kind": "zero"},
        {"label": "Stock-Based Compensation",                 "kind": "dollar_driver",
         "driver": ("CF DRIVERS", "SBC $")},
        {"label": "(Benefit) Provision for Deferred Income Taxes", "kind": "zero"},
        {"label": "Foreign Currency Gain (Loss)",             "kind": "zero"},
        {"label": "Gain (Loss) on Lease Cancellations",       "kind": "zero"},
        {"label": "Other Operating Items",                    "kind": "zero"},
        {"label": "Accounts Receivable",                      "kind": "cf_wc_asset",
         "bs_label": "Accounts Receivable"},
        {"label": "Note Receivable",                          "kind": "cf_wc_asset",
         "bs_label": "Note Receivable - Current"},
        {"label": "Inventories",                              "kind": "cf_wc_asset",
         "bs_label": "Inventories"},
        {"label": "Prepaid Expenses",                         "kind": "cf_wc_asset",
         "bs_label": "Prepaid Expenses"},
        {"label": "Accounts Payable",                         "kind": "cf_wc_liability",
         "bs_label": "Accounts Payable"},
        {"label": "Accrued Expenses",                         "kind": "cf_wc_liability",
         "bs_label": "Accrued Expenses"},
        {"label": "Other Current Liabilities",                "kind": "cf_wc_liability",
         "bs_label": "Other Current Liabilities"},
        {"label": "Accrued Promotional Allowance",            "kind": "cf_wc_liability",
         "bs_label": "Accrued Promotional Allowance"},
        {"label": "Accrued Distributor Termination",          "kind": "cf_wc_liability",
         "bs_label": "Accrued Distributor Termination Fees"},
        {"label": "ROU & Lease Liability, Net",               "kind": "zero"},
        {"label": "Deferred Revenue",                         "kind": "cf_wc_combined",
         "bs_labels": ["Deferred Revenue - Current", "Deferred Revenue - Non-Current"],
         "sign": 1},
        {"label": "Other Non-Current Assets",                 "kind": "cf_wc_asset",
         "bs_label": "Other Non-Current Assets"},
        {"label": "Cash Flow from Operations",                "kind": "subtotal"},
        {"label": "Collections from Note Receivable",         "kind": "flat"},
        {"label": "Purchase of PP&E",                         "kind": "capex",
         "driver": ("CF DRIVERS", "CapEx % of PP&E")},
        {"label": "Purchase of Non-Marketable Equity Securities", "kind": "skip"},
        {"label": "Acquisition of Big Beverages",             "kind": "skip"},
        {"label": "Cash Flow from Investing",                 "kind": "subtotal"},
        {"label": "Finance Lease Payments",                   "kind": "zero"},
        {"label": "Proceeds from Exercise of Stock Options",  "kind": "flat"},
        {"label": "Proceeds from Issuance of Preferred Stock","kind": "flat"},
        {"label": "Preferred Dividends",                      "kind": "dividends_preferred",
         "driver": ("CF DRIVERS", "Preferred Dividends % of Preferred Balance")},
        {"label": "Proceeds from Issuance of Common Stock",   "kind": "flat"},
        {"label": "Share Repurchases",                        "kind": "dollar_driver",
         "driver": ("CF DRIVERS", "Share Repurchases $")},
        {"label": "Cash Flow from Financing",                 "kind": "subtotal"},
        {"label": "FX Effect on Cash",                        "kind": "flat"},
        {"label": "Net Change in Cash",                       "kind": "cf_net_change"},
        {"label": "Cash at Beginning of Period",              "kind": "cash_beg"},
        {"label": "Cash at End of Period",                    "kind": "cash_end"},
    ],
}


# ============================================================================
# Statement forecast builder
# ============================================================================

def _all_sheet_maps(wb) -> tuple[dict[str, dict[str, int]], dict[str, dict[str, int]]]:
    """Build period-col and row maps for ALL sheets in the workbook, not just
    SOURCE_SHEETS. Needed to resolve driver-sheet references."""
    period_col: dict[str, dict[str, int]] = {}
    row_map: dict[str, dict[str, int]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        periods: dict[str, int] = {}
        for col in range(2, (ws.max_column or 1) + 1):
            label = ws.cell(row=1, column=col).value
            if label:
                periods[str(label).strip()] = col
        period_col[sheet_name] = periods
        rows: dict[str, int] = {}
        for r in range(2, (ws.max_row or 1) + 1):
            label = ws.cell(row=r, column=1).value
            if label:
                rows[str(label).strip()] = r
        row_map[sheet_name] = rows
    return period_col, row_map


def _resolve_across(
    ref: tuple[str, str],
    period_label: str,
    period_col: dict[str, dict[str, int]],
    row_map: dict[str, dict[str, int]],
) -> tuple[str, int, int] | None:
    """Like _resolve, but for any sheet in the workbook."""
    sheet, label = ref
    rows = row_map.get(sheet, {})
    periods = period_col.get(sheet, {})
    row = rows.get(label)
    col = periods.get(period_label)
    if row is None or col is None:
        return None
    return (sheet, row, col)


def build_statement_forecast_formula(
    spec: dict,
    sheet: str,
    row: int,
    period_label: str,
    prev_period_label: str | None,
    period_col: dict[str, dict[str, int]],
    row_map: dict[str, dict[str, int]],
    assumptions_rows: dict[str, int],
) -> str | None:
    """Return a forecast Excel formula for this (sheet, row, period) or None
    to leave the cell blank. prev_period_label is None for the first forecast year."""
    kind = spec["kind"]

    # --- Helpers ---
    # CRITICAL: different sheets may have different historical column counts
    # (e.g. BS has 3 historicals, P&L/CF have 4), which offsets every column.
    # ALWAYS look up columns by period_label per target sheet — never use the
    # current sheet's column index for a cross-sheet reference.
    col_prev_here = period_col[sheet].get(prev_period_label) if prev_period_label else None

    def here_on(target_sheet: str) -> int | None:
        return period_col.get(target_sheet, {}).get(period_label)

    def prev_on(target_sheet: str) -> int | None:
        if prev_period_label is None:
            return None
        return period_col.get(target_sheet, {}).get(prev_period_label)

    def rowref(sheet_name: str, label: str, col: int | None) -> str | None:
        if col is None:
            return None
        r = row_map.get(sheet_name, {}).get(label)
        if r is None:
            return None
        return _cell_ref(sheet_name, r, col)

    def days_ref() -> str:
        return assumption_ref("Days in Year", assumptions_rows)

    def driver_ref_here() -> str | None:
        driver = spec.get("driver")
        if driver is None:
            return None
        sht, lbl = driver
        return rowref(sht, lbl, here_on(sht))

    if kind == "subtotal" or kind == "skip":
        return None

    if kind == "zero":
        return "=0"

    if kind == "flat":
        if col_prev_here is None:
            return None
        return f"={get_column_letter(col_prev_here)}{row}"

    if kind == "revenue_growth":
        if prev_period_label is None:
            return None
        prev_rev = rowref("ANNL P&L", "Net Sales / Revenue", prev_on("ANNL P&L"))
        drv = driver_ref_here()
        if prev_rev is None or drv is None:
            return None
        return f"={prev_rev}*(1+{drv})"

    if kind == "ratio_of_rev":
        drv = driver_ref_here()
        rev = rowref("ANNL P&L", "Net Sales / Revenue", here_on("ANNL P&L"))
        if drv is None or rev is None:
            return None
        return f"={drv}*{rev}"

    if kind == "ratio_of_cogs":
        drv = driver_ref_here()
        cogs = rowref("ANNL P&L", "COGS", here_on("ANNL P&L"))
        if drv is None or cogs is None:
            return None
        return f"={drv}*{cogs}"

    if kind == "days_driven_rev":
        drv = driver_ref_here()
        rev = rowref("ANNL P&L", "Net Sales / Revenue", here_on("ANNL P&L"))
        if drv is None or rev is None:
            return None
        return f"={drv}*{rev}/{days_ref()}"

    if kind == "days_driven_cogs":
        drv = driver_ref_here()
        cogs = rowref("ANNL P&L", "COGS", here_on("ANNL P&L"))
        if drv is None or cogs is None:
            return None
        return f"={drv}*{cogs}/{days_ref()}"

    if kind == "dollar_driver":
        drv = driver_ref_here()
        if drv is None:
            return None
        return f"={drv}"

    if kind == "tax":
        drv = driver_ref_here()
        pretax = rowref("ANNL P&L", "Pre-Tax Income (Loss)", here_on("ANNL P&L"))
        if drv is None or pretax is None:
            return None
        return f"={pretax}*{drv}"

    if kind == "cf_from_pl":
        pl = rowref("ANNL P&L", spec["pl_label"], here_on("ANNL P&L"))
        return f"={pl}" if pl else None

    if kind == "cf_from_cf":
        cf = rowref("CASH FLOW", spec["cf_label"], here_on("CASH FLOW"))
        return f"={cf}" if cf else None

    if kind == "cash_rollforward":
        if col_prev_here is None:
            return None
        prior = f"{get_column_letter(col_prev_here)}{row}"
        net_change = rowref("CASH FLOW", "Net Change in Cash", here_on("CASH FLOW"))
        if net_change is None:
            return None
        return f"={prior}+{net_change}"

    if kind == "pp_e_rollforward":
        if col_prev_here is None:
            return None
        prior = f"{get_column_letter(col_prev_here)}{row}"
        capex = rowref("CASH FLOW", "Purchase of PP&E", here_on("CASH FLOW"))
        da = rowref("CASH FLOW", "Depreciation & Amortization", here_on("CASH FLOW"))
        if capex is None or da is None:
            return None
        # PP&E[t] = prior - CF!CapEx (negative outflow subtracts to add back) - CF!D&A
        return f"={prior}-{capex}-{da}"

    if kind == "apic_rollforward":
        if col_prev_here is None:
            return None
        prior = f"{get_column_letter(col_prev_here)}{row}"
        sbc = rowref("CASH FLOW", "Stock-Based Compensation", here_on("CASH FLOW"))
        stock_opt = rowref("CASH FLOW", "Proceeds from Exercise of Stock Options", here_on("CASH FLOW"))
        terms = [prior]
        if sbc: terms.append(sbc)
        if stock_opt: terms.append(stock_opt)
        return "=" + "+".join(terms)

    if kind == "re_rollforward":
        if col_prev_here is None:
            return None
        prior = f"{get_column_letter(col_prev_here)}{row}"
        ni = rowref("ANNL P&L", "Net Income (Loss)", here_on("ANNL P&L"))
        pref = rowref("CASH FLOW", "Preferred Dividends", here_on("CASH FLOW"))
        common = rowref("CASH FLOW", "Common Dividends", here_on("CASH FLOW"))
        terms = [prior]
        if ni: terms.append(ni)
        if pref: terms.append(pref)  # Preferred stored negative; natural add reduces RE
        if common: terms.append(common)
        return "=" + "+".join(terms)

    if kind == "amortize":
        if col_prev_here is None:
            return None
        prior = f"{get_column_letter(col_prev_here)}{row}"
        amort = rowref("CASH FLOW", "Amortization of Deferred Other Costs", here_on("CASH FLOW"))
        if amort is None:
            return f"={prior}"
        return f"={prior}-{amort}"

    if kind == "ni_attrib_common":
        ni = rowref("ANNL P&L", "Net Income (Loss)", here_on("ANNL P&L"))
        pref = rowref("ANNL P&L", "Preferred Dividends", here_on("ANNL P&L"))
        alloc = rowref("ANNL P&L", "Income Allocated to Participating Preferred", here_on("ANNL P&L"))
        if ni is None:
            return None
        parts = [ni]
        if pref: parts.append(pref)
        if alloc: parts.append(alloc)
        return "=" + "+".join(parts)

    if kind == "cf_wc_asset":
        if prev_period_label is None:
            return None
        bs_label = spec["bs_label"]
        cur = rowref("BALANCE SHEET", bs_label, here_on("BALANCE SHEET"))
        prev = rowref("BALANCE SHEET", bs_label, prev_on("BALANCE SHEET"))
        if cur is None or prev is None:
            return None
        return f"=-({cur}-{prev})"

    if kind == "cf_wc_liability":
        if prev_period_label is None:
            return None
        bs_label = spec["bs_label"]
        cur = rowref("BALANCE SHEET", bs_label, here_on("BALANCE SHEET"))
        prev = rowref("BALANCE SHEET", bs_label, prev_on("BALANCE SHEET"))
        if cur is None or prev is None:
            return None
        return f"=({cur}-{prev})"

    if kind == "cf_wc_combined":
        if prev_period_label is None:
            return None
        cur_refs = []
        prev_refs = []
        for lbl in spec["bs_labels"]:
            cur = rowref("BALANCE SHEET", lbl, here_on("BALANCE SHEET"))
            prev = rowref("BALANCE SHEET", lbl, prev_on("BALANCE SHEET"))
            if cur and prev:
                cur_refs.append(cur)
                prev_refs.append(prev)
        if not cur_refs:
            return None
        sign = spec.get("sign", 1)
        lead = "=" if sign == 1 else "=-"
        return f"{lead}(({'+'.join(cur_refs)})-({'+'.join(prev_refs)}))"

    if kind == "capex":
        drv = driver_ref_here()
        pp_e_prev = rowref("BALANCE SHEET", "Net PP&E", prev_on("BALANCE SHEET"))
        if drv is None or pp_e_prev is None:
            return None
        return f"={drv}*{pp_e_prev}"

    if kind == "d_a":
        drv = driver_ref_here()
        pp_e_prev = rowref("BALANCE SHEET", "Net PP&E", prev_on("BALANCE SHEET"))
        if drv is None or pp_e_prev is None:
            return None
        return f"={drv}*{pp_e_prev}"

    if kind == "dividends_preferred":
        drv = driver_ref_here()
        pref = rowref("BALANCE SHEET", "Convertible Preferred Stock", here_on("BALANCE SHEET"))
        if drv is None or pref is None:
            return None
        return f"={drv}*{pref}"

    if kind == "dividends_common":
        drv = driver_ref_here()
        ni = rowref("ANNL P&L", "Net Income (Loss)", here_on("ANNL P&L"))
        if drv is None or ni is None:
            return None
        return f"={drv}*{ni}"

    if kind == "cf_net_change":
        cfo = rowref("CASH FLOW", "Cash Flow from Operations", here_on("CASH FLOW"))
        cfi = rowref("CASH FLOW", "Cash Flow from Investing", here_on("CASH FLOW"))
        cff = rowref("CASH FLOW", "Cash Flow from Financing", here_on("CASH FLOW"))
        fx = rowref("CASH FLOW", "FX Effect on Cash", here_on("CASH FLOW"))
        parts = [p for p in (cfo, cfi, cff, fx) if p is not None]
        if not parts:
            return None
        return "=" + "+".join(parts)

    if kind == "cash_beg":
        if prev_period_label is None:
            return None
        end = rowref("CASH FLOW", "Cash at End of Period", prev_on("CASH FLOW"))
        return f"={end}" if end else None

    if kind == "cash_end":
        beg = rowref("CASH FLOW", "Cash at Beginning of Period", here_on("CASH FLOW"))
        net_change = rowref("CASH FLOW", "Net Change in Cash", here_on("CASH FLOW"))
        if beg is None or net_change is None:
            return None
        return f"={beg}+{net_change}"

    if kind == "aoci_rollforward":
        if col_prev_here is None:
            return None
        prior = f"{get_column_letter(col_prev_here)}{row}"
        fx = rowref("CASH FLOW", "FX Effect on Cash", here_on("CASH FLOW"))
        if fx is None:
            return f"={prior}"
        return f"={prior}+{fx}"

    return None


def write_statement_forecasts(
    wb,
    assumptions_rows: dict[str, int],
) -> dict[str, int]:
    """Walk each statement's forecast columns and write forecast formulas per
    FORECAST_STATEMENT_SPECS. Returns per-sheet count of formulas written."""
    period_col, row_map = _all_sheet_maps(wb)

    counts: dict[str, int] = {}
    for sheet_name, specs in FORECAST_STATEMENT_SPECS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        periods = period_col.get(sheet_name, {})
        # Chronologically ordered labels in this sheet.
        # Historicals come first; forecast labels are FY####E.
        label_cols = sorted(periods.items(), key=lambda kv: kv[1])
        forecast_pairs = [(lbl, col) for lbl, col in label_cols if lbl.endswith("E")]
        if not forecast_pairs:
            continue

        # Build prior-period lookup: for each forecast label, the immediately-prior label in this sheet.
        prev_of: dict[str, str] = {}
        all_labels = [lbl for lbl, _col in label_cols]
        for i, lbl in enumerate(all_labels):
            if i > 0 and lbl.endswith("E"):
                prev_of[lbl] = all_labels[i - 1]
            elif lbl.endswith("E"):
                prev_of[lbl] = ""  # first forecast has no in-sheet prior; rare, won't happen normally

        # Row map for this specific sheet.
        sheet_rows = row_map.get(sheet_name, {})

        # Last historical column on this sheet — used to source number_format
        # so forecast cells render with the same accounting format as historicals.
        last_hist_col = None
        for lbl, c in label_cols:
            if not lbl.endswith("E"):
                last_hist_col = c

        n_written = 0
        for spec in specs:
            label = spec["label"]
            if label not in sheet_rows:
                continue
            row = sheet_rows[label]
            hist_fmt = None
            if last_hist_col is not None:
                fmt = ws.cell(row=row, column=last_hist_col).number_format
                if fmt and fmt != "General":
                    hist_fmt = fmt
            for period_label, col in forecast_pairs:
                # Propagate accounting format to every forecast cell in this row,
                # including subtotal rows whose formulas were written by model-write.
                if hist_fmt is not None:
                    ws.cell(row=row, column=col).number_format = hist_fmt
                prev_period_label = prev_of.get(period_label)
                if not prev_period_label:
                    prev_period_label = None
                formula = build_statement_forecast_formula(
                    spec, sheet_name, row, period_label, prev_period_label,
                    period_col, row_map, assumptions_rows,
                )
                if formula is None:
                    continue
                cell = ws.cell(row=row, column=col)
                # Preserve existing fill (e.g. FORECAST_FILL) but set the formula
                cell.value = formula
                n_written += 1
        counts[sheet_name] = n_written
    return counts


# ============================================================================
# Workbook driver
# ============================================================================

def reorder_sheets(wb) -> None:
    """Put ASSUMPTIONS first, then source sheets, then driver sheets."""
    desired = [ASSUMPTIONS_SHEET, *SOURCE_SHEETS, *DRIVER_SHEETS]
    ordered = [wb[name] for name in desired if name in wb.sheetnames]
    extras = [ws for ws in wb._sheets if ws not in ordered]
    wb._sheets = ordered + extras


def add_drivers(in_path: Path, out_path: Path) -> dict:
    wb = load_workbook(in_path)

    for src in SOURCE_SHEETS:
        if src not in wb.sheetnames:
            print(f"ERROR: source sheet {src!r} missing from {in_path}", file=sys.stderr)
            sys.exit(2)

    assumptions_rows = build_assumptions_sheet(wb)
    period_col_by_sheet, row_by_sheet = build_source_maps(wb)
    driver_periods = build_driver_period_labels(period_col_by_sheet)

    per_sheet: dict[str, int] = {}
    for sheet_name in DRIVER_SHEETS:
        specs = DRIVER_SPECS.get(sheet_name, [])
        written = write_driver_sheet(
            wb, sheet_name, specs, driver_periods,
            period_col_by_sheet, row_by_sheet, assumptions_rows,
        )
        per_sheet[sheet_name] = written

    # Statement forecasts — written AFTER driver sheets so driver cells exist
    # for reference lookups.
    forecast_counts = write_statement_forecasts(wb, assumptions_rows)

    reorder_sheets(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {
        "out_path": str(out_path),
        "period_labels": driver_periods,
        "per_sheet": per_sheet,
        "forecast_counts": forecast_counts,
        "assumptions_rows": assumptions_rows,
    }


# ============================================================================
# Main
# ============================================================================

def load_config(ticker_root: Path) -> dict:
    return json.loads((ticker_root / "config.json").read_text(encoding="utf-8"))


def main():
    parser = argparse.ArgumentParser(
        description="Add ASSUMPTIONS + IS/BS/CF driver tabs with historical + projection formulas.",
    )
    parser.add_argument("--ticker-root", required=True, type=Path)
    parser.add_argument("--in", dest="in_path", required=True, type=Path,
                        help="Input xlsx produced by model-write.")
    parser.add_argument("--out", dest="out_path", type=Path, default=None,
                        help="Output xlsx path. Defaults to --in (in-place).")
    args = parser.parse_args()

    config = load_config(args.ticker_root)
    _ = config.get("ticker")

    in_path = args.in_path
    out_path = args.out_path or in_path
    if not in_path.exists():
        print(f"ERROR: input xlsx not found: {in_path}", file=sys.stderr)
        sys.exit(2)

    report = add_drivers(in_path, out_path)

    print(f"Added ASSUMPTIONS + driver tabs + statement forecasts to {out_path}")
    print(f"  ASSUMPTIONS   : {len(report['assumptions_rows'])} constants")
    for sheet, n in report["per_sheet"].items():
        print(f"  {sheet:14s}: {n} driver formula cells")
    print(f"  Statement forecast formulas:")
    for sheet, n in report.get("forecast_counts", {}).items():
        print(f"    {sheet:14s}: {n} cells")
    print(f"  Period columns: {report['period_labels']}")


if __name__ == "__main__":
    main()
