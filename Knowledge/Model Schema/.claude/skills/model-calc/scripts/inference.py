"""
Driver-kind inference engine for the quarterly model-calc rebuild.

Walks every row in the workbook's QTR P&L / QTR BS / QTR CF sheets, looks up
each row's `model_label` in the merged library (generic + per-ticker), and
assigns a `DriverKind` based on `filing_section` + label heuristics.

The point of this module is to make adding a new ticker zero-config: as long
as a row's label is in the library (or carries a `filing_section` via the
ticker ledger), the inference engine picks the right driver kind. No
per-ticker DRIVER_SPECS dicts.

Heuristic policy: this module DOES use label-text matching (e.g. "Cash" /
"Receivable" / "Payable") because filing_section alone is too coarse to
distinguish an AR canonical from a Cash canonical when both sit in
`current_assets`. The label-text checks here match against canonical
`model_label`s only — these are stable strings under our control, NOT raw
filer rendering. So this is structural matching against the library, not
heuristic patching of filer idiosyncrasies (per the no-heuristic policy in
`feedback_structural_over_heuristic.md`).
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Optional

from openpyxl.worksheet.worksheet import Worksheet

from driver_models import DriverKind, DriverSpec


# ============================================================================
# Subtotal labels (DERIVED — workbook already has cascade/SUM formulas)
# Per-sheet, because the same label can mean different things across sheets.
# Notably "Net Income (Loss)" is a derived cascade subtotal on the IS but the
# starting line on CF (where it's a LINK_TO_IS, not derived).
# ============================================================================

DERIVED_LABELS_IS: set[str] = {
    "Gross Profit (Loss)",
    "Income (Loss) from Operations",
    "Pre-Tax Income (Loss)",
    "Net Income (Loss)",
    "Net Income (Loss) Including NCI",
    "Net Income (Loss) Less NCI",
    "Net Income (Loss) Attributable to Common Shareholders",
}

DERIVED_LABELS_BS: set[str] = {
    "Total Current Assets",
    "Total Non-Current Assets",
    "Total Assets",
    "Total Current Liabilities",
    "Total Non-Current Liabilities",
    "Total Liabilities",
    "Total Stockholders' Equity",
    "Total Mezzanine Equity",
    "Total Liabilities, Mezzanine & Stockholders' Equity",
}

DERIVED_LABELS_CF: set[str] = {
    "Cash Flow from Operations",
    "Cash Flow from Investing",
    "Cash Flow from Financing",
    "Net Change in Cash",
    "Cash at Beginning of Period",
    "Cash at End of Period",
}


# ============================================================================
# Library loader — merge generic + ticker
# ============================================================================

def _sheet_group(model_sheet: str | None) -> str:
    """Map a library entry's `model_sheet` to one of {"IS","BS","CF","?"}.
    Same convention as financials_schema.lookup._sheet_group, duplicated here
    to avoid importing the schema package."""
    if not model_sheet:
        return "?"
    ml = model_sheet.lower()
    if "balance sheet" in ml or ml == "qtr bs":
        return "BS"
    if "cash flow" in ml or ml == "qtr cf":
        return "CF"
    if "p&l" in ml:
        return "IS"
    return "?"


def load_label_section_map(library_path: Path, ticker_root: Path) -> dict[tuple[str, str], dict]:
    """Build a {(sheet_group, model_label): {...}} map merging the generic
    library and the ticker decisions ledger.

    Sheet group keys (IS / BS / CF) avoid the cross-sheet collision class:
    "Accounts Receivable" exists as both a BS canonical and a CF working-
    capital canonical — keying by sheet keeps each side correct.

    Ticker entries are read from both `mappings[]` (which uses `model_label`)
    and `new_rows[]` (which uses `new_row_label`)."""
    out: dict[tuple[str, str], dict] = {}

    def _add(entry: dict, label: str, source: str) -> None:
        sg = _sheet_group(entry.get("model_sheet"))
        # Generic library uses `filing_section`; ticker decisions ledger
        # historically uses `section`. Read both so ticker `new_rows[]`
        # entries (Distributor Termination Fees, Juice Transaction, etc.)
        # carry their section tag through to driver inference.
        out[(sg, label)] = {
            "filing_section":  entry.get("filing_section") or entry.get("section"),
            "model_sheet":     entry.get("model_sheet"),
            "sheet_group":     sg,
            "memo":            entry.get("memo", False),
            "sign_convention": entry.get("sign_convention"),
            "source":          source,
            "rule_id":         entry.get("rule_id"),
            # cf_delta_target: structural BS-row → CF-row coupling. Declared on
            # BS canonicals whose CF partner has a different canonical label
            # (e.g. PEP's BS "Accrued Expenses" routes ΔBS to CF "Accounts
            # Payable"). When unset, BS_DELTA wiring falls back to exact-label
            # match on the CF side, then to RESIDUAL_PLUG.
            "cf_delta_target": entry.get("cf_delta_target"),
            # parent_canonical: structural detail-of relationship. When set,
            # this canonical is a child whose forecast = ratio × parent_forecast
            # (RATIO_OF_PARENT). Read by infer_bs_kind to override the section
            # default. Resolved to a parent label downstream.
            "parent_canonical": entry.get("parent_canonical"),
        }

    if library_path.exists():
        gen = json.loads(library_path.read_text(encoding="utf-8"))
        for entry in gen.get("mappings", []):
            label = entry.get("model_label")
            if not label:
                continue
            _add(entry, label, "generic")

    ledger_path = ticker_root / "Financial Statements" / "decisions_ledger.json"
    if ledger_path.exists():
        ledger = json.loads(ledger_path.read_text(encoding="utf-8"))
        for entry in ledger.get("mappings", []):
            label = entry.get("model_label")
            if not label:
                continue
            _add(entry, label, "ticker_mapping")
        for entry in ledger.get("new_rows", []):
            label = entry.get("new_row_label") or entry.get("model_label")
            if not label:
                continue
            _add(entry, label, "ticker_new_row")

    # Resolve parent_canonical (rule_id) → parent's model_label, so downstream
    # forecast formulas can reference the parent row by label without another
    # rule_id lookup.
    rid_to_label: dict[str, str] = {}
    for v in out.values():
        rid = v.get("rule_id")
        if rid:
            rid_to_label[rid] = None
    # Reverse-walk: scan the same library/ledger sources to populate rid → label
    if library_path.exists():
        gen = json.loads(library_path.read_text(encoding="utf-8"))
        for entry in gen.get("mappings", []):
            if entry.get("rule_id"):
                rid_to_label[entry["rule_id"]] = entry.get("model_label")
    if ledger_path.exists():
        ledger = json.loads(ledger_path.read_text(encoding="utf-8"))
        for entry in ledger.get("mappings", []):
            if entry.get("rule_id"):
                rid_to_label[entry["rule_id"]] = entry.get("model_label")
        for entry in ledger.get("new_rows", []):
            if entry.get("rule_id"):
                rid_to_label[entry["rule_id"]] = entry.get("new_row_label") or entry.get("model_label")

    for v in out.values():
        parent_rid = v.get("parent_canonical")
        if parent_rid:
            v["parent_label"] = rid_to_label.get(parent_rid)

    return out


# ============================================================================
# Per-statement inference rules
# ============================================================================

def _is_skip_label(label: str) -> bool:
    """EPS and share-count rows skipped from forecasting."""
    s = label.lower()
    if " per share" in s or "per share" == s.strip() or s.endswith("eps"):
        return True
    # Match share-count rows (Weighted Avg Shares Outstanding) but NOT "Stock-Based Compensation"
    if "shares outstanding" in s or "shares issued" in s:
        return True
    return False


def _label_contains(label: str, *needles: str) -> bool:
    s = label.lower()
    return any(n.lower() in s for n in needles)


def infer_is_kind(label: str, section: Optional[str]) -> tuple[DriverKind, str]:
    """IS rows."""
    if label in DERIVED_LABELS_IS:
        return DriverKind.DERIVED, "IS subtotal/cascade row"
    if _is_skip_label(label):
        return DriverKind.SKIP, "EPS / share-count row — not forecast in v1"

    if section == "revenue_cost":
        if _label_contains(label, "revenue", "net sales"):
            return DriverKind.GROWTH, "Top-line revenue → YoY growth %"
        if _label_contains(label, "cogs", "cost of goods", "cost of sales"):
            return DriverKind.RATIO_OF_REV, "COGS → % of Revenue"
        return DriverKind.RATIO_OF_REV, "revenue_cost section default"

    if section == "operating_expenses":
        return DriverKind.RATIO_OF_REV, "Opex → % of Revenue"

    if section == "non_operating":
        return DriverKind.DOLLAR_INPUT, "Non-operating items → analyst dollar input per quarter"

    if section == "tax":
        return DriverKind.TAX_RATE, "Income Tax → effective rate × Pre-Tax"

    if section == "post_ni_deduction":
        return DriverKind.DOLLAR_INPUT, "NI attribution to NCI → dollar input"

    if section == "eps":
        # eps section contains both per-share metrics (skipped above by
        # _is_skip_label) AND dollar-value rows like Preferred Dividends and
        # Income Allocated to Participating Preferred. Those need driver kinds.
        if label == "Preferred Dividends":
            # Mirror the CF Pref Div forecast — single driver lives on the CF
            # side (analyst overrides once, IS row reads through). The IS row
            # is the deduction-from-common rendering of the same cash payment.
            return DriverKind.LINK_TO_CF, "IS Pref Div mirrors CF Pref Div (single shared driver)"
        if label == "Income Allocated to Participating Preferred":
            # Two-class-method allocation; not a cash flow, no CF counterpart.
            # HOLD_LAST default; analyst overrides per quarter via driver tab.
            return DriverKind.HOLD_LAST, "Income Allocated to Participating Preferred — HOLD_LAST default"
        return DriverKind.SKIP, "EPS section default — share counts / per-share metrics"

    return DriverKind.RATIO_OF_REV, f"IS fallback (section={section})"


def infer_bs_kind(label: str, section: Optional[str]) -> tuple[DriverKind, str]:
    """BS rows."""
    if label in DERIVED_LABELS_BS:
        return DriverKind.DERIVED, "BS subtotal row"

    if section == "current_assets":
        if _label_contains(label, "cash & cash equivalents", "cash and cash equivalents"):
            return DriverKind.ROLLFORWARD, "Cash → BS rollforward (prev + CF Net Change)"
        if _label_contains(label, "restricted cash"):
            # Restricted Cash held flat — only the primary Cash row rollforwards
            # via Net Change. Filers that combine Cash + Restricted on the CF
            # would otherwise double-count Net Change here.
            return DriverKind.HOLD_LAST, "Restricted Cash → hold at last historical (avoid Net Change double-count with primary Cash row)"
        if _label_contains(label, "short-term investment", "short term investment", "marketable securit"):
            return DriverKind.HOLD_LAST, "Short-Term Investments → hold at last historical"
        if _label_contains(label, "receivable"):
            return DriverKind.DSO_RATIO, "Receivable → DSO × Revenue"
        if _label_contains(label, "inventor"):
            return DriverKind.DIO_RATIO, "Inventory → DIO × COGS"
        return DriverKind.RATIO_OF_REV, "current_assets default → % of Revenue"

    if section == "non_current_assets":
        if _label_contains(label, "property, plant", "ppe", "pp&e"):
            return DriverKind.ROLLFORWARD, "PP&E → BS rollforward (prev + CapEx − D&A)"
        if _label_contains(label, "deferred tax asset"):
            return DriverKind.ROLLFORWARD, "DTA → BS rollforward (prev − CF Deferred Tax Provision)"
        if _label_contains(label, "goodwill", "intangible", "equity method"):
            return DriverKind.HOLD_LAST, "Long-lived asset → hold at last historical"
        return DriverKind.RATIO_OF_REV, "non_current_assets default → % of Revenue"

    if section == "current_liabilities":
        if _label_contains(label, "accounts payable") and not _label_contains(label, "accrued"):
            return DriverKind.DPO_RATIO, "Accounts Payable → DPO × COGS"
        if _label_contains(label, "debt", "borrowing", "current portion of long-term", "commercial paper"):
            return DriverKind.HOLD_LAST, "Debt obligation → hold at last historical (debt schedule out of v1 scope)"
        return DriverKind.RATIO_OF_REV, "current_liabilities default → % of Revenue"

    if section == "non_current_liabilities":
        if _label_contains(label, "long-term debt", "deferred tax liab", "pension", "opeb", "operating lease liab"):
            return DriverKind.HOLD_LAST, "Long-lived liability → hold at last historical"
        return DriverKind.RATIO_OF_REV, "non_current_liabilities default → % of Revenue"

    if section == "mezzanine":
        return DriverKind.HOLD_LAST, "Mezzanine equity → hold at last historical"

    if section == "equity":
        if _label_contains(label, "retained earnings", "accumulated deficit"):
            return DriverKind.ROLLFORWARD, "Retained Earnings → BS rollforward (prev + NI − Dividends)"
        if _label_contains(label, "additional paid-in capital", "apic"):
            return DriverKind.ROLLFORWARD, "APIC → BS rollforward (prev + SBC) — captures equity offset of SBC expense"
        if _label_contains(label, "accumulated other comprehensive"):
            return DriverKind.ROLLFORWARD, "AOCI → BS rollforward (prev + FX Effect) — captures translation gains/losses"
        return DriverKind.HOLD_LAST, "Equity component → hold at last historical (rollforwards beyond RE/APIC/AOCI deferred)"

    return DriverKind.HOLD_LAST, f"BS fallback (section={section})"


def infer_cf_kind(
    label: str,
    section: Optional[str],
    bs_wc_assets:   set[str] | None = None,
    bs_wc_liab:     set[str] | None = None,
) -> tuple[DriverKind, str]:
    """CF rows.

    Path B (BS-driven CF closure, 2026-04-28). Default for non-special CF
    rows is ZERO — not HOLD_LAST. Reasoning: every CF entry must be
    accounted for by a BS movement; CF rows held flat at last historical
    while their BS counterpart is HOLD_LAST violate ΔA = ΔL + ΔE. Only the
    following CF rows can be non-zero in forecast:

      - LINK_TO_IS: NI line.
      - DERIVED: section subtotals.
      - RATIO_OF_REV: D&A / SBC / ROU Amortization (non-cash add-backs that
        scale with business; matched by exact canonical labels) and CapEx
        (PP&E rollforward input).
      - PAYOUT_RATIO: Common Dividends (RE rollforward input).
      - BS_DELTA: matches a BS row in `bs_wc_assets` / `bs_wc_liab` (set of
        non-flat BS rows the engine tagged as ratio-driven).
      - RESIDUAL_PLUG: the "Other Operating Items" canonical absorbs leftover
        ΔBS for any non-flat BS rows that didn't get a 1:1 BS_DELTA partner.
      - Everything else → ZERO.

    `bs_wc_assets` / `bs_wc_liab` are the workbook's actual non-flat BS
    label sets (built once by infer_drivers and passed in)."""
    bs_wc_assets = bs_wc_assets or set()
    bs_wc_liab   = bs_wc_liab   or set()

    # LINK_TO_IS check FIRST: "Net Income (Loss)" appears as a subtotal label
    # on IS but is the *starting* line on CF.
    if _label_contains(label, "net income"):
        return DriverKind.LINK_TO_IS, "CF Net Income → equals IS Net Income for the quarter"

    if label in DERIVED_LABELS_CF:
        return DriverKind.DERIVED, "CF subtotal row"

    if section == "operating":
        if label == "Depreciation & Amortization":
            return DriverKind.RATIO_OF_REV, "D&A → % of Revenue (non-cash add-back, PP&E rollforward input)"
        if label == "Stock-Based Compensation":
            return DriverKind.RATIO_OF_REV, "SBC → % of Revenue (non-cash add-back)"
        if label == "Operating Lease ROU Asset Amortization":
            return DriverKind.RATIO_OF_REV, "ROU Amortization → % of Revenue (non-cash add-back)"
        # Structural NWC detection — match against the workbook's non-flat BS
        # label sets. Sign comes from which set matched (asset vs liability).
        if label in bs_wc_assets:
            return DriverKind.BS_DELTA, f"NWC asset (BS {label!r} non-flat) → -(BS[Q]-BS[Q-1])"
        if label in bs_wc_liab:
            return DriverKind.BS_DELTA, f"NWC liability (BS {label!r} non-flat) → +(BS[Q]-BS[Q-1])"
        # Residual plug: absorbs ΔBS of any non-flat BS rows that didn't get
        # a 1:1 BS_DELTA partner. Either "Other Operating Items" (GEN-CF-010,
        # used by PEP/PG/CELH) or "Net Change in Other Working Capital"
        # (GEN-CF-051, used by MNST) qualifies — both are filer-rendered
        # catch-alls for unspecified WC changes. Sources are attached in
        # `infer_drivers` after both sides are walked.
        if label in PLUG_LABELS:
            return DriverKind.RESIDUAL_PLUG, f"{label} → residual ΔBS plug for BS-balance closure"
        # Operating row with no BS coupling and no special role → ZERO. If a
        # non-zero forecast is needed, the BS row must be ROLLFORWARD or
        # ratio-driven so its delta has a CF counterpart.
        return DriverKind.ZERO, "Operating non-WC item with no BS coupling → 0 (BS closure)"

    if section == "investing":
        if label == "Purchase of PP&E":
            return DriverKind.RATIO_OF_REV, "CapEx → % of Revenue (PP&E rollforward input)"
        # All other investing rows (Acquisitions, Sale of Assets, ST
        # Investments, etc.) → ZERO unless their BS counterpart is
        # ROLLFORWARD. v1: HOLD_LAST BS rows ⇒ ZERO CF.
        return DriverKind.ZERO, "Investing non-CapEx → 0 (BS closure; HOLD_LAST BS counterpart)"

    if section == "financing":
        if label == "Common Dividends":
            return DriverKind.PAYOUT_RATIO, "Common Dividends → % of Net Income (RE rollforward input)"
        if label == "Preferred Dividends":
            return DriverKind.DOLLAR_INPUT, "Preferred Dividends → analyst dollar input (RE rollforward input)"
        # All other financing rows (Issue/Repay LTD, ST debt, share
        # repurchases, exercise of options, etc.) → ZERO. Non-zero forecasts
        # here would require BS LT Debt / Treasury / etc. to be ROLLFORWARD,
        # which is a v2 concern (debt schedule, buyback program).
        return DriverKind.ZERO, "Financing non-dividend → 0 (BS closure; HOLD_LAST BS counterpart)"

    if section == "cash_other":
        # FX Effect on Cash: held flat at last historical (small enough not
        # to break BS materially; proper treatment via AOCI rollforward
        # deferred to v2).
        return DriverKind.HOLD_LAST, "Cash-other (FX effect) → hold at last historical"

    return DriverKind.ZERO, f"CF fallback (section={section}) → 0 (BS closure)"


# ============================================================================
# Top-level: walk a workbook and emit DriverSpecs
# ============================================================================

SHEET_INFER = {
    "QTR P&L": ("IS", infer_is_kind),
    "QTR BS":  ("BS", infer_bs_kind),
    "QTR CF":  ("CF", infer_cf_kind),
}

# BS row kinds whose forecast value moves quarter-to-quarter via a ratio.
# These are the BS rows that *need* a CF counterpart to absorb their ΔBS —
# either via a 1:1 BS_DELTA match or via the RESIDUAL_PLUG.
BS_WC_RATIO_KINDS = {
    DriverKind.DSO_RATIO,
    DriverKind.DIO_RATIO,
    DriverKind.DPO_RATIO,
    DriverKind.RATIO_OF_REV,
}

ASSET_SECTIONS    = {"current_assets", "non_current_assets"}
LIAB_SECTIONS     = {"current_liabilities", "non_current_liabilities"}

# CF rows that act as the BS-closure residual plug — either canonical absorbs
# ΔBS for any non-flat BS rows that didn't get a 1:1 BS_DELTA partner.
PLUG_LABELS       = {"Other Operating Items", "Net Change in Other Working Capital"}


def _walk_sheet(wb, sheet_name: str) -> list[tuple[int, str]]:
    """Yield (excel_row, label) for every populated label cell on a sheet."""
    out: list[tuple[int, str]] = []
    if sheet_name not in wb.sheetnames:
        return out
    ws: Worksheet = wb[sheet_name]
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        label = str(v).strip()
        if label:
            out.append((r, label))
    return out


def _build_spec(
    sheet_name: str, sheet_group: str, excel_row: int, label: str,
    label_section_map: dict[tuple[str, str], dict],
    bs_wc_assets: set[str], bs_wc_liab: set[str],
    growth_basis_default: str = "yoy",
) -> DriverSpec:
    """Resolve one workbook row to a DriverSpec. Handles ticker `new_rows[]`
    override (ZERO default), then dispatches to the per-statement inference
    function. CF inference also receives the BS WC label sets."""
    entry = label_section_map.get((sheet_group, label), {})
    section = entry.get("filing_section")
    parent_label = entry.get("parent_label")

    # Ticker `new_rows[]` entries default to ZERO (per user 2026-04-27):
    # these are by construction ticker-specific one-offs (Juice Transaction,
    # Product Recall, TCJ Act, etc.). Hard structural linkages still win.
    if entry.get("source") == "ticker_new_row":
        kind, note = DriverKind.ZERO, "Ticker-specific line → forecast = 0 (non-recurring default)"
    elif parent_label:
        # Detail-of-parent (e.g. Raw Materials → Inventories Net): forecast =
        # ratio × parent_forecast. Parent keeps its own driver (DIO for Net).
        kind, note = (
            DriverKind.RATIO_OF_PARENT,
            f"Detail of {parent_label!r} → ratio × parent forecast",
        )
    elif sheet_group == "IS":
        kind, note = infer_is_kind(label, section)
    elif sheet_group == "BS":
        kind, note = infer_bs_kind(label, section)
    elif sheet_group == "CF":
        kind, note = infer_cf_kind(label, section, bs_wc_assets, bs_wc_liab)
    else:
        kind, note = DriverKind.HOLD_LAST, "unknown sheet group"

    spec = DriverSpec(
        sheet=sheet_name, excel_row=excel_row, label=label,
        kind=kind, filing_section=section, note=note,
        growth_basis=(growth_basis_default if kind == DriverKind.GROWTH else None),
    )

    # RATIO_OF_PARENT anchor: parent row on the same sheet, looked up by label.
    if kind == DriverKind.RATIO_OF_PARENT and parent_label:
        spec.parent_source = (sheet_name, parent_label)

    # LINK_TO_IS anchor (CF NI row).
    if kind == DriverKind.LINK_TO_IS:
        spec.is_link_source = ("QTR P&L", "Net Income (Loss)")

    # LINK_TO_CF anchor (IS Pref Div mirrors CF Pref Div).
    if kind == DriverKind.LINK_TO_CF:
        spec.cf_link_source = ("QTR CF", "Preferred Dividends")

    # ROLLFORWARD anchors. Sign convention: `inputs` are added, `outputs`
    # are subtracted. We rely on the FILER's signs on the CF (negative for
    # outflows, positive for add-backs) so the math composes naturally:
    #   Cash[Q] = Cash[Q-1] + CF_NetChange[Q]
    #   PP&E[Q] = PP&E[Q-1] - CF_CapEx[Q] - CF_DA[Q]
    #   RE[Q]   = RE[Q-1]   + IS_NI[Q] + CF_CommonDiv[Q] + CF_PreferredDiv[Q]
    #   DTA[Q]  = DTA[Q-1]  - CF_Provision[Q]
    if kind == DriverKind.ROLLFORWARD:
        if _label_contains(label, "cash"):
            spec.rollforward_inputs = [("QTR CF", "Net Change in Cash")]
        elif _label_contains(label, "property, plant", "ppe", "pp&e"):
            spec.rollforward_outputs = [
                ("QTR CF", "Purchase of PP&E"),
                ("QTR CF", "Depreciation & Amortization"),
            ]
        elif _label_contains(label, "retained earnings", "accumulated deficit"):
            spec.rollforward_inputs = [
                ("QTR P&L", "Net Income (Loss)"),
                ("QTR CF", "Common Dividends"),
                ("QTR CF", "Preferred Dividends"),
            ]
        elif _label_contains(label, "deferred tax asset"):
            spec.rollforward_outputs = [("QTR CF", "(Benefit) Provision for Deferred Income Taxes")]
        elif _label_contains(label, "additional paid-in capital", "apic"):
            # APIC[Q] = APIC[Q-1] + CF SBC. SBC on CF is positive add-back; the
            # corresponding BS movement is APIC growing by the same amount,
            # offsetting the NI hit in RE so total equity stays whole.
            spec.rollforward_inputs = [("QTR CF", "Stock-Based Compensation")]
        elif _label_contains(label, "accumulated other comprehensive"):
            # AOCI[Q] = AOCI[Q-1] + CF FX Effect on Cash. v1 simplification:
            # full OCI statement deferred (per `feedback_oci_separate_statement.md`),
            # so we proxy AOCI movement with FX Effect on Cash. Keeps Cash and
            # AOCI moving together; a proper fix lands when OCI is broken out.
            spec.rollforward_inputs = [("QTR CF", "FX Effect on Cash")]

    return spec


def _wire_bs_delta_and_plug(
    bs_specs: list[DriverSpec],
    cf_specs: list[DriverSpec],
    label_section_map: dict[tuple[str, str], dict],
) -> None:
    """Path B closure pass — runs after both sides are inferred.

    For every non-flat BS row (kind in BS_WC_RATIO_KINDS):
      1. Try the library's `cf_delta_target` declaration (BS canonical names
         the CF canonical that absorbs its delta — handles cases like PEP
         where BS lumps "Accrued Expenses" but CF splits "AP" + "Other CL").
      2. Fall back to exact-label match on the CF side.
      3. If matched, force the CF spec's kind to BS_DELTA and wire its
         bs_delta_source / bs_delta_is_liability fields.
      4. If unmatched, append the BS row to the RESIDUAL_PLUG spec's
         residual_plug_sources list (the "Other Operating Items" CF row
         absorbs the leftover ΔBS to keep the BS in balance)."""
    cf_specs_by_label: dict[str, DriverSpec] = {s.label: s for s in cf_specs}
    matched_bs_labels: set[str] = set()
    matched_cf_labels: set[str] = set()  # CF rows already claimed by a BS_DELTA

    # Two passes over BS rows: (1) exact label matches first — they're the
    # "natural" couplings and must claim their CF row before a cf_delta_target
    # declaration can fall back to that same CF row. (2) cf_delta_target
    # fallback for BS rows whose canonical label has no exact CF counterpart.
    for pass_kind in ("exact", "declared"):
        for bs in bs_specs:
            if bs.kind not in BS_WC_RATIO_KINDS:
                continue
            if bs.label in matched_bs_labels:
                continue
            is_liab  = bs.filing_section in LIAB_SECTIONS
            bs_entry = label_section_map.get(("BS", bs.label), {})

            cf_match: DriverSpec | None = None
            note_src = ""
            if pass_kind == "exact" and bs.label in cf_specs_by_label:
                cf_match = cf_specs_by_label[bs.label]
                note_src = "exact label match"
            elif pass_kind == "declared":
                cf_target_label = bs_entry.get("cf_delta_target")
                if cf_target_label and cf_target_label in cf_specs_by_label:
                    cf_match = cf_specs_by_label[cf_target_label]
                    note_src = f"cf_delta_target → {cf_target_label!r}"

            if cf_match is None:
                continue
            # Skip if this CF row is already claimed by another BS row in pass 1.
            if cf_match.label in matched_cf_labels:
                continue
            # Don't overwrite a CF row that's already a structurally-distinct kind
            # (LINK_TO_IS, DERIVED, ROLLFORWARD anchor for someone else). Only
            # override rows currently tagged ZERO, HOLD_LAST, or the preliminary
            # BS_DELTA tag from infer_cf_kind.
            if cf_match.kind not in (DriverKind.ZERO, DriverKind.HOLD_LAST, DriverKind.BS_DELTA):
                continue
            cf_match.kind = DriverKind.BS_DELTA
            cf_match.bs_delta_source = ("QTR BS", bs.label)
            cf_match.bs_delta_is_liability = is_liab
            cf_match.note = (
                f"BS_DELTA from BS {bs.label!r} ({note_src}) → "
                f"{'+' if is_liab else '-'}(BS[Q]-BS[Q-1])"
            )
            matched_bs_labels.add(bs.label)
            matched_cf_labels.add(cf_match.label)

    # Protect rollforward-anchor CF rows from being ZEROed. Any CF label
    # referenced as input or output of a BS ROLLFORWARD spec must carry a
    # meaningful forecast value — otherwise the rollforward formula reads 0
    # and the BS row stays flat (which silently breaks the financial story
    # even when the BS arithmetic still balances). Upgrade ZERO → HOLD_LAST.
    cf_rollforward_anchors: set[str] = set()
    for bs in bs_specs:
        if bs.kind != DriverKind.ROLLFORWARD:
            continue
        for (sheet, lbl) in (list(bs.rollforward_inputs) + list(bs.rollforward_outputs)):
            if sheet == "QTR CF":
                cf_rollforward_anchors.add(lbl)
    for cf in cf_specs:
        if cf.label in cf_rollforward_anchors and cf.kind == DriverKind.ZERO:
            cf.kind = DriverKind.HOLD_LAST
            cf.note = (
                f"Rollforward anchor (input/output of a BS ROLLFORWARD spec) → "
                f"HOLD_LAST (was ZERO default; would have nullified the BS rollforward)"
            )

    # Non-cash CF add-backs (RATIO_OF_REV in operating section) without a BS
    # rollforward partner break BS balance: NI is reduced by the expense
    # (RE drops), CFO adds it back (Cash rises), but no BS row absorbs the
    # equity-side or asset-side counterpart → +Δ gap each quarter. PEP's ROU
    # Amortization is the canonical example (PEP doesn't break out ROU asset
    # on BS; it's folded into Other NCA which is RATIO_OF_REV not rollforward).
    # Demote to ZERO so the expense flows through Cash naturally without a
    # phantom add-back.
    for cf in cf_specs:
        if (cf.filing_section == "operating"
                and cf.kind == DriverKind.RATIO_OF_REV
                and cf.label not in cf_rollforward_anchors):
            cf.kind = DriverKind.ZERO
            cf.note = (
                f"Non-cash add-back without BS rollforward partner → ZERO "
                f"(would otherwise create phantom cash inflow each quarter)"
            )

    # Attach unmatched non-flat BS rows to the RESIDUAL_PLUG spec.
    plug = next((s for s in cf_specs if s.kind == DriverKind.RESIDUAL_PLUG), None)
    if plug is None:
        # No "Other Operating Items" canonical present — BS will not balance
        # for tickers with unmatched WC rows. (Handled with a stderr warning
        # by the caller; not raised here so the rest of the pipeline can run.)
        return
    sources: list[tuple[str, str, bool]] = []
    for bs in bs_specs:
        if bs.kind not in BS_WC_RATIO_KINDS:
            continue
        if bs.label in matched_bs_labels:
            continue
        sources.append(("QTR BS", bs.label, bs.filing_section in LIAB_SECTIONS))
    plug.residual_plug_sources = sources
    if sources:
        plug.note = (
            f"Residual plug — absorbs ΔBS for {len(sources)} unmatched non-flat BS rows: "
            + ", ".join(f"{lbl}{'(L)' if is_liab else '(A)'}" for (_, lbl, is_liab) in sources)
        )


def infer_drivers(wb, label_section_map: dict[tuple[str, str], dict],
                  growth_basis_default: str = "yoy") -> list[DriverSpec]:
    """Walk the QTR sheets and emit a DriverSpec per data row.

    Order matters now: BS specs FIRST so we can build the WC label sets
    before CF inference; then CF inference uses those sets to identify
    BS_DELTA candidates. After both sides are walked, run the Path B
    closure pass to wire `cf_delta_target`-declared couplings and attach
    residual plug sources for unmatched non-flat BS rows."""
    # --- Pass 1: BS specs ---
    bs_specs: list[DriverSpec] = []
    for r, label in _walk_sheet(wb, "QTR BS"):
        bs_specs.append(_build_spec(
            "QTR BS", "BS", r, label, label_section_map,
            bs_wc_assets=set(), bs_wc_liab=set(),
            growth_basis_default=growth_basis_default,
        ))

    bs_wc_assets: set[str] = {
        s.label for s in bs_specs
        if s.kind in BS_WC_RATIO_KINDS and s.filing_section in ASSET_SECTIONS
    }
    bs_wc_liab: set[str] = {
        s.label for s in bs_specs
        if s.kind in BS_WC_RATIO_KINDS and s.filing_section in LIAB_SECTIONS
    }

    # --- Pass 2: IS specs ---
    is_specs: list[DriverSpec] = []
    for r, label in _walk_sheet(wb, "QTR P&L"):
        is_specs.append(_build_spec(
            "QTR P&L", "IS", r, label, label_section_map,
            bs_wc_assets, bs_wc_liab,
            growth_basis_default=growth_basis_default,
        ))

    # --- Pass 3: CF specs (preliminary kinds; BS_DELTA finalized in Pass 4) ---
    cf_specs: list[DriverSpec] = []
    for r, label in _walk_sheet(wb, "QTR CF"):
        cf_specs.append(_build_spec(
            "QTR CF", "CF", r, label, label_section_map,
            bs_wc_assets, bs_wc_liab,
            growth_basis_default=growth_basis_default,
        ))

    # --- Pass 4: closure — wire BS_DELTA via cf_delta_target + residual plug ---
    _wire_bs_delta_and_plug(bs_specs, cf_specs, label_section_map)

    # Final pass: BS rows whose kind is not ratio-driven or rollforward but
    # still have a non-trivial ΔBS (most importantly, ticker_new_row → ZERO
    # rows that step down from a non-zero last historical to 0 in forecast
    # Q1) need their delta absorbed somewhere. Add them to the residual plug
    # as well — HOLD_LAST contributes Δ=0 by construction (no harm); ZERO
    # contributes a one-time boundary delta that the plug can absorb.
    plug = next((s for s in cf_specs if s.kind == DriverKind.RESIDUAL_PLUG), None)
    if plug is not None:
        already = {(sh, lbl) for (sh, lbl, _) in plug.residual_plug_sources}
        for bs in bs_specs:
            if bs.kind in (DriverKind.DERIVED, DriverKind.ROLLFORWARD):
                continue
            if bs.kind in BS_WC_RATIO_KINDS:
                continue  # already considered in main plug pass
            key = ("QTR BS", bs.label)
            if key in already:
                continue
            if bs.filing_section not in (ASSET_SECTIONS | LIAB_SECTIONS | {"equity", "mezzanine"}):
                continue
            # Equity / mezzanine ZERO+HOLD_LAST rows go on the liability side of
            # the BS identity (ΔA = ΔL + ΔE, so equity drops absorb like liab
            # drops in the plug). Catches things like PG's ESOP reserve and
            # CELH's mezzanine memo rows that step from non-zero historical
            # to 0 in forecast Q1.
            is_liab = bs.filing_section in LIAB_SECTIONS or bs.filing_section in {"equity", "mezzanine"}
            plug.residual_plug_sources.append(("QTR BS", bs.label, is_liab))

    return is_specs + bs_specs + cf_specs


