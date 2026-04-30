"""
model-calc — quarterly-first forecasting layer over a model-write workbook.

Fresh rebuild (2026-04-27 PEP-onboarding follow-on). The prior calc.py is
preserved at `_calc_legacy.py` for reference.

Architecture:
  - Drivers + forecasts operate on the QTR sheets (post-model-qtr-derive).
  - ANNL forecast cells aggregate from QTR: =SUM(Q1+Q2+Q3+Q4) for IS/CF;
    =QTR Q4 for BS.
  - Per-canonical driver inference (inference.py) replaces hand-curated
    DRIVER_SPECS. Each row gets one DriverKind from the inference engine.
  - User inputs live as yellow cells on driver tabs (per forecast quarter).
  - Forecast horizon: 5 fiscal years from the year of the last historical
    quarter. Quarters from (last_historical + 1) through Q4 of (year + 4).

CLI:
    python calc.py --in <workbook.xlsx> --ticker-root <path>            # full build
    python calc.py --in <workbook.xlsx> --ticker-root <path> --inspect  # dry-run
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Force utf-8 stdout for Δ / em-dash safety on Windows cp1252.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# Local imports (sibling modules in the same scripts/ directory)
sys.path.insert(0, str(Path(__file__).parent))
from inference import infer_drivers, load_label_section_map     # noqa: E402
from driver_models import DriverKind, DriverSpec                # noqa: E402

USER_HOME = Path.home()
DEFAULT_LIBRARY = USER_HOME / "Desktop" / "Brain" / "Knowledge" / "Model Schema" / "pattern_libraries" / "generic_line_item_mappings.json"

# ============================================================================
# Constants
# ============================================================================

N_FORECAST_YEARS = 5
DAYS_PER_QUARTER = 91

ASSUMPTIONS_SHEET = "ASSUMPTIONS"
DRIVER_SHEET_FOR_GROUP = {"IS": "IS DRIVERS", "BS": "BS DRIVERS", "CF": "CF DRIVERS"}
SOURCE_SHEET_FOR_GROUP = {"IS": "QTR P&L", "BS": "QTR BS", "CF": "QTR CF"}
YTD_SHEET_FOR_GROUP    = {"IS": "YTD P&L", "BS": "YTD BS", "CF": "YTD CF"}
ANNL_SHEET_FOR_GROUP   = {"IS": "ANNL P&L", "BS": "BALANCE SHEET", "CF": "CASH FLOW"}

# Format constants
LINE_ITEM_FMT  = '#,##0;(#,##0);"--"'
SUBTOTAL_FMT   = '$#,##0_);($#,##0);"$--"_)'
PERCENT_FMT    = '0.0%;(0.0%);"--"'
DAYS_FMT       = '0.0_);(0.0)'
DOLLAR_FMT     = '$#,##0_);($#,##0)'

HEADER_FONT    = Font(bold=True, color="FFFFFF")
HEADER_FILL    = PatternFill(fill_type="solid", start_color="203864", end_color="203864")
HEADER_ALIGN   = Alignment(horizontal="center")
LABEL_FONT     = Font(bold=True)
INPUT_FILL     = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")  # yellow user-input
FORECAST_FILL  = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")  # gray forecast tint

# Header parsers
QTR_HDR_RE  = re.compile(r"^Q([1-4])\s+FY(\d{4})E?$")
ANNL_HDR_RE = re.compile(r"^FY(\d{4})E?$")


def _prior_qtr(y: int, q: int) -> tuple[int, int]:
    """Period (year, quarter) immediately preceding (y, q)."""
    return (y, q - 1) if q > 1 else (y - 1, 4)

# Every non-SKIP / non-DERIVED kind gets a row on the appropriate DRIVERS tab.
# Source-sheet forecast cells then just lookup `=DRIVERS!{col}{driver_row}` —
# all forecast logic centralizes on the driver tabs (single place to override).
DRIVER_TAB_KINDS = {
    DriverKind.GROWTH,
    DriverKind.RATIO_OF_REV,
    DriverKind.RATIO_OF_COGS,
    DriverKind.DSO_RATIO,
    DriverKind.DIO_RATIO,
    DriverKind.DPO_RATIO,
    DriverKind.RATIO_OF_PARENT,
    DriverKind.TAX_RATE,
    DriverKind.PAYOUT_RATIO,
    # Value/derivation kinds — also go on driver tab so user has one place
    DriverKind.HOLD_LAST,
    DriverKind.ZERO,
    DriverKind.DOLLAR_INPUT,
    DriverKind.LINK_TO_IS,
    DriverKind.LINK_TO_CF,
    DriverKind.BS_DELTA,
    DriverKind.ROLLFORWARD,
    DriverKind.RESIDUAL_PLUG,
}

# Kinds where the forecast value is computed by a FORMULA on the driver tab
# (no user input — derived from other cells). Cell is not yellow-tinted.
COMPUTED_KINDS = {
    DriverKind.LINK_TO_IS,
    DriverKind.LINK_TO_CF,
    DriverKind.BS_DELTA,
    DriverKind.ROLLFORWARD,
    DriverKind.RESIDUAL_PLUG,
}

# Kinds with a user-overridable forecast (yellow input cells).
# All ratio/growth kinds + HOLD_LAST (default = last historical = straight-line)
# + ZERO + DOLLAR_INPUT.
USER_INPUT_KINDS = {
    DriverKind.GROWTH,
    DriverKind.RATIO_OF_REV,
    DriverKind.RATIO_OF_COGS,
    DriverKind.DSO_RATIO,
    DriverKind.DIO_RATIO,
    DriverKind.DPO_RATIO,
    DriverKind.RATIO_OF_PARENT,
    DriverKind.TAX_RATE,
    DriverKind.PAYOUT_RATIO,
    DriverKind.HOLD_LAST,
    DriverKind.ZERO,
    DriverKind.DOLLAR_INPUT,
}


# ============================================================================
# Workbook indexers
# ============================================================================

def parse_qtr_header_cols(ws) -> dict[tuple[int, int], int]:
    """Return {(year, quarter): col_idx} from a QTR-style header row."""
    out: dict[tuple[int, int], int] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        m = QTR_HDR_RE.match(str(cell.value).strip())
        if m:
            q = int(m.group(1))
            y = int(m.group(2))
            out[(y, q)] = cell.column
    return out


def parse_annl_header_cols(ws) -> dict[int, int]:
    """Return {fiscal_year: col_idx} for ANNL-style header (handles FY2024 + FY2024E)."""
    out: dict[int, int] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        s = str(cell.value).strip()
        m = ANNL_HDR_RE.match(s)
        if m:
            y = int(m.group(1))
            # E suffix wins (we want forecast-column slot for years that have both)
            if s.endswith("E") or y not in out:
                out[y] = cell.column
    return out


def find_row_by_label(ws, target_label: str) -> int | None:
    """Find row index whose column-A label equals target_label."""
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() == target_label:
            return r
    return None


# ============================================================================
# Forecast horizon computation
# ============================================================================

def compute_forecast_periods(wb, n_years: int = N_FORECAST_YEARS) -> list[tuple[int, int]]:
    """Return chronological list of forecast (year, quarter) tuples.

    Anchor: the latest historical quarter on QTR P&L (or any QTR sheet that
    has data). Forecast covers from (anchor + 1 quarter) through Q4 of
    (anchor_year + n_years). For a Q4-anchor ticker (CELH, last hist =
    Q4 FY2025) this gives a clean 5 full forecast years = 20 quarters
    (Q1 FY2026E .. Q4 FY2030E). For mid-fiscal-year anchors (PEP last hist
    Q1 FY2026, PG Q2 FY2026), the forecast covers the partial fiscal year +
    5 full years (PEP=23 quarters Q2 FY2026E..Q4 FY2031E; PG=22 quarters
    Q3 FY2026E..Q4 FY2031E). Last year always ends at Q4 so ANNL aggregation
    (=SUM(Q1..Q4)) produces a complete fiscal year."""
    last_year, last_q = 0, 0
    for sheet in ("QTR P&L", "QTR BS", "QTR CF"):
        if sheet not in wb.sheetnames:
            continue
        cols = parse_qtr_header_cols(wb[sheet])
        for (y, q) in cols:
            if (y, q) > (last_year, last_q):
                last_year, last_q = y, q
    if last_year == 0:
        return []

    last_forecast_year = last_year + n_years
    out: list[tuple[int, int]] = []
    y, q = last_year, last_q
    while True:
        q += 1
        if q > 4:
            q = 1
            y += 1
        if y > last_forecast_year:
            break
        out.append((y, q))
    return out


# ============================================================================
# Forecast column extension on QTR sheets
# ============================================================================

def extend_qtr_with_forecast_cols(ws, forecast_periods: list[tuple[int, int]]) -> dict[tuple[int, int], int]:
    """Append forecast quarter columns at the right edge of `ws`. Returns
    {(year, quarter): excel_col} for forecast cols only."""
    existing = parse_qtr_header_cols(ws)
    next_col = ws.max_column + 1 if ws.max_column else 2
    out: dict[tuple[int, int], int] = {}
    for (y, q) in forecast_periods:
        if (y, q) in existing:
            out[(y, q)] = existing[(y, q)]
            continue
        c = ws.cell(row=1, column=next_col, value=f"Q{q} FY{y}E")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(next_col)].width = 13
        out[(y, q)] = next_col
        next_col += 1
    return out


# ============================================================================
# ASSUMPTIONS tab
# ============================================================================

def build_assumptions_tab(wb) -> dict[str, int]:
    """Create / overwrite the ASSUMPTIONS sheet. Returns {label: row}."""
    if ASSUMPTIONS_SHEET in wb.sheetnames:
        del wb[ASSUMPTIONS_SHEET]
    ws = wb.create_sheet(ASSUMPTIONS_SHEET, 0)  # leftmost

    # Headers
    for col_idx, label in enumerate(("Assumption", "Value", "Note"), start=1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN

    rows: dict[str, int] = {}
    entries = [
        ("Days in Quarter", DAYS_PER_QUARTER, "Used by DSO/DIO/DPO ratios. Approximation; varies 89-92 by quarter."),
    ]
    for i, (label, value, note) in enumerate(entries, start=2):
        ws.cell(row=i, column=1, value=label).font = LABEL_FONT
        v = ws.cell(row=i, column=2, value=value)
        v.fill = INPUT_FILL
        ws.cell(row=i, column=3, value=note)
        rows[label] = i

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A2"
    return rows


def assumption_ref(label: str, rows: dict[str, int]) -> str:
    return f"{ASSUMPTIONS_SHEET}!$B${rows[label]}"


# ============================================================================
# Driver tab builder
# ============================================================================

def build_driver_tab(
    wb,
    group: str,
    specs: list[DriverSpec],
    qtr_hist_cols_by_sheet: dict[str, dict[tuple[int, int], int]],
    qtr_fcst_cols_by_sheet: dict[str, dict[tuple[int, int], int]],
    rev_row: int | None,
    cogs_row: int | None,
    pretax_row: int | None,
    ni_row_is: int | None,
    div_row_cf: int | None,
    assump_rows: dict[str, int],
) -> dict[str, int]:
    """Build IS DRIVERS / BS DRIVERS / CF DRIVERS sheet for one group.

    Every non-SKIP / non-DERIVED row on the source sheet gets a corresponding
    driver-tab row. For:
      - ratio/growth/days kinds: historical = computed ratio formula; forecast
        = yellow input (default = last historical for straight-line)
      - HOLD_LAST / ZERO / DOLLAR_INPUT: historical = mirror of source value;
        forecast = yellow input (default = last historical for HOLD_LAST,
        zero for ZERO, blank for DOLLAR_INPUT)
      - LINK_TO_IS / BS_DELTA / ROLLFORWARD: historical = mirror of source;
        forecast = derivation formula (no user input — computed)

    Returns ({spec_label: driver_row}, {(year, quarter): driver_col}) so
    source-sheet writers can lookup `=DRIVERS!{col}{driver_row}` for their
    forecast cells. The col map is needed because IS DRIVERS adds a Basis
    column at B, shifting period cols by one relative to the source sheet."""
    sheet_name = DRIVER_SHEET_FOR_GROUP[group]
    src_sheet  = SOURCE_SHEET_FOR_GROUP[group]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    qtr_hist_cols = qtr_hist_cols_by_sheet.get(src_sheet, {})
    qtr_fcst_cols = qtr_fcst_cols_by_sheet.get(src_sheet, {})

    group_specs = [s for s in specs if s.sheet == src_sheet and s.kind in DRIVER_TAB_KINDS]
    if not group_specs:
        return {}

    # --- Layout: A=label; (optional) B=Basis for GROWTH-bearing groups;
    #     period cols start at column 3 (or 2 if no Basis col). ---
    has_basis_col = any(s.kind == DriverKind.GROWTH for s in group_specs)
    basis_col_idx = 2 if has_basis_col else None
    period_start_col = 3 if has_basis_col else 2

    all_cols = sorted(qtr_hist_cols.keys()) + sorted(qtr_fcst_cols.keys())
    col_idx_by_period: dict[tuple[int, int], int] = {}
    ws.cell(row=1, column=1, value="Driver").font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    if has_basis_col:
        bc = ws.cell(row=1, column=basis_col_idx, value="Basis")
        bc.font = HEADER_FONT
        bc.fill = HEADER_FILL
        bc.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(basis_col_idx)].width = 8
    for i, (y, q) in enumerate(all_cols, start=period_start_col):
        is_fcst = (y, q) in qtr_fcst_cols
        label = f"Q{q} FY{y}E" if is_fcst else f"Q{q} FY{y}"
        c = ws.cell(row=1, column=i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        col_idx_by_period[(y, q)] = i
        ws.column_dimensions[get_column_letter(i)].width = 13

    # Single dropdown validator covering every GROWTH row's Basis cell.
    basis_dv = None
    if has_basis_col:
        basis_dv = DataValidation(type="list", formula1='"YoY,QoQ"', allow_blank=False)
        basis_dv.error = "Basis must be YoY or QoQ"
        basis_dv.errorTitle = "Invalid basis"
        ws.add_data_validation(basis_dv)

    # --- One row per spec ---
    driver_row_by_label: dict[str, int] = {}
    days_ref = assumption_ref("Days in Quarter", assump_rows)
    last_hist_period = max(qtr_hist_cols.keys()) if qtr_hist_cols else None
    last_hist_col = col_idx_by_period[last_hist_period] if last_hist_period else None

    # Lookup table for RATIO_OF_PARENT children: parent's row on the source
    # sheet (e.g. Inventories Net's QTR BS row) — looked up by label.
    parent_src_row_by_label: dict[str, int] = {
        ps.label: ps.excel_row for ps in group_specs
    }

    for r_offset, s in enumerate(group_specs, start=2):
        driver_label = _driver_row_label(s)
        ws.cell(row=r_offset, column=1, value=driver_label).font = LABEL_FONT
        driver_row_by_label[s.label] = r_offset

        # Basis cell (column B) — populated for GROWTH rows only.
        basis_cell_ref: str | None = None
        if has_basis_col and s.kind == DriverKind.GROWTH:
            basis_letter = get_column_letter(basis_col_idx)
            display = "QoQ" if (s.growth_basis or "yoy").lower() == "qoq" else "YoY"
            bc = ws.cell(row=r_offset, column=basis_col_idx, value=display)
            bc.alignment = HEADER_ALIGN
            bc.fill = INPUT_FILL  # yellow — analyst-editable
            if basis_dv is not None:
                basis_dv.add(f"{basis_letter}{r_offset}")
            basis_cell_ref = f"${basis_letter}{r_offset}"

        # --- Historical cells: ratio kinds compute; everything else mirrors
        #     the source-sheet value so the historical driver row reads as the
        #     actual historical value (= the source-sheet historical cell).
        # P&L period→col lookup (for ratio refs to Revenue / COGS / Pre-Tax /
        # NI on QTR P&L). Required because BS and P&L can have different
        # historical period coverage — e.g. CELH BS starts Q4 FY2022 while
        # P&L starts Q1 FY2022, so reusing the BS row's col index for a P&L
        # reference would land on a quarter 3 cells too early.
        pl_hist_cols = qtr_hist_cols_by_sheet.get("QTR P&L", {})
        for (y, q), col in qtr_hist_cols.items():
            target_col = col_idx_by_period.get((y, q))
            if target_col is None:
                continue
            prev_qtr_period = _prior_qtr(y, q)
            parent_src_row = None
            if s.kind == DriverKind.RATIO_OF_PARENT and s.parent_source:
                parent_src_row = parent_src_row_by_label.get(s.parent_source[1])
            formula, fmt = _historical_driver_formula(
                s.kind, s.excel_row, col, src_sheet,
                rev_row=rev_row, cogs_row=cogs_row, pretax_row=pretax_row,
                ni_row_is=ni_row_is, div_row_cf=div_row_cf,
                days_ref=days_ref,
                prior_year_col=qtr_hist_cols.get((y - 1, q)),
                prior_qtr_col=qtr_hist_cols.get(prev_qtr_period),
                basis_cell_ref=basis_cell_ref,
                pl_col=pl_hist_cols.get((y, q)),
                parent_src_row=parent_src_row,
            )
            if formula is None:
                continue
            cell = ws.cell(row=r_offset, column=target_col, value=formula)
            cell.number_format = fmt

        # --- Forecast cells ---
        for (y, q), _ in sorted(qtr_fcst_cols.items()):
            target_col = col_idx_by_period[(y, q)]
            target_letter = get_column_letter(target_col)

            if s.kind in COMPUTED_KINDS:
                # Pure derivation — no user input. Compute via cross-sheet
                # references (link to IS NI, BS delta, or rollforward chain).
                formula = _computed_forecast_formula(
                    s, y, q, target_col, col_idx_by_period, wb,
                    qtr_hist_cols_by_sheet=qtr_hist_cols_by_sheet,
                    qtr_fcst_cols_by_sheet=qtr_fcst_cols_by_sheet,
                    ni_row_is=ni_row_is,
                    driver_row=r_offset,
                )
                if formula is not None:
                    cell = ws.cell(row=r_offset, column=target_col, value=formula)
                    cell.number_format = _driver_format(s.kind)
            elif s.kind == DriverKind.ZERO:
                # Yellow input defaulted to 0 — user can override.
                cell = ws.cell(row=r_offset, column=target_col, value=0)
                cell.fill = INPUT_FILL
                cell.number_format = _driver_format(s.kind)
            elif s.kind == DriverKind.DOLLAR_INPUT:
                # Yellow input chained off immediately prior period so a single
                # override propagates forward (analyst types Q1 once, Q2..Qn
                # follow until another override).
                prev_col = target_col - 1
                if prev_col >= period_start_col:
                    cell = ws.cell(row=r_offset, column=target_col,
                                   value=f"={get_column_letter(prev_col)}{r_offset}")
                else:
                    cell = ws.cell(row=r_offset, column=target_col, value=0)
                cell.fill = INPUT_FILL
                cell.number_format = _driver_format(s.kind)
            else:
                # Ratio / growth / HOLD_LAST kinds: yellow input default =
                # immediately prior period's driver cell. Chains forward so
                # overriding any one cell carries through every subsequent
                # cell until another override.
                prev_col = target_col - 1
                if prev_col >= period_start_col:
                    formula = f"={get_column_letter(prev_col)}{r_offset}"
                else:
                    formula = 0
                cell = ws.cell(row=r_offset, column=target_col, value=formula)
                cell.fill = INPUT_FILL
                cell.number_format = _driver_format(s.kind)

    ws.column_dimensions["A"].width = 38
    ws.freeze_panes = f"{get_column_letter(period_start_col)}2"
    return driver_row_by_label, col_idx_by_period


def _driver_row_label(s: DriverSpec) -> str:
    """Label that appears in column A on the driver tab. Ratio/growth/days
    kinds get a suffix describing what the driver IS (since they're a derived
    metric, not the line item itself). Value/derived kinds use the canonical
    label as-is — the driver row reads as the actual line item."""
    suffix_map = {
        DriverKind.GROWTH:        "Growth %",
        DriverKind.RATIO_OF_REV:  "% of Revenue",
        DriverKind.RATIO_OF_COGS: "% of COGS",
        DriverKind.DSO_RATIO:     "DSO (days)",
        DriverKind.DIO_RATIO:     "DIO (days)",
        DriverKind.DPO_RATIO:     "DPO (days)",
        DriverKind.RATIO_OF_PARENT: "% of Parent",
        DriverKind.TAX_RATE:      "Effective Tax Rate %",
        DriverKind.PAYOUT_RATIO:  "% of Net Income",
    }
    if s.kind in suffix_map:
        return f"{s.label} — {suffix_map[s.kind]}"
    return s.label


def _driver_format(kind: DriverKind) -> str:
    if kind in (DriverKind.GROWTH, DriverKind.RATIO_OF_REV, DriverKind.RATIO_OF_COGS,
                DriverKind.RATIO_OF_PARENT,
                DriverKind.TAX_RATE, DriverKind.PAYOUT_RATIO):
        return PERCENT_FMT
    if kind in (DriverKind.DSO_RATIO, DriverKind.DIO_RATIO, DriverKind.DPO_RATIO):
        return DAYS_FMT
    return LINE_ITEM_FMT


def _historical_driver_formula(
    kind: DriverKind, src_row: int, src_col: int, src_sheet: str,
    rev_row: int | None, cogs_row: int | None, pretax_row: int | None,
    ni_row_is: int | None, div_row_cf: int | None,
    days_ref: str,
    prior_year_col: int | None,
    prior_qtr_col: int | None = None,
    basis_cell_ref: str | None = None,
    pl_col: int | None = None,
    parent_src_row: int | None = None,
) -> tuple[str | None, str]:
    """Return (formula, number_format) for a historical-period driver cell.

    `pl_col` is the QTR P&L column for the same (year, quarter) as src_col.
    Required for ratio kinds whose source sheet (BS/CF) may have different
    historical period coverage than P&L. If None, the formula is skipped
    (no P&L data for this period — happens when BS history starts later)."""
    L = get_column_letter
    src_ref = f"'{src_sheet}'!{L(src_col)}{src_row}"

    if kind == DriverKind.GROWTH:
        py_ref = f"'{src_sheet}'!{L(prior_year_col)}{src_row}" if prior_year_col else None
        qq_ref = f"'{src_sheet}'!{L(prior_qtr_col)}{src_row}" if prior_qtr_col else None
        if py_ref and qq_ref and basis_cell_ref:
            base = f'IF({basis_cell_ref}="QoQ",{qq_ref},{py_ref})'
            return f"=IFERROR({src_ref}/{base}-1,\"\")", PERCENT_FMT
        # Fall back to whichever single ref we have (earliest historical
        # quarter / single-Q history). Honors the basis flag if set, but only
        # the available ref is reachable.
        ref = py_ref or qq_ref
        if ref is None:
            return None, PERCENT_FMT
        return f"=IFERROR({src_ref}/{ref}-1,\"\")", PERCENT_FMT

    if kind == DriverKind.RATIO_OF_REV:
        if rev_row is None or pl_col is None: return None, PERCENT_FMT
        rev_ref = f"'QTR P&L'!{L(pl_col)}{rev_row}"
        return f"=IFERROR({src_ref}/{rev_ref},\"\")", PERCENT_FMT

    if kind == DriverKind.RATIO_OF_COGS:
        if cogs_row is None or pl_col is None: return None, PERCENT_FMT
        cogs_ref = f"'QTR P&L'!{L(pl_col)}{cogs_row}"
        return f"=IFERROR({src_ref}/{cogs_ref},\"\")", PERCENT_FMT

    if kind == DriverKind.DSO_RATIO:
        if rev_row is None or pl_col is None: return None, DAYS_FMT
        rev_ref = f"'QTR P&L'!{L(pl_col)}{rev_row}"
        return f"=IFERROR({src_ref}/{rev_ref}*{days_ref},\"\")", DAYS_FMT

    if kind in (DriverKind.DIO_RATIO, DriverKind.DPO_RATIO):
        if cogs_row is None or pl_col is None: return None, DAYS_FMT
        cogs_ref = f"'QTR P&L'!{L(pl_col)}{cogs_row}"
        return f"=IFERROR({src_ref}/{cogs_ref}*{days_ref},\"\")", DAYS_FMT

    if kind == DriverKind.RATIO_OF_PARENT:
        if parent_src_row is None: return None, PERCENT_FMT
        parent_ref = f"'{src_sheet}'!{L(src_col)}{parent_src_row}"
        return f"=IFERROR({src_ref}/{parent_ref},\"\")", PERCENT_FMT

    if kind == DriverKind.TAX_RATE:
        if pretax_row is None or pl_col is None: return None, PERCENT_FMT
        pt_ref = f"'QTR P&L'!{L(pl_col)}{pretax_row}"
        return f"=IFERROR({src_ref}/{pt_ref},\"\")", PERCENT_FMT

    if kind == DriverKind.PAYOUT_RATIO:
        if ni_row_is is None or pl_col is None: return None, PERCENT_FMT
        ni_ref = f"'QTR P&L'!{L(pl_col)}{ni_row_is}"
        return f"=IFERROR({src_ref}/{ni_ref},\"\")", PERCENT_FMT

    # Value kinds (HOLD_LAST/ZERO/DOLLAR_INPUT) and computed kinds
    # (LINK_TO_IS/BS_DELTA/ROLLFORWARD): historical driver = mirror of the
    # source-sheet historical value. The driver row reads as the actual line
    # item across all periods; user sees one continuous time series.
    if kind in (DriverKind.HOLD_LAST, DriverKind.ZERO, DriverKind.DOLLAR_INPUT,
                DriverKind.LINK_TO_IS, DriverKind.BS_DELTA, DriverKind.ROLLFORWARD,
                DriverKind.RESIDUAL_PLUG):
        return f"={src_ref}", LINE_ITEM_FMT

    return None, LINE_ITEM_FMT


# ============================================================================
# Computed forecast formulas (for COMPUTED_KINDS — written into driver tab)
# ============================================================================

def _computed_forecast_formula(
    s: DriverSpec, year: int, quarter: int, target_col: int,
    col_idx_by_period: dict[tuple[int, int], int], wb,
    qtr_hist_cols_by_sheet: dict, qtr_fcst_cols_by_sheet: dict,
    ni_row_is: int | None,
    driver_row: int,
) -> str | None:
    """Build a forecast formula for one cell on the DRIVER tab — used for
    LINK_TO_IS / BS_DELTA / ROLLFORWARD where the value is a derivation, not
    a user input. References the source QTR sheets (which themselves equal
    DRIVER!cell in forecast quarters via lookup, so the chain composes)."""
    L = get_column_letter

    def _col_on(sheet_name: str, period: tuple[int, int]) -> int | None:
        return (qtr_hist_cols_by_sheet.get(sheet_name, {}).get(period)
                or qtr_fcst_cols_by_sheet.get(sheet_name, {}).get(period))

    def _prior_period(y: int, q: int) -> tuple[int, int]:
        return (y, q - 1) if q > 1 else (y - 1, 4)

    if s.kind == DriverKind.LINK_TO_IS:
        if ni_row_is is None: return None
        is_col = _col_on("QTR P&L", (year, quarter))
        if is_col is None: return None
        return f"='QTR P&L'!{L(is_col)}{ni_row_is}"

    if s.kind == DriverKind.LINK_TO_CF:
        if not s.cf_link_source: return None
        cf_sheet, cf_label = s.cf_link_source
        cf_row = find_row_by_label(wb[cf_sheet], cf_label) if cf_sheet in wb.sheetnames else None
        cf_col = _col_on(cf_sheet, (year, quarter))
        if cf_row is None or cf_col is None: return None
        return f"='{cf_sheet}'!{L(cf_col)}{cf_row}"

    if s.kind == DriverKind.BS_DELTA:
        if not s.bs_delta_source: return None
        bs_sheet, bs_label = s.bs_delta_source
        bs_row = find_row_by_label(wb[bs_sheet], bs_label)
        if bs_row is None: return None
        cur_col  = _col_on(bs_sheet, (year, quarter))
        prev_col = _col_on(bs_sheet, _prior_period(year, quarter))
        if cur_col is None or prev_col is None: return None
        sign = "" if s.bs_delta_is_liability else "-"
        return f"={sign}('{bs_sheet}'!{L(cur_col)}{bs_row}-'{bs_sheet}'!{L(prev_col)}{bs_row})"

    if s.kind == DriverKind.RESIDUAL_PLUG:
        # Sum of ±(BS[t] - BS[t-1]) over every BS row this plug absorbs.
        # Sign: liabilities → +Δ (liab drop = cash use, plug subtracts cash);
        # assets → -Δ (asset drop = cash source, plug adds cash). The plug
        # row's value flows into CFO via the SUM(S2:S24)-style subtotal,
        # which then propagates to Cash via CFO+CFI+CFF+CashOther.
        if not s.residual_plug_sources:
            return "=0"
        parts: list[str] = []
        for (bs_sheet, bs_label, is_liab) in s.residual_plug_sources:
            if bs_sheet not in wb.sheetnames:
                continue
            bs_row = find_row_by_label(wb[bs_sheet], bs_label)
            cur_col  = _col_on(bs_sheet, (year, quarter))
            prev_col = _col_on(bs_sheet, _prior_period(year, quarter))
            if bs_row is None or cur_col is None or prev_col is None:
                continue
            sign = "+" if is_liab else "-"
            parts.append(f"{sign}('{bs_sheet}'!{L(cur_col)}{bs_row}-'{bs_sheet}'!{L(prev_col)}{bs_row})")
        return "=" + "".join(parts) if parts else "=0"

    if s.kind == DriverKind.ROLLFORWARD:
        # prev driver-row cell = previous period's value on this same row.
        # Driver tab columns are contiguous (B+ = historicals then forecasts),
        # so prev = target_col - 1 on the driver tab itself.
        prev_driver_col = target_col - 1
        if prev_driver_col < 2:
            return None
        prev_ref = f"{L(prev_driver_col)}{driver_row}"
        parts = [prev_ref]
        for (in_sheet, in_label) in s.rollforward_inputs:
            row = find_row_by_label(wb[in_sheet], in_label) if in_sheet in wb.sheetnames else None
            col = _col_on(in_sheet, (year, quarter))
            if row is None or col is None: continue
            parts.append(f"+'{in_sheet}'!{L(col)}{row}")
        for (out_sheet, out_label) in s.rollforward_outputs:
            row = find_row_by_label(wb[out_sheet], out_label) if out_sheet in wb.sheetnames else None
            col = _col_on(out_sheet, (year, quarter))
            if row is None or col is None: continue
            parts.append(f"-'{out_sheet}'!{L(col)}{row}")
        return "=" + "".join(parts)

    return None


# ============================================================================
# Forecast cell formula builders (write into source sheets)
# ============================================================================

RATIO_KINDS = {
    DriverKind.GROWTH, DriverKind.RATIO_OF_REV, DriverKind.RATIO_OF_COGS,
    DriverKind.DSO_RATIO, DriverKind.DIO_RATIO, DriverKind.DPO_RATIO,
    DriverKind.RATIO_OF_PARENT,
    DriverKind.TAX_RATE, DriverKind.PAYOUT_RATIO,
}


def write_forecast_cells(
    wb,
    specs: list[DriverSpec],
    qtr_hist_cols_by_sheet: dict[str, dict[tuple[int, int], int]],
    qtr_fcst_cols_by_sheet: dict[str, dict[tuple[int, int], int]],
    driver_row_by_label_by_group: dict[str, dict[str, int]],
    driver_col_by_period_by_group: dict[str, dict[tuple[int, int], int]],
    rev_row: int | None,
    cogs_row: int | None,
    pretax_row: int | None,
    ni_row_is: int | None,
    days_ref: str,
) -> int:
    """Write forecast formulas on each QTR sheet.

    Two patterns:
      - Value / computed kinds (HOLD_LAST/ZERO/DOLLAR_INPUT/LINK_TO_IS/
        BS_DELTA/ROLLFORWARD): driver tab cell IS the value → source cell is
        a pure `=DRIVERS!{col}{driver_row}` lookup.
      - Ratio kinds (GROWTH/RATIO_OF_*/DSO/DIO/DPO/TAX_RATE/PAYOUT_RATIO):
        driver tab cell is the RATE → source cell computes the dollar value
        (e.g. `=Rev × DRIVERS!{rate_cell}`). Driver remains the single place
        the user overrides; source sheet just applies the rate.
    """
    L = get_column_letter
    cells_written = 0
    group_for_sheet = {"QTR P&L": "IS", "QTR BS": "BS", "QTR CF": "CF"}

    # (sheet, label) → excel_row on that source sheet — for RATIO_OF_PARENT
    # children to reference their parent row's forecast cell.
    parent_src_row_by_sheet_label: dict[tuple[str, str], int] = {
        (sp.sheet, sp.label): sp.excel_row for sp in specs
    }

    def _col_on(sheet_name: str, period: tuple[int, int]) -> int | None:
        return (qtr_hist_cols_by_sheet.get(sheet_name, {}).get(period)
                or qtr_fcst_cols_by_sheet.get(sheet_name, {}).get(period))

    for s in specs:
        if s.kind in (DriverKind.SKIP, DriverKind.DERIVED):
            continue
        if s.sheet not in wb.sheetnames:
            continue
        group = group_for_sheet[s.sheet]
        driver_row = driver_row_by_label_by_group.get(group, {}).get(s.label)
        if driver_row is None:
            continue
        driver_sheet = DRIVER_SHEET_FOR_GROUP[group]
        ws = wb[s.sheet]
        fcst_cols = qtr_fcst_cols_by_sheet.get(s.sheet, {})

        # Sample number_format / font / border from the last populated
        # historical cell on this row (mirrors what model-write put there)
        # so forecast cells render identically. Compute once per row.
        sample_fmt = LINE_ITEM_FMT
        sample_font = None
        sample_border = None
        hist_cols = qtr_hist_cols_by_sheet.get(s.sheet, {})
        for (_, _), hcol in sorted(hist_cols.items(), reverse=True):
            sample = ws.cell(row=s.excel_row, column=hcol)
            if sample.value is not None:
                if sample.number_format and sample.number_format != "General":
                    sample_fmt = sample.number_format
                if sample.font is not None:
                    sample_font = copy(sample.font)
                if sample.border is not None:
                    sample_border = copy(sample.border)
                break

        driver_col_map = driver_col_by_period_by_group.get(group, {})
        for (y, q), fcst_col in sorted(fcst_cols.items()):
            fcst_letter = L(fcst_col)
            # Driver-tab col may differ from source-sheet col (IS DRIVERS adds
            # a Basis column at B). Fall back to source col if not in map.
            drv_col = driver_col_map.get((y, q), fcst_col)
            drv_ref = f"'{driver_sheet}'!{L(drv_col)}{driver_row}"
            formula = None

            if s.kind not in RATIO_KINDS:
                # Value / computed kinds: driver tab cell IS the value.
                formula = f"={drv_ref}"
            elif s.kind == DriverKind.GROWTH:
                # Basis cell on IS DRIVERS column B picks YoY (prev-year same
                # quarter) vs QoQ (prev quarter). Both refs resolve to whichever
                # of historical/forecast contains the period — sorted forecast
                # iteration ensures upstream cells are already populated.
                py_col = _col_on(s.sheet, (y - 1, q))
                qq_col = _col_on(s.sheet, _prior_qtr(y, q))
                py_ref = f"'{s.sheet}'!{L(py_col)}{s.excel_row}" if py_col is not None else None
                qq_ref = f"'{s.sheet}'!{L(qq_col)}{s.excel_row}" if qq_col is not None else None
                basis_cell = f"'{driver_sheet}'!$B{driver_row}"
                if py_ref and qq_ref:
                    base = f'IF({basis_cell}="QoQ",{qq_ref},{py_ref})'
                    formula = f"={base}*(1+{drv_ref})"
                elif py_ref or qq_ref:
                    formula = f"={py_ref or qq_ref}*(1+{drv_ref})"
            elif s.kind == DriverKind.RATIO_OF_REV:
                rev_col = _col_on("QTR P&L", (y, q))
                if rev_col is not None and rev_row is not None:
                    formula = f"='QTR P&L'!{L(rev_col)}{rev_row}*{drv_ref}"
            elif s.kind == DriverKind.RATIO_OF_COGS:
                cogs_col = _col_on("QTR P&L", (y, q))
                if cogs_col is not None and cogs_row is not None:
                    formula = f"='QTR P&L'!{L(cogs_col)}{cogs_row}*{drv_ref}"
            elif s.kind == DriverKind.DSO_RATIO:
                rev_col = _col_on("QTR P&L", (y, q))
                if rev_col is not None and rev_row is not None:
                    formula = f"='QTR P&L'!{L(rev_col)}{rev_row}*{drv_ref}/{days_ref}"
            elif s.kind in (DriverKind.DIO_RATIO, DriverKind.DPO_RATIO):
                cogs_col = _col_on("QTR P&L", (y, q))
                if cogs_col is not None and cogs_row is not None:
                    formula = f"='QTR P&L'!{L(cogs_col)}{cogs_row}*{drv_ref}/{days_ref}"
            elif s.kind == DriverKind.RATIO_OF_PARENT:
                # Child forecast = parent_forecast_cell × ratio_cell.
                if s.parent_source:
                    parent_sheet, parent_label = s.parent_source
                    parent_row = parent_src_row_by_sheet_label.get((parent_sheet, parent_label))
                    parent_col = _col_on(parent_sheet, (y, q))
                    if parent_row is not None and parent_col is not None:
                        parent_ref = f"'{parent_sheet}'!{L(parent_col)}{parent_row}"
                        formula = f"={parent_ref}*{drv_ref}"
            elif s.kind == DriverKind.TAX_RATE:
                pt_col = _col_on("QTR P&L", (y, q))
                if pt_col is not None and pretax_row is not None:
                    formula = f"='QTR P&L'!{L(pt_col)}{pretax_row}*{drv_ref}"
            elif s.kind == DriverKind.PAYOUT_RATIO:
                ni_col = _col_on("QTR P&L", (y, q))
                if ni_col is not None and ni_row_is is not None:
                    formula = f"='QTR P&L'!{L(ni_col)}{ni_row_is}*{drv_ref}"

            if formula is None:
                continue
            cell = ws.cell(row=s.excel_row, column=fcst_col, value=formula)
            cell.number_format = sample_fmt
            if sample_font is not None:
                cell.font = sample_font
            if sample_border is not None:
                cell.border = sample_border
            if cell.fill.fill_type is None:
                cell.fill = FORECAST_FILL
            cells_written += 1
    return cells_written


# ============================================================================
# Subtotal pattern propagation for forecast columns
# ============================================================================

_COL_REF_RE = re.compile(r"(?<![A-Za-z'!])([A-Z]{1,3})(\d+)")  # standalone cell refs

def copy_subtotal_pattern_to_forecast(
    qtr_ws, ytd_ws,
    derived_rows: set[int],
    fcst_cols_by_period: dict[tuple[int, int], int],
):
    """For each DERIVED row on the QTR sheet, find the corresponding YTD row's
    SUM/+ formula and translate the column letters to each forecast column.
    Used so subtotals on QTR sheets compute correctly in forecast quarters."""
    if not fcst_cols_by_period:
        return
    fcst_cols = sorted(fcst_cols_by_period.values())
    for r in derived_rows:
        # Find a SUM/+ formula in any column on YTD for this row.
        # Track the cell so we can also lift its border / number_format /
        # font onto each forecast cell — historical YTD subtotal cells have
        # the top border (from model-write); without copying, forecast
        # subtotals render flat.
        ytd_formula = None
        ytd_letter  = None
        sample_fmt = SUBTOTAL_FMT
        sample_font = Font(bold=True)
        sample_border = None
        for ytd_col in range(2, ytd_ws.max_column + 1):
            sample = ytd_ws.cell(row=r, column=ytd_col)
            v = sample.value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            if "SUM" in v.upper() or "+" in v[1:]:
                ytd_formula = v
                ytd_letter = get_column_letter(ytd_col)
                if sample.number_format and sample.number_format != "General":
                    sample_fmt = sample.number_format
                if sample.font is not None:
                    sample_font = copy(sample.font)
                if sample.border is not None:
                    sample_border = copy(sample.border)
                break
        if not ytd_formula:
            continue
        for qcol in fcst_cols:
            qletter = get_column_letter(qcol)
            new_formula = _COL_REF_RE.sub(
                lambda m: f"{qletter}{m.group(2)}" if m.group(1) == ytd_letter else m.group(0),
                ytd_formula,
            )
            cell = qtr_ws.cell(row=r, column=qcol, value=new_formula)
            cell.number_format = sample_fmt
            cell.font = sample_font
            if sample_border is not None:
                cell.border = sample_border


# ============================================================================
# ANNL forecast aggregation
# ============================================================================

def aggregate_annl_forecasts(
    wb,
    qtr_all_cols_by_sheet: dict[str, dict[tuple[int, int], int]],
):
    """Replace the empty forecast cells on ANNL P&L / BALANCE SHEET / CASH FLOW
    with formulas: SUM(Q1+Q2+Q3+Q4 of that fiscal year on the matching QTR
    sheet) for IS/CF; QTR Q4 of that year for BS."""
    L = get_column_letter
    for group, annl_sheet in ANNL_SHEET_FOR_GROUP.items():
        if annl_sheet not in wb.sheetnames:
            continue
        annl_ws = wb[annl_sheet]
        qtr_sheet = SOURCE_SHEET_FOR_GROUP[group]
        qtr_cols = qtr_all_cols_by_sheet.get(qtr_sheet, {})
        if not qtr_cols:
            continue

        # Find ANNL forecast columns (those ending with "E")
        for cell in annl_ws[1]:
            if cell.value is None:
                continue
            s = str(cell.value).strip()
            if not s.endswith("E"):
                continue
            m = ANNL_HDR_RE.match(s)
            if not m:
                continue
            year = int(m.group(1))
            annl_col = cell.column

            # Find QTR cols for this fiscal year
            year_qtr_cols = {q: qtr_cols[(year, q)] for q in (1, 2, 3, 4) if (year, q) in qtr_cols}
            if not year_qtr_cols:
                continue

            for r in range(2, annl_ws.max_row + 1):
                label = annl_ws.cell(row=r, column=1).value
                if not label:
                    continue
                lbl_str = str(label).strip()
                # Skip EPS / share-count rows: per-share metrics aren't linearly
                # additive across quarters (share count drifts within the year).
                # Leave the forecast cell blank for these.
                lower = lbl_str.lower()
                if "per share" in lower or "shares outstanding" in lower or "shares issued" in lower:
                    continue
                # Map this ANNL row to the matching QTR row by label
                qtr_row = find_row_by_label(wb[qtr_sheet], lbl_str)
                if qtr_row is None:
                    continue

                if group == "BS":
                    # BS = point-in-time. Use Q4 if available, else max quarter.
                    q = 4 if 4 in year_qtr_cols else max(year_qtr_cols)
                    qcol = year_qtr_cols[q]
                    formula = f"='{qtr_sheet}'!{L(qcol)}{qtr_row}"
                else:
                    # IS / CF = sum of available quarters
                    parts = [f"'{qtr_sheet}'!{L(c)}{qtr_row}" for q, c in sorted(year_qtr_cols.items())]
                    formula = "=" + "+".join(parts)

                # Only overwrite if existing cell is empty / 0 / a literal (don't blow away formulas)
                existing = annl_ws.cell(row=r, column=annl_col).value
                if existing is None or existing == 0 or isinstance(existing, (int, float)):
                    new_cell = annl_ws.cell(row=r, column=annl_col, value=formula)
                    # Mirror the formatting of the last historical cell on this
                    # row so forecasts render identically — number format, font
                    # (bold for subtotals), and top border (which model-write
                    # puts on every subtotal row's historical cells).
                    last_hist_col = annl_col - 1
                    while last_hist_col >= 2:
                        sample = annl_ws.cell(row=r, column=last_hist_col)
                        sv = sample.value
                        if sv is not None and (isinstance(sv, (int, float)) or
                                                (isinstance(sv, str) and sv.startswith("="))):
                            if sample.number_format and sample.number_format != "General":
                                new_cell.number_format = sample.number_format
                            if sample.font is not None:
                                new_cell.font = copy(sample.font)
                            if sample.border is not None:
                                new_cell.border = copy(sample.border)
                            break
                        last_hist_col -= 1


# ============================================================================
# Orchestrator
# ============================================================================

def _load_growth_basis(ticker_root: Path) -> str:
    """Read `growth_basis` from ticker config.json. Returns "yoy" (default) or
    "qoq". Per-row override happens in Excel via the Basis column on IS DRIVERS.
    """
    cfg_path = ticker_root / "config.json"
    if not cfg_path.exists():
        return "yoy"
    try:
        cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return "yoy"
    val = (cfg.get("growth_basis") or "yoy").lower()
    return val if val in ("yoy", "qoq") else "yoy"


def build(workbook_path: Path, ticker_root: Path, library_path: Path) -> dict:
    wb = load_workbook(workbook_path)
    label_map = load_label_section_map(library_path, ticker_root)
    growth_basis = _load_growth_basis(ticker_root)
    specs = infer_drivers(wb, label_map, growth_basis_default=growth_basis)

    # 1. Forecast period horizon
    forecast_periods = compute_forecast_periods(wb)
    if not forecast_periods:
        raise RuntimeError("No QTR sheet found / no historical quarters detected.")

    # 2. ASSUMPTIONS tab
    assump_rows = build_assumptions_tab(wb)
    days_ref = assumption_ref("Days in Quarter", assump_rows)

    # 3. Extend QTR sheets with forecast columns
    qtr_hist_cols_by_sheet: dict[str, dict[tuple[int, int], int]] = {}
    qtr_fcst_cols_by_sheet: dict[str, dict[tuple[int, int], int]] = {}
    qtr_all_cols_by_sheet:  dict[str, dict[tuple[int, int], int]] = {}
    for sheet in ("QTR P&L", "QTR BS", "QTR CF"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hist = parse_qtr_header_cols(ws)
        qtr_hist_cols_by_sheet[sheet] = hist
        fcst = extend_qtr_with_forecast_cols(ws, forecast_periods)
        qtr_fcst_cols_by_sheet[sheet] = fcst
        qtr_all_cols_by_sheet[sheet] = {**hist, **fcst}

    # 4. Index canonical row positions on QTR sheets (used by formula builders)
    pl_ws = wb["QTR P&L"] if "QTR P&L" in wb.sheetnames else None
    bs_ws = wb["QTR BS"]   if "QTR BS"   in wb.sheetnames else None
    cf_ws = wb["QTR CF"]   if "QTR CF"   in wb.sheetnames else None
    rev_row    = find_row_by_label(pl_ws, "Net Sales / Revenue") if pl_ws else None
    cogs_row   = find_row_by_label(pl_ws, "COGS") if pl_ws else None
    pretax_row = find_row_by_label(pl_ws, "Pre-Tax Income (Loss)") if pl_ws else None
    ni_row_is  = find_row_by_label(pl_ws, "Net Income (Loss)") if pl_ws else None
    div_row_cf = find_row_by_label(cf_ws, "Common Dividends") if cf_ws else None
    ar_row     = find_row_by_label(bs_ws, "Accounts Receivable") if bs_ws else None
    inv_row    = find_row_by_label(bs_ws, "Inventories") if bs_ws else None
    ap_row     = find_row_by_label(bs_ws, "Accounts Payable") if bs_ws else None

    # 5. Build driver tabs (one per group)
    driver_row_by_label_by_group: dict[str, dict[str, int]] = {}
    driver_col_by_period_by_group: dict[str, dict[tuple[int, int], int]] = {}
    for group in ("IS", "BS", "CF"):
        rows, cols = build_driver_tab(
            wb, group, specs,
            qtr_hist_cols_by_sheet=qtr_hist_cols_by_sheet,
            qtr_fcst_cols_by_sheet=qtr_fcst_cols_by_sheet,
            rev_row=rev_row, cogs_row=cogs_row, pretax_row=pretax_row,
            ni_row_is=ni_row_is, div_row_cf=div_row_cf,
            assump_rows=assump_rows,
        )
        driver_row_by_label_by_group[group] = rows
        driver_col_by_period_by_group[group] = cols

    # 6. Forecast cells on QTR sheets — lookups for value/computed kinds,
    #    rate-applied formulas for ratio kinds.
    cells_written = write_forecast_cells(
        wb, specs,
        qtr_hist_cols_by_sheet=qtr_hist_cols_by_sheet,
        qtr_fcst_cols_by_sheet=qtr_fcst_cols_by_sheet,
        driver_row_by_label_by_group=driver_row_by_label_by_group,
        driver_col_by_period_by_group=driver_col_by_period_by_group,
        rev_row=rev_row, cogs_row=cogs_row, pretax_row=pretax_row,
        ni_row_is=ni_row_is, days_ref=days_ref,
    )

    # 7. Subtotal pattern propagation for DERIVED rows in forecast columns
    derived_rows_by_sheet: dict[str, set[int]] = {}
    for s in specs:
        if s.kind == DriverKind.DERIVED:
            derived_rows_by_sheet.setdefault(s.sheet, set()).add(s.excel_row)
    for sheet, drv_rows in derived_rows_by_sheet.items():
        ytd_sheet = {"QTR P&L": "YTD P&L", "QTR BS": "YTD BS", "QTR CF": "YTD CF"}[sheet]
        if ytd_sheet not in wb.sheetnames:
            continue
        copy_subtotal_pattern_to_forecast(
            wb[sheet], wb[ytd_sheet], drv_rows, qtr_fcst_cols_by_sheet.get(sheet, {})
        )

    # 8. ANNL forecast aggregation
    aggregate_annl_forecasts(wb, qtr_all_cols_by_sheet)

    # 9. Forecast tint on QTR forecast cells (rows that didn't get a formula)
    for sheet, fcst_cols in qtr_fcst_cols_by_sheet.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for col in fcst_cols.values():
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col)
                if cell.fill.fill_type is None:
                    cell.fill = FORECAST_FILL

    wb.save(workbook_path)
    return {
        "workbook":         str(workbook_path),
        "forecast_periods": [f"Q{q} FY{y}E" for (y, q) in forecast_periods],
        "n_forecast_q":     len(forecast_periods),
        "specs":            len(specs),
        "cells_written":    cells_written,
    }


# ============================================================================
# Inspect mode (Pass 1)
# ============================================================================

def run_inspect(workbook_path: Path, ticker_root: Path, library_path: Path,
                json_out: Path | None = None) -> None:
    wb = load_workbook(workbook_path)
    label_map = load_label_section_map(library_path, ticker_root)
    specs = infer_drivers(wb, label_map)

    by_sheet: dict[str, list] = {}
    by_kind:  dict[str, int]  = {}
    for s in specs:
        by_sheet.setdefault(s.sheet, []).append(s)
        by_kind[s.kind.value] = by_kind.get(s.kind.value, 0) + 1

    n_gen = sum(1 for v in label_map.values() if v.get('source') == 'generic')
    n_tic = sum(1 for v in label_map.values() if v.get('source','').startswith('ticker'))
    print(f"\nDriver inference for {workbook_path.name}")
    print(f"  Library:     {library_path.name}  ({n_gen} generic entries)")
    print(f"  Ticker root: {ticker_root.name}   ({n_tic} ticker entries)")
    print(f"  Total rows:  {len(specs)}")
    print(f"  By kind:     {sorted(by_kind.items(), key=lambda x: -x[1])}\n")

    for sheet in ("QTR P&L", "QTR BS", "QTR CF"):
        rows = by_sheet.get(sheet, [])
        if not rows:
            continue
        print(f"=== {sheet} ({len(rows)} rows) ===")
        print(f"  {'r':>3}  {'label':38s}  {'section':22s}  {'kind':14s}  note")
        for s in rows:
            print(f"  {s.excel_row:>3}  {s.label[:38]:38s}  {(s.filing_section or '-'):22s}  {s.kind.value:14s}  {s.note or ''}")
        print()

    if json_out:
        payload = [json.loads(s.model_dump_json()) for s in specs]
        json_out.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        print(f"Wrote {json_out}  ({len(payload)} specs)")


# ============================================================================
# CLI
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="model-calc — quarterly-first forecasting.")
    parser.add_argument("--in", dest="in_path", required=True, type=Path)
    parser.add_argument("--ticker-root", required=True, type=Path)
    parser.add_argument("--library", type=Path, default=DEFAULT_LIBRARY)
    parser.add_argument("--inspect", action="store_true", help="Dry-run: print inferred drivers, no mutations.")
    parser.add_argument("--json-out", type=Path, default=None)
    args = parser.parse_args()

    if not args.in_path.exists():
        print(f"ERROR: {args.in_path} not found", file=sys.stderr)
        sys.exit(2)
    if not args.ticker_root.exists():
        print(f"ERROR: {args.ticker_root} not found", file=sys.stderr)
        sys.exit(2)

    if args.inspect:
        run_inspect(args.in_path, args.ticker_root, args.library, args.json_out)
        return

    report = build(args.in_path, args.ticker_root, args.library)
    print(f"\nmodel-calc: built quarterly forecast layer on {report['workbook']}")
    print(f"  Forecast horizon: {report['n_forecast_q']} quarters ({report['forecast_periods'][0]} … {report['forecast_periods'][-1]})")
    print(f"  Specs inferred:   {report['specs']}")
    print(f"  Forecast cells written on QTR sheets: {report['cells_written']}")


if __name__ == "__main__":
    main()
