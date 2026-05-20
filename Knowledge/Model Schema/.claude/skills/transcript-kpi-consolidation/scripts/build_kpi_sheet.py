"""STEP 3 — build the 'KPI Consolidated' sheet in {TICKER}_disclosures.xlsx.

Usage:  python build_kpi_sheet.py {TICKER}

Reads the per-transcript digest JSONs, normalizes metric + period, and writes a
metric x period matrix. ALL quantitative datapoints are kept — the cell shows
the earnings-call figure (else earliest), and the Shift+F2 note lists every
source. Prior-year values are backfilled from the .md STEP-5 PriorYearValue.

Canonicalization comes from the generic transcript_kpi_library.json plus the
per-ticker transcript_kpi_overrides.json (aliases + corrections).
"""
import json, glob, re, os, sys
from collections import Counter
from urllib.parse import quote
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

if len(sys.argv) < 2:
    sys.exit("usage: python build_kpi_sheet.py {TICKER}")
TICKER = sys.argv[1].upper()

HERE = os.path.dirname(os.path.abspath(__file__))
MODEL_SCHEMA = os.path.abspath(os.path.join(HERE, "..", "..", "..", ".."))
BRAIN_ROOT = os.path.abspath(os.path.join(MODEL_SCHEMA, "..", ".."))
VAULT = os.path.basename(BRAIN_ROOT)                       # Obsidian vault name
SOURCES = os.path.join(BRAIN_ROOT, "Sources", TICKER)
TICKER_LIB = os.path.join(MODEL_SCHEMA, "Ticker Libraries", TICKER, "MDA and Other")
DIGESTS = os.path.join(TICKER_LIB, "transcript_digests")
WB_PATH = os.path.join(BRAIN_ROOT, "Knowledge", "Model Outputs", TICKER,
                       f"{TICKER}_disclosures.xlsx")
GENERIC_LIB = os.path.join(MODEL_SCHEMA, "pattern_libraries", "MDA and Other",
                           "transcript_kpi_library.json")
SHEET = "KPI Consolidated"

# --- load libraries ---------------------------------------------------------
with open(GENERIC_LIB, encoding='utf-8') as fh:
    _lib = json.load(fh)
CANON = {k.lower(): v for k, v in _lib['canonical_aliases'].items()}
CORE = list(_lib['core_order'])
CORRECTIONS = []
_ov_path = os.path.join(TICKER_LIB, "transcript_kpi_overrides.json")
if os.path.exists(_ov_path):
    with open(_ov_path, encoding='utf-8') as fh:
        _ov = json.load(fh)
    for k, v in _ov.get('aliases', {}).items():
        CANON[k.lower()] = v                       # per-ticker aliases extend generic
    CORRECTIONS = _ov.get('corrections', [])

PER_RE = re.compile(
    r'\b(Q[1-4])\s*[\'’]?\s*(20\d\d)\b'
    r'|\bFY\s*(20\d\d)\b'
    r'|\b(9M)\s*(20\d\d)\b'
    r'|\b(H[12])\s*(20\d\d)\b'
    r'|\bDec(?:ember)?\s*31?,?\s*(20\d\d)\b',
    re.I)
ORD = {'Q1': 1, 'Q2': 2, 'H1': 2.5, 'Q3': 3, '9M': 3.5, 'H2': 3.5, 'Q4': 4, 'FY': 5}
UNIT_RE = re.compile(
    r'\s*\(\s*(?:in\s*)?'
    r'(?:\$\s*[MBK]?|\$?\s*000s?|[MBK]|%|bps|pts?|pp|x|'
    r'millions?|billions?|thousands?)'
    r'\s*\)\s*$', re.I)
SCALE_RE = re.compile(r'\(\s*(?:in\s*)?\$?\s*(B|K|000|millions?|billions?|thousands?)\s*\)\s*$', re.I)
LEAD_PER = re.compile(r'^\s*(?:Q[1-4]|FY|9M|H[12]|1H|2H|YTD)\b[\s.:/-]*', re.I)
TRAIL_PER = re.compile(r'[\s.:/-]*\b(?:Q[1-4]|FY|9M|H[12]|1H|2H|YTD)\s*$', re.I)
_PAREN_PERIOD = re.compile(
    r'\s*\((?:Q[1-4]|FY|H[12]|9M|Full[ -]?Year|Nine Months(?: YTD)?|'
    r'Six Months(?: YTD)?|Three Months(?: YTD)?|YTD|First Half|Second Half|'
    r'Quarter|Annual)\)\s*', re.I)
QUANT_RE = re.compile(
    r'[~≈<>]?\s*[+-]?\$?[\d,]+\.?\d*\s*(?:%|pp|ppt|pts?|bps|bp|x)?\+?\s*$', re.I)
HEADER_TOKENS = {'prior yr', 'prior year', 'current', 'change', 'yoy', 'yoy %',
                 'value', 'period', 'qoq', 'qoq %', 'metric'}
_MD_CACHE = {}
_FNAME_PERIOD = re.compile(r'(20\d\d)\s*(Q[1-4]|FY)', re.I)
_EVENT_PERIOD = re.compile(r'\b(Q[1-4])\b.{0,14}?\b(20\d\d)\b')
_MD_RE = re.compile(re.escape(TICKER) + r'_(\d{4}-\d{2}-\d{2})_(.+)\.md$')


def label_scale(label):
    m = SCALE_RE.search(str(label))
    if not m:
        return 1.0
    t = m.group(1).lower()
    if t in ('b', 'billion', 'billions'):
        return 1000.0
    if t in ('k', '000', 'thousand', 'thousands'):
        return 0.001
    return 1.0


def parse_period(text):
    if text is None:
        return None
    m = PER_RE.search(str(text))
    if not m:
        return None
    g = m.groups()
    if g[0]:
        q, y = g[0].upper(), int(g[1])
        return (f"{q} {y}", y * 10 + ORD[q])
    if g[2]:
        y = int(g[2]); return (f"FY{y}", y * 10 + ORD['FY'])
    if g[3]:
        y = int(g[4]); return (f"9M {y}", y * 10 + ORD['9M'])
    if g[5]:
        h, y = g[5].upper(), int(g[6])
        return (f"{h} {y}", y * 10 + ORD.get(h, 2.5))
    if g[7]:
        y = int(g[7]); return (f"FY{y}", y * 10 + ORD['FY'])
    return None


def is_period_cell(x):
    s = str(x).strip()
    p = PER_RE.search(s)
    if not p:
        return False
    remainder = (s[:p.start()] + s[p.end():]).strip(" .,;:()-–—'\"E")
    return len(remainder) <= 1


def norm_metric(label):
    s = str(label)
    s = PER_RE.sub('', s)
    s = _PAREN_PERIOD.sub(' ', s)
    s = LEAD_PER.sub('', s)
    s = TRAIL_PER.sub('', s)
    s = UNIT_RE.sub('', s)
    s = re.sub(r'\bex[- ]?(China|Asia|one-time|outbound freight)\b', '', s, flags=re.I)
    s = re.sub(r'\s{2,}', ' ', s).strip(' -/,')
    return s


def finalize(v):
    if isinstance(v, (int, float)):
        return (float(v), False)
    if not isinstance(v, str):
        return None
    s = v.strip()
    if not s or not QUANT_RE.fullmatch(s):
        return None
    m = re.search(r'[+-]?[\d,]+\.?\d*', s)
    if not m:
        return None
    try:
        num = float(m.group().replace(',', ''))
    except ValueError:
        return None
    if '%' in s:
        return (num / 100.0, True)
    return (num, False)


def clean_value(v):
    if isinstance(v, (int, float)):
        return v
    if not isinstance(v, str):
        return v
    s = v.strip().lstrip('~≈').strip().replace('$', '').strip()
    m = re.fullmatch(r'([+-]?[\d,]+\.?\d*)\s*([MmBbKk]?)', s)
    if m:
        try:
            num = float(m.group(1).replace(',', '').lstrip('+'))
            suf = m.group(2).upper()
            if suf == 'B':
                num *= 1000
            elif suf == 'K':
                num /= 1000
            return num
        except ValueError:
            pass
    return v.strip()


def canon_metric(raw):
    m = norm_metric(raw)
    return CANON.get(m.lower(), m)


def md_value(s):
    s = str(s).strip()
    m = re.search(r'[+-]?\$?\s*[+-]?[\d,]*\.?\d+', s)
    if not m:
        return None
    raw = m.group()
    had_dollar, had_comma = '$' in raw, ',' in raw
    try:
        num = float(raw.replace('$', '').replace(' ', '').replace(',', '').lstrip('+'))
    except ValueError:
        return None
    tail = s[m.end():m.end() + 1].upper()
    is_pct = (tail == '%')
    if tail == 'B':
        num *= 1000
    elif tail == 'K':
        num /= 1000
    elif is_pct:
        num /= 100.0
    elif tail != 'M' and had_dollar and had_comma:
        num /= 1_000_000.0
    return (num, is_pct)


def parse_md_kpis(path):
    if path in _MD_CACHE:
        return _MD_CACHE[path]
    kpis, step5_line, ev, dt = [], None, '', ''
    try:
        lines = open(path, encoding='utf-8').read().split('\n')
    except OSError:
        _MD_CACHE[path] = (kpis, step5_line, ev, dt)
        return _MD_CACHE[path]
    in5 = False
    for i, ln in enumerate(lines, 1):
        if not ev and ln.startswith('event_title:'):
            ev = ln.split(':', 1)[1].strip()
        if not dt and ln.startswith('date:'):
            dt = ln.split(':', 1)[1].strip()
        if re.match(r'##\s+STEP\s*5\b', ln):
            in5, step5_line = True, i
            continue
        if in5 and re.match(r'##\s+STEP\s*6\b', ln):
            break
        if in5 and ln.startswith('##### '):
            mm = re.match(r'#####\s+\*\*(.+?)\*\*', ln)
            if not mm:
                continue
            vm = re.search(r'\*\*Value\*\*:\s*([^|]+)', ln)
            pvm = re.search(r'PriorYearValue:\s*([^|]+)', ln)
            ym = re.search(r'YoYChangePct:\s*([^|]+)', ln)
            yoy = None
            if ym:
                yn = re.search(r'[+-]?\d[\d.]*', ym.group(1))
                if yn:
                    try:
                        yoy = float(yn.group())
                    except ValueError:
                        pass
            fy = re.search(r'FiscalYear:\s*(\d{4})', ln)
            fq = re.search(r'FiscalQuarter:\s*([A-Za-z0-9]+)', ln)
            period = None
            if fy and fq:
                y, q = fy.group(1), fq.group(1).upper()
                if q in ('Q1', 'Q2', 'Q3', 'Q4', 'H1', 'H2'):
                    period = f"{q} {y}"
                elif q == 'FY':
                    period = f"FY{y}"
                elif q == '9M':
                    period = f"9M {y}"
            mv = md_value(vm.group(1)) if vm else None
            pv = md_value(pvm.group(1)) if pvm else None
            kpis.append({'line': i, 'metric': canon_metric(mm.group(1)),
                         'metric_raw': mm.group(1).strip(), 'period': period,
                         'value': mv[0] if mv else None,
                         'prior': pv[0] if pv else None,
                         'yoy': yoy,
                         'pct': mv[1] if mv else False})
    _MD_CACHE[path] = (kpis, step5_line, ev, dt)
    return _MD_CACHE[path]


def paren_period(label, tp):
    if not tp:
        return None
    m = _PAREN_PERIOD.search(str(label))
    if not m:
        return None
    tok = m.group(0).strip().strip('()').strip().lower()
    year = tp[0]
    if re.fullmatch(r'q[1-4]', tok):
        return f"{tok.upper()} {year}"
    if tok in ('fy', 'full year', 'full-year', 'annual'):
        return f"FY{year}"
    if tok in ('9m', 'nine months', 'nine months ytd', 'ytd'):
        return f"9M {year}"
    if tok in ('h1', 'first half', 'six months', 'six months ytd'):
        return f"H1 {year}"
    if tok in ('h2', 'second half'):
        return f"H2 {year}"
    return None


def prior_consistent(value, prior, yoy):
    if value is None or prior is None or yoy is None:
        return True
    denom = 1.0 + yoy / 100.0
    if abs(denom) < 0.05:
        return True
    expected = value / denom
    if abs(expected) < 1e-6:
        return True
    return abs(prior - expected) / abs(expected) <= 0.4


def prior_year_period(plabel):
    s = str(plabel).strip()
    m = re.match(r'(Q[1-4]|9M|H[12])\s*(20\d\d)$', s)
    if m:
        return f"{m.group(1)} {int(m.group(2)) - 1}"
    m = re.match(r'FY\s*(20\d\d)$', s)
    if m:
        return f"FY{int(m.group(1)) - 1}"
    return None


def kpi_line(path, metric, period, value):
    kpis, step5, _, _ = parse_md_kpis(path)
    cands = [k for k in kpis if k['period'] == period]
    section = (step5, 'section' if step5 else 'none')
    if not cands:
        return section

    def vclose(k):
        if k['value'] is None or value is None:
            return False
        return abs(k['value'] - value) <= 0.02 * max(abs(k['value']), abs(value), 1e-9) + 1e-6

    ml = str(metric).lower()
    vmatch = [k for k in cands if vclose(k)]
    if vmatch:
        if len(vmatch) == 1:
            return (vmatch[0]['line'], 'kpi')
        exact = [k for k in vmatch if k['metric'].lower() == ml]
        if exact:
            return (exact[0]['line'], 'kpi')
        mt = set(ml.split())
        best = max(vmatch, key=lambda k: len(mt & set(k['metric'].lower().split())))
        return (best['line'], 'kpi')
    exact = [k for k in cands if k['metric'].lower() == ml]
    if len(exact) == 1:
        return (exact[0]['line'], 'kpi')
    return section


def obsidian_uri(path, line):
    rel = os.path.relpath(path, BRAIN_ROOT).replace(os.sep, '/')
    uri = f"obsidian://adv-uri?vault={quote(VAULT)}&filepath={quote(rel)}"
    if line:
        uri += f"&line={line}"
    return uri


def transcript_period(fname, event):
    m = _FNAME_PERIOD.search(os.path.basename(fname))
    if m:
        return (int(m.group(1)), m.group(2).upper())
    m = _EVENT_PERIOD.search(event or '')
    if m:
        return (int(m.group(2)), m.group(1).upper())
    return None


def fallback_period(label, tp):
    year, tq = tp
    lm = LEAD_PER.match(str(label))
    if lm:
        tk = re.match(r'(Q[1-4]|FY|9M|H[12]|1H|2H|YTD)', lm.group(0).strip(), re.I)
        tok = tk.group(1).upper() if tk else ''
        if tok in ('Q1', 'Q2', 'Q3', 'Q4'):
            return f"{tok} {year}"
        if tok == 'FY':
            return f"FY{year}"
        if tok == '9M':
            return f"9M {year}"
        if tok in ('H1', '1H'):
            return f"H1 {year}"
        if tok in ('H2', '2H'):
            return f"H2 {year}"
        if tok == 'YTD':
            return {'Q2': f"H1 {year}", 'Q3': f"9M {year}",
                    'Q4': f"FY{year}", 'FY': f"FY{year}"}.get(tq, f"{tq} {year}")
    return f"FY{year}" if tq == 'FY' else f"{tq} {year}"


def tp_label(tp):
    """Convert a transcript period tuple (year, 'Q1'/'FY'/...) to a label."""
    if not tp:
        return None
    return f"FY{tp[0]}" if tp[1] == 'FY' else f"{tp[1]} {tp[0]}"


_YEAR_IN_TEXT = re.compile(r'\b(20\d{2})\b')


def freeform_period_key(label):
    """Sort key for free-form period labels like 'Summer 2026' or 'balance of 2026'.

    Returns key ~= year * 10 + 5 (between Q4 and FY of that year), or 0 if no year."""
    m = _YEAR_IN_TEXT.search(str(label))
    return int(m.group(1)) * 10 + 5 if m else 0


SKIP_COL_TOKENS = {'change', 'yoy', 'yoy %', 'qoq', 'qoq %', 'pct change',
                   '% change', 'delta', 'metric'}


def build_column_specs(low_cells, per_in_hdr, tp):
    """Decide what each header column means. Returns a list of spec dicts.

    Each spec drives one matrix-cell write per data row:
      {col, plabel} for fixed-period columns (or 'PRIOR_YR' / 'CURRENT' sentinels);
      {col, period_from_col, fallback_to_label_or_tp} for value/period tables.
    """
    specs = []
    value_col = None
    period_col = None
    for i, tok in enumerate(low_cells):
        col_1based = i + 1
        if per_in_hdr[i]:
            specs.append({'col': col_1based, 'plabel': per_in_hdr[i][0]})
        elif tok == 'value':
            value_col = col_1based
        elif tok == 'period':
            period_col = col_1based
        elif tok == 'current':
            label = tp_label(tp)
            if label:
                specs.append({'col': col_1based, 'plabel': label})
        elif tok in ('prior yr', 'prior year'):
            label = tp_label(tp)
            prior = prior_year_period(label) if label else None
            if prior:
                specs.append({'col': col_1based, 'plabel': prior})
        elif tok in SKIP_COL_TOKENS:
            pass  # explicit skip
        # unknown header cells: ignored (no spec written)
    if value_col and period_col:
        specs.append({'col': value_col, 'period_from_col': period_col})
    elif value_col:
        # Single-value table — period from row label or transcript period
        specs.append({'col': value_col, 'fallback_to_label_or_tp': True})
    return specs


def _spec_to_plabel(spec, label, tp, row):
    """Resolve the period label for one (spec, data row) pair."""
    if 'plabel' in spec:
        return spec['plabel']
    if spec.get('period_from_col') is not None:
        pcol = spec['period_from_col']
        if pcol < len(row):
            raw = str(row[pcol]).strip()
            if not raw:
                return None
            per = parse_period(raw)
            if per:
                return per[0]
            # accept free-form ('Summer 2026', 'balance of 2026') only if a year is present
            if len(raw) <= 40 and _YEAR_IN_TEXT.search(raw):
                return raw
        return None
    if spec.get('fallback_to_label_or_tp'):
        per = parse_period(label)
        if per:
            return per[0]
        plabel = paren_period(label, tp)
        if plabel:
            return plabel
        if tp:
            return fallback_period(label, tp)
    return None


def main():
    files = sorted(glob.glob(os.path.join(DIGESTS, "*.json")))
    if not files:
        sys.exit(f"no digests at {DIGESTS} — run STEP 1 (extract) first")
    if not os.path.exists(WB_PATH):
        sys.exit(f"workbook not found: {WB_PATH}")
    matrix = {}
    periods = {}
    skipped_no_period = fallback_used = dropped_text = placed = 0

    for f in files:
        d = json.load(open(f, encoding='utf-8'))
        rows = d.get('rows', [])
        event = str(d.get('event', ''))
        date = str(d.get('date', ''))
        if rows and len(rows[0]) > 1 and str(rows[0][0]).strip().lower() == 'event':
            event = str(rows[0][1]).strip()
        if len(rows) > 1 and len(rows[1]) > 1 and str(rows[1][0]).strip().lower() == 'date':
            date = str(rows[1][1]).strip()
        is_earn = 'earning' in event.lower()
        tp = transcript_period(f, event)
        in_quant = False
        col_specs = []   # active column specs for the current subgroup

        for r in rows:
            if not r:
                continue
            c0 = str(r[0]).strip()
            if c0 == 'QUANTITATIVE':
                in_quant = True
                col_specs = []
                continue
            if c0.startswith('QUALITATIVE') or c0.startswith('Q&A'):
                in_quant = False
                col_specs = []
                continue
            if not in_quant or len(r) < 2:
                continue

            vals = r[1:]
            low = [str(x).strip().lower() for x in vals]
            is_header = any(t in HEADER_TOKENS for t in low)
            per_in_hdr = [parse_period(x) if is_period_cell(x) else None for x in vals]
            has_hdr_period = any(per_in_hdr)

            if is_header or has_hdr_period:
                col_specs = build_column_specs(low, per_in_hdr, tp)
                continue

            # data row — write one matrix cell per spec
            if not col_specs:
                # No active header; try transcript-period fallback as a single-cell write
                if not tp:
                    continue
                col_specs_local = [{'col': 1, 'fallback_to_label_or_tp': True}]
            else:
                col_specs_local = col_specs

            for spec in col_specs_local:
                col = spec['col']
                if col >= len(r):
                    continue
                raw = r[col]
                if raw in (None, '', 'n/a', 'N/A', 'n/d', '-', '--'):
                    continue
                value = clean_value(raw)
                if isinstance(value, (int, float)) and not (
                        isinstance(raw, str) and re.search(r'[MBK]\s*$', raw.strip(), re.I)):
                    value *= label_scale(c0)
                fin = finalize(value)
                if fin is None:
                    dropped_text += 1
                    continue
                value, is_pct = fin

                plabel = _spec_to_plabel(spec, c0, tp, r)
                if not plabel:
                    # secondary fallback: row label / paren / transcript period
                    per = parse_period(c0)
                    plabel = per[0] if per else paren_period(c0, tp)
                    if not plabel and tp:
                        plabel = fallback_period(c0, tp)
                        fallback_used += 1
                if not plabel:
                    skipped_no_period += 1
                    continue
                pp = parse_period(plabel)
                pkey = pp[1] if pp else freeform_period_key(plabel)

                metric = norm_metric(c0)
                if not metric:
                    continue
                periods[plabel] = pkey
                rec = {'value': value, 'pct': is_pct, 'earn': is_earn, 'event': event,
                       'date': date, 'label': c0}
                matrix.setdefault(metric, {}).setdefault(plabel, []).append(rec)
                placed += 1

    # --- consolidate near-duplicate metric rows ---
    groups = {}
    for m in matrix:
        groups.setdefault(m.lower(), []).append(m)
    casefold = {}
    for low, variants in groups.items():
        if len(variants) == 1:
            casefold[low] = variants[0]
        else:
            score = Counter({v: sum(len(r) for r in matrix[v].values()) for v in variants})
            casefold[low] = score.most_common(1)[0][0]

    def canon(m):
        return CANON.get(m.lower(), casefold.get(m.lower(), m))

    n_before = len(matrix)
    merged = {}
    for m, pm in matrix.items():
        dst = merged.setdefault(canon(m), {})
        for p, recs in pm.items():
            dst.setdefault(p, []).extend(recs)
    matrix = merged
    n_after = len(matrix)

    # --- backfill prior-year values from .md STEP 5 into otherwise-empty cells ---
    backfilled = computed_cnt = bad_prior = 0
    matrix_lc = {k.lower(): k for k in matrix}
    for mdp in glob.glob(os.path.join(SOURCES, "**", "transcripts", f"{TICKER}_*.md"),
                         recursive=True):
        kpis, _, ev, dt = parse_md_kpis(mdp)
        for k in kpis:
            if k['prior'] is None or not k['period']:
                continue
            pp = prior_year_period(k['period'])
            cmk = matrix_lc.get(k['metric'].lower())
            if not pp or cmk is None or pp in matrix[cmk]:
                continue
            base = {'pct': k['pct'], 'event': ev, 'date': dt,
                    'label': k['metric_raw'], 'md': mdp, 'line': k['line']}
            if k['pct'] or prior_consistent(k['value'], k['prior'], k['yoy']):
                matrix[cmk][pp] = [dict(base, value=k['prior'], earn=False, derived=True)]
                backfilled += 1
            elif (k['yoy'] is not None and k['period'] in matrix[cmk]
                  and any(not r.get('derived') and not r.get('computed')
                          for r in matrix[cmk][k['period']])):
                matrix[cmk][pp] = [dict(base, computed=True,
                                        current_period=k['period'], yoy=k['yoy'])]
                computed_cnt += 1
            else:
                bad_prior += 1
                continue
            pk = parse_period(pp)
            periods.setdefault(pp, pk[1] if pk else 0)

    # --- apply per-ticker manual corrections ---
    corrected_cnt = 0
    for corr in CORRECTIONS:
        m, p = corr.get('metric'), corr.get('period')
        if not m or not p or m not in matrix:
            continue
        if corr.get('skip'):
            matrix[m].pop(p, None)
            corrected_cnt += 1
        elif 'value' in corr:
            fin = finalize(corr['value'])
            if fin:
                matrix[m][p] = [{'value': fin[0], 'pct': fin[1], 'corrected': True,
                                 'note': corr.get('note', ''), 'event': '', 'date': '',
                                 'label': m}]
                pk = parse_period(p)
                periods.setdefault(p, pk[1] if pk else 0)
                corrected_cnt += 1

    # --- order axes ---
    ordered_periods = sorted(periods, key=lambda p: periods[p])
    core_present = [m for m in CORE if m in matrix]
    rest = sorted((m for m in matrix if m not in CORE), key=str.lower)

    # --- write sheet ---
    wb = openpyxl.load_workbook(WB_PATH)
    tab_map = {}
    for sn in wb.sheetnames:
        if sn == SHEET:
            continue
        sh = wb[sn]
        if str(sh.cell(1, 1).value).strip().lower() != 'event':
            continue
        tab_map[(str(sh.cell(1, 2).value).strip(),
                 str(sh.cell(2, 2).value).strip()[:10])] = sn

    md_by_date = {}
    for p in glob.glob(os.path.join(SOURCES, "**", "transcripts", f"{TICKER}_*.md"),
                       recursive=True):
        mm = _MD_RE.match(os.path.basename(p))
        if mm:
            md_by_date.setdefault(mm.group(1), []).append((mm.group(2), p))

    def find_md(event, date):
        cands = md_by_date.get(str(date).strip()[:10], [])
        if not cands:
            return None
        if len(cands) == 1:
            return cands[0][1]
        ev = str(event).lower()
        return max(cands, key=lambda c: sum(w in ev for w in c[0].lower().split()))[1]

    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, 0)

    hdr_fill = PatternFill('solid', fgColor='1F4E78')
    hdr_font = Font(bold=True, color='FFFFFF')
    core_font = Font(bold=True)
    reported_font = Font(color='2F6FA8')
    prior_font = Font(italic=True, color='808080')
    computed_font = Font(italic=True)
    corrected_font = Font(color='C55A11', bold=True)

    ws.cell(row=1, column=1,
            value=f"{TICKER} Transcript KPI Consolidation — reported value per period")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.cell(row=2, column=1,
            value=f"Every quantitative datapoint disclosed across {len(files)} transcripts. "
                  "Blue = directly reported; gray italic = prior-year comparative from .md; "
                  "black italic = reconstructed formula; orange = manual correction. "
                  "Shift+F2 note lists every source; cells deep-link to the source .md.")
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color='808080')

    hrow = 4
    ws.cell(row=hrow, column=1, value="Metric")
    for j, p in enumerate(ordered_periods):
        ws.cell(row=hrow, column=2 + j, value=p)
    for c in ws[hrow]:
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center')

    unlocated = set()
    no_md = []
    stats = {'kpi': 0, 'section': 0, 'none': 0, 'prior': 0, 'computed': 0, 'corrected': 0}

    def tab_of(rec):
        return tab_map.get((str(rec['event']).strip(), str(rec['date']).strip()[:10]))

    def vfmt(r):
        return f"{r['value']:.1%}" if r['pct'] else f"{r['value']:g}"

    def place(row, j, metric, period, recs):
        recs_s = sorted(recs, key=lambda r: str(r['date']))
        if recs_s[0].get('corrected'):
            r = recs_s[0]
            c = ws.cell(row=row, column=2 + j, value=r['value'])
            if r['pct']:
                c.number_format = '0.0%'
            c.font = corrected_font
            c.comment = Comment("Manual correction (transcript_kpi_overrides.json):\n"
                                + str(r['note']), "transcript consolidation")
            stats['corrected'] += 1
            return
        if recs_s[0].get('derived'):
            r = recs_s[0]
            c = ws.cell(row=row, column=2 + j, value=r['value'])
            if r['pct']:
                c.number_format = '0.0%'
            c.font = prior_font
            c.hyperlink = obsidian_uri(r['md'], r['line'])
            stats['prior'] += 1
            c.comment = Comment(
                "Prior-year comparative — no transcript reported this period "
                f"directly.\nFigure as stated in {r['event']} ({r['date']}),\n"
                f"STEP-5 KPI “{r['label']}” (PriorYearValue field).",
                "transcript consolidation")
            return
        if recs_s[0].get('computed'):
            r = recs_s[0]
            ref = f"{get_column_letter(2 + ordered_periods.index(r['current_period']))}{row}"
            c = ws.cell(row=row, column=2 + j, value=f"={ref}/(1+{r['yoy'] / 100:g})")
            c.font = computed_font
            c.hyperlink = obsidian_uri(r['md'], r['line'])
            stats['computed'] += 1
            c.comment = Comment(
                f"Computed prior-year value: ={ref}/(1+{r['yoy']:g}%).\n"
                "The .md PriorYearValue conflicted with the line's stated YoY%, so "
                f"it is reconstructed from the {r['current_period']} figure.\n"
                f"YoY source: {r['event']} ({r['date']}) — “{r['label']}”.",
                "transcript consolidation")
            return
        prim = min(recs_s, key=lambda r: (
            0 if r['earn'] else 1,
            1 if 'common' in str(r['label']).lower() else 0,
            str(r['date'])))
        c = ws.cell(row=row, column=2 + j, value=prim['value'])
        c.font = reported_font
        if prim['pct']:
            c.number_format = '0.0%'
        md = find_md(prim['event'], prim['date'])
        if md:
            line, kind = kpi_line(md, metric, period, prim['value'])
            c.hyperlink = obsidian_uri(md, line)
            stats[kind] = stats.get(kind, 0) + 1
        else:
            no_md.append((prim['event'], prim['date']))
        for r in recs_s:
            if tab_of(r) is None:
                unlocated.add((r['event'], r['date']))
        if len(recs_s) == 1:
            r = recs_s[0]
            txt = (f"Source: {r['event']}\nDate: {r['date']}\n"
                   f"Workbook tab: {tab_of(r) or '(see Event/Date)'}\n"
                   f"Digest label: “{r['label']}”")
        else:
            lines = [f"Reported in {len(recs_s)} transcripts (► = figure shown):"]
            for r in recs_s:
                mark = "►" if r is prim else "•"
                lines.append(f"{mark} {vfmt(r)}  —  {r['event']} ({r['date']})")
                lines.append(f"   tab: {tab_of(r) or '(see Event/Date)'}  |  “{r['label']}”")
            txt = "\n".join(lines)
        c.comment = Comment(txt, "transcript consolidation")

    div_fill = PatternFill('solid', fgColor='D9D9D9')
    rr = hrow
    for m in core_present:
        rr += 1
        ws.cell(row=rr, column=1, value=m).font = core_font
        for j, p in enumerate(ordered_periods):
            recs = matrix[m].get(p)
            if recs:
                place(rr, j, m, p, recs)
    rr += 1
    dc = ws.cell(row=rr, column=1, value="— Other metrics —")
    dc.font = Font(bold=True, italic=True, size=9)
    for j in range(len(ordered_periods) + 1):
        ws.cell(row=rr, column=1 + j).fill = div_fill
    for m in rest:
        rr += 1
        ws.cell(row=rr, column=1, value=m)
        for j, p in enumerate(ordered_periods):
            recs = matrix[m].get(p)
            if recs:
                place(rr, j, m, p, recs)

    ws.freeze_panes = "B5"
    ws.column_dimensions['A'].width = 38
    for j in range(len(ordered_periods)):
        ws.column_dimensions[ws.cell(row=hrow, column=2 + j).column_letter].width = 11

    transcript_tabs, other_tabs = [], []
    for sn in wb.sheetnames:
        if sn == SHEET:
            continue
        sh = wb[sn]
        if str(sh.cell(1, 1).value).strip().lower() == 'event':
            mm = re.match(r'(\d+)', sn)
            transcript_tabs.append((int(mm.group(1)) if mm else 0, sn))
        else:
            other_tabs.append(sn)
    transcript_tabs.sort(key=lambda x: x[0], reverse=True)
    wb._sheets = ([wb[SHEET]] + [wb[sn] for _, sn in transcript_tabs]
                  + [wb[sn] for sn in other_tabs])

    wb.save(WB_PATH)
    n_cells = sum(len(pm) for pm in matrix.values())
    print(f"[{TICKER}] wrote '{SHEET}': {len(matrix)} metrics x {len(ordered_periods)} periods")
    print(f"  metric rows consolidated: {n_before} -> {n_after}")
    print(f"  numeric datapoints kept: {placed}; non-numeric dropped: {dropped_text}")
    print(f"  prior-year: {backfilled} backfilled, {computed_cnt} formulas, {bad_prior} blank")
    print(f"  manual corrections applied: {corrected_cnt}")
    print(f"  rows via transcript-period fallback: {fallback_used}; "
          f"skipped (no period): {skipped_no_period}")
    print(f"  matrix cells: {n_cells}; no .md found: {len(no_md)}; "
          f"no tab match: {len(unlocated)}")


if __name__ == "__main__":
    main()
