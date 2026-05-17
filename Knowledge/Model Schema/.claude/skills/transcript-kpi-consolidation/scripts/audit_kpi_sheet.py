"""STEP 4 — audit the 'KPI Consolidated' sheet against the .md STEP-5 KPIs.

Usage:  python audit_kpi_sheet.py {TICKER}

For every directly-reported cell (blue, non-italic — not a prior-year backfill,
formula, or correction), finds the source transcript from the cell's Shift+F2
note, re-parses that transcript's .md STEP-5 KPIs, and checks the cell value
against them. Writes transcript_kpi_audit_report.md in the ticker library.

Catches: digest mis-copies, wrong-column / wrong-period placement, value-vs-
prior-year swaps. Cannot catch errors in the .md itself vs the source PDF.
"""
import os
import re
import sys
import glob
import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_kpi_sheet as B          # reuses the .md parser + ticker paths
import openpyxl

REPORT = os.path.join(B.TICKER_LIB, "transcript_kpi_audit_report.md")
TOL = 0.025


def vmatch(a, b):
    if not isinstance(a, (int, float)) or not isinstance(b, (int, float)):
        return False
    if a == 0 and b == 0:
        return True
    return abs(a - b) <= TOL * max(abs(a), abs(b)) + 1e-6


def main():
    wb = openpyxl.load_workbook(B.WB_PATH)
    if B.SHEET not in wb.sheetnames:
        sys.exit(f"no '{B.SHEET}' sheet — run build_kpi_sheet.py first")
    ws = wb[B.SHEET]
    hdr = [str(c.value) if c.value is not None else '' for c in ws[4]]

    md_by_date = {}
    for p in glob.glob(os.path.join(B.SOURCES, "**", "transcripts", f"{B.TICKER}_*.md"),
                       recursive=True):
        mm = re.match(re.escape(B.TICKER) + r'_(\d{4}-\d{2}-\d{2})_', os.path.basename(p))
        if mm:
            md_by_date.setdefault(mm.group(1), []).append(p)

    counts = {'ok': 0, 'ok_loose': 0, 'mismatch': 0, 'misplaced': 0,
              'no_metric': 0, 'no_period': 0, 'no_source': 0}
    discrepancies, unverifiable = [], []

    for row in ws.iter_rows(min_row=5):
        metric = row[0].value
        if not metric or str(metric).startswith('—'):
            continue
        metric = str(metric)
        for c in row[1:]:
            if c.value is None:
                continue
            if isinstance(c.value, str) and c.value.startswith('='):
                continue                                    # computed formula
            if c.font is not None and c.font.italic:
                continue                                    # prior-year / derived
            if c.font is not None and c.font.color is not None \
                    and getattr(c.font.color, 'rgb', None) == '00C55A11':
                continue                                    # manual correction
            period = hdr[c.column - 1]
            ct = c.comment.text if c.comment else ''
            dm = re.search(r'Date:\s*(\d{4}-\d{2}-\d{2})', ct) \
                or re.search(r'►.*?\((\d{4}-\d{2}-\d{2})\)', ct)
            mds = md_by_date.get(dm.group(1), []) if dm else []
            if not mds:
                counts['no_source'] += 1
                unverifiable.append((metric, period, c.value, 'no source .md'))
                continue

            per_kpis = []
            for md in mds:
                kpis, _, _, _ = B.parse_md_kpis(md)
                for k in kpis:
                    if k['period'] == period:
                        per_kpis.append(dict(k, _md=md))
            if not per_kpis:
                counts['no_period'] += 1
                unverifiable.append((metric, period, c.value,
                                     'period not in source STEP 5'))
                continue

            mk = [k for k in per_kpis if k['metric'].lower() == metric.lower()]
            if mk:
                if any(vmatch(c.value, k['value']) for k in mk):
                    counts['ok'] += 1
                elif any(vmatch(c.value, k['prior']) for k in mk):
                    counts['misplaced'] += 1
                    discrepancies.append(('MISPLACED — cell holds the prior-year figure',
                                          metric, period, c.value, mk[0]['value'],
                                          mk[0]['_md'], mk[0]['line']))
                else:
                    counts['mismatch'] += 1
                    discrepancies.append(('VALUE MISMATCH', metric, period, c.value,
                                          mk[0]['value'], mk[0]['_md'], mk[0]['line']))
            else:
                vk = [k for k in per_kpis if vmatch(c.value, k['value'])]
                pk = [k for k in per_kpis if vmatch(c.value, k['prior'])]
                if vk:
                    counts['ok_loose'] += 1
                elif pk:
                    counts['misplaced'] += 1
                    discrepancies.append(('MISPLACED — value equals a prior-year figure',
                                          metric, period, c.value, pk[0]['value'],
                                          pk[0]['_md'], pk[0]['line']))
                else:
                    counts['no_metric'] += 1
                    unverifiable.append((metric, period, c.value,
                                         'metric not in STEP 5 for this period'))

    total = sum(counts.values())
    checked_ok = counts['ok'] + counts['ok_loose']
    flagged = counts['mismatch'] + counts['misplaced']

    L = [f"# {B.TICKER} KPI Consolidated — Audit vs `.md` STEP-5 KPIs", "",
         f"Generated: {datetime.date.today().isoformat()}  ·  "
         f"{total} directly-reported cells checked", "",
         "## Summary", "", "| Result | Count |", "|---|---|",
         f"| Confirmed — value matches a STEP-5 KPI | {counts['ok']} |",
         f"| Confirmed — value matches, metric-name differs | {counts['ok_loose']} |",
         f"| **VALUE MISMATCH** | {counts['mismatch']} |",
         f"| **MISPLACED** — cell holds a prior-year figure | {counts['misplaced']} |",
         f"| Unverifiable — metric not in STEP 5 for that period | {counts['no_metric']} |",
         f"| Unverifiable — period not in the source STEP 5 | {counts['no_period']} |",
         f"| No source `.md` found | {counts['no_source']} |", "",
         f"**{checked_ok} confirmed · {flagged} flagged · "
         f"{total - checked_ok - flagged} unverifiable.**", ""]
    if discrepancies:
        L += [f"## Flagged — review these ({len(discrepancies)})", ""]
        for kind, met, per, sv, mv, md, ln in sorted(discrepancies):
            rel = os.path.relpath(md, B.BRAIN_ROOT).replace(os.sep, '/')
            L.append(f"- **{kind}** — `{met}` / {per}: sheet = `{sv}`, "
                     f".md STEP-5 = `{mv}`  ·  {rel}:{ln}")
        L.append("")
    if unverifiable:
        L += [f"## Unverifiable ({len(unverifiable)})", "",
              "Mostly channel / distribution / market-share metrics not formalized "
              "as STEP-5 KPIs.", ""]
        for met, per, sv, why in sorted(unverifiable):
            L.append(f"- `{met}` / {per} (`{sv}`) — {why}")
        L.append("")
    L += ["## Limits", "",
          "- Cannot detect errors in the `.md` itself vs the source PDF.",
          "- STEP 5 omits channel / distribution / market-share metrics.",
          "- Prior-year, formula and correction cells are not audited (they are derived)."]

    with open(REPORT, 'w', encoding='utf-8') as fh:
        fh.write("\n".join(L) + "\n")

    print(f"[{B.TICKER}] audited {total} directly-reported cells")
    print(f"  confirmed: {checked_ok}  (exact {counts['ok']}, value-only {counts['ok_loose']})")
    print(f"  FLAGGED:   {flagged}  (mismatch {counts['mismatch']}, "
          f"misplaced {counts['misplaced']})")
    print(f"  unverifiable: {counts['no_metric'] + counts['no_period']}; "
          f"no source: {counts['no_source']}")
    print(f"  report -> {REPORT}")


if __name__ == "__main__":
    main()
