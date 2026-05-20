"""STEP 2 — add one worksheet tab per transcript to {TICKER}_disclosures.xlsx.

Usage:  python build_transcript_tabs.py {TICKER}

Reads the per-transcript digest JSONs and writes one worksheet each (4-section
layout: Event/Date header, QUANTITATIVE, QUALITATIVE, Q&A). Idempotent — any
existing transcript tabs (A1 == 'Event') are removed and rebuilt; the KPI
Consolidated sheet, the MDA Disclosures tab and any user tabs are preserved.

Every populated cell on a transcript tab is decorated with an obsidian://adv-uri
hyperlink to the matching line in the source .md (Brain\\Sources\\{TICKER}\\...).
Anchor resolution by row kind:
  * Event / Date header rows         -> top of .md (line 1)
  * Section markers                  -> ## STEP 5 / 6 / 10 header
  * QUANTITATIVE subheader rows      -> #### subsection inside STEP 5
  * QUANTITATIVE data rows           -> ##### KPI line in STEP 5 (label+value match)
  * QUALITATIVE takeaway rows        -> best fuzzy match in STEPS 4 + 6..9
  * Q&A question-header rows         -> #### Question N header in STEP 10
  * Q&A 'Q' / 'A' rows               -> last seen Question N header
"""
import json, glob, re, os, sys
from urllib.parse import quote
import openpyxl
from openpyxl.styles import Font, PatternFill

if len(sys.argv) < 2:
    sys.exit("usage: python build_transcript_tabs.py {TICKER}")
TICKER = sys.argv[1].upper()

HERE = os.path.dirname(os.path.abspath(__file__))
MODEL_SCHEMA = os.path.abspath(os.path.join(HERE, "..", "..", "..", ".."))
BRAIN_ROOT = os.path.abspath(os.path.join(MODEL_SCHEMA, "..", ".."))
VAULT = os.path.basename(BRAIN_ROOT)
SOURCES = os.path.join(BRAIN_ROOT, "Sources", TICKER)
DIGESTS = os.path.join(MODEL_SCHEMA, "Ticker Libraries", TICKER, "MDA and Other",
                       "transcript_digests")
WB_PATH = os.path.join(BRAIN_ROOT, "Knowledge", "Model Outputs", TICKER,
                       f"{TICKER}_disclosures.xlsx")

ILLEGAL = re.compile(r'[\[\]:\*\?/\\]')
_MD_RE = re.compile(re.escape(TICKER) + r'_(\d{4}-\d{2}-\d{2})_(.+)\.md$')

# ----- helpers --------------------------------------------------------------


def clean_label(sn):
    s = re.sub(r'^\d{4}-\d{2}-\d{2}\s+', '', str(sn))
    s = re.sub(r'^\d{1,3}\s+', '', s)
    return s.strip()


def cell(v):
    return v if v is None or isinstance(v, (str, int, float)) else str(v)


def is_section(v):
    return isinstance(v, str) and v.strip().upper().split()[0:1] in (
        ['QUANTITATIVE'], ['QUALITATIVE'], ['Q&A'])


def obsidian_uri(path, line):
    rel = os.path.relpath(path, BRAIN_ROOT).replace(os.sep, '/')
    uri = f"obsidian://adv-uri?vault={quote(VAULT)}&filepath={quote(rel)}"
    if line:
        uri += f"&line={line}"
    return uri


def build_md_index():
    """Return {date_str: [(slug, path), ...]} for every {TICKER}_*.md transcript."""
    out = {}
    for p in glob.glob(os.path.join(SOURCES, "**", "transcripts", f"{TICKER}_*.md"),
                       recursive=True):
        m = _MD_RE.match(os.path.basename(p))
        if m:
            out.setdefault(m.group(1), []).append((m.group(2), p))
    return out


def find_md(md_by_date, event, date):
    cands = md_by_date.get(str(date).strip()[:10], [])
    if not cands:
        return None
    if len(cands) == 1:
        return cands[0][1]
    ev = str(event).lower()
    return max(cands, key=lambda c: sum(w in ev for w in c[0].lower().split()))[1]


# ----- .md parsing -----------------------------------------------------------

_MD_CACHE = {}
_STEP_HDR = re.compile(r'^##\s+STEP\s+(\d+)\b', re.I)
_SUB_HDR = re.compile(r'^####\s+(?!#)(.+?)\s*$')   # `#### ...` but not `#####`
_KPI_HDR = re.compile(r'^#####\s+\*\*(.+?)\*\*')
_KPI_VAL = re.compile(r'\*\*Value\*\*:\s*([^|]+)')
_KPI_FY = re.compile(r'FiscalYear:\s*(\d{4})')
_KPI_FQ = re.compile(r'FiscalQuarter:\s*([A-Za-z0-9]+)')
_NUM_RE = re.compile(r'[+-]?\$?[\d,]+\.?\d*')
_STOP = {'the', 'a', 'an', 'of', 'to', 'in', 'for', 'on', 'with', 'and', 'or',
         'is', 'was', 'as', 'by', 'at', 'from', 'that', 'this', 'it', 'be',
         'are', 'were', 'will', 'has', 'have', 'had', 'our', 'we', 'its',
         'their', 'they', 'than', 'over', 'into', 'up', 'down', 'about',
         'vs', 'q', 'fy', 'h1', 'h2', '9m'}


def _tokens(text):
    return {t for t in re.findall(r"[a-z0-9]+", str(text).lower())
            if t not in _STOP and len(t) > 1}


def _md_value(s):
    """Best-effort numeric extraction from a STEP-5 Value string."""
    if s is None:
        return None
    raw = str(s).strip()
    m = _NUM_RE.search(raw)
    if not m:
        return None
    try:
        v = float(m.group().replace('$', '').replace(',', ''))
    except ValueError:
        return None
    tail = raw[m.end():].lstrip().upper()[:1]
    if tail == 'B':
        v *= 1_000
    elif tail == 'K':
        v /= 1_000
    elif '%' in raw[m.end():]:
        v /= 100.0
    return v


def parse_md(path):
    """Parse one transcript .md once and cache.

    Returns dict with:
      steps:           {step_no: header_line}
      step_end:        {step_no: end_line}  (exclusive — start of next step)
      kpis:            list[{label, label_lc, value, period, line}] from STEP 5
      sub_by_step:     {step_no: [(line, title_lc, tokens), ...]}  for `####`
      lines_by_step:   {step_no: [(line, tokens), ...]}  for non-empty content
    """
    if path in _MD_CACHE:
        return _MD_CACHE[path]
    info = {'steps': {}, 'step_end': {}, 'kpis': [],
            'sub_by_step': {}, 'lines_by_step': {}}
    try:
        text = open(path, encoding='utf-8').read()
    except OSError:
        _MD_CACHE[path] = info
        return info
    lines = text.split('\n')

    # pass 1: locate STEP headers
    for i, ln in enumerate(lines, 1):
        m = _STEP_HDR.match(ln)
        if m:
            info['steps'][int(m.group(1))] = i
    # compute end-of-step (exclusive)
    sorted_steps = sorted(info['steps'].items())
    for idx, (n, ln) in enumerate(sorted_steps):
        info['step_end'][n] = (sorted_steps[idx + 1][1] if idx + 1 < len(sorted_steps)
                               else len(lines) + 1)

    def step_for(i):
        for n, ln in sorted_steps:
            if ln <= i < info['step_end'][n]:
                return n
        return 0

    # pass 2: subsections + content lines + STEP-5 KPIs
    for i, ln in enumerate(lines, 1):
        st = step_for(i)
        if not st:
            continue
        sub = _SUB_HDR.match(ln)
        if sub:
            title = sub.group(1).strip().lstrip('—').strip()
            info['sub_by_step'].setdefault(st, []).append(
                (i, title.lower(), _tokens(title)))
            continue
        kpi = _KPI_HDR.match(ln)
        if kpi and st == 5:
            label = kpi.group(1).strip()
            vm = _KPI_VAL.search(ln)
            fy = _KPI_FY.search(ln)
            fq = _KPI_FQ.search(ln)
            period = None
            if fy and fq:
                y, q = fy.group(1), fq.group(1).upper()
                if q in ('Q1', 'Q2', 'Q3', 'Q4', 'H1', 'H2'):
                    period = f"{q} {y}"
                elif q == 'FY':
                    period = f"FY{y}"
                elif q == '9M':
                    period = f"9M {y}"
            info['kpis'].append({
                'line': i, 'label': label, 'label_lc': label.lower(),
                'label_tokens': _tokens(label),
                'value': _md_value(vm.group(1)) if vm else None,
                'period': period})
        body = ln.strip()
        if body and not body.startswith('#'):
            info['lines_by_step'].setdefault(st, []).append((i, _tokens(body)))

    _MD_CACHE[path] = info
    return info


# ----- anchor resolution ----------------------------------------------------

_PERIOD_TOKEN = re.compile(
    r'^\s*(?:Q[1-4]|H[12]|FY|9M|YTD|1H|2H)\s*\'?20\d{2}\b|'
    r'^\s*(?:Prior\s*Yr|Prior\s*Year|Current|YoY|YoY\s*%|QoQ|QoQ\s*%|Value|Period|Change)\s*$',
    re.I)


def _is_period_header_row(row):
    """Heuristic — row[1:] cells are period labels / header tokens, not data."""
    rest = [c for c in row[1:] if c not in (None, '')]
    if not rest:
        return False
    hits = sum(1 for c in rest if isinstance(c, str) and _PERIOD_TOKEN.match(c.strip()))
    return hits >= max(1, len(rest) // 2)


def _values_in_row(row):
    """Pull numeric tokens out of row cells for KPI value matching."""
    out = []
    for c in row[1:]:
        if c is None or c == '':
            continue
        if isinstance(c, (int, float)):
            out.append(float(c))
        else:
            m = _NUM_RE.search(str(c))
            if m:
                try:
                    out.append(float(m.group().replace('$', '').replace(',', '')))
                except ValueError:
                    pass
    return out


def _vclose(a, b):
    if a is None or b is None:
        return False
    return abs(a - b) <= 0.02 * max(abs(a), abs(b), 1e-9) + 1e-6


def _best_fuzzy(tokens, candidates, min_overlap=2):
    """Pick the (line, ...) candidate whose tokens overlap most with `tokens`.

    candidates: iterable of (line, ..., token_set) — final element is the set.
    """
    if not tokens:
        return None
    best = None
    best_score = min_overlap - 1
    for c in candidates:
        score = len(tokens & c[-1])
        if score > best_score:
            best_score = score
            best = c[0]
    return best


def _resolve_quant_subheader(row, md):
    """QUANTITATIVE subheader row -> closest #### inside STEP 5."""
    subs = md['sub_by_step'].get(5, [])
    if not subs:
        return md['steps'].get(5)
    label_tokens = _tokens(row[0])
    return (_best_fuzzy(label_tokens, [(s[0], s[2]) for s in subs])
            or md['steps'].get(5))


def _resolve_quant_data(row, md):
    """QUANTITATIVE data row -> STEP-5 KPI line by label+value."""
    if not md['kpis']:
        return md['steps'].get(5)
    label_tokens = _tokens(row[0])
    label_lc = str(row[0]).lower()
    vals = _values_in_row(row)

    # value-close candidates first
    vmatch = [k for k in md['kpis']
              if any(_vclose(k['value'], v) for v in vals)]
    if vmatch:
        if len(vmatch) == 1:
            return vmatch[0]['line']
        # tie-break by label overlap, then exact label
        exact = [k for k in vmatch if k['label_lc'] == label_lc]
        if exact:
            return exact[0]['line']
        scored = sorted(vmatch,
                        key=lambda k: -len(label_tokens & k['label_tokens']))
        return scored[0]['line']

    # label-only
    exact = [k for k in md['kpis'] if k['label_lc'] == label_lc]
    if len(exact) == 1:
        return exact[0]['line']
    scored = [(len(label_tokens & k['label_tokens']), k) for k in md['kpis']]
    scored.sort(key=lambda t: -t[0])
    if scored and scored[0][0] >= 2:
        return scored[0][1]['line']
    return md['steps'].get(5)


def _resolve_qual(row, md):
    """QUALITATIVE takeaway row -> best line across STEPS 4 + 6..9."""
    pool = []
    for n in (4, 6, 7, 8, 9):
        pool.extend(md['lines_by_step'].get(n, []))
    text = ' '.join(str(c) for c in row if c not in (None, ''))
    hit = _best_fuzzy(_tokens(text), pool, min_overlap=3)
    return hit or md['steps'].get(7) or md['steps'].get(4)


def _resolve_qa_question(row, md):
    """Q&A header row ['Q1','Headline','Speaker'] -> #### Question N header line."""
    subs = md['sub_by_step'].get(10, [])
    if not subs:
        return md['steps'].get(10)
    qn = str(row[0]).strip().upper()
    headline = str(row[1] if len(row) > 1 else '').strip()
    # match by leading 'Question N' in the #### title, then by headline tokens
    n_match = re.match(r'Q(\d+)$', qn)
    if n_match:
        wanted = f"question {n_match.group(1)}"
        cands = [s for s in subs if s[1].startswith(wanted)]
        if len(cands) == 1:
            return cands[0][0]
        if len(cands) > 1:
            ht = _tokens(headline)
            return _best_fuzzy(ht, [(c[0], c[2]) for c in cands]) or cands[0][0]
    # fallback fuzzy on headline alone
    ht = _tokens(headline)
    return (_best_fuzzy(ht, [(s[0], s[2]) for s in subs])
            or md['steps'].get(10))


def resolve_anchor(row, md, ctx):
    """Return (line, kind) for a row; updates ctx in place.

    ctx fields: section ('quant'|'qual'|'qa'|None), qa_line (last Q# header line),
                subhdr_line (last subheader line in current section).
    """
    if not row:
        return None, 'blank'
    first = str(row[0]).strip() if row[0] is not None else ''
    head = first.upper().split()[:1]

    if first in ('Event', 'Date'):
        return 1, 'header'
    if head == ['QUANTITATIVE']:
        ctx['section'] = 'quant'
        ctx['qa_line'] = None
        ctx['subhdr_line'] = None
        return md['steps'].get(5), 'section'
    if head == ['QUALITATIVE']:
        ctx['section'] = 'qual'
        ctx['qa_line'] = None
        ctx['subhdr_line'] = None
        return md['steps'].get(6) or md['steps'].get(7), 'section'
    if head == ['Q&A']:
        ctx['section'] = 'qa'
        ctx['qa_line'] = None
        ctx['subhdr_line'] = None
        return md['steps'].get(10), 'section'

    section = ctx.get('section')

    # Q&A
    if section == 'qa':
        if re.match(r'Q\d+$', first):
            line = _resolve_qa_question(row, md)
            ctx['qa_line'] = line
            return line, 'qa-q'
        if first.upper() in ('Q', 'A'):
            return ctx.get('qa_line') or md['steps'].get(10), 'qa-body'
        # unanchored — fall back to section
        return ctx.get('qa_line') or md['steps'].get(10), 'qa-other'

    # QUALITATIVE
    if section == 'qual':
        return _resolve_qual(row, md), 'qual'

    # QUANTITATIVE
    if section == 'quant':
        if _is_period_header_row(row):
            line = _resolve_quant_subheader(row, md)
            ctx['subhdr_line'] = line
            return line, 'quant-sub'
        return _resolve_quant_data(row, md), 'quant-data'

    # before any section marker
    return 1, 'header'


# ----- main ----------------------------------------------------------------


def main():
    files = sorted(glob.glob(os.path.join(DIGESTS, "*.json")))
    if not files:
        sys.exit(f"no digests at {DIGESTS} — run STEP 1 (extract) first")
    if not os.path.exists(WB_PATH):
        sys.exit(f"workbook not found: {WB_PATH}")

    frags = [json.load(open(f, encoding='utf-8')) for f in files]
    frags.sort(key=lambda d: (str(d.get('date', '')), d.get('index', 0)))

    wb = openpyxl.load_workbook(WB_PATH)
    for sn in list(wb.sheetnames):
        if str(wb[sn].cell(1, 1).value).strip().lower() == 'event':
            del wb[sn]
    kept = list(wb.sheetnames)

    md_by_date = build_md_index()
    hdr_fill = PatternFill('solid', fgColor='D9D9D9')
    used, new_titles = set(), []
    no_md, link_stats = [], {}

    for r, d in enumerate(frags, 1):
        name = ILLEGAL.sub('', f"{r:02d} {clean_label(d.get('sheet_name', d.get('event', '')))}")[:31].strip()
        base, k = name, 2
        while name.lower() in used:
            suf = f" ({k})"
            name = base[:31 - len(suf)] + suf
            k += 1
        used.add(name.lower())
        new_titles.append(name)
        ws = wb.create_sheet(title=name)

        md_path = find_md(md_by_date, d.get('event', ''), d.get('date', ''))
        md = parse_md(md_path) if md_path else None
        if not md_path:
            no_md.append((d.get('date', ''), d.get('event', '')))

        ctx = {'section': None, 'qa_line': None, 'subhdr_line': None}

        for row in d.get('rows', []):
            cells = [cell(x) for x in row] if row else []
            ws.append(cells)
            if not row or not md:
                continue
            line, kind = resolve_anchor(row, md, ctx)
            if not line:
                continue
            link_stats[kind] = link_stats.get(kind, 0) + 1
            written_row = ws.max_row
            uri = obsidian_uri(md_path, line)
            for col_idx, val in enumerate(cells, 1):
                if val is None or val == '':
                    continue
                ws.cell(row=written_row, column=col_idx).hyperlink = uri

        for cr in ws.iter_rows():
            v = cr[0].value
            if is_section(v):
                for c in cr:
                    c.font = Font(bold=True, size=11)
                    c.fill = hdr_fill
            elif v in ('Event', 'Date'):
                cr[0].font = Font(bold=True)
        ws.column_dimensions['A'].width = 46
        for col in 'BCDEFGH':
            ws.column_dimensions[col].width = 15

    wb._sheets = [wb[t] for t in kept] + [wb[t] for t in new_titles]
    wb.save(WB_PATH)
    total_links = sum(link_stats.values())
    print(f"[{TICKER}] wrote {len(new_titles)} transcript tabs; "
          f"total sheets: {len(wb.sheetnames)}; "
          f"hyperlinked rows: {total_links}")
    if link_stats:
        breakdown = ', '.join(f"{k}={v}" for k, v in sorted(link_stats.items()))
        print(f"[{TICKER}] anchor kinds: {breakdown}")
    if no_md:
        print(f"[{TICKER}] WARNING: {len(no_md)} transcripts could not be matched to a .md:")
        for dt, ev in no_md[:5]:
            print(f"    {dt}  {ev}")


if __name__ == "__main__":
    main()
