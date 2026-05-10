"""Build {TICKER}_disclosures.xlsx from iXBRL filings + generic disclosure library + per-ticker overrides.

Axis-driven extraction — no label heuristics, no static taxonomies. Each disclosure table reads from the
XBRL axes the filer already tagged. Output shape mirrors whatever members the filer disclosed.

Usage:
    python build_disclosures_workbook.py --ticker CELH

Auto-discovers 10-K and 10-Q iXBRL htm files under Brain\\Sources\\{TICKER}\\{YYYY-FY|YYYY-Qn}\\filings\\.
Loads generic library from pattern_libraries/MDA and Other/generic_disclosure_tables.json.
Loads per-ticker overrides from Ticker Libraries/{TICKER}/MDA and Other/disclosure_overrides.json (if exists).
Loads per-ticker narrative from Ticker Libraries/{TICKER}/MDA and Other/mda_narrative.json (if exists).
Output: Brain\\Knowledge\\Model Outputs\\{TICKER}\\{TICKER}_disclosures.xlsx.
"""
import argparse
import json
from collections import defaultdict
from datetime import date
from pathlib import Path

from lxml import etree
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

# Path resolution
SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
MODEL_SCHEMA = SKILL_DIR.parent.parent.parent  # .../.claude/skills/mda-disaggregation/scripts -> Model Schema
KNOWLEDGE = MODEL_SCHEMA.parent  # Model Schema -> Knowledge
BRAIN = KNOWLEDGE.parent  # Knowledge -> Brain

GENERIC_LIB_PATH = MODEL_SCHEMA / "pattern_libraries" / "MDA and Other" / "generic_disclosure_tables.json"
TICKER_LIBS = MODEL_SCHEMA / "Ticker Libraries"
SOURCES = BRAIN / "Sources"
MODEL_OUTPUTS = KNOWLEDGE / "Model Outputs"

# iXBRL namespaces
NS = {
    "xbrli": "http://www.xbrl.org/2003/instance",
    "xbrldi": "http://xbrl.org/2006/xbrldi",
    "ix": "http://www.xbrl.org/2013/inlineXBRL",
    "xlink": "http://www.w3.org/1999/xlink",
    "link": "http://www.xbrl.org/2003/linkbase",
}

# Workbook styling
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
PERIOD_HDR = Font(bold=True, size=10)
LABEL_FONT = Font(size=10)
LABEL_BOLD = Font(bold=True, size=10)
LABEL_INDENT = Font(size=10, color="404040")
METRIC_FONT = Font(bold=True, size=10, color="595959")
TOP_BORDER = Border(top=Side(border_style="thin", color="000000"))
BOTTOM_BORDER = Border(bottom=Side(border_style="thin", color="000000"))


# ============================================================================
# iXBRL parsing
# ============================================================================

def localname(el):
    """Namespace-agnostic local name. Handles HTMLParser stripping (returns 'ix:nonfraction' or
    similar) and XMLParser ('{ns}nonFraction')."""
    tag = el.tag
    if not isinstance(tag, str):
        return ""
    if "}" in tag:
        return tag.split("}", 1)[1]
    if ":" in tag:
        return tag.split(":", 1)[1]
    return tag


def find_first_local(parent, name):
    """First descendant whose local name == name (case-insensitive — HTMLParser lowercases)."""
    target = name.lower()
    for el in parent.iter():
        if localname(el).lower() == target:
            return el
    return None


def find_all_local(parent, name):
    """All descendants whose local name == name."""
    target = name.lower()
    return [el for el in parent.iter() if localname(el).lower() == target]


def quarter_label(end_d):
    """Map any end_date to its containing quarter (Dec fiscal). Returns 'Q{n} {yyyy}'."""
    q = ((end_d.month - 1) // 3) + 1
    return f"Q{q} {end_d.year}"


def parse_period(ctx_el):
    """Extract period label from a context element. Returns (period_label, end_date).

    Labels: FYyyyy (instant Dec-end or ~365d duration), Qn yyyy (any non-FY context maps
    to its containing quarter via month % 3), H1/H2 yyyy (~180d duration), 9M yyyy (~270d).
    Acquisition stub contexts (e.g., 2025-04 close date) get quarter-mapped to Q2 2025.
    Assumes Dec fiscal year — PG (June fiscal) needs a fiscal-year-anchor in a future iteration."""
    period = find_first_local(ctx_el, "period")
    if period is None:
        return None, None
    instant = find_first_local(period, "instant")
    if instant is not None and instant.text:
        end = date.fromisoformat(instant.text.strip())
        # Instant = balance-sheet point. Dec-end -> FY. Else map to containing quarter.
        if end.month == 12 and end.day >= 28:
            return f"FY{end.year}", end
        return quarter_label(end), end
    start_el = find_first_local(period, "startDate")
    end_el = find_first_local(period, "endDate")
    if start_el is not None and end_el is not None and start_el.text and end_el.text:
        start = date.fromisoformat(start_el.text.strip())
        end = date.fromisoformat(end_el.text.strip())
        days = (end - start).days
        if days >= 350:
            return f"FY{end.year}", end
        if days >= 250:
            return f"9M {end.year}", end
        if days >= 150:
            return f"H{1 if end.month <= 6 else 2} {end.year}", end
        # Any shorter duration (quarter, acquisition stub, etc.) maps to its containing quarter
        return quarter_label(end), end
    return None, None


def parse_dimensions(ctx_el):
    """Extract dimensions as list of (axis_qname, member_qname) tuples."""
    dims = []
    for dim_el in find_all_local(ctx_el, "explicitMember"):
        axis = dim_el.get("dimension")
        member = (dim_el.text or "").strip()
        if axis and member:
            dims.append((axis, member))
    return dims


def parse_value(fact_el):
    """Apply scale, sign, decimals to text value. Returns float or None.
    HTMLParser-tolerant: text may live in fact_el.text OR in a nested span (iXBRL inline format)."""
    text = (fact_el.text or "").strip().replace(",", "")
    if not text:
        # iXBRL often wraps the visible value in nested spans for display; collect all descendant text
        text = "".join(fact_el.itertext()).strip().replace(",", "")
    if not text or text == "-":
        return None
    try:
        val = float(text)
    except ValueError:
        return None
    scale = int(fact_el.get("scale", "0") or "0")
    if scale:
        val *= 10 ** scale
    sign = fact_el.get("sign", "")
    if sign == "-":
        val = -val
    return val


def strip_concept_ns(name):
    """'us-gaap:Revenues' -> 'Revenues'. Preserves namespace prefix if non-us-gaap (extension concepts)."""
    if not name:
        return name
    if name.startswith("us-gaap:"):
        return name.split(":", 1)[1]
    return name  # keep filer-extension namespace (e.g. 'celh:Foo')


def parse_filing(htm_path):
    """Parse iXBRL htm. Returns list of fact dicts: concept, period_label, end_date, dimensions, value, ctx_id."""
    parser = etree.HTMLParser(recover=True)
    tree = etree.parse(str(htm_path), parser)
    root = tree.getroot()

    # Build context lookup (xbrli:context elements)
    contexts = {}
    for ctx in find_all_local(root, "context"):
        cid = ctx.get("id")
        if not cid:
            continue
        period_label, end_date = parse_period(ctx)
        dims = parse_dimensions(ctx)
        contexts[cid] = (period_label, end_date, dims)

    # Extract numeric facts (ix:nonFraction). HTMLParser lowercases attribute names too.
    facts = []
    for fact_el in find_all_local(root, "nonFraction"):
        ctx_ref = fact_el.get("contextref") or fact_el.get("contextRef")
        if not ctx_ref or ctx_ref not in contexts:
            continue
        period_label, end_date, dims = contexts[ctx_ref]
        if period_label is None:
            continue
        value = parse_value(fact_el)
        if value is None:
            continue
        facts.append({
            "concept": fact_el.get("name"),
            "period_label": period_label,
            "end_date": end_date,
            "dimensions": dims,
            "value": value,
            "ctx_id": ctx_ref,
        })
    return facts


# ============================================================================
# Disclosure-table extraction
# ============================================================================

def extract_disclosure_table(facts, table_def, overrides):
    """For a given disclosure table from the generic library, bucket facts by (metric, member, period).

    Returns dict: {metric: {member: {period_label: (value, source_note)}}} + set of novel members.
    """
    required_axes = set(table_def["extraction"]["axes"])
    concept_priority = dict(table_def["extraction"]["concepts"])  # copy so per-ticker extensions don't mutate library

    # Apply per-ticker extension_concepts: append filer-extension concepts to a metric's priority list.
    # E.g. {"DISC-SEG-PNL": {"extension_concepts": {"depreciation_amortization": ["mnst:DepreciationDepletionAndAmortizationExcludingRightOfUseAssets"]}}}
    extension_concepts = (overrides or {}).get("extension_concepts", {})
    for metric, ext_list in extension_concepts.items():
        if metric in concept_priority:
            concept_priority[metric] = list(concept_priority[metric]) + list(ext_list)
        else:
            concept_priority[metric] = list(ext_list)

    member_aliases = (overrides or {}).get("member_aliases", {})

    buckets = defaultdict(lambda: defaultdict(dict))  # metric -> member -> period -> (value, src)
    novels = set()

    for fact in facts:
        # Filter to facts whose context carries at least one required axis
        fact_axes = {axis for axis, _ in fact["dimensions"]}
        if not (fact_axes & required_axes):
            continue
        # Find which axis it's on; pick the first matching required axis
        member = None
        for axis, mem in fact["dimensions"]:
            if axis in required_axes:
                member = mem
                break
        if member is None:
            continue

        concept_local = strip_concept_ns(fact["concept"])
        # Resolve to metric (first-match priority)
        matched_metric = None
        for metric, concept_list in concept_priority.items():
            if concept_local in concept_list:
                matched_metric = metric
                break
        if matched_metric is None:
            continue  # concept not declared for this disclosure — skip silently

        # Apply member alias normalization
        canonical_member = member_aliases.get(member, member)
        # Track novels: member never aliased AND not yet seen as canonical
        if member not in member_aliases and member not in member_aliases.values():
            novels.add(member)

        period = fact["period_label"]
        # Prefer facts with FEWER dimensions — a fact tagged only on the disclosure's required
        # axis is the "pure" disclosure value; a fact with additional axes (e.g. segment ×
        # geography cross-cut) is a sub-slice that shouldn't replace the pure value.
        # Tie-break: existing first-fact-wins (deterministic).
        fact_dims = len(fact["dimensions"])
        existing = buckets[matched_metric][canonical_member].get(period)
        if existing is None or existing[2] > fact_dims:
            buckets[matched_metric][canonical_member][period] = (
                fact["value"], f"{concept_local} @ {period}", fact_dims,
            )

    return buckets, novels


# ============================================================================
# Workbook rendering
# ============================================================================

def render_disclosure_section(ws, start_row, table_id, table_def, buckets, all_periods, display, period_to_date):
    """Render one disclosure table. Returns row index after the section.

    Sort rules:
      - Periods chronological by end_date (period_to_date).
      - Members ordered by total magnitude on the FIRST metric in metric_rows (largest first).
        If first metric absent for a member, falls back to total across metrics.
    """
    name = table_def["name"]
    metric_rows = table_def["render"]["metric_rows"]
    metric_labels = table_def["render"]["metric_labels"]
    value_type = table_def["render"].get("value_type", "dollar")
    if value_type == "percent":
        fmt = table_def["render"]["number_format_percent"]
        divisor = 1  # raw decimals (0.602 = 60.2%) — no unit conversion
        unit_suffix = "(% of revenue)"
    else:
        fmt = table_def["render"]["number_format_dollar_thousands"]
        divisor = display.get("unit_divisor", 1)
        unit_suffix = f"($ in {display.get('unit_label', 'dollars')})"

    # Section header
    sr = start_row
    header = f"{table_def['section_order']}. {name.upper()}  {unit_suffix}"
    ws.cell(row=sr, column=1, value=header).font = SECTION_FONT
    ws.cell(row=sr, column=1).border = TOP_BORDER

    # Period header row — sort by actual end_date (chronological)
    pr = sr + 1
    ws.cell(row=pr, column=1, value="Metric / Segment").font = PERIOD_HDR
    sorted_periods = sorted(all_periods, key=lambda p: period_to_date.get(p, date.min))
    for ci, period in enumerate(sorted_periods):
        c = ws.cell(row=pr, column=2 + ci, value=period)
        c.font = PERIOD_HDR
        c.alignment = Alignment(horizontal="center")
        c.border = BOTTOM_BORDER

    # Determine member sort order — use first metric (typically revenue) sum across periods desc
    primary_metric = metric_rows[0] if metric_rows else None
    all_members = set()
    for metric in metric_rows:
        all_members.update(buckets.get(metric, {}).keys())

    def member_magnitude(member):
        if primary_metric:
            data = buckets.get(primary_metric, {}).get(member, {})
            if data:
                return -sum(abs(t[0]) for t in data.values())  # negative for descending
        total = 0
        for metric in metric_rows:
            for t in buckets.get(metric, {}).get(member, {}).values():
                total += abs(t[0])
        return -total

    sorted_members = sorted(all_members, key=member_magnitude)

    # Metric blocks: each metric gets a label row + N segment rows
    cur = pr + 1
    for metric in metric_rows:
        metric_data = buckets.get(metric, {})
        if not metric_data:
            continue
        ws.cell(row=cur, column=1, value=metric_labels.get(metric, metric)).font = METRIC_FONT
        cur += 1
        for member in sorted_members:
            if member not in metric_data:
                continue
            mc = ws.cell(row=cur, column=1, value=f"  {member}")
            mc.font = LABEL_INDENT
            for ci, period in enumerate(sorted_periods):
                entry = metric_data[member].get(period)
                if entry is None:
                    continue
                val, src, _dims = entry
                cell = ws.cell(row=cur, column=2 + ci, value=val / divisor)
                cell.number_format = fmt
                cell.font = LABEL_FONT
                cell.comment = Comment(f"{src}; raw={val:,.0f}", "MDA-extract")
            cur += 1
        cur += 1

    return cur


def render_narrative_section(ws, start_row, narrative):
    """Render filer-specific narrative content. Always labeled 'Not Standardized Cross-Ticker'."""
    cur = start_row
    ws.cell(row=cur, column=1, value="X. FILER-SPECIFIC (NOT STANDARDIZED CROSS-TICKER)").font = SECTION_FONT
    ws.cell(row=cur, column=1).border = TOP_BORDER
    cur += 1
    ws.cell(row=cur, column=1, value="Hand-curated content from mda_narrative.json. Filer-specific by design.").font = Font(italic=True, size=9, color="595959")
    cur += 2
    if not narrative:
        ws.cell(row=cur, column=1, value="(empty — populate Ticker Libraries/{TICKER}/MDA and Other/mda_narrative.json)").font = LABEL_INDENT
        return cur + 1
    for section_key, section_content in narrative.items():
        ws.cell(row=cur, column=1, value=section_key).font = LABEL_BOLD
        cur += 1
        ws.cell(row=cur, column=1, value=json.dumps(section_content, indent=2)[:500]).font = LABEL_FONT
        cur += 2
    return cur


# ============================================================================
# Main
# ============================================================================

def discover_filings(ticker):
    """Find all 10-K/10-Q iXBRL htm files under Brain\\Sources\\{TICKER}\\."""
    ticker_src = SOURCES / ticker
    if not ticker_src.exists():
        return []
    htms = list(ticker_src.glob("*/filings/*_10-K.htm")) + list(ticker_src.glob("*/filings/*_10-Q.htm"))
    return sorted(htms)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ticker", required=True)
    ap.add_argument("--filings", nargs="*", help="Specific filing htm paths (optional; defaults to auto-discover)")
    ap.add_argument("--out", help="Output xlsx path (default: Model Outputs/{TICKER}/{TICKER}_disclosures.xlsx)")
    args = ap.parse_args()

    ticker = args.ticker.upper()

    # Load generic library
    generic_lib = json.loads(GENERIC_LIB_PATH.read_text(encoding="utf-8"))
    tables = generic_lib["disclosure_tables"]

    # Load per-ticker overrides + narrative
    ticker_lib = TICKER_LIBS / ticker / "MDA and Other"
    overrides_path = ticker_lib / "disclosure_overrides.json"
    narrative_path = ticker_lib / "mda_narrative.json"
    overrides = json.loads(overrides_path.read_text(encoding="utf-8")) if overrides_path.exists() else {}
    narrative = json.loads(narrative_path.read_text(encoding="utf-8")) if narrative_path.exists() else {}

    # Discover or load filings
    filings = [Path(p) for p in args.filings] if args.filings else discover_filings(ticker)
    if not filings:
        raise SystemExit(f"No filings found for {ticker}")
    print(f"[mda-disaggregation] {ticker}: {len(filings)} filings to process")

    # Parse all filings into a single fact pool
    all_facts = []
    for htm in filings:
        print(f"  parsing {htm.name}...")
        try:
            facts = parse_filing(htm)
            print(f"    {len(facts)} numeric facts")
            all_facts.extend(facts)
        except Exception as e:
            print(f"    SKIPPED: {e}")

    # Extract each disclosure table
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Disclosures")
    ws["A1"] = f"{ticker} -- Structural Disclosure Tables"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Source: {ticker} 10-K + 10-Q iXBRL filings. Axis-driven extraction. $ in filer's reporting unit."
    ws["A2"].font = Font(italic=True, size=9, color="595959")

    cur_row = 4
    all_novels = {}
    display = overrides.get("_display", {})

    # Build period_label -> end_date map for chronological sort
    period_to_date = {}
    for f in all_facts:
        period_to_date.setdefault(f["period_label"], f["end_date"])

    # Drop only pure aggregation periods (H1/H2 = 6-month overlap; 9M YTD = 9-month overlap)
    # because they duplicate Q1+Q2 / Q1+Q2+Q3 data already present as single quarters.
    # Keep FY, Q1/Q2/Q3, and any non-aggregation periods (acquisition stub durations,
    # custom close-date contexts) — these carry unique data (e.g., M&A pro forma).
    def is_canonical(label):
        if label.startswith("H1 ") or label.startswith("H2 ") or label.startswith("9M "):
            return False
        return True

    for table_id, table_def in tables.items():
        table_overrides = overrides.get(table_id, {})
        if table_overrides.get("skip"):
            continue
        buckets, novels = extract_disclosure_table(all_facts, table_def, table_overrides)
        # Filter buckets to canonical periods only
        for metric_key, metric_data in buckets.items():
            for member, period_data in list(metric_data.items()):
                metric_data[member] = {p: v for p, v in period_data.items() if is_canonical(p)}
                if not metric_data[member]:
                    del metric_data[member]
        all_novels[table_id] = novels
        all_periods = set()
        for metric_data in buckets.values():
            for member_data in metric_data.values():
                all_periods.update(member_data.keys())
        if not buckets or not all_periods:
            continue
        cur_row = render_disclosure_section(ws, cur_row, table_id, table_def, buckets, all_periods, display, period_to_date) + 2

    # Render narrative section
    cur_row = render_narrative_section(ws, cur_row, narrative)

    # Column widths + freeze
    ws.column_dimensions["A"].width = 36
    for ci in range(2, 20):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.freeze_panes = "B5"

    # Save
    out_path = Path(args.out) if args.out else (MODEL_OUTPUTS / ticker / f"{ticker}_disclosures.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[mda-disaggregation] wrote {out_path}")

    # Print novel summary
    for table_id, novels in all_novels.items():
        if novels:
            print(f"  NOVELS in {table_id}: {sorted(novels)[:10]}")


if __name__ == "__main__":
    main()
