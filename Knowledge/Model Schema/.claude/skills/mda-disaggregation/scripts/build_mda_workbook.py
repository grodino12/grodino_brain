"""Build {TICKER}_MDA.xlsx from data/{TICKER}.json.

Usage: python build_mda_workbook.py --ticker CELH
"""
import argparse
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ---- styles ----
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(italic=True, size=9, color='595959')
SECTION_FONT = Font(bold=True, size=11, color='1F4E79')
PERIOD_HDR = Font(bold=True, size=10)
LABEL_FONT = Font(size=10)
LABEL_BOLD = Font(bold=True, size=10)
LABEL_INDENT = Font(size=10, color='404040')
SUBTOTAL_FONT = Font(bold=True, size=10, color='1F4E79')
FORMULA_FONT = Font(size=10, italic=True, color='595959')
SOURCE_FONT = Font(size=9, color='808080', italic=True)
ND_FONT = Font(size=10, color='808080', italic=True)
TOP_BORDER = Border(top=Side(border_style='thin', color='000000'))
BOTTOM_BORDER = Border(bottom=Side(border_style='thin', color='000000'))
ACCT = '#,##0;(#,##0);"--"'
PCT = '0.0%;(0.0%);"--"'

# Canonical SG&A walk taxonomy (drivers split into M&S vs G&A buckets + special)
MS_DRIVERS = [
    ('mkt_invest', 'Marketing investments / campaigns'),
    ('storage', 'Storage / Distribution'),
    ('employee_ms', 'Sales/Marketing Employee'),
    # acquired-brand M&S rows are inserted dynamically per ticker
    ('other_selling', 'Other selling'),
]
GA_DRIVERS = [
    ('admin', 'Administrative expenses'),
    ('acq_integ', 'Acquisition / Integration'),
    # acquired-brand G&A rows inserted dynamically
    ('contingent', 'Contingent Consideration'),
    ('legal_accrual', 'Legal Accrual'),
    ('stock_comp', 'Stock-based Compensation'),
    ('other_admin', 'Other admin'),
]

GP_DRIVER_LABELS = [
    ('raw', 'Raw Mat / Pkg'),
    ('promo', 'Promo Spend'),
    ('freight', 'Freight'),
    ('mix', 'Channel/Pack Mix'),
    ('brand', 'Brand Mix Drag'),
    ('inv', 'Inv Step-Up'),
    ('tariffs', 'Tariffs'),
]

OI_DRIVER_LABELS = [
    ('int_inc', 'Interest Income Δ'),
    ('int_exp', 'Interest Expense Δ'),
    ('rockstar_agency', 'Rockstar Agency Inc Δ'),  # generic: "<acquired_brand>_agency" if needed
    ('fx_other', 'FX / Other Δ'),
]


def build_workbook(ticker: str, data: dict, out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('MD&A')

    periods = data['periods']
    p_cols = {p: get_column_letter(2 + i) for i, p in enumerate(periods)}
    source_col = get_column_letter(2 + len(periods))

    # ---- helpers ----
    def hardcode(ref, val, comment_text=None, fmt=ACCT, font=None):
        c = ws[ref]
        c.value = val
        c.number_format = fmt
        c.font = font or LABEL_FONT
        if comment_text:
            c.comment = Comment(comment_text, 'Source')

    def formula(ref, expr, comment_text=None, fmt=ACCT, font=None):
        c = ws[ref]
        c.value = expr
        c.number_format = fmt
        c.font = font or FORMULA_FONT
        if comment_text:
            c.comment = Comment(comment_text, 'Source')

    def nd(ref, comment_text):
        c = ws[ref]
        c.value = 'n/d'
        c.alignment = Alignment(horizontal='right')
        c.font = ND_FONT
        c.comment = Comment(comment_text, 'Source')

    def section_header(row, text):
        c = ws.cell(row=row, column=1, value=text)
        c.font = SECTION_FONT
        c.border = TOP_BORDER

    def write_period_headers(row):
        ws.cell(row=row, column=1, value='Metric').font = PERIOD_HDR
        for p in periods:
            c = ws.cell(row=row, column=ord(p_cols[p][0]) - ord('A') + 1, value=p)
            c.font = PERIOD_HDR
            c.alignment = Alignment(horizontal='center')
            c.border = BOTTOM_BORDER
            if p.startswith('Q4 '):
                c.font = Font(bold=True, italic=True, size=10, color='1F4E79')
            if p.startswith('FY'):
                c.font = Font(bold=True, size=10, color='1F4E79')
        ws.cell(row=row, column=ord(source_col) - ord('A') + 1, value='Notes').font = PERIOD_HDR

    def label(row, text, bold=False, indent=False):
        c = ws.cell(row=row, column=1, value=('  ' + text) if indent else text)
        c.font = LABEL_BOLD if bold else (LABEL_INDENT if indent else LABEL_FONT)

    def is_quarter(p): return len(p.split()) == 2 and p.split()[0].startswith('Q')
    def is_q4(p): return p.startswith('Q4 ')
    def is_fy(p): return p.startswith('FY')

    def prior_period(p):
        """Same period, prior year (e.g., 'Q3 2025' -> 'Q3 2024')."""
        if is_quarter(p):
            qn, yr = p.split()
            return f'{qn} {int(yr) - 1}'
        if is_fy(p):
            yr = int(p.replace('FY', ''))
            return f'FY{yr - 1}'
        return None

    # ===== TITLE =====
    ws['A1'] = f'{ticker} -- MD&A Revenue & P&L Disaggregation'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = (f'Source: {ticker} 10-Q + 10-K filings. $ in thousands. '
                f'Q4 derived as FY - 9M YTD where derivable. Hover any cell for source note.')
    ws['A2'].font = SUBTITLE_FONT

    # ===========================================================
    # SECTION 1: GEOGRAPHY
    # ===========================================================
    SR = 4
    section_header(SR, '1. GEOGRAPHY -- Revenue ($ thousands, single-period)')
    write_period_headers(SR + 1)

    geo_data = data['geography']
    regions = geo_data['regions']
    geo_first = SR + 2
    for ri, region in enumerate(regions):
        r = geo_first + ri
        label(r, region)
        for p in periods:
            col = p_cols[p]
            if is_q4(p):
                year = p.split()[1]
                fy_col = p_cols[f'FY{year}']
                q1c, q2c, q3c = p_cols[f'Q1 {year}'], p_cols[f'Q2 {year}'], p_cols[f'Q3 {year}']
                formula(f'{col}{r}', f'={fy_col}{r}-{q3c}{r}-{q2c}{r}-{q1c}{r}',
                        f'Q4 {region} derived as FY - Q1 - Q2 - Q3.')
            elif p in geo_data['data'].get(region, {}):
                v, src = geo_data['data'][region][p]
                hardcode(f'{col}{r}', v, src)

    geo_total_row = geo_first + len(regions)
    label(geo_total_row, 'Total Revenue', bold=True)
    for p in periods:
        col = p_cols[p]
        formula(f'{col}{geo_total_row}',
                f'=SUM({col}{geo_first}:{col}{geo_total_row - 1})',
                'Total = sum of regions', font=LABEL_BOLD)
        ws[f'{col}{geo_total_row}'].border = TOP_BORDER
    ws[f'A{geo_total_row}'].border = TOP_BORDER

    # ===========================================================
    # SECTION 2: CUSTOMER CONCENTRATION
    # ===========================================================
    CR = geo_total_row + 3
    section_header(CR, '2. CUSTOMER CONCENTRATION -- % of single-period revenue (only customers > 10% disclosed). Q4 NOT derivable from FY/9M %.')
    write_period_headers(CR + 1)

    cust_data = data.get('customer_concentration')
    if cust_data:
        customers = cust_data['customers']
        cust_first = CR + 2
        for ci, cust in enumerate(customers):
            r = cust_first + ci
            label(r, cust)
            cust_period_data = cust_data['data'].get(cust, {})
            for p in periods:
                col = p_cols[p]
                if is_q4(p):
                    nd(f'{col}{r}', 'Q4 NOT derivable: percentages cannot be subtracted.')
                elif p in cust_period_data:
                    v, src = cust_period_data[p]
                    hardcode(f'{col}{r}', v, src, fmt=PCT)
                else:
                    nd(f'{col}{r}', f'{cust} below 10% disclosure threshold for this period.')

        cust_total_row = cust_first + len(customers)
        label(cust_total_row, 'Total', bold=True)
        for p in periods:
            col = p_cols[p]
            if is_q4(p):
                ws[f'{col}{cust_total_row}'].value = 'n/d'
                ws[f'{col}{cust_total_row}'].font = ND_FONT
                ws[f'{col}{cust_total_row}'].alignment = Alignment(horizontal='right')
            else:
                formula(f'{col}{cust_total_row}',
                        f'=SUM({col}{cust_first}:{col}{cust_total_row - 1})',
                        'Should sum to 100%; n/d cells treated as 0.', fmt=PCT, font=LABEL_BOLD)
            ws[f'{col}{cust_total_row}'].border = TOP_BORDER
        ws[f'A{cust_total_row}'].border = TOP_BORDER
        FR = cust_total_row + 3
    else:
        FR = CR + 3

    # ===========================================================
    # SECTION 3: FUNCTIONAL / PRODUCT CONCENTRATION
    # ===========================================================
    section_header(FR, '3. FUNCTIONAL/PRODUCT CONCENTRATION -- % of revenue. Q4 NOT derivable.')
    write_period_headers(FR + 1)

    fn_data = data.get('functional_concentration', {})
    fed_first = FR + 2
    fed_rows = []
    if 'single_period' in fn_data:
        fed_rows.append(('Single-Period %', fn_data['single_period']))
    if 'ytd' in fn_data:
        fed_rows.append(('YTD %', fn_data['ytd']))

    for fi, (rlbl, rdata) in enumerate(fed_rows):
        r = fed_first + fi
        label(r, rlbl)
        for p in periods:
            col = p_cols[p]
            if is_q4(p):
                nd(f'{col}{r}', 'Q4 NOT derivable from period %.')
            elif p in rdata:
                v, src = rdata[p]
                hardcode(f'{col}{r}', v, src, fmt=PCT)

    BR = (fed_first + len(fed_rows) if fed_rows else FR + 1) + 2

    # ===========================================================
    # SECTION 4: BRAND CONTRIBUTION
    # ===========================================================
    brand_data = data.get('brand_contribution')
    if brand_data:
        section_header(BR, '4. BRAND CONTRIBUTION -- Revenue ($ thousands). Pre-acquisition periods = 100% legacy brand.')
        write_period_headers(BR + 1)

        active_periods = set(brand_data.get('active_periods', []))
        brand_first = BR + 2
        residual_lbl = brand_data.get('residual_brand_label', 'Legacy (residual)')
        acquired = brand_data.get('acquired_brands', [])

        # Row 0: residual brand
        label(brand_first, residual_lbl)
        # Acquired-brand rows
        acq_rows = {}  # name -> row index
        for bi, b in enumerate(acquired):
            r = brand_first + 1 + bi
            acq_rows[b['name']] = r
            label(r, b['name'])

        for p in periods:
            col = p_cols[p]
            if p not in active_periods:
                continue
            # Acquired brands first
            for b in acquired:
                r = acq_rows[b['name']]
                if is_q4(p):
                    year = p.split()[1]
                    fy_col = p_cols[f'FY{year}']
                    q1c, q2c, q3c = p_cols[f'Q1 {year}'], p_cols[f'Q2 {year}'], p_cols[f'Q3 {year}']
                    formula(f'{col}{r}', f'={fy_col}{r}-{q1c}{r}-{q2c}{r}-{q3c}{r}',
                            f'Q4 {b["name"]} derived as FY - Q1 - Q2 - Q3.')
                elif p in b['data']:
                    v, src = b['data'][p]
                    hardcode(f'{col}{r}', v, src)
            # Residual = Total - sum(acquired)
            acq_terms = '-'.join(f'{col}{acq_rows[b["name"]]}' for b in acquired)
            formula(f'{col}{brand_first}', f'={col}{geo_total_row}-{acq_terms}',
                    f'Residual = Total Revenue (Section 1) - acquired brands.')

        # Total link row
        brand_total_row = brand_first + 1 + len(acquired)
        label(brand_total_row, 'Total Revenue (link)', bold=True)
        for p in periods:
            col = p_cols[p]
            if p not in active_periods:
                continue
            formula(f'{col}{brand_total_row}', f'={col}{geo_total_row}',
                    f'Links to Section 1 Total ({col}{geo_total_row}).', font=LABEL_BOLD)
            ws[f'{col}{brand_total_row}'].border = TOP_BORDER
        ws[f'A{brand_total_row}'].border = TOP_BORDER

        # % of first acquired brand
        if acquired:
            pct_row = brand_total_row + 1
            label(pct_row, f'% {acquired[0]["name"]} of Revenue')
            for p in periods:
                col = p_cols[p]
                if p not in active_periods:
                    continue
                formula(f'{col}{pct_row}',
                        f'=IFERROR({col}{acq_rows[acquired[0]["name"]]}/{col}{brand_total_row}, 0)',
                        f'% {acquired[0]["name"]}.', fmt=PCT)
            PR = pct_row + 3
        else:
            PR = brand_total_row + 3
    else:
        PR = BR

    # ===========================================================
    # SECTION 5: PRO FORMA
    # ===========================================================
    pf_data = data.get('pro_forma')
    if pf_data:
        section_header(PR, '5. PRO FORMA vs AS-REPORTED -- Revenue ($ thousands). Pro forma = as if acquisitions closed at start of prior year.')
        ws.cell(row=PR + 1, column=1, value='Metric').font = PERIOD_HDR
        pf_periods = pf_data['periods']
        for i, p in enumerate(pf_periods):
            c = ws.cell(row=PR + 1, column=2 + i, value=p)
            c.font = PERIOD_HDR
            c.alignment = Alignment(horizontal='center')
            c.border = BOTTOM_BORDER
            if p.startswith('FY'):
                c.font = Font(bold=True, size=10, color='1F4E79')

        ws.cell(row=PR + 2, column=1, value='As-Reported Revenue').font = LABEL_FONT
        ws.cell(row=PR + 3, column=1, value='Pro Forma Revenue').font = LABEL_FONT
        ws.cell(row=PR + 4, column=1, value='Implied "Missing"').font = LABEL_FONT
        ws.cell(row=PR + 5, column=1, value='Pro Forma Y/Y %').font = LABEL_FONT
        for i, p in enumerate(pf_periods):
            cl = get_column_letter(2 + i)
            actual_v, actual_src = pf_data['as_reported'][p]
            pf_v, pf_src = pf_data['pro_forma_values'][p]
            hardcode(f'{cl}{PR+2}', actual_v, f'{actual_src}: actual reported revenue.')
            hardcode(f'{cl}{PR+3}', pf_v, f'{pf_src}: pro forma revenue.')
            formula(f'{cl}{PR+4}', f'={cl}{PR+3}-{cl}{PR+2}',
                    'Pro forma - actual = revenue from acquisitions in periods not yet owned.')
            # Y/Y for the latest of each pair (Q3 2025 vs Q3 2024 etc.)
            if i > 0:
                prior_cl = get_column_letter(2 + i - 1)
                # Heuristic: pair years (i=1 with i=0, i=3 with i=2, i=5 with i=4)
                if i % 2 == 1:
                    formula(f'{cl}{PR+5}', f'={cl}{PR+3}/{prior_cl}{PR+3}-1',
                            'Pro forma YoY growth (apples-to-apples).', fmt=PCT)
        SG = PR + 8
    else:
        SG = PR

    # ===========================================================
    # SECTION 6: SG&A WALK (hierarchical M&S vs G&A)
    # ===========================================================
    sga = data['sga_walk']
    section_header(SG, '6. SG&A WALK -- $ thousands. Driver Δ rows are period-over-period change. Q4 derived from FY - 9M YTD where both walks disclose the bucket.')
    write_period_headers(SG + 1)

    # Determine acquired-brand keys (e.g., alani_ms / alani_ga / rockstar_ms / rockstar_ga)
    ms_acq_keys = set()
    ga_acq_keys = set()
    for period_dict in list(sga.get('data', {}).values()) + list(sga.get('ytd_data', {}).values()):
        for k in period_dict.keys():
            if k.endswith('_ms') and k != 'employee_ms':
                ms_acq_keys.add(k)
            if k.endswith('_ga'):
                ga_acq_keys.add(k)

    # Build full ms_drivers / ga_drivers including dynamic acquired-brand keys
    ms_drivers = list(MS_DRIVERS)
    # insert acquired-brand M&S rows just before "other_selling"
    for k in sorted(ms_acq_keys):
        brand_name = k[:-3].replace('_', ' ').title()
        ms_drivers.insert(-1, (k, f'{brand_name} attributable (M&S)'))
    ga_drivers = list(GA_DRIVERS)
    # insert acquired-brand G&A rows just after "acq_integ"
    insert_at = next((i for i, (k, _) in enumerate(ga_drivers) if k == 'acq_integ'), 1) + 1
    for k in sorted(ga_acq_keys):
        brand_name = k[:-3].replace('_', ' ').title()
        ga_drivers.insert(insert_at, (k, f'{brand_name} attributable (G&A)'))

    # Helper to derive Q4 walk from FY - 9M YTD
    walk_data = sga.get('data', {})
    ytd_walk_data = sga.get('ytd_data', {})

    def q4_walk_value(year, key):
        fy = walk_data.get(f'FY{year}', {}).get(key)
        ytd = ytd_walk_data.get(f'9M {year}', {}).get(key)
        if fy is None and ytd is None:
            return None
        if fy is None:
            return -ytd
        if ytd is None:
            return fy  # all in Q4 (assumes 9M didn't break it out)
        return fy - ytd

    # Row: Total SG&A
    sga_total_r = SG + 2
    label(sga_total_r, 'Total SG&A', bold=True)
    for p in periods:
        col = p_cols[p]
        if is_q4(p):
            year = p.split()[1]
            fy_col = p_cols[f'FY{year}']
            q1c, q2c, q3c = p_cols[f'Q1 {year}'], p_cols[f'Q2 {year}'], p_cols[f'Q3 {year}']
            formula(f'{col}{sga_total_r}', f'={fy_col}{sga_total_r}-{q3c}{sga_total_r}-{q2c}{sga_total_r}-{q1c}{sga_total_r}',
                    f'Q4 {year} SG&A = FY - Q1 - Q2 - Q3.', font=Font(bold=True, italic=True, size=10, color='595959'))
        elif p in sga.get('totals', {}):
            v, src = sga['totals'][p]
            hardcode(f'{col}{sga_total_r}', v, src, font=LABEL_BOLD)

    # Row: YoY Total Δ (formula vs prior period same)
    sga_yoy_r = sga_total_r + 1
    label(sga_yoy_r, 'YoY Total Δ')
    for p in periods:
        col = p_cols[p]
        prior = prior_period(p)
        if prior and prior in periods:
            prior_col = p_cols[prior]
            formula(f'{col}{sga_yoy_r}', f'={col}{sga_total_r}-{prior_col}{sga_total_r}',
                    f'YoY Δ vs {prior}.')

    # ---- M&S subtotal block ----
    ms_subtotal_r = sga_yoy_r + 2
    label(ms_subtotal_r, 'MARKETING & SELLING Δ', bold=True)
    ws[f'A{ms_subtotal_r}'].font = SUBTOTAL_FONT
    ws[f'A{ms_subtotal_r}'].border = TOP_BORDER

    ms_first_sub = ms_subtotal_r + 1
    for di, (k, lbl) in enumerate(ms_drivers):
        r = ms_first_sub + di
        label(r, lbl, indent=True)
        for p in periods:
            col = p_cols[p]
            if is_q4(p):
                year = p.split()[1]
                v = q4_walk_value(year, k)
                if v is not None:
                    fy_walk = walk_data.get(f'FY{year}', {})
                    ytd_walk = ytd_walk_data.get(f'9M {year}', {})
                    src = (f'Q4 {year} {lbl} derived as FY - 9M YTD. '
                           f'FY src: {fy_walk.get("src", "")}; 9M src: {ytd_walk.get("src", "")}.')
                    hardcode(f'{col}{r}', v, src)
            elif p in walk_data and k in walk_data[p]:
                v = walk_data[p][k]
                hardcode(f'{col}{r}', v, walk_data[p].get('src', ''))

    # M&S subtotal formula (sum of M&S sub-rows for each period)
    for p in periods:
        col = p_cols[p]
        formula(f'{col}{ms_subtotal_r}',
                f'=SUM({col}{ms_first_sub}:{col}{ms_first_sub + len(ms_drivers) - 1})',
                'M&S subtotal = sum of M&S driver rows.', font=SUBTOTAL_FONT)

    # ---- G&A subtotal block ----
    ga_subtotal_r = ms_first_sub + len(ms_drivers) + 1
    label(ga_subtotal_r, 'GENERAL & ADMIN Δ', bold=True)
    ws[f'A{ga_subtotal_r}'].font = SUBTOTAL_FONT
    ws[f'A{ga_subtotal_r}'].border = TOP_BORDER

    ga_first_sub = ga_subtotal_r + 1
    for di, (k, lbl) in enumerate(ga_drivers):
        r = ga_first_sub + di
        label(r, lbl, indent=True)
        for p in periods:
            col = p_cols[p]
            if is_q4(p):
                year = p.split()[1]
                v = q4_walk_value(year, k)
                if v is not None:
                    fy_walk = walk_data.get(f'FY{year}', {})
                    ytd_walk = ytd_walk_data.get(f'9M {year}', {})
                    src = (f'Q4 {year} {lbl} derived as FY - 9M YTD. '
                           f'FY src: {fy_walk.get("src", "")}; 9M src: {ytd_walk.get("src", "")}.')
                    hardcode(f'{col}{r}', v, src)
            elif p in walk_data and k in walk_data[p]:
                v = walk_data[p][k]
                hardcode(f'{col}{r}', v, walk_data[p].get('src', ''))

    for p in periods:
        col = p_cols[p]
        formula(f'{col}{ga_subtotal_r}',
                f'=SUM({col}{ga_first_sub}:{col}{ga_first_sub + len(ga_drivers) - 1})',
                'G&A subtotal = sum of G&A driver rows.', font=SUBTOTAL_FONT)

    # ---- Distributor Term Δ (special, in-SG&A era) ----
    distrib_r = ga_first_sub + len(ga_drivers) + 1
    label(distrib_r, 'Distributor Term Δ (in-SG&A era)', bold=True)
    ws[f'A{distrib_r}'].border = TOP_BORDER
    for p in periods:
        col = p_cols[p]
        if is_q4(p):
            year = p.split()[1]
            v = q4_walk_value(year, 'distrib_term')
            if v is not None:
                src = f'Q4 {year} distrib term derived as FY - 9M YTD.'
                hardcode(f'{col}{distrib_r}', v, src)
        elif p in walk_data and 'distrib_term' in walk_data[p]:
            v = walk_data[p]['distrib_term']
            hardcode(f'{col}{distrib_r}', v, walk_data[p].get('src', ''))

    # ---- Σ Buckets check ----
    sigma_r = distrib_r + 1
    label(sigma_r, 'Σ Buckets (check)')
    ws[f'A{sigma_r}'].border = TOP_BORDER
    for p in periods:
        col = p_cols[p]
        formula(f'{col}{sigma_r}',
                f'={col}{ms_subtotal_r}+{col}{ga_subtotal_r}+{col}{distrib_r}',
                'Σ check = M&S + G&A + Distributor Term Δ; should approximate YoY Total Δ row above.')

    # ---- Distributor Term Fees (separate IS line, post-relocation era) ----
    dt_sep = sga.get('distributor_term_separate', {})
    if dt_sep:
        dt_sep_r = sigma_r + 2
        label(dt_sep_r, 'Distributor Term Fees (IS line, post-2025)', bold=True)
        for p in periods:
            col = p_cols[p]
            if p in dt_sep:
                v, src = dt_sep[p]
                hardcode(f'{col}{dt_sep_r}', v, src, font=LABEL_BOLD)
            elif is_q4(p):
                year = p.split()[1]
                fy_p = f'FY{year}'
                q3_p = f'Q3 {year}'
                if fy_p in dt_sep and q3_p in dt_sep:
                    fy_col = p_cols[fy_p]
                    q3_col = p_cols[q3_p]
                    formula(f'{col}{dt_sep_r}', f'={fy_col}{dt_sep_r}-{q3_col}{dt_sep_r}',
                            f'Q4 {year} = FY - Q3 (assumes Q1+Q2 = 0).')
        GR = dt_sep_r + 3
    else:
        GR = sigma_r + 3

    # ===========================================================
    # SECTION 7: GROSS PROFIT & MARGIN
    # ===========================================================
    gp = data.get('gp', {})
    section_header(GR, '7. GROSS PROFIT & MARGIN -- $ thousands + %. Drivers qualitative (CELH does NOT disclose $ bridges for GP).')
    write_period_headers(GR + 1)

    gp_dollar_r = GR + 2
    label(gp_dollar_r, 'Gross Profit $')
    for p in periods:
        col = p_cols[p]
        if is_q4(p):
            year = p.split()[1]
            fy_col = p_cols[f'FY{year}']
            q1c, q2c, q3c = p_cols[f'Q1 {year}'], p_cols[f'Q2 {year}'], p_cols[f'Q3 {year}']
            formula(f'{col}{gp_dollar_r}', f'={fy_col}{gp_dollar_r}-{q3c}{gp_dollar_r}-{q2c}{gp_dollar_r}-{q1c}{gp_dollar_r}',
                    'Q4 GP $ = FY - Q1 - Q2 - Q3.')
        elif p in gp.get('totals', {}):
            v, src = gp['totals'][p]
            hardcode(f'{col}{gp_dollar_r}', v, src)

    gp_margin_r = gp_dollar_r + 1
    label(gp_margin_r, 'GP Margin %', bold=True)
    for p in periods:
        col = p_cols[p]
        formula(f'{col}{gp_margin_r}', f'=IFERROR({col}{gp_dollar_r}/{col}{geo_total_row}, 0)',
                f'GP Margin = Gross Profit $ / Total Revenue (row {geo_total_row}).',
                fmt=PCT, font=Font(bold=True, size=10, italic=True, color='595959'))

    gp_yoy_r = gp_margin_r + 1
    label(gp_yoy_r, 'YoY Δ pts')
    for p in periods:
        col = p_cols[p]
        prior = prior_period(p)
        if prior and prior in periods:
            prior_col = p_cols[prior]
            formula(f'{col}{gp_yoy_r}', f'={col}{gp_margin_r}-{prior_col}{gp_margin_r}',
                    f'YoY pts Δ vs {prior}.', fmt=PCT)

    gp_drivers = gp.get('driver_data', {})
    gp_driver_first = gp_yoy_r + 1
    for di, (k, lbl) in enumerate(GP_DRIVER_LABELS):
        r = gp_driver_first + di
        label(r, lbl)
        for p in periods:
            col = p_cols[p]
            if is_q4(p):
                continue
            if p in gp_drivers and k in gp_drivers[p]:
                sym = gp_drivers[p][k]
                c = ws[f'{col}{r}']
                c.value = sym
                c.alignment = Alignment(horizontal='center')
                c.font = Font(bold=True, size=11, color='2E7D32' if sym == '+' else 'C62828')
                quote = gp_drivers[p].get('q', '')
                src = gp_drivers[p].get('src', '')
                c.comment = Comment(f'{src}: {quote}', 'Source')

    OI = gp_driver_first + len(GP_DRIVER_LABELS) + 2

    # ===========================================================
    # SECTION 8: OTHER INC/EXP WALK
    # ===========================================================
    oi = data.get('other_inc_exp', {})
    section_header(OI, '8. OTHER (EXPENSE) / INCOME WALK -- $ thousands.')
    write_period_headers(OI + 1)

    oi_total_r = OI + 2
    label(oi_total_r, 'Total Other Inc/(Exp)', bold=True)
    for p in periods:
        col = p_cols[p]
        if is_q4(p):
            year = p.split()[1]
            fy_col = p_cols[f'FY{year}']
            q1c, q2c, q3c = p_cols[f'Q1 {year}'], p_cols[f'Q2 {year}'], p_cols[f'Q3 {year}']
            formula(f'{col}{oi_total_r}', f'={fy_col}{oi_total_r}-{q3c}{oi_total_r}-{q2c}{oi_total_r}-{q1c}{oi_total_r}',
                    'Q4 = FY - Q1 - Q2 - Q3.')
        elif p in oi.get('totals', {}):
            v, src = oi['totals'][p]
            hardcode(f'{col}{oi_total_r}', v, src, font=LABEL_BOLD)

    oi_walk = oi.get('data', {})
    oi_walk_first = oi_total_r + 1
    for wi, (k, lbl) in enumerate(OI_DRIVER_LABELS):
        r = oi_walk_first + wi
        label(r, lbl)
        for p in periods:
            col = p_cols[p]
            if p in oi_walk and k in oi_walk[p]:
                v = oi_walk[p][k]
                hardcode(f'{col}{r}', v, oi_walk[p].get('src', ''))

    # ---- column widths + freeze ----
    ws.column_dimensions['A'].width = 32
    for p in periods:
        ws.column_dimensions[p_cols[p]].width = 12
    ws.column_dimensions[source_col].width = 40
    ws.freeze_panes = 'B6'

    # save
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--ticker', required=True)
    parser.add_argument('--data-dir', default=None,
                        help='Path to data/ folder. Defaults to ../data relative to this script.')
    parser.add_argument('--out', default=None,
                        help=r'Output xlsx path. Defaults to Brain\Knowledge\Model Outputs\{TICKER}\{TICKER}_MDA.xlsx')
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    data_dir = Path(args.data_dir) if args.data_dir else script_dir.parent / 'data'
    json_path = data_dir / f'{args.ticker}.json'
    if not json_path.exists():
        raise SystemExit(f'JSON not found: {json_path}')

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if args.out:
        out = Path(args.out)
    else:
        # default: Brain\Knowledge\Model Outputs\{TICKER}\{TICKER}_MDA.xlsx
        # script lives at ...\Model Schema\.claude\skills\mda-disaggregation\scripts\
        # so go up 4 levels to Model Schema, then over to Model Outputs
        model_schema = script_dir.parent.parent.parent.parent
        out = model_schema.parent / 'Model Outputs' / args.ticker / f'{args.ticker}_MDA.xlsx'

    build_workbook(args.ticker, data, out)
    print(f'Built {out}')
    print(f'Size: {out.stat().st_size} bytes')


if __name__ == '__main__':
    main()
