"""
model-qtr-derive — add 3 single-quarter sheets to a workbook produced by
model-write. Renames the existing YTD-shaped quarterly sheets to
`YTD P&L` / `YTD BS` / `YTD CF` and writes new `QTR P&L` / `QTR BS` / `QTR CF`
sheets containing cell-reference formulas that decompose YTD into single-
quarter values.

CLI:
    python build.py --in <workbook.xlsx>

Importable:
    from build import derive_quarterly
    derive_quarterly(Path("workbook.xlsx"))
"""
from __future__ import annotations

import argparse
import re
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Force utf-8 stdout so PG / PEP / CELH labels with em-dashes don't crash
# Windows cp1252.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ============================================================================
# Sheet name conventions
# ============================================================================

ANNL_PL    = "ANNL P&L"
ANNL_BS    = "BALANCE SHEET"
ANNL_CF    = "CASH FLOW"

YTD_PL     = "YTD P&L"
YTD_BS     = "YTD BS"
YTD_CF     = "YTD CF"

QTR_PL_OLD = "QTR P&L"
QTR_BS_OLD = "QTR BS"
QTR_CF_OLD = "QTR CF"

# New sheets reuse the QTR_*_OLD names (clean name to the user-facing sheet).

STATEMENT_TRIPLES = [
    # (annl_sheet, ytd_sheet, new_qtr_sheet, statement_kind)
    (ANNL_PL, YTD_PL, "QTR P&L", "is"),
    (ANNL_BS, YTD_BS, "QTR BS",  "bs"),
    (ANNL_CF, YTD_CF, "QTR CF",  "cf"),
]

# ============================================================================
# Format constants — kept in sync with model-write/scripts/write.py
# ============================================================================

LINE_ITEM_FMT  = '#,##0;(#,##0);"--"'
SUBTOTAL_FMT   = '$#,##0_);($#,##0);"$--"_)'
EPS_FMT        = '$#,##0.00_);($#,##0.00)'

HEADER_FONT    = Font(bold=True, color="FFFFFF")
HEADER_FILL    = PatternFill(fill_type="solid", start_color="203864", end_color="203864")
HEADER_ALIGN   = Alignment(horizontal="center")
LABEL_FONT     = Font(bold=True)

# ============================================================================
# Header parsers
# ============================================================================

ANNL_HDR_RE = re.compile(r"^FY(\d{4})$")              # "FY2023" — historical only
QTR_HDR_RE  = re.compile(r"^Q([1-4])\s+FY(\d{4})$")   # "Q2 FY2024"


def parse_annl_columns(ws: Worksheet) -> dict[int, int]:
    """Map fiscal_year -> excel column index for ANNL sheets. Forecast columns
    (e.g. FY2025E) skipped — we only derive Q4 from real historical data."""
    out: dict[int, int] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        m = ANNL_HDR_RE.match(str(cell.value).strip())
        if m:
            out[int(m.group(1))] = cell.column
    return out


def parse_ytd_columns(ws: Worksheet) -> dict[tuple[int, int], int]:
    """Map (fiscal_year, fiscal_quarter) -> excel column index for YTD sheets."""
    out: dict[tuple[int, int], int] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        m = QTR_HDR_RE.match(str(cell.value).strip())
        if m:
            q = int(m.group(1))
            y = int(m.group(2))
            out[(y, q)] = cell.column
    return out


# ============================================================================
# EPS / share-count detection (skip these rows on the new sheets)
# ============================================================================

def is_eps_or_share_label(label: str | None) -> bool:
    if not label:
        return False
    s = label.lower()
    if "shares" in s:
        return True   # "Weighted Average Shares Outstanding"
    if "per share" in s or "eps" in s:
        return True
    return False


# ============================================================================
# Target column resolution
# ============================================================================

def resolve_target_columns(
    annl_years: set[int],
    ytd_qcols: set[tuple[int, int]],
    statement_kind: str,
) -> list[tuple[int, int]]:
    """Return the chronological list of (year, quarter) tuples to emit on the
    new sheet. A column is included iff its formula dependencies exist:

      Q1 needs YTD[(y,1)]
      Q2 needs YTD[(y,1)] AND YTD[(y,2)]
      Q3 needs YTD[(y,2)] AND YTD[(y,3)]
      Q4 IS/CF needs ANNL[y] AND YTD[(y,3)]
      Q4 BS needs ANNL[y]              (BS = point-in-time, no subtraction)
      Q1/Q2/Q3 BS needs YTD[(y,N)]     (instant balance carries through)
    """
    years = sorted({y for (y, _) in ytd_qcols} | annl_years)
    out: list[tuple[int, int]] = []
    for y in years:
        if statement_kind == "bs":
            for q in (1, 2, 3):
                if (y, q) in ytd_qcols:
                    out.append((y, q))
            if y in annl_years:
                out.append((y, 4))
        else:
            # IS / CF
            if (y, 1) in ytd_qcols:
                out.append((y, 1))
            if (y, 2) in ytd_qcols and (y, 1) in ytd_qcols:
                out.append((y, 2))
            if (y, 3) in ytd_qcols and (y, 2) in ytd_qcols:
                out.append((y, 3))
            if y in annl_years and (y, 3) in ytd_qcols:
                out.append((y, 4))
    return out


# ============================================================================
# Per-cell formula builder
# ============================================================================

def build_formula(
    statement_kind: str,
    year: int,
    quarter: int,
    ytd_row: int,
    annl_row: int | None,
    ytd_sheet_name: str,
    annl_sheet_name: str,
    ytd_qcols: dict[tuple[int, int], int],
    annl_years: dict[int, int],
) -> str | None:
    """Return the formula string for one cell, or None if dependencies missing.

    `ytd_row` is the row index on the YTD sheet (matches the new QTR sheet by
    construction, since the new sheet mirrors YTD's row layout).

    `annl_row` is the corresponding row index on the ANNL sheet, looked up BY
    LABEL by the caller. May be None when the line item appears in YTD-only
    statements (e.g. PEP's "Product Recall Impact" — quarterly disclosure that
    was rolled into a different annual line). Q4 cells return None when
    `annl_row is None` because we can't decompose without an annual anchor."""

    def ytd_ref(q: int) -> str:
        col = ytd_qcols.get((year, q))
        if col is None:
            return ""
        return f"'{ytd_sheet_name}'!{get_column_letter(col)}{ytd_row}"

    def annl_ref() -> str:
        if annl_row is None:
            return ""
        col = annl_years.get(year)
        if col is None:
            return ""
        return f"'{annl_sheet_name}'!{get_column_letter(col)}{annl_row}"

    if statement_kind == "bs":
        if quarter in (1, 2, 3):
            ref = ytd_ref(quarter)
            return f"={ref}" if ref else None
        if quarter == 4:
            ref = annl_ref()
            return f"={ref}" if ref else None
        return None

    # IS / CF
    if quarter == 1:
        ref = ytd_ref(1)
        return f"={ref}" if ref else None
    if quarter == 2:
        a, b = ytd_ref(2), ytd_ref(1)
        return f"={a}-{b}" if (a and b) else None
    if quarter == 3:
        a, b = ytd_ref(3), ytd_ref(2)
        return f"={a}-{b}" if (a and b) else None
    if quarter == 4:
        a, b = annl_ref(), ytd_ref(3)
        return f"={a}-{b}" if (a and b) else None
    return None


# ============================================================================
# Main per-statement build
# ============================================================================

def build_quarterly_sheet(
    wb,
    annl_sheet_name: str,
    ytd_sheet_name: str,
    new_sheet_name: str,
    statement_kind: str,
) -> dict | None:
    """Build one new single-quarter sheet referencing ANNL + YTD. Returns a
    summary dict, or None if YTD sheet is empty / missing."""

    if ytd_sheet_name not in wb.sheetnames:
        return None
    ytd_ws  = wb[ytd_sheet_name]
    # ANNL sheet may legitimately be missing on a 10-Q-only run; in that case
    # only Q1/Q2/Q3 derivations work (no Q4).
    annl_ws = wb[annl_sheet_name] if annl_sheet_name in wb.sheetnames else None

    annl_years = parse_annl_columns(annl_ws) if annl_ws is not None else {}
    ytd_qcols  = parse_ytd_columns(ytd_ws)

    if not ytd_qcols and not annl_years:
        return None

    # Build label -> row map on the ANNL sheet so Q4 formulas can look up the
    # right ANNL row for each YTD label. ANNL and YTD often have different row
    # layouts (e.g. PEP's CASH FLOW has "Indirect Tax Impact" only on annual,
    # plus reordered rows around it) — using YTD's row index against ANNL would
    # subtract the wrong line item.
    annl_label_to_row: dict[str, int] = {}
    if annl_ws is not None:
        for r in range(2, annl_ws.max_row + 1):
            v = annl_ws.cell(row=r, column=1).value
            if v is None:
                continue
            label = str(v).strip()
            if label and label not in annl_label_to_row:
                annl_label_to_row[label] = r

    target_cols = resolve_target_columns(set(annl_years.keys()), set(ytd_qcols.keys()), statement_kind)
    if not target_cols:
        return None

    # If a sheet with this name already exists (re-run), drop it.
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]
    new_ws = wb.create_sheet(new_sheet_name)

    # --- Header row ---
    new_ws.cell(row=1, column=1, value="").fill = HEADER_FILL
    new_ws.cell(row=1, column=1).font = HEADER_FONT
    target_col_idx: dict[tuple[int, int], int] = {}
    for col_idx, (y, q) in enumerate(target_cols, start=2):
        c = new_ws.cell(row=1, column=col_idx, value=f"Q{q} FY{y}")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        target_col_idx[(y, q)] = col_idx

    # --- Walk YTD sheet's rows; mirror to new sheet with formulas ---
    # Iterate the full row range — column A holds the row label; rows 2+ are
    # data rows (subtotals included). Row 1 is the header (already written).
    ytd_max_row = ytd_ws.max_row
    rows_written = 0

    for row_idx in range(2, ytd_max_row + 1):
        label_cell = ytd_ws.cell(row=row_idx, column=1)
        label = label_cell.value
        if label is None or str(label).strip() == "":
            # No label: skip — model-write doesn't emit content rows without a
            # column-A label.
            continue

        # Mirror the label + label font.
        new_label = new_ws.cell(row=row_idx, column=1, value=label)
        if label_cell.font is not None:
            new_label.font = copy(label_cell.font)

        if is_eps_or_share_label(str(label)):
            # Skip cell formulas; leave blank. Per user direction (ignore EPS
            # in v1) and because per-share metrics aren't linearly subtractable.
            continue

        # Sample number format from a populated YTD cell on this row, so we
        # can reapply the same format on the new sheet (line item vs subtotal
        # vs EPS — though EPS is already handled above).
        sample_fmt = LINE_ITEM_FMT
        sample_font = None
        sample_border = None
        for ytd_col in ytd_qcols.values():
            sample_cell = ytd_ws.cell(row=row_idx, column=ytd_col)
            if sample_cell.value is not None:
                if sample_cell.number_format and sample_cell.number_format != "General":
                    sample_fmt = sample_cell.number_format
                if sample_cell.font is not None and sample_cell.font.bold:
                    sample_font = copy(sample_cell.font)
                if sample_cell.border is not None:
                    sample_border = copy(sample_cell.border)
                break

        annl_row = annl_label_to_row.get(str(label).strip())

        for (y, q), col_idx in target_col_idx.items():
            formula = build_formula(
                statement_kind, y, q,
                ytd_row=row_idx,
                annl_row=annl_row,
                ytd_sheet_name=ytd_sheet_name,
                annl_sheet_name=annl_sheet_name,
                ytd_qcols=ytd_qcols,
                annl_years=annl_years,
            )
            if formula is None:
                continue
            cell = new_ws.cell(row=row_idx, column=col_idx, value=formula)
            cell.number_format = sample_fmt
            if sample_font is not None:
                cell.font = sample_font
            if sample_border is not None:
                cell.border = sample_border

        rows_written += 1

    # --- Column widths + freeze panes ---
    new_ws.column_dimensions["A"].width = 42
    for col_idx in target_col_idx.values():
        new_ws.column_dimensions[get_column_letter(col_idx)].width = 13
    new_ws.freeze_panes = "B2"

    return {
        "sheet": new_sheet_name,
        "rows": rows_written,
        "columns": len(target_cols),
        "target_columns": [f"Q{q} FY{y}" for (y, q) in target_cols],
    }


# ============================================================================
# Orchestrator
# ============================================================================

def derive_quarterly(workbook_path: Path) -> dict:
    """Open the workbook, rename the existing QTR sheets to YTD, and add 3 new
    single-quarter sheets. Saves in place. Idempotent — a re-run drops and
    rebuilds the new sheets (the rename step is skipped on the second pass
    because the QTR_*_OLD sheets no longer exist by their old names)."""
    wb = load_workbook(workbook_path)

    # Step 1: rename QTR -> YTD (skip if already renamed).
    rename_pairs = [
        (QTR_PL_OLD, YTD_PL),
        (QTR_BS_OLD, YTD_BS),
        (QTR_CF_OLD, YTD_CF),
    ]
    for old_name, new_name in rename_pairs:
        if old_name in wb.sheetnames and new_name not in wb.sheetnames:
            wb[old_name].title = new_name

    # Step 1.5: hide the YTD sheets — they're still referenced by QTR formulas
    # but shouldn't clutter the workbook UI. 'hidden' is user-unhideable via
    # right-click; 'veryHidden' would require VBA. Idempotent.
    for ytd_name in (YTD_PL, YTD_BS, YTD_CF):
        if ytd_name in wb.sheetnames:
            wb[ytd_name].sheet_state = "hidden"

    # Step 2: build the 3 new sheets.
    summaries = []
    for annl, ytd, new_qtr, kind in STATEMENT_TRIPLES:
        summary = build_quarterly_sheet(wb, annl, ytd, new_qtr, kind)
        if summary is not None:
            summaries.append(summary)

    wb.save(workbook_path)
    return {"workbook": str(workbook_path), "sheets": summaries}


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Add single-quarter sheets to a model-write workbook.",
    )
    parser.add_argument("--in", dest="in_path", required=True, type=Path,
                        help="Workbook xlsx (edited in place)")
    args = parser.parse_args()

    if not args.in_path.exists():
        print(f"ERROR: {args.in_path} not found", file=sys.stderr)
        sys.exit(2)

    report = derive_quarterly(args.in_path)
    print(f"model-qtr-derive: added single-quarter sheets to {report['workbook']}")
    for s in report["sheets"]:
        cols_preview = ", ".join(s["target_columns"][:3])
        if len(s["target_columns"]) > 3:
            cols_preview += f", … ({len(s['target_columns'])} total)"
        print(f"  {s['sheet']:8s}: {s['rows']} rows x {s['columns']} quarters  [{cols_preview}]")


if __name__ == "__main__":
    main()
