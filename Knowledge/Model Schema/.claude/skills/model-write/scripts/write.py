"""model-write CLI — build a fresh xlsx financial model from ValidatedFiling JSONs."""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from financials_schema import (
    FilingType,
    MappedLineItem,
    Period,
    StatementType,
    Unit,
    ValidatedFiling,
    keep_statement_for_pipeline,
)


# ============================================================================
# Unit normalization
# ============================================================================
#
# The workbook's canonical monetary scale is THOUSANDS. Filers report at
# different scales (CELH filed 2021 10-Qs in whole dollars then switched to
# thousands; PG/PEP report in millions). The extractor records the true scale
# per statement in `stmt.unit`; this is the single place that normalizes every
# monetary value onto the thousands scale before it lands in a cell. Without
# this, a workbook assembled from mixed-scale filings (e.g. CELH) carries
# columns that differ by 1000x.

_UNIT_TO_THOUSANDS: dict[Unit, Decimal] = {
    Unit.ACTUAL:    Decimal("0.001"),
    Unit.THOUSANDS: Decimal(1),
    Unit.MILLIONS:  Decimal(1000),
    Unit.BILLIONS:  Decimal(1_000_000),
}


def _is_eps_or_share_label(label: str | None) -> bool:
    """EPS and share-count rows are governed by `eps_unit` / `share_unit`, not
    the statement's monetary `unit` — they must NOT be rescaled to thousands."""
    if not label:
        return False
    s = label.lower()
    return "shares" in s or "per share" in s or "eps" in s


# ============================================================================
# Constants
# ============================================================================

N_FORECAST_YEARS = 6  # forecast horizon: 6 fiscal years past the last historical

ANNL_SHEETS = ["ANNL P&L", "BALANCE SHEET", "CASH FLOW"]
QTR_SHEETS = ["QTR P&L", "QTR BS", "QTR CF"]
V1_SHEETS = ANNL_SHEETS + QTR_SHEETS

# CF subtotal rows: label on the xlsx → item section whose rows sum to it.
CF_SUBTOTAL_LABELS = {
    "Cash Flow from Operations": "operating",
    "Cash Flow from Investing":  "investing",
    "Cash Flow from Financing":  "financing",
}

# BS subtotal rows inserted at section boundaries. Each spec:
#   (section_key, label, formula_type, cascade_from_section)
# formula_type: "sum" = SUM of this section only;
#               "cascade" = prior_section_subtotal + SUM of this section;
#               "grand" = TL + mezzanine + TSE (one-off, fields not used).
BS_SUBTOTAL_SPECS = [
    ("current_assets",         "Total Current Assets",         "sum",     None),
    ("non_current_assets",     "Total Assets",                 "cascade", "current_assets"),
    ("current_liabilities",    "Total Current Liabilities",    "sum",     None),
    ("non_current_liabilities","Total Liabilities",            "cascade", "current_liabilities"),
    ("equity",                 "Total Stockholders' Equity",   "sum",     None),
    ("__grand_total__",        "Total Liabilities, Mezzanine & Stockholders' Equity", "grand", None),
]

BS_SECTION_ORDER = [
    "current_assets", "non_current_assets",
    "current_liabilities", "non_current_liabilities",
    "mezzanine", "equity",
]

# IS subtotals, top-to-bottom. Each cascades from the prior subtotal using a
# sign rule appropriate to the section in between:
#   Gross Profit (Loss)          = Revenue − SUM(cost rows)            (revenue_cost, stored +)
#   Income (Loss) from Operations = GP − SUM(opex)                      (operating_expenses, stored +)
#   Pre-Tax Income (Loss)        = OP + SUM(non-op, signed)             (non_operating, natural signs)
#   Net Income (Loss)            = PT + SUM(tax, signed)                (tax, natural signs — + benefit, − expense)
IS_SUBTOTAL_LABELS_IN_ORDER = [
    "Gross Profit (Loss)",
    "Income (Loss) from Operations",
    "Pre-Tax Income (Loss)",
    "Net Income (Loss)",
]

SUBTOTAL_BORDER = Border(top=Side(style="thin", color="000000"))

# Number formats — line items vs subtotals.
# The third section (after the second ;) is the zero-display format, so any
# cell (literal 0 or formula that evaluates to 0) renders as "--" everywhere.
LINE_ITEM_FMT = '#,##0;(#,##0);"--"'
SUBTOTAL_FMT = '$#,##0_);($#,##0);"$--"_)'
EPS_FMT = '$#,##0.00_);($#,##0.00)'

# Ledger model_sheet → canonical sheet name in the output workbook.
# Entries with model_sheet we don't support (e.g. "_subtotal") return None.
LEDGER_SHEET_MAP = {
    "BALANCE SHEET":         "BALANCE SHEET",
    "CASH FLOW":             "CASH FLOW",
    "ANNL P&L":              "ANNL P&L",
    "ANNL P&L / QTR P&L":    "ANNL P&L",   # legacy dual-sheet entries route annual
    "QTR P&L":               "QTR P&L",
    "QTR BS":                "QTR BS",
    "QTR CF":                "QTR CF",
}


def stmt_to_sheet(stmt_type: StatementType, filing_type: FilingType) -> str:
    """Route a (statement_type, filing_type) to its target sheet in the workbook.

    10-K / 8-K / press-release filings feed the annual sheet family.
    10-Q filings feed the parallel quarterly sheet family. Reconcile has already
    dropped YTD 6mo / 9mo IS+CF statements from 10-Qs, so callers can trust that
    any 10-Q statement arriving here is a 3-month duration (IS/CF) or a
    point-in-time snapshot (BS).
    """
    is_qtr = filing_type == FilingType.TEN_Q
    return {
        StatementType.BALANCE_SHEET:    "QTR BS"  if is_qtr else "BALANCE SHEET",
        StatementType.CASH_FLOW:        "QTR CF"  if is_qtr else "CASH FLOW",
        StatementType.INCOME_STATEMENT: "QTR P&L" if is_qtr else "ANNL P&L",
    }[stmt_type]


def _is_qtr_sheet(sheet: str) -> bool:
    return sheet in QTR_SHEETS


# `_keep_statement` was a local copy of reconcile's filter. Both now live as
# `keep_statement_for_pipeline` in financials_schema — single source of truth.
_keep_statement = keep_statement_for_pipeline

# Light styling
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", start_color="203864", end_color="203864")
HEADER_ALIGN = Alignment(horizontal="center")
LABEL_FONT = Font(bold=True)
FORECAST_FILL = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ============================================================================
# Period utilities
# ============================================================================

def period_label(p: Period) -> str:
    """Canonical column label from a Period. 10-K → FY2023; 10-Q → Q3 FY2023."""
    if p.fiscal_quarter is None:
        return f"FY{p.fiscal_year}"
    return f"Q{p.fiscal_quarter} FY{p.fiscal_year}"


# ============================================================================
# Loading
# ============================================================================

def load_ledger(ticker_root: Path) -> dict:
    return json.loads((ticker_root / "Financial Statements" / "decisions_ledger.json").read_text(encoding="utf-8"))


def load_generic_library(ticker_root: Path, explicit_path: Path | None = None) -> dict:
    """Load the cross-ticker generic_line_item_mappings.json. Auto-resolves to
    <ticker-root>/../../pattern_libraries/generic_line_item_mappings.json
    (ticker folders live at Model Schema/Ticker Libraries/<TICKER>/, while
    pattern_libraries/ sits at Model Schema/pattern_libraries/) if not passed
    explicitly. Returns an empty-mappings stub if the file is absent
    (ticker-only mode)."""
    path = explicit_path
    if path is None:
        path = ticker_root.parent.parent / "pattern_libraries" / "generic_line_item_mappings.json"
    if not path.exists():
        return {"mappings": []}
    return json.loads(path.read_text(encoding="utf-8"))


def load_config(ticker_root: Path) -> dict:
    return json.loads((ticker_root / "Financial Statements" / "config.json").read_text(encoding="utf-8"))


def load_validated(path: Path) -> ValidatedFiling:
    return ValidatedFiling.model_validate_json(path.read_text(encoding="utf-8"))


# ============================================================================
# Row layout — derive from filing document order (post-Phase-4 migration)
# ============================================================================

def _collect_parent_children(generic: dict) -> tuple[dict[str, list[str]], dict[str, str]]:
    """Build (parent_to_children, child_to_parent) from the generic library.
    A child canonical declares `parent_canonical` pointing to the parent's
    rule_id. Parent renders as `=SUM(child cells)` in historical when any
    child has data; children render as plain line_items beneath. See
    `feedback_inventory_breakdown.md` (RM/WIP/FG → GEN-BS-005)."""
    p2c: dict[str, list[str]] = {}
    c2p: dict[str, str] = {}
    for e in generic.get("mappings", []):
        parent = e.get("parent_canonical")
        if not parent:
            continue
        rid = e["rule_id"]
        c2p[rid] = parent
        p2c.setdefault(parent, []).append(rid)
    return p2c, c2p


def _collect_filing_metadata(
    ledger: dict, generic: dict,
) -> tuple[set[str], set[str], dict[str, str], dict[str, str], dict[str, str]]:
    """Returns (superseded_rule_ids, memo_rule_ids, canonical_label_by_rule_id,
    section_hint_by_rule_id, sign_by_rule_id).

    Memo items (Comprehensive Income, FCT, Cash Paid for Interest/Taxes) are
    reconciled but not rendered on the xlsx — they get no row.

    `section_hint` carries the ledger/generic entry's `filing_section` when
    present — used by the BS section bucketer to place a row correctly even
    when the extract mis-tagged the item's section (e.g. ROU operating
    labeled current_assets by the extractor, but known to be non-current).

    `sign_by_rule_id` maps rule_id → 'negative' / 'positive'. Applied at
    render time as `value = ±abs(value)`, so the xlsx cell is always the
    intended sign regardless of how the filer reported it (handles both
    "filer reported positive" and "filer reported negative" cases without
    an extra sign-flip-detection step).
    """
    superseded = {
        e["rule_id"]
        for e in list(ledger.get("mappings", [])) + list(ledger.get("new_rows", []))
        if e.get("superseded_by")
    }
    memo = set()
    for e in generic.get("mappings", []):
        if e.get("memo"):
            memo.add(e["rule_id"])
    for e in list(ledger.get("mappings", [])) + list(ledger.get("new_rows", [])):
        if e.get("memo"):
            memo.add(e["rule_id"])

    # Canonical label: ticker-specific overrides beat generic, matching reconcile's tier precedence.
    canonical: dict[str, str] = {}
    section_hint: dict[str, str] = {}
    sign_by_rule: dict[str, str] = {}
    for e in generic.get("mappings", []):
        canonical[e["rule_id"]] = e["model_label"]
        if e.get("filing_section"):
            section_hint[e["rule_id"]] = e["filing_section"]
        if e.get("sign_convention") in ("negative", "positive"):
            sign_by_rule[e["rule_id"]] = e["sign_convention"]
    for e in ledger.get("mappings", []):
        if not e.get("superseded_by"):
            canonical[e["rule_id"]] = e.get("model_label") or e["rule_id"]
            if e.get("filing_section"):
                section_hint[e["rule_id"]] = e["filing_section"]
            if e.get("sign_convention") in ("negative", "positive"):
                sign_by_rule[e["rule_id"]] = e["sign_convention"]
    for e in ledger.get("new_rows", []):
        if not e.get("superseded_by"):
            canonical[e["rule_id"]] = e.get("new_row_label") or e["rule_id"]
            if e.get("filing_section"):
                section_hint[e["rule_id"]] = e["filing_section"]
            if e.get("sign_convention") in ("negative", "positive"):
                sign_by_rule[e["rule_id"]] = e["sign_convention"]
    return superseded, memo, canonical, section_hint, sign_by_rule


def resolve_row_positions(
    ledger: dict,
    generic: dict,
    filings: list[ValidatedFiling],
) -> tuple[dict[str, dict[str, int]], dict[str, dict[int, str]], dict[str, dict[int, str]]]:
    """Derive row layout from the LATEST filing's document order. Older-only
    items append at the end, preserving their prior filing's relative order.

    Returns:
      rule_to_excel: {sheet: {rule_id: excel_row}}
      row_labels:    {sheet: {excel_row: canonical_label}}
      row_section:   {sheet: {excel_row: section_tag}}  # for BS section bucketing

    Skipped at row-allocation time:
      - Items with ledger_rule_id=None (novels — shouldn't exist at this stage)
      - Items mapped to superseded ledger entries
      - Items mapped to memo entries (carried in MappedFiling for audit but not rendered)
      - The "_subtotal" carry-through items (subtotal formulas inserted separately)
    """
    superseded, memo, canonical, section_hint, _sign_by_rule = _collect_filing_metadata(ledger, generic)

    # Iterate filings newest-first so the latest filing drives row order.
    filings_desc = sorted(filings, key=lambda vf: vf.mapped.raw.filing_date, reverse=True)

    # Per-sheet ordered list of rule_ids + lookup of their positions.
    # When an older filing surfaces a rule_id not yet placed, we insert it
    # immediately AFTER the most-recently-seen placed rule_id from the same
    # filing's document order — so an older-only item lands next to its
    # neighbor in the filing where it last appeared, not at the end of the sheet.
    order: dict[str, list[str]] = {s: [] for s in V1_SHEETS}
    positions: dict[str, dict[str, int]] = {s: {} for s in V1_SHEETS}
    rid_label: dict[str, dict[str, str]] = {s: {} for s in V1_SHEETS}
    rid_section: dict[str, dict[str, str]] = {s: {} for s in V1_SHEETS}

    for vf in filings_desc:
        idx = 0
        all_stmts = vf.mapped.raw.statements
        for stmt in all_stmts:
            count = len(stmt.line_items)
            # Advance idx by count ONLY if reconcile kept this statement's items
            # in mapped_line_items. If reconcile dropped the statement (10-Q
            # short-duration duplicate), its items aren't in mapped — skip
            # without idx advance to stay aligned with mapped_line_items.
            if not _keep_statement(stmt, vf.mapped.raw.filing_type, all_stmts):
                continue
            group = vf.mapped.mapped_line_items[idx : idx + count]
            idx += count

            sheet = stmt_to_sheet(stmt.statement_type, vf.mapped.raw.filing_type)
            if sheet not in V1_SHEETS:
                continue

            # Walk this filing's items in document order. Track the most recent
            # already-placed rule_id as the insertion anchor for any NEW rule_id
            # encountered in this statement.
            last_placed_idx: int | None = None
            for grp_i, item in enumerate(group):
                if item.model_sheet in ("_subtotal", "_memo"):
                    continue
                rid = item.ledger_rule_id
                if rid is None or rid in superseded or rid in memo:
                    continue

                if rid in positions[sheet]:
                    # Anchor advances, never regresses. When the same rule_id
                    # appears twice in a filing's document order (e.g. both
                    # "Income taxes payable" and "Other current liabilities"
                    # alias GEN-CF-017), the second occurrence at an earlier
                    # combined-order position must not pull the anchor back —
                    # otherwise subsequent NEW items get inserted between the
                    # first placement and items already inserted after it,
                    # pushing them toward the end of the sheet.
                    cand = positions[sheet][rid]
                    if last_placed_idx is None or cand > last_placed_idx:
                        last_placed_idx = cand
                    continue

                # New rule_id — preferred insertion: right after the prev
                # already-placed rule_id in this filing's document order. If no
                # prev anchor (this is the first NEW item AND nothing earlier in
                # this filing was already on the sheet), look FORWARD for the
                # next already-placed rule_id and insert immediately BEFORE it —
                # that keeps the new item near its filing-order neighbors.
                # Without the look-forward, atypical filings that lead with an
                # uncommon line (e.g. PG's Q2/Q3 FY2024 10-Qs reporting
                # Impairment of Intangibles as the very first CF item) get
                # appended past the operating subtotal, breaking the CFO sum.
                if last_placed_idx is not None:
                    insert_at = last_placed_idx + 1
                else:
                    next_placed_idx: int | None = None
                    for forward_item in group[grp_i + 1 :]:
                        if forward_item.model_sheet in ("_subtotal", "_memo"):
                            continue
                        f_rid = forward_item.ledger_rule_id
                        if f_rid is None or f_rid in superseded or f_rid in memo:
                            continue
                        if f_rid in positions[sheet]:
                            next_placed_idx = positions[sheet][f_rid]
                            break
                    insert_at = next_placed_idx if next_placed_idx is not None else len(order[sheet])
                order[sheet].insert(insert_at, rid)
                positions[sheet] = {r: i for i, r in enumerate(order[sheet])}
                last_placed_idx = insert_at

                rid_label[sheet][rid] = canonical.get(rid) or item.model_label
                item_section_val = item.section.value if hasattr(item.section, "value") else str(item.section)
                rid_section[sheet][rid] = section_hint.get(rid) or item_section_val

    # Mirror annual-only rule_ids onto the parallel QTR sheet. Reason: filers
    # often present a fully-classified balance sheet on the 10-K but a
    # condensed one on the 10-Q (Reg S-X Article 10 allows it), so a row like
    # CELH's Treasury Stock appears on BALANCE SHEET but not QTR BS. Without
    # this mirror, model-calc's forecast subtotal formulas — which translate
    # YTD column letters to forecast columns — read row layouts that don't
    # include the annual-only row, and the BS gap shows up as a constant
    # boundary residual exactly equal to that row's value. Mirrored rule_ids
    # arrive on QTR with no historical Q1-Q3 data; their Q4 cell cross-refs
    # the annual sheet (handled by the existing Q4-from-annual logic), and
    # forecast cells inherit HOLD_LAST behavior from inference.
    ANNL_TO_QTR = [("ANNL P&L", "QTR P&L"), ("BALANCE SHEET", "QTR BS"), ("CASH FLOW", "QTR CF")]
    for annl_sheet, qtr_sheet in ANNL_TO_QTR:
        if not order.get(annl_sheet) or qtr_sheet not in order:
            continue
        qtr_rids = set(positions[qtr_sheet].keys())
        # Walk annual rule_ids in their canonical order; for each missing on
        # QTR, find the most recent annual rid that IS on QTR (the "anchor")
        # and insert immediately after it. If the missing rid is before any
        # shared anchor, fall back to the position of the next shared anchor.
        for ai, rid in enumerate(order[annl_sheet]):
            if rid in qtr_rids:
                continue
            prev_anchor = None
            for back_rid in reversed(order[annl_sheet][:ai]):
                if back_rid in qtr_rids:
                    prev_anchor = back_rid
                    break
            if prev_anchor is not None:
                insert_at = positions[qtr_sheet][prev_anchor] + 1
            else:
                next_anchor = next((r for r in order[annl_sheet][ai + 1:] if r in qtr_rids), None)
                insert_at = positions[qtr_sheet][next_anchor] if next_anchor else len(order[qtr_sheet])
            order[qtr_sheet].insert(insert_at, rid)
            positions[qtr_sheet] = {r: i for i, r in enumerate(order[qtr_sheet])}
            qtr_rids.add(rid)
            rid_label[qtr_sheet][rid] = rid_label[annl_sheet][rid]
            rid_section[qtr_sheet][rid] = rid_section[annl_sheet].get(rid, "unclassified")

    # Inject missing IS cascade subtotal rows (Gross Profit, Op Income, Pre-Tax,
    # Net Income) at section transitions on every IS sheet. The cascade rows
    # always render — even when the filer doesn't break them out (e.g. PG omits
    # Gross Profit) — and always get a SUM formula via write_is_subtotals().
    # Skipped at any section boundary where the filer has already placed the
    # cascade subtotal as a line of their own.
    IS_SECTION_TO_CASCADE = {
        "revenue_cost":        "Gross Profit (Loss)",
        "operating_expenses":  "Income (Loss) from Operations",
        "non_operating":       "Pre-Tax Income (Loss)",
        "tax":                 "Net Income (Loss)",
    }
    for sheet in V1_SHEETS:
        if "P&L" not in sheet:
            continue
        # Pre-compute which cascade anchors the filer ALREADY rendered as
        # their own canonical row (anywhere in document order). Skip
        # injection for those — the filer's row IS the anchor, and the
        # by-canonical-label dedup at row-allocation time will land both the
        # cascade-vrid and the real rid on the same excel row, producing a
        # duplicate placement when sections inside a bucket aren't strictly
        # monotonic (PEP's `non_operating` Juice Gain rendered between SG&A
        # and Impairment, both `operating_expenses` — without this guard the
        # opex→non_op→opex pseudo-transition fires the IFO cascade BEFORE
        # the filer's real IFO row in document order). Per
        # `feedback_no_duplicate_anchor_subtotals.md`.
        already_rendered_cascades = {
            rid_label[sheet].get(rid)
            for rid in order[sheet]
            if rid_label[sheet].get(rid) in IS_SECTION_TO_CASCADE.values()
        }
        new_order: list[str] = []
        prev_section: str | None = None
        for rid in order[sheet]:
            section = rid_section[sheet].get(rid, "unclassified")
            current_label = rid_label[sheet].get(rid)
            # Section just changed → maybe inject cascade for the prev section.
            if (prev_section
                    and prev_section != section
                    and prev_section in IS_SECTION_TO_CASCADE):
                cascade_label = IS_SECTION_TO_CASCADE[prev_section]
                if (cascade_label not in already_rendered_cascades
                        and current_label != cascade_label):
                    vrid = f"_CASCADE_{prev_section}"
                    rid_label[sheet][vrid] = cascade_label
                    rid_section[sheet][vrid] = "subtotal"
                    new_order.append(vrid)
            new_order.append(rid)
            prev_section = section
        # Tail: cascade for the final section (typically tax → Net Income).
        if prev_section in IS_SECTION_TO_CASCADE:
            cascade_label = IS_SECTION_TO_CASCADE[prev_section]
            if cascade_label not in already_rendered_cascades:
                vrid = f"_CASCADE_{prev_section}"
                rid_label[sheet][vrid] = cascade_label
                rid_section[sheet][vrid] = "subtotal"
                new_order.append(vrid)
        order[sheet] = new_order
        positions[sheet] = {r: i for i, r in enumerate(new_order)}

    rule_to_excel: dict[str, dict[str, int]] = {}
    row_labels: dict[str, dict[int, str]] = {}
    row_section: dict[str, dict[int, str]] = {}
    for sheet in V1_SHEETS:
        rmap: dict[str, int] = {}
        lmap: dict[int, str] = {}
        smap: dict[int, str] = {}
        # Consolidate rule_ids that share a canonical label onto a single
        # excel row. Two distinct rule_ids legitimately map to the same row
        # when a ticker override and the generic library entry both produce
        # the same canonical (PG MAP-IS-001 + GEN-IS-009 → "Pre-Tax Income
        # (Loss)"). Without consolidation, each rule_id gets its own row,
        # producing duplicate "Pre-Tax Income (Loss)" rows on the IS — and
        # the IS cascade SUM range then includes the filer-extracted Pre-Tax
        # row as a component, doubling the cascade subtotal.
        label_to_row: dict[str, int] = {}
        next_row = 2
        for rid in order[sheet]:
            label = rid_label[sheet].get(rid)
            section = rid_section[sheet].get(rid)
            if label and label in label_to_row:
                rmap[rid] = label_to_row[label]
                continue
            excel_row = next_row
            rmap[rid] = excel_row
            lmap[excel_row] = label
            smap[excel_row] = section
            if label:
                label_to_row[label] = excel_row
            next_row += 1
        rule_to_excel[sheet] = rmap
        row_labels[sheet] = lmap
        row_section[sheet] = smap

    return rule_to_excel, row_labels, row_section


# ============================================================================
# Column layout — periods across filings, dedup with newer filing winning
# ============================================================================

def build_column_layout(
    filings: list[ValidatedFiling],
) -> dict[str, list[tuple[str, date | None]]]:
    """Per sheet, return chronological list of (label, period_end_date | None).
    None marks forecast columns.

    Annual sheets get `FY{year}` labels + annual forecast columns
    (FY2025E..FY2030E). Quarterly sheets get `Q{N} FY{year}` labels and no
    forecast columns in v1 — quarterly forecasts are deferred to model-calc."""
    # Collect (period_end_date, quarter_label) per sheet. Storing the label
    # alongside the date so we can disambiguate Q1/Q2/Q3 within a fiscal year
    # on the QTR sheets (where `FY{d.year}` would collide across quarters).
    per_sheet_hist: dict[str, dict[date, str]] = {s: {} for s in V1_SHEETS}
    for vf in filings:
        all_stmts = vf.mapped.raw.statements
        for stmt in all_stmts:
            if not _keep_statement(stmt, vf.mapped.raw.filing_type, all_stmts):
                continue
            sheet = stmt_to_sheet(stmt.statement_type, vf.mapped.raw.filing_type)
            if sheet not in V1_SHEETS:
                continue
            period = stmt.period
            if _is_qtr_sheet(sheet) and period.fiscal_quarter is not None:
                label = f"Q{period.fiscal_quarter} FY{period.fiscal_year}"
            else:
                label = f"FY{period.fiscal_year}"
            # First writer wins per date; duplicates from multiple filings for
            # the same period use the same label anyway.
            per_sheet_hist[sheet].setdefault(period.period_end_date, label)

    out: dict[str, list[tuple[str, date | None]]] = {}
    for sheet, hist_map in per_sheet_hist.items():
        hist = [(hist_map[d], d) for d in sorted(hist_map)]
        if _is_qtr_sheet(sheet):
            # Quarterly forecast columns deferred to model-calc; don't pre-fill.
            out[sheet] = hist
        else:
            # Forecast labels start at (last historical fiscal year + 1) and
            # extend N_FORECAST_YEARS forward. Computed per sheet because each
            # statement may have a different latest historical year (e.g. an
            # ANNL P&L can have FY2025 as latest while BALANCE SHEET only has
            # FY2024 if the latest 10-K hasn't been parsed for BS yet).
            last_hist_year = max(d.year for d in hist_map) if hist_map else 2024
            # Use fiscal year from the last label in case period_end_date.year
            # disagrees with fiscal_year (June FYE filers).
            last_label = hist_map[max(hist_map)] if hist_map else f"FY{last_hist_year}"
            try:
                last_fy = int(last_label.replace("FY", ""))
            except ValueError:
                last_fy = last_hist_year
            fcst_labels = [f"FY{y}E" for y in range(last_fy + 1, last_fy + 1 + N_FORECAST_YEARS)]
            fcst = [(lbl, None) for lbl in fcst_labels]
            out[sheet] = hist + fcst
    return out


# ============================================================================
# Cell writes — walk mapped items, dedupe by (sheet, rule, period) with newer wins
# ============================================================================

def collect_writes(
    filings: list[ValidatedFiling],
    superseded_rule_ids: set[str],
    sign_by_rule_id: dict[str, str],
) -> list[tuple[date, str, str, date, Decimal]]:
    """Return every non-superseded mapped write as
    (filing_date, sheet, ledger_rule_id, period_end_date, value).

    `sign_by_rule_id` maps rule_id → 'negative' / 'positive'. Applied as
    ±abs(value) so the xlsx cell always carries the intended sign regardless
    of how the filer reported it. The MappedFiling itself is unchanged
    (validators rely on the natural-sign value + sign_convention semantics).

    Aggregation across writes happens later in build_workbook:
      - Within a single filing, multiple rule_ids landing on the same
        (sheet, excel_row, period) SUM (genuine sibling ledger entries).
      - Across filings, for the same (sheet, excel_row, period), the
        NEWEST filing wins (same conceptual row reported by both filings).
    """
    out: list[tuple[date, str, str, date, Decimal]] = []
    for vf in filings:
        filing_date = vf.mapped.raw.filing_date
        idx = 0
        all_stmts = vf.mapped.raw.statements
        for stmt in all_stmts:
            count = len(stmt.line_items)
            if not _keep_statement(stmt, vf.mapped.raw.filing_type, all_stmts):
                continue
            group = vf.mapped.mapped_line_items[idx : idx + count]
            idx += count

            sheet = stmt_to_sheet(stmt.statement_type, vf.mapped.raw.filing_type)
            if sheet not in V1_SHEETS:
                continue

            for item in group:
                if item.model_sheet in ("_subtotal", "_memo"):
                    continue
                if item.ledger_rule_id is None:
                    continue
                if item.ledger_rule_id in superseded_rule_ids:
                    continue
                value = item.value
                # Normalize the monetary value onto the workbook's canonical
                # thousands scale using the statement's reported unit. Skip
                # EPS/share rows — those carry per-share / share-count units.
                if not _is_eps_or_share_label(
                    item.canonical_label or item.raw_filing_label
                ):
                    scale = _UNIT_TO_THOUSANDS.get(stmt.unit)
                    if scale is not None and scale != 1:
                        value = value * scale
                # sign_convention overlays apply ONLY to IS expense lines and
                # BS contra accounts. CF values come out of extract already
                # matching the filer's visual sign (the iXBRL extractor honors
                # presentation linkbase preferredLabel negation). Re-applying
                # sign_convention to CF items would double-flip.
                if "CF" not in sheet.upper() and "CASH FLOW" not in sheet.upper():
                    item_sign = item.sign_convention
                    item_sc = item_sign.value if hasattr(item_sign, "value") else item_sign
                    forced_sign = item_sc if item_sc in ("negative", "positive") else sign_by_rule_id.get(item.ledger_rule_id)
                    if forced_sign == "negative":
                        value = -abs(value)
                    elif forced_sign == "positive":
                        value = abs(value)
                out.append((filing_date, sheet, item.ledger_rule_id,
                            stmt.period.period_end_date, value))
    return out


# ============================================================================
# Workbook build
# ============================================================================

def insert_bs_subtotal_slots(
    bs_rule_to_excel: dict[str, int],
    bs_labels: dict[int, str],
    bs_row_section: dict[int, str],
) -> tuple[dict[str, int], dict[int, str], list[dict]]:
    """Rebuild BS layout in canonical section order (CA→NCA→CL→NCL→Mezz→Equity),
    inserting subtotal rows at section boundaries + grand total at the end.

    Rows whose section is known get re-bucketed into their correct section (so
    strays like NEW-BS-009 Accrued Distributor Termination land back in CL).
    Unknown-section rows appended at the end.

    Returns (new_rule_to_excel, new_labels, subtotal_specs_resolved).
    """
    bucketed: dict[str, list[int]] = {s: [] for s in BS_SECTION_ORDER}
    unclassified: list[int] = []
    for xr in sorted(bs_labels.keys()):
        sec = bs_row_section.get(xr, "unclassified")
        # Per user directive (2026-04-26): convertible / redeemable preferred
        # stock is rendered in mezzanine equity per ASC 480 but folded into
        # the Stockholders' Equity section for the workbook layout. Mezzanine
        # rows are bucketed into equity here so they appear under the SE
        # subtotal alongside Common Stock / APIC / RE / etc.
        if sec == "mezzanine":
            sec = "equity"
        if sec in bucketed:
            bucketed[sec].append(xr)
        else:
            unclassified.append(xr)

    new_labels: dict[int, str] = {}
    old_to_new: dict[int, int] = {}
    section_item_rows: dict[str, list[int]] = {s: [] for s in BS_SECTION_ORDER}
    subtotal_rows: dict[str, int] = {}
    grand_total_row: int | None = None

    next_row = 2
    for section in BS_SECTION_ORDER:
        for old_xr in bucketed[section]:
            old_to_new[old_xr] = next_row
            new_labels[next_row] = bs_labels[old_xr]
            section_item_rows[section].append(next_row)
            next_row += 1
        for spec_sec, label, _ftype, _cascade in BS_SUBTOTAL_SPECS:
            if spec_sec == section and section_item_rows[section]:
                new_labels[next_row] = label
                subtotal_rows[section] = next_row
                next_row += 1
                break

    # Grand total after everything, if we have TL and TSE
    if subtotal_rows.get("non_current_liabilities") and subtotal_rows.get("equity"):
        grand_total_label = "Total Liabilities, Mezzanine & Stockholders' Equity"
        new_labels[next_row] = grand_total_label
        grand_total_row = next_row
        next_row += 1

    # Unclassified rows (safety valve)
    for old_xr in unclassified:
        old_to_new[old_xr] = next_row
        new_labels[next_row] = bs_labels[old_xr]
        next_row += 1

    new_rule_to_excel: dict[str, int] = {
        rid: old_to_new[old_xr]
        for rid, old_xr in bs_rule_to_excel.items()
        if old_xr in old_to_new
    }

    # Resolve subtotal specs into formula-ready entries
    resolved: list[dict] = []
    for spec_sec, label, ftype, cascade_sec in BS_SUBTOTAL_SPECS:
        if spec_sec == "__grand_total__":
            if grand_total_row is None:
                continue
            resolved.append({
                "row": grand_total_row,
                "type": "grand",
                "tl_row": subtotal_rows.get("non_current_liabilities"),
                "mezz_rows": section_item_rows.get("mezzanine", []),
                "tse_row": subtotal_rows.get("equity"),
            })
            continue
        sub_row = subtotal_rows.get(spec_sec)
        item_rows = section_item_rows.get(spec_sec, [])
        if sub_row is None or not item_rows:
            continue
        resolved.append({
            "row": sub_row,
            "type": ftype,
            "min": min(item_rows),
            "max": max(item_rows),
            "cascade_row": subtotal_rows.get(cascade_sec) if cascade_sec else None,
        })

    return new_rule_to_excel, new_labels, resolved


def _is_subtotal_formula(
    label: str,
    col: str,
    range_start: int,
    range_end: int,
    prev_row: int | None,
) -> str | None:
    """Build the IS subtotal formula for one (row, column)."""
    empty_range = range_start > range_end
    single_cell = range_start == range_end

    def _sum_expr() -> str | None:
        if empty_range:
            return None
        if single_cell:
            return f"{col}{range_start}"
        return f"SUM({col}{range_start}:{col}{range_end})"

    # All cascade subtotals use PLUS-SUM. Items are stored with NATURAL signs
    # (expenses negative via sign_convention="negative", income positive),
    # so summing yields the correct contribution to the subtotal directly.
    # Example: Gross Profit = Revenue + (-COGS) = Revenue - |COGS|.
    if label == "Gross Profit (Loss)":
        # First row in range = Revenue; rest are cost items stored negative.
        if empty_range:
            return None
        if single_cell:
            return f"={col}{range_start}"
        cost_start = range_start + 1
        if cost_start == range_end:
            return f"={col}{range_start}+{col}{cost_start}"
        return f"={col}{range_start}+SUM({col}{cost_start}:{col}{range_end})"

    if label in ("Income (Loss) from Operations",
                 "Pre-Tax Income (Loss)",
                 "Net Income (Loss)"):
        if prev_row is None:
            return None
        sum_expr = _sum_expr()
        if sum_expr is None:
            return f"={col}{prev_row}"
        return f"={col}{prev_row}+{sum_expr}"

    return None


def write_is_subtotals(
    ws,
    labels_for_sheet: dict[int, str],
    all_cols: list[int],
) -> None:
    """Overwrite IS subtotal rows with live formulas. Subtotals are matched by
    canonical label. Skips subtotals not present on the sheet."""
    label_to_row: dict[str, int] = {
        label: xr
        for xr, label in labels_for_sheet.items()
        if label in IS_SUBTOTAL_LABELS_IN_ORDER
    }

    prev_row: int | None = None
    section_start = 2
    for label in IS_SUBTOTAL_LABELS_IN_ORDER:
        xr = label_to_row.get(label)
        if xr is None:
            continue
        range_end = xr - 1
        wrote_any = False
        for col_idx in all_cols:
            col = get_column_letter(col_idx)
            formula = _is_subtotal_formula(label, col, section_start, range_end, prev_row)
            if formula is None:
                continue
            cell = ws.cell(row=xr, column=col_idx, value=formula)
            cell.number_format = SUBTOTAL_FMT
            cell.font = Font(bold=True)
            cell.border = SUBTOTAL_BORDER
            wrote_any = True
        if wrote_any:
            ws.cell(row=xr, column=1).font = Font(bold=True)
        prev_row = xr
        section_start = xr + 1

    # Net Income (Loss) Less NCI = Net Income (Loss) − NI Attributable to NCI.
    # Doesn't fit the cascade SUM pattern (it's a 2-row delta), so handled here.
    ni_row = label_to_row.get("Net Income (Loss)")
    less_nci_row = next(
        (xr for xr, lbl in labels_for_sheet.items() if lbl == "Net Income (Loss) Less NCI"),
        None,
    )
    nci_attrib_row = next(
        (xr for xr, lbl in labels_for_sheet.items()
         if lbl == "Net Income (Loss) Attributable to Noncontrolling Interest"),
        None,
    )
    if ni_row and less_nci_row and nci_attrib_row:
        for col_idx in all_cols:
            col = get_column_letter(col_idx)
            cell = ws.cell(
                row=less_nci_row, column=col_idx,
                value=f"={col}{ni_row}-{col}{nci_attrib_row}",
            )
            cell.number_format = SUBTOTAL_FMT
            cell.font = Font(bold=True)
            cell.border = SUBTOTAL_BORDER
        ws.cell(row=less_nci_row, column=1).font = Font(bold=True)

    # Net Income (Loss) Attributable to Common Shareholders =
    #   Net Income (Loss) + Preferred Dividends + Income Allocated to
    #   Participating Preferred. Components carry their natural filer signs
    #   (Pref Div / Alloc typically negative on IS), so plain summation gives
    #   the deduction. Live formula replaces the filer-extracted hardcoded
    #   value so the row updates with NI / PrefDiv / Alloc changes.
    nic_row = next(
        (xr for xr, lbl in labels_for_sheet.items()
         if lbl == "Net Income (Loss) Attributable to Common Shareholders"),
        None,
    )
    pref_div_row = next(
        (xr for xr, lbl in labels_for_sheet.items() if lbl == "Preferred Dividends"),
        None,
    )
    alloc_row = next(
        (xr for xr, lbl in labels_for_sheet.items()
         if lbl == "Income Allocated to Participating Preferred"),
        None,
    )
    if ni_row and nic_row:
        for col_idx in all_cols:
            col = get_column_letter(col_idx)
            parts = [f"{col}{ni_row}"]
            if pref_div_row:
                parts.append(f"{col}{pref_div_row}")
            if alloc_row:
                parts.append(f"{col}{alloc_row}")
            cell = ws.cell(row=nic_row, column=col_idx, value="=" + "+".join(parts))
            cell.number_format = SUBTOTAL_FMT
            cell.font = Font(bold=True)
            cell.border = SUBTOTAL_BORDER
        ws.cell(row=nic_row, column=1).font = Font(bold=True)


# Thresholds are in the workbook's canonical scale — THOUSANDS (collect_writes
# normalizes every filing to thousands via stmt.unit). $2M = 2000, $4M = 4000.
CF_TIE_OUT_WARN_MIN = Decimal("2000")   # gaps ≤ $2M are silent (filer rounding)
CF_TIE_OUT_ERROR_MIN = Decimal("4000")  # gaps > $4M are validation ERRORS (raise)


def check_cf_section_subtotals(
    sheet_cells: dict[tuple[str, int, date], Decimal],
    sheet_name: str,
    labels_for_sheet: dict[int, str],
    sheet_sections: dict[int, str],
    section_subtotal_rows: list[int],
    date_to_col: dict[date, int],
) -> tuple[list[str], list[str]]:
    """For each (period × section) on a CF sheet, verify that the dedup-
    picked component rows sum to the dedup-picked subtotal row. Returns
    (warnings, errors) where:
      - gaps in (CF_TIE_OUT_WARN_MIN, CF_TIE_OUT_ERROR_MIN] are warnings
      - gaps > CF_TIE_OUT_ERROR_MIN are errors (caller must raise)
      - gaps ≤ CF_TIE_OUT_WARN_MIN are silent (filer-rounding noise).

    Section ranges are bounded by `section_subtotal_rows` (sorted ascending):
    rows in (prev_subtotal, current_subtotal) belong to that section.
    """
    if not section_subtotal_rows:
        return [], []
    warnings: list[str] = []
    errors: list[str] = []
    sub_rows_sorted = sorted(section_subtotal_rows)
    # Map subtotal row -> section key, via canonical-label lookup
    xr_to_section: dict[int, str] = {}
    for xr in sub_rows_sorted:
        lbl = labels_for_sheet.get(xr)
        sec = CF_SUBTOTAL_LABELS.get(lbl)
        if sec:
            xr_to_section[xr] = sec
    # Group cells by period; for each period verify each section
    periods = sorted({d for (s, _, d) in sheet_cells if s == sheet_name})
    for period in periods:
        prev_sub = 1  # rows start at 2 — section starts at prev_sub+1
        for sub_xr in sub_rows_sorted:
            sec = xr_to_section.get(sub_xr)
            if sec is None:
                prev_sub = sub_xr
                continue
            # Sum component rows (prev_sub, sub_xr) exclusive of subtotals,
            # excluding any that belong to a different (typically off-section)
            # tag like cash_other (shouldn't be inside the CFO/CFI/CFF range
            # but be defensive).
            comp_sum = Decimal("0")
            for xr in range(prev_sub + 1, sub_xr):
                if xr in xr_to_section:
                    continue  # nested subtotal — shouldn't happen but skip
                row_sec = sheet_sections.get(xr)
                if row_sec and row_sec != sec:
                    # Row tagged for a different section landed inside this
                    # section's SUM range. Layered containment check (added
                    # earlier) should have caught this; flag separately for
                    # diagnostics.
                    pass
                comp_sum += sheet_cells.get((sheet_name, xr, period), Decimal("0"))
            sub_val = sheet_cells.get((sheet_name, sub_xr, period))
            if sub_val is None:
                prev_sub = sub_xr
                continue
            gap = comp_sum - sub_val
            abs_gap = abs(gap)
            if abs_gap > CF_TIE_OUT_ERROR_MIN:
                errors.append(
                    f"  {period.isoformat()} {sec}: components sum to "
                    f"{comp_sum} but filer-reported subtotal at row {sub_xr} "
                    f"is {sub_val} (gap {gap})"
                )
            elif abs_gap > CF_TIE_OUT_WARN_MIN:
                warnings.append(
                    f"  {period.isoformat()} {sec}: components sum to "
                    f"{comp_sum} but filer-reported subtotal at row {sub_xr} "
                    f"is {sub_val} (gap {gap})"
                )
            prev_sub = sub_xr
    return warnings, errors


def aggregate_cell_totals(
    writes: list[tuple[date, str, str, date, Decimal]],
    row_map: dict[str, dict[str, int]],
) -> dict[tuple[str, int, date], Decimal]:
    """Collapse raw writes into one value per cell — first-filing-wins per
    (sheet, period), entire row set.

    Three-stage aggregation:
      1. Per (sheet, excel_row, period, filing_date): SUM all rule_ids in that
         filing that landed on the same row. Genuine sibling ledger entries
         (e.g. MAP-IS-010 net interest + MAP-IS-011 note receivable interest)
         both contribute their share of the row's total within one filing.
      2. For each (sheet, period), find the OLDEST filing that reported it —
         this is the "authoritative filing" for that period.
      3. Emit cells only from the authoritative filing. All rows in a given
         (sheet, period) column come from the same filing — never mixed.

    Why per-period (not per-row): when a filer changes us-gaap concept names
    between filings (e.g. PG's 2024-Q3 10-Q used the hybrid concept
    `PaymentsForProceedsFromBusinessesAndOtherInvestingActivities` rolling
    Acquisitions in with Other Investing, then switched to the pure
    `PaymentsToAcquireBusinessesNetOfCashAcquired` concept in the 2025-Q3
    comparative for the same 2024-03-31 period), per-row dedup picks both
    the original combined row AND the split-out Acquisitions row, double-
    counting the same economic item. Per-period dedup uses the original
    filing's complete row breakdown for that period, sidestepping the
    concept-rename problem entirely. Within-filing CF-2/CF-3/CF-4 internal
    consistency carries through to the workbook automatically.
    """
    per_filing: dict[tuple[str, int, date, date], Decimal] = {}
    for fdate, sheet, rule_id, period, value in writes:
        excel_row = row_map.get(sheet, {}).get(rule_id)
        if excel_row is None:
            continue
        key = (sheet, excel_row, period, fdate)
        per_filing[key] = per_filing.get(key, Decimal("0")) + value

    # For each (sheet, period), find the oldest filing that reported it.
    period_first_fdate: dict[tuple[str, date], date] = {}
    for (sheet, _xr, period, fdate) in per_filing:
        sp_key = (sheet, period)
        cur = period_first_fdate.get(sp_key)
        if cur is None or fdate < cur:
            period_first_fdate[sp_key] = fdate

    # Emit cells only from the authoritative filing per (sheet, period).
    final: dict[tuple[str, int, date], Decimal] = {}
    for (sheet, xr, period, fdate), total in per_filing.items():
        if fdate == period_first_fdate.get((sheet, period)):
            final[(sheet, xr, period)] = total
    return final


def _collect_filer_mezzanine_sums(
    filings: list[ValidatedFiling],
) -> dict[tuple[str, date], Decimal]:
    """Sum mezzanine items per (sheet, period_end_date), oldest-filing-wins.
    Used to back out the mezzanine→equity workbook fold when comparing the
    workbook's TSE section sum to the filer's TSE subtotal — the workbook
    deliberately folds mezzanine into equity (per
    feedback_convertible_preferred_to_equity.md) but the filer's TSE does
    not, so direct comparison would always show a gap = mezzanine balance."""
    out: dict[tuple[str, date], tuple[date, Decimal]] = {}
    for vf in filings:
        ftype = vf.mapped.raw.filing_type
        fdate = vf.mapped.raw.filing_date
        for stmt in vf.mapped.raw.statements:
            if stmt.statement_type != StatementType.BALANCE_SHEET:
                continue
            if not _keep_statement(stmt, ftype, vf.mapped.raw.statements):
                continue
            sheet = stmt_to_sheet(stmt.statement_type, ftype)
            period_end = stmt.period.period_end_date
            mezz_sum = Decimal("0")
            for li in stmt.line_items:
                sec = li.section.value if hasattr(li.section, "value") else li.section
                if sec != "mezzanine":
                    continue
                rt = li.row_type.value if hasattr(li.row_type, "value") else li.row_type
                if rt in ("memo", "subtotal"):
                    continue
                mezz_sum += Decimal(li.value)
            key = (sheet, period_end)
            cur = out.get(key)
            if cur is None or fdate < cur[0]:
                out[key] = (fdate, mezz_sum)
    return {k: v for k, (_, v) in out.items()}


def _collect_filer_subtotals(
    filings: list[ValidatedFiling],
) -> dict[tuple[str, date, str], Decimal]:
    """Map (sheet, period_end_date, canonical_subtotal_label) → filer's
    reported subtotal value. When multiple filings cover one period, oldest
    wins (mirrors `aggregate_cell_totals`'s first-filing-wins dedup so the
    tie-out check compares the workbook's value against the SAME filing
    that supplied its line items)."""
    BS_PATTERNS = [
        (re.compile(r"^total\s+current\s+assets\b", re.I),                          "Total Current Assets"),
        (re.compile(r"^total\s+(liabilities\s+and|liabilities,?\s*mezz)", re.I),    "Total L+SE"),
        (re.compile(r"^total\s+assets\b", re.I),                                    "Total Assets"),
        (re.compile(r"^total\s+current\s+liab", re.I),                              "Total Current Liabilities"),
        (re.compile(r"^total\s+liab", re.I),                                        "Total Liabilities"),
        (re.compile(r"^total\s+(stockholders|shareholders)['’]?\s*equity", re.I),   "Total Stockholders' Equity"),
    ]
    CF_PATTERNS = [
        (re.compile(r"^(net\s+cash\s+(provided\s+by|used\s+in)?.*operat|total\s+operating\s+activ)", re.I), "Cash Flow from Operations"),
        (re.compile(r"^(net\s+cash\s+(provided\s+by|used\s+in)?.*invest|total\s+investing\s+activ)", re.I), "Cash Flow from Investing"),
        (re.compile(r"^(net\s+cash\s+(provided\s+by|used\s+in)?.*financ|total\s+financing\s+activ)", re.I), "Cash Flow from Financing"),
    ]

    out: dict[tuple[str, date, str], tuple[date, Decimal]] = {}
    for vf in filings:
        ftype = vf.mapped.raw.filing_type
        fdate = vf.mapped.raw.filing_date
        for stmt in vf.mapped.raw.statements:
            if not _keep_statement(stmt, ftype, vf.mapped.raw.statements):
                continue
            sheet = stmt_to_sheet(stmt.statement_type, ftype)
            period_end = stmt.period.period_end_date
            if stmt.statement_type == StatementType.BALANCE_SHEET:
                patterns = BS_PATTERNS
            elif stmt.statement_type == StatementType.CASH_FLOW:
                patterns = CF_PATTERNS
            else:
                continue  # IS subtotals (cascade with sign rules) deferred — BS/CF cover the bug
            for li in stmt.line_items:
                if li.row_type != "subtotal":
                    continue
                lbl = (li.raw_filing_label or "").strip()
                matched = None
                for pat, c in patterns:
                    if pat.match(lbl):
                        matched = c
                        break
                # Canonical label fallback (e.g. PG's CF "TOTAL OPERATING ACTIVITIES"
                # carries canonical_label="Cash Flow from Operations" via library).
                if matched is None and li.canonical_label in {p[1] for p in patterns}:
                    matched = li.canonical_label
                if matched is None:
                    continue
                key = (sheet, period_end, matched)
                cur = out.get(key)
                if cur is None or fdate < cur[0]:
                    out[key] = (fdate, Decimal(li.value))
    return {k: v for k, (_, v) in out.items()}


def _compute_section_displacement(
    filings: list[ValidatedFiling],
    row_section: dict[str, dict[int, str]],
    row_map: dict[str, dict[str, int]],
    superseded_rule_ids: set[str],
) -> dict[tuple[str, date, str], Decimal]:
    """For each item where the filer's `item.section` differs from the row's
    canonical filing_section (a legitimate cross-section route via the
    canonical's `accepted_sections` opt-in), return a map of net inflow to
    each (sheet, period, section): positive = workbook section gained value
    from a filer-classified-elsewhere item; negative = workbook section lost
    value to an item whose canonical lives in another section.

    The tie-out validator adds these adjustments to filer-expected subtotals
    so workbook section sums tie when cross-section routing is structural
    (e.g. MNST `Prepaid income taxes` filer-classified `current_assets` but
    routed to `Deferred Tax Assets` canonical (NCA): subtract the value
    from the period's expected TCA, add it to expected TNCA)."""
    displacement: dict[tuple[str, date, str], Decimal] = {}
    for vf in filings:
        ftype = vf.mapped.raw.filing_type
        idx = 0
        all_stmts = vf.mapped.raw.statements
        for stmt in all_stmts:
            count = len(stmt.line_items)
            if not _keep_statement(stmt, ftype, all_stmts):
                idx += count
                continue
            group = vf.mapped.mapped_line_items[idx : idx + count]
            idx += count
            sheet = stmt_to_sheet(stmt.statement_type, ftype)
            if sheet not in V1_SHEETS:
                continue
            period_end = stmt.period.period_end_date
            for item in group:
                if item.model_sheet in ("_subtotal", "_memo"):
                    continue
                if item.ledger_rule_id is None or item.ledger_rule_id in superseded_rule_ids:
                    continue
                excel_row = row_map.get(sheet, {}).get(item.ledger_rule_id)
                if excel_row is None:
                    continue
                target_section = row_section.get(sheet, {}).get(excel_row)
                if target_section is None:
                    continue
                source_section = item.section.value if hasattr(item.section, "value") else str(item.section)
                if source_section == target_section:
                    continue
                value = Decimal(item.value)
                # Source section LOST this value (it's a workbook-row in target section)
                key_src = (sheet, period_end, source_section)
                displacement[key_src] = displacement.get(key_src, Decimal("0")) - value
                # Target section GAINED this value
                key_tgt = (sheet, period_end, target_section)
                displacement[key_tgt] = displacement.get(key_tgt, Decimal("0")) + value
    return displacement


def validate_workbook_ties(
    sheet_cells: dict[tuple[str, int, date], Decimal],
    row_section: dict[str, dict[int, str]],
    row_labels: dict[str, dict[int, str]],
    filer_subs: dict[tuple[str, date, str], Decimal],
    filer_mezz: dict[tuple[str, date], Decimal] | None = None,
    section_displacement: dict[tuple[str, date, str], Decimal] | None = None,
    child_to_parent_row: dict[tuple[str, int], int] | None = None,
    tolerance: Decimal = Decimal("5"),
) -> list[str]:
    """Compare the workbook's section-sum totals to the filer's reported
    subtotals for the same period. Catches structural errors model-write can
    introduce (e.g. a line item routed to the wrong canonical → wrong section
    → wrong side of the BS) that the data-side BS-1..5 / IS-1..5 / CF-1..4
    validators cannot see — those rules verified the FILING was internally
    consistent before the workbook was built."""
    SECTION_DIRECT = {
        # (sheet, item_section) → canonical subtotal label
        ("BALANCE SHEET", "current_assets"):           "Total Current Assets",
        ("BALANCE SHEET", "current_liabilities"):      "Total Current Liabilities",
        ("BALANCE SHEET", "equity"):                   "Total Stockholders' Equity",
        ("QTR BS",        "current_assets"):           "Total Current Assets",
        ("QTR BS",        "current_liabilities"):      "Total Current Liabilities",
        ("QTR BS",        "equity"):                   "Total Stockholders' Equity",
        ("CASH FLOW",     "operating"):                "Cash Flow from Operations",
        ("CASH FLOW",     "investing"):                "Cash Flow from Investing",
        ("CASH FLOW",     "financing"):                "Cash Flow from Financing",
        ("QTR CF",        "operating"):                "Cash Flow from Operations",
        ("QTR CF",        "investing"):                "Cash Flow from Investing",
        ("QTR CF",        "financing"):                "Cash Flow from Financing",
    }

    # Labels of section subtotal rows that sit INSIDE their section bucket
    # (per row_section). Their values are the section's filer-reported
    # subtotal — including them in the section sum would double-count.
    SUBTOTAL_LABELS_TO_SKIP = {
        "Cash Flow from Operations",
        "Cash Flow from Investing",
        "Cash Flow from Financing",
        "Net Change in Cash",
    }

    child_to_parent_row = child_to_parent_row or {}

    section_sums: dict[tuple[str, str, date], Decimal] = {}
    for (sheet, row, period), value in sheet_cells.items():
        section = row_section.get(sheet, {}).get(row)
        if section is None:
            continue
        label = row_labels.get(sheet, {}).get(row, "")
        if label in SUBTOTAL_LABELS_TO_SKIP:
            continue
        # Detail-of-parent rows (e.g. RM/WIP/FG → Inventories Net): if the
        # parent has data for this same (sheet, period), skip the child here
        # — the parent value is the section's authoritative contributor and
        # children would double-count. When the parent has no data for this
        # period (filing didn't disclose Net even with detail), include the
        # child.
        parent_row = child_to_parent_row.get((sheet, row))
        if parent_row is not None:
            parent_val = sheet_cells.get((sheet, parent_row, period))
            if parent_val is not None:
                continue
        # Mezzanine items are folded into equity at workbook layout (per
        # feedback_convertible_preferred_to_equity.md). Mirror that here so
        # the equity section sum lines up with filer's TSE.
        if section == "mezzanine":
            section = "equity"
        key = (sheet, section, period)
        section_sums[key] = section_sums.get(key, Decimal("0")) + value

    errors: list[str] = []

    filer_mezz = filer_mezz or {}
    section_displacement = section_displacement or {}

    # Direct section subtotal checks
    for (sheet, section), canon in SECTION_DIRECT.items():
        for (s, sec, period), wb_val in section_sums.items():
            if (s, sec) != (sheet, section):
                continue
            filer_val = filer_subs.get((sheet, period, canon))
            if filer_val is None:
                continue
            # The workbook intentionally folds mezzanine into equity. To tie
            # workbook TSE against the filer (whose TSE excludes mezzanine),
            # add the period's mezzanine balance back into the expected.
            adjusted = filer_val
            if sheet in ("BALANCE SHEET", "QTR BS") and section == "equity":
                adjusted = filer_val + filer_mezz.get((sheet, period), Decimal("0"))
            # Cross-section displacement: items routed via `accepted_sections`
            # opt-in shift value across workbook section buckets without
            # changing TA/TL/TLSE totals. Adjust expected per-section sum by
            # the net inflow to this section for this period.
            adjusted = adjusted + section_displacement.get((sheet, period, section), Decimal("0"))
            if abs(wb_val - adjusted) > tolerance:
                errors.append(
                    f"[{sheet}] {canon} @ {period}: workbook=${wb_val:,.0f} "
                    f"vs filer=${adjusted:,.0f} (gap ${wb_val - adjusted:+,.0f})"
                )

    # BS cumulative subtotals (Total Assets, Total Liabilities)
    for sheet in ("BALANCE SHEET", "QTR BS"):
        periods = {p for (s, _sec, p) in section_sums if s == sheet}
        for period in periods:
            ca = section_sums.get((sheet, "current_assets", period), Decimal("0"))
            nca = section_sums.get((sheet, "non_current_assets", period), Decimal("0"))
            filer_ta = filer_subs.get((sheet, period, "Total Assets"))
            if filer_ta is not None and abs((ca + nca) - filer_ta) > tolerance:
                errors.append(
                    f"[{sheet}] Total Assets @ {period}: workbook=${ca + nca:,.0f} "
                    f"vs filer=${filer_ta:,.0f} (gap ${(ca + nca) - filer_ta:+,.0f})"
                )
            cl = section_sums.get((sheet, "current_liabilities", period), Decimal("0"))
            ncl = section_sums.get((sheet, "non_current_liabilities", period), Decimal("0"))
            filer_tl = filer_subs.get((sheet, period, "Total Liabilities"))
            if filer_tl is not None and abs((cl + ncl) - filer_tl) > tolerance:
                errors.append(
                    f"[{sheet}] Total Liabilities @ {period}: workbook=${cl + ncl:,.0f} "
                    f"vs filer=${filer_tl:,.0f} (gap ${(cl + ncl) - filer_tl:+,.0f})"
                )

    return errors


def build_workbook(
    ledger: dict,
    generic: dict,
    filings: list[ValidatedFiling],
    out_path: Path,
) -> dict:
    superseded, memo, _canonical, _section_hint, sign_by_rule = _collect_filing_metadata(ledger, generic)
    # Skip both superseded and memo items when collecting per-cell writes.
    skip_rule_ids = superseded | memo
    parent_to_children, child_to_parent = _collect_parent_children(generic)

    row_map, row_labels, row_section = resolve_row_positions(ledger, generic, filings)
    col_layout = build_column_layout(filings)            # {sheet: [(label, date|None)]}
    raw_writes = collect_writes(filings, skip_rule_ids, sign_by_rule)

    # Snapshot pre-mutation row_map + row_section. The per-sheet loop calls
    # `insert_bs_subtotal_slots` which mutates row_map for BS sheets (shifting
    # line-item excel_rows to make room for subtotal slots) without updating
    # row_section. The end-of-build tie-out validator needs the consistent
    # pair, so we hold onto it here.
    pre_mutation_row_map = {s: dict(m) for s, m in row_map.items()}
    pre_mutation_row_section = {s: dict(m) for s, m in row_section.items()}

    wb = Workbook()
    wb.remove(wb.active)  # drop default Sheet

    cells_written = 0
    summary: dict[str, dict] = {}

    for sheet_name in V1_SHEETS:
        # Skip sheets with no content — relevant when no 10-Q input is provided
        # (ANNL-only run shouldn't emit empty QTR P&L / QTR BS / QTR CF tabs)
        # or vice versa for a 10-Q-only run.
        sheet_has_rows = bool(row_map.get(sheet_name, {}))
        sheet_has_cols = any(d is not None for _, d in col_layout.get(sheet_name, []))
        if not sheet_has_rows and not sheet_has_cols:
            continue
        ws = wb.create_sheet(sheet_name)
        cols = col_layout.get(sheet_name, [])

        # BS gets a rebuilt layout with subtotal rows inserted at section boundaries.
        bs_subtotals: list[dict] = []
        if sheet_name in ("BALANCE SHEET", "QTR BS"):
            new_rules, new_labels, bs_subtotals = insert_bs_subtotal_slots(
                row_map.get(sheet_name, {}),
                row_labels.get(sheet_name, {}),
                row_section.get(sheet_name, {}),
            )
            row_map[sheet_name] = new_rules
            row_labels[sheet_name] = new_labels

        rows = row_map.get(sheet_name, {})

        # --- Header row (row 1) ---
        ws.cell(row=1, column=1, value="").fill = HEADER_FILL
        ws.cell(row=1, column=1).font = HEADER_FONT
        date_to_col: dict[date, int] = {}
        forecast_cols: list[int] = []
        for col_idx, (label, d) in enumerate(cols, start=2):
            c = ws.cell(row=1, column=col_idx, value=label)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = HEADER_ALIGN
            if d is not None:
                date_to_col[d] = col_idx
            else:
                forecast_cols.append(col_idx)

        # --- Column A row labels ---
        labels_for_sheet = row_labels.get(sheet_name, {})
        for excel_row, label in labels_for_sheet.items():
            lc = ws.cell(row=excel_row, column=1, value=label)
            lc.font = LABEL_FONT

        # --- Value cells ---
        # Aggregate: within each filing, sibling rule_ids on the same row SUM;
        # across filings, for each (sheet, period), the oldest filing's entire
        # row set wins (per-period dedup — see aggregate_cell_totals docstring).
        sheet_cells = aggregate_cell_totals(raw_writes, row_map)
        for (s, excel_row, d), total in sheet_cells.items():
            if s != sheet_name:
                continue
            col_idx = date_to_col.get(d)
            if col_idx is None:
                continue
            c = ws.cell(row=excel_row, column=col_idx, value=float(total))
            c.number_format = LINE_ITEM_FMT
            cells_written += 1

        last_row = 1 + len(labels_for_sheet)

        # --- Zero-fill empty HISTORICAL cells (forecast stays blank) ---
        for r in range(2, last_row + 1):
            for col_idx in date_to_col.values():
                cell = ws.cell(row=r, column=col_idx)
                if cell.value is None:
                    cell.value = 0
                    cell.number_format = LINE_ITEM_FMT

        # --- Parent/child SUM rendering (e.g. Inventories Net = SUM of RM/WIP/FG) ---
        # When a parent canonical and ANY of its children are both placed on
        # this sheet, overwrite the parent's HISTORICAL cells with =SUM(child
        # rows) and apply subtotal styling (bold + top border + subtotal fmt).
        # Per-period gating: only flip the parent for periods where at least
        # one child has a non-zero value, so filings that don't disclose the
        # detail keep the filer-reported parent value.
        rid_to_row = row_map.get(sheet_name, {})
        flipped_any: set[int] = set()
        for parent_rid, child_rids in parent_to_children.items():
            parent_row = rid_to_row.get(parent_rid)
            if parent_row is None:
                continue
            child_rows = [rid_to_row[c] for c in child_rids if c in rid_to_row]
            if not child_rows:
                continue
            # Per-period decision based on data-side sheet_cells (filer values).
            for d, col_idx in date_to_col.items():
                has_child_data = any(
                    sheet_cells.get((sheet_name, cr, d), Decimal("0")) != 0
                    for cr in child_rows
                )
                if not has_child_data:
                    continue
                col_letter = get_column_letter(col_idx)
                cell_refs = ",".join(f"{col_letter}{cr}" for cr in sorted(child_rows))
                cell = ws.cell(
                    row=parent_row, column=col_idx,
                    value=f"=SUM({cell_refs})",
                )
                cell.number_format = SUBTOTAL_FMT
                cell.font = Font(bold=True)
                cell.border = SUBTOTAL_BORDER
                flipped_any.add(parent_row)
            # Bold the parent's column-A label if it flipped for any period.
            if parent_row in flipped_any:
                ws.cell(row=parent_row, column=1).font = Font(bold=True)

        # --- CF subtotal SUM formulas (operating/investing/financing) ---
        # Walk rows top-to-bottom; each subtotal sums the rows since the
        # previous subtotal (or from row 2 for the first one). Stops after
        # the last CF subtotal, so below-the-line rows (FX, Net Change,
        # Cash at Beg/End) don't get swept in.
        if sheet_name in ("CASH FLOW", "QTR CF"):
            subtotal_labels = set(CF_SUBTOTAL_LABELS.keys())
            all_cols = list(date_to_col.values()) + forecast_cols
            section_start = 2
            section_subtotal_rows: list[int] = []
            for xr in sorted(labels_for_sheet.keys()):
                if labels_for_sheet[xr] not in subtotal_labels:
                    continue
                if xr - 1 >= section_start:
                    for col_idx in all_cols:
                        col_letter = get_column_letter(col_idx)
                        cell = ws.cell(
                            row=xr, column=col_idx,
                            value=f"=SUM({col_letter}{section_start}:{col_letter}{xr - 1})",
                        )
                        cell.number_format = SUBTOTAL_FMT
                        cell.font = Font(bold=True)
                        cell.border = SUBTOTAL_BORDER
                    ws.cell(row=xr, column=1).font = Font(bold=True)
                section_subtotal_rows.append(xr)
                section_start = xr + 1

            # Cross-filing CFO/CFI/CFF subtotal tie-out check. For each period,
            # the workbook's SUM(operating-rows) cell value comes from rows
            # the dedup picked across (potentially) different filings. If a
            # filer re-categorized line items between filings, the per-row
            # picks may not sum back to the filer-reported subtotal — the
            # subtotal cell is itself a row picked by dedup with a known
            # value. CF-2/CF-3/CF-4 in validate.py catch internal-to-filing
            # inconsistencies; THIS check catches the post-dedup workbook-
            # level mismatch that arises when filers' line breakdowns differ
            # across filings (PG's H1 FY2025 'Other Noncash' was 135 in the
            # 2025-Q2 10-Q but redistributed to a different breakdown in the
            # 2026-Q2 10-Q comparative).
            cross_warnings, cross_errors = check_cf_section_subtotals(
                sheet_cells, sheet_name, labels_for_sheet, row_section.get(sheet_name, {}),
                section_subtotal_rows, date_to_col,
            )
            # Warnings: gaps in (CF_TIE_OUT_WARN_MIN, CF_TIE_OUT_ERROR_MIN]. Print
            # but don't block — filer comparatives routinely redistribute a few $M
            # between Other-* rows between filings without changing the section
            # subtotal; that's signal worth surfacing but not worth blocking on.
            if cross_warnings:
                print(
                    f"\n[!] {sheet_name} cross-filing CF section/subtotal "
                    f"warnings (workbook formula sum won't match filer-"
                    f"reported subtotal after dedup):"
                )
                for w in cross_warnings:
                    print(w)
            # Errors: gaps > CF_TIE_OUT_ERROR_MIN. Raise — these indicate the
            # workbook is materially wrong (per-row dedup picked rows from
            # different filings whose concept-level alignment shifted, e.g. a
            # filer using "PaymentsForProceedsFromBusinessesAndOtherInvesting"
            # in one filing and the pure "PaymentsToAcquireBusinesses" concept
            # in the comparative — the same economic item gets counted twice
            # under two different rows). Block the build until resolved.
            if cross_errors:
                raise RuntimeError(
                    f"[{sheet_name}] cross-filing CF section/subtotal tie-out "
                    f"failed — gap > ${CF_TIE_OUT_ERROR_MIN}M (workbook formula "
                    f"sum won't match filer-reported subtotal after dedup):\n"
                    + "\n".join(cross_errors)
                )

            # Net Change in Cash = sum of all 4 section subtotals (operating +
            # investing + financing + cash_other). cash_other typically has no
            # explicit subtotal row — just one or more line items (FX Effect on
            # Cash etc.) — so we sum the section's individual rows. Per user
            # directive (2026-04-26): the row must always be present even when
            # no filing emitted it explicitly, so we synthesize one if missing.
            sheet_sections = row_section.get(sheet_name, {})
            cash_other_rows = [xr for xr, sec in sheet_sections.items() if sec == "cash_other"]
            net_change_row = next(
                (xr for xr, lbl in labels_for_sheet.items() if lbl == "Net Change in Cash"),
                None,
            )
            if net_change_row is None and section_subtotal_rows:
                # Synthesize: append at the end of the CF data block (after the
                # last existing row).
                net_change_row = max(
                    list(labels_for_sheet.keys()) + section_subtotal_rows + cash_other_rows
                ) + 1
                labels_for_sheet[net_change_row] = "Net Change in Cash"
                ws.cell(row=net_change_row, column=1, value="Net Change in Cash").font = Font(bold=True)
            if net_change_row and section_subtotal_rows:
                # Exclude net_change_row itself from the SUM range — when the
                # walker tags Net Change in Cash with section=cash_other, the
                # row would otherwise reference itself (circular formula that
                # Excel can't resolve, displays blank).
                cash_other_rows_no_self = [r for r in cash_other_rows if r != net_change_row]
                for col_idx in all_cols:
                    col_letter = get_column_letter(col_idx)
                    parts = [f"{col_letter}{r}" for r in section_subtotal_rows + cash_other_rows_no_self]
                    cell = ws.cell(
                        row=net_change_row, column=col_idx,
                        value="=" + "+".join(parts),
                    )
                    cell.number_format = SUBTOTAL_FMT
                    cell.font = Font(bold=True)
                    cell.border = SUBTOTAL_BORDER
                ws.cell(row=net_change_row, column=1).font = Font(bold=True)

            # CF section/subtotal containment check. Each section-tagged row
            # (operating / investing / financing) must fall within the SUM
            # range of its corresponding section subtotal — otherwise the
            # subtotal under-counts or double-counts. Catches the class of bug
            # where a filer's iXBRL document order or a row-layout edge case
            # parks an operating item past the CFO subtotal, leaving it out
            # of the SUM(...) range. Underlying validate runs against the
            # MappedFiling data, not the workbook formulas, so this check is
            # the only line of defense against a wrong CFO/CFI/CFF in the xlsx.
            # Map subtotal canonical labels back to their section keys.
            # CF_SUBTOTAL_LABELS = {label: section}; we want the inverse.
            sub_xr_by_section: dict[str, int] = {}
            for xr in sorted(labels_for_sheet.keys()):
                lbl = labels_for_sheet[xr]
                sec = CF_SUBTOTAL_LABELS.get(lbl)
                if sec and sec not in sub_xr_by_section:
                    sub_xr_by_section[sec] = xr
            sub_rows_sorted = sorted(sub_xr_by_section.values())
            xr_to_section = {xr: sec for sec, xr in sub_xr_by_section.items()}
            subtotal_xrs = set(sub_xr_by_section.values())
            containment_errors: list[str] = []
            valid_sections = set(CF_SUBTOTAL_LABELS.values())
            for xr, sec in sheet_sections.items():
                if sec not in valid_sections:
                    continue
                # Skip the subtotal rows themselves — they ARE the section
                # boundary, not content rows to be contained within it.
                if xr in subtotal_xrs:
                    continue
                next_sub_xr = next((s for s in sub_rows_sorted if s > xr), None)
                if next_sub_xr is None:
                    containment_errors.append(
                        f"  row {xr} ({labels_for_sheet.get(xr)!r}, section={sec!r}) "
                        f"lands after every section subtotal — would not be summed"
                    )
                    continue
                expected = xr_to_section[next_sub_xr]
                if expected != sec:
                    containment_errors.append(
                        f"  row {xr} ({labels_for_sheet.get(xr)!r}) tagged section={sec!r} "
                        f"but its enclosing SUM range belongs to {expected!r} (subtotal at row {next_sub_xr})"
                    )
            if containment_errors:
                raise RuntimeError(
                    f"[{sheet_name}] CF section/subtotal containment check failed — "
                    f"row layout puts a section item outside its subtotal's SUM range:\n"
                    + "\n".join(containment_errors)
                )

        # --- IS subtotal formulas (Gross Profit, Op Profit, Pre-Tax, NI) ---
        if sheet_name in ("ANNL P&L", "QTR P&L"):
            all_cols = list(date_to_col.values()) + forecast_cols
            write_is_subtotals(ws, labels_for_sheet, all_cols)

            # EPS rows: force dollar+decimal format on any row whose canonical
            # label contains "per Share" (matches "Basic Earnings (Loss) per
            # Share", "Diluted...", "EPS - Basic", etc. while excluding the
            # share-count "Weighted Average Shares Outstanding" rows).
            for xr, label in labels_for_sheet.items():
                if not label:
                    continue
                lbl_lower = label.lower()
                if ("per share" in lbl_lower or "eps" in lbl_lower) and "shares" not in lbl_lower:
                    for col_idx in all_cols:
                        ws.cell(row=xr, column=col_idx).number_format = EPS_FMT

        # --- BS subtotal SUM / cascade / grand-total formulas ---
        if sheet_name in ("BALANCE SHEET", "QTR BS"):
            all_cols = list(date_to_col.values()) + forecast_cols
            for spec in bs_subtotals:
                sub_row = spec["row"]
                for col_idx in all_cols:
                    col = get_column_letter(col_idx)
                    if spec["type"] == "sum":
                        formula = f"=SUM({col}{spec['min']}:{col}{spec['max']})"
                    elif spec["type"] == "cascade":
                        prev = spec.get("cascade_row")
                        if prev:
                            formula = f"={col}{prev}+SUM({col}{spec['min']}:{col}{spec['max']})"
                        else:
                            formula = f"=SUM({col}{spec['min']}:{col}{spec['max']})"
                    elif spec["type"] == "grand":
                        tl, tse = spec.get("tl_row"), spec.get("tse_row")
                        mezz = spec.get("mezz_rows", [])
                        if not (tl and tse):
                            continue
                        parts = [f"{col}{tl}"] + [f"{col}{m}" for m in mezz] + [f"{col}{tse}"]
                        formula = "=" + "+".join(parts)
                    else:
                        continue
                    cell = ws.cell(row=sub_row, column=col_idx, value=formula)
                    cell.number_format = SUBTOTAL_FMT
                    cell.font = Font(bold=True)
                    cell.border = SUBTOTAL_BORDER
                ws.cell(row=sub_row, column=1).font = Font(bold=True)

        # --- Forecast column tint (visual only) ---
        for fc in forecast_cols:
            for r in range(2, last_row + 1):
                cell = ws.cell(row=r, column=fc)
                if cell.fill.fill_type is None:
                    cell.fill = FORECAST_FILL

        # --- Column widths ---
        ws.column_dimensions["A"].width = 42
        for col_idx, _ in enumerate(cols, start=2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 13

        # --- Freeze first row + first column ---
        ws.freeze_panes = "B2"

        summary[sheet_name] = {
            "rows": len(labels_for_sheet),
            "periods": len([c for c in cols if c[1] is not None]),
            "forecast_cols": len(forecast_cols),
        }

    # --- Workbook tie-out validator (BS / CF subtotals must equal filer's
    # reported value for each period). Catches structural errors that the
    # data-side BS-1..5 / IS-1..5 / CF-1..4 validators can't see — those run
    # against `item.section` from the walker; this runs against the actual
    # workbook layout, which can disagree (e.g. PG's "DEFERRED INCOME TAXES"
    # tagged section=non_current_liabilities by walker but routed to the
    # `Deferred Tax Assets` canonical, placed on the asset side at write
    # time). Raises before save so a wrong workbook never lands on disk.
    all_sheet_cells = aggregate_cell_totals(raw_writes, pre_mutation_row_map)
    filer_subs = _collect_filer_subtotals(filings)
    filer_mezz = _collect_filer_mezzanine_sums(filings)
    section_displacement = _compute_section_displacement(
        filings, pre_mutation_row_section, pre_mutation_row_map, skip_rule_ids,
    )
    # Build (sheet, child_row) → parent_row from child_to_parent rule_id map
    # using pre-mutation row positions (tie-out runs against pre-mutation
    # layout, like the rest of validate_workbook_ties).
    child_to_parent_row: dict[tuple[str, int], int] = {}
    for sheet, rid_to_row in pre_mutation_row_map.items():
        for child_rid, parent_rid in child_to_parent.items():
            child_row = rid_to_row.get(child_rid)
            parent_row = rid_to_row.get(parent_rid)
            if child_row is not None and parent_row is not None:
                child_to_parent_row[(sheet, child_row)] = parent_row
    tie_errors = validate_workbook_ties(
        all_sheet_cells, pre_mutation_row_section, row_labels, filer_subs, filer_mezz,
        section_displacement=section_displacement,
        child_to_parent_row=child_to_parent_row,
    )
    if tie_errors:
        raise RuntimeError(
            "Workbook tie-out failed — section sums do not match filer-reported "
            "subtotals. Fix the underlying canonical/section mapping before "
            "writing the xlsx:\n  " + "\n  ".join(tie_errors)
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    # Auto-chain to model-qtr-derive: rename QTR -> YTD and append the 3
    # single-quarter sheets (QTR P&L / QTR BS / QTR CF) that decompose YTD
    # via cell-reference formulas. Sibling skill, dynamic import so model-write
    # still runs if the skill isn't installed (warns instead of failing).
    qtr_summary = _auto_derive_quarterly(out_path)

    return {
        "out_path": str(out_path),
        "cells_written": cells_written,
        "historical_labels": [lbl for lbl, d in col_layout[V1_SHEETS[0]] if d is not None],
        "forecast_labels": [lbl for lbl, d in col_layout[V1_SHEETS[0]] if d is None],
        "per_sheet": summary,
        "qtr_derive": qtr_summary,
    }


def _auto_derive_quarterly(out_path: Path) -> dict | None:
    """Dynamically load model-qtr-derive's `derive_quarterly` and run it on the
    just-saved workbook. Tolerant of the skill being absent.

    Resolves model-qtr-derive's location relative to this file, so it works
    both at user-level (~/.claude/skills/) and project-level
    (<project>/.claude/skills/) without hardcoding either path."""
    try:
        import importlib.util
        # this file: .../<skills_root>/model-write/scripts/write.py
        # sibling:   .../<skills_root>/model-qtr-derive/scripts/build.py
        skills_root = Path(__file__).resolve().parent.parent.parent
        skill_path = skills_root / "model-qtr-derive" / "scripts" / "build.py"
        if not skill_path.exists():
            # Fallback to user-level for the case where someone's running
            # model-write from a project that doesn't co-locate model-qtr-derive.
            skill_path = Path.home() / ".claude" / "skills" / "model-qtr-derive" / "scripts" / "build.py"
        if not skill_path.exists():
            print(f"[!] model-qtr-derive not found (checked sibling + ~/.claude/skills) — skipping single-quarter sheet derivation")
            return None
        spec = importlib.util.spec_from_file_location("model_qtr_derive_build", skill_path)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        report = module.derive_quarterly(out_path)
        for s in report.get("sheets", []):
            print(f"  {s['sheet']:8s}: {s['rows']} rows x {s['columns']} quarters (single-quarter)")
        return report
    except Exception as e:
        print(f"[!] model-qtr-derive auto-chain failed: {e}")
        return None


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Build a fresh xlsx financial model from ValidatedFiling JSONs.",
    )
    parser.add_argument("--ticker-root", required=True, type=Path)
    parser.add_argument("--in", dest="in_paths", action="append", required=True,
                        type=Path, help="ValidatedFiling JSON (repeatable)")
    parser.add_argument("--out", required=True, type=Path,
                        help="Output .xlsx path")
    parser.add_argument("--generic-library", type=Path, default=None,
                        help="Path to cross-ticker generic_line_item_mappings.json. "
                             "If omitted, auto-resolves to <ticker-root>/../../pattern_libraries/generic_line_item_mappings.json")
    args = parser.parse_args()

    # CLI ticker guard
    config = load_config(args.ticker_root)
    expected_ticker = config["ticker"]

    filings: list[ValidatedFiling] = []
    for p in args.in_paths:
        vf = load_validated(p)
        if vf.mapped.raw.ticker != expected_ticker:
            print(f"ERROR: {p} ticker={vf.mapped.raw.ticker!r} "
                  f"does not match config ticker={expected_ticker!r}",
                  file=sys.stderr)
            sys.exit(2)
        filings.append(vf)

    ledger = load_ledger(args.ticker_root)
    generic = load_generic_library(args.ticker_root, args.generic_library)

    report = build_workbook(ledger, generic, filings, args.out)

    # Stdout report
    print(f"Built {args.out}")
    for sheet, s in report["per_sheet"].items():
        print(f"  {sheet:14s}: {s['rows']} rows x {s['periods']} periods "
              f"(+ {s['forecast_cols']} forecast)")
    print(f"  Historical columns: {report['historical_labels']}")
    print(f"  Forecast columns:   {report['forecast_labels']}")
    print(f"  Cells written:      {report['cells_written']}")


if __name__ == "__main__":
    main()
